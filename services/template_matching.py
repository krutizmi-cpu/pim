from __future__ import annotations

from io import BytesIO
from pathlib import Path
import posixpath
import re
import zipfile
import json

import pandas as pd
from openpyxl import load_workbook

from services.attribute_service import list_attribute_definitions, list_channel_mapping_rules, set_product_attribute_value
from services.source_priority import can_overwrite_field
from services.source_tracking import save_field_source
from services.transforms import apply_transform, infer_transform_rule


BASE_PRODUCT_FIELD_ALIASES = {
    "supplier_article": ["идентификатор номенклатуры поставщика", "внутренний код поставщика", "артикул поставщика", "supplier article", "vendor code", "vendor_code"],
    "article": ["артикул", "sku", "код товара", "article"],
    "internal_article": ["id sap", "id_sap", "внутренний код", "internal article"],
    "name": ["название", "наименование", "товар", "name", "title", "name on site", "name_on_site"],
    "barcode": ["штрихкод", "ean", "barcode", "bar code", "bar_code"],
    "brand": ["бренд", "brand", "марка"],
    "description": ["описание", "description"],
    "uom": ["единица измерения", "ед. изм", "uom"],
    "weight": ["вес", "weight"],
    "length": ["длина", "length"],
    "width": ["ширина", "width"],
    "height": ["высота", "height"],
    "package_length": ["длина упаковки", "упаковка длина", "глубина упаковки"],
    "package_width": ["ширина упаковки", "упаковка ширина", "packing width", "packing_width"],
    "package_height": ["высота упаковки", "упаковка высота", "packing height", "packing_height"],
    "gross_weight": ["вес брутто", "вес в упаковке", "gross weight", "package weight", "package_weight"],
    "image_url": ["фото", "изображение", "image", "main image", "image links", "image_links"],
    "tnved_code": ["tnved", "тнвэд", "тн вэд", "код тнвэд"],
}


def _get_public_media_base_url(conn) -> str:
    try:
        row = conn.execute(
            "SELECT value FROM system_settings WHERE key = 'media.public_base_url' LIMIT 1"
        ).fetchone()
    except Exception:
        return ""
    if not row:
        return ""
    try:
        value = row["value"]
    except Exception:
        value = row[0] if len(row) > 0 else ""
    return str(value or "").strip().rstrip("/")


def _looks_like_local_media_path(value: str | None) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    low = text.lower()
    if text.startswith("\\\\") and re.search(r"\.(jpg|jpeg|png|webp|gif)$", low):
        return True
    if re.match(r"^[a-zA-Z]:\\", text) and re.search(r"\.(jpg|jpeg|png|webp|gif)$", low):
        return True
    return False


def _normalize_media_reference(value: object, public_base_url: str = "") -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.startswith("//"):
        return "https:" + text
    if text.lower().startswith(("http://", "https://")):
        return text
    if re.match(r"^[a-z0-9][a-z0-9\\.\\-]+\\.[a-z]{2,}.*$", text.lower()) and " " not in text:
        return f"https://{text}"
    if not public_base_url or not _looks_like_local_media_path(text):
        return text
    normalized = text.replace("/", "\\")
    if normalized.startswith("\\\\"):
        parts = [p for p in normalized.split("\\") if p]
        rel_parts = parts[1:] if len(parts) >= 2 else parts
    elif re.match(r"^[a-zA-Z]:\\", normalized):
        rel_parts = [p for p in normalized[3:].split("\\") if p]
    else:
        rel_parts = [p for p in normalized.split("\\") if p]
    if not rel_parts:
        return text
    rel_path = posixpath.join(*rel_parts)
    return f"{public_base_url}/{rel_path}"


def _parse_media_value(raw_value: object, public_base_url: str = "") -> list[str]:
    if raw_value in (None, ""):
        return []
    if isinstance(raw_value, (list, tuple, set)):
        values = list(raw_value)
    else:
        text = str(raw_value).strip()
        if not text:
            return []
        try:
            loaded = json.loads(text)
            if isinstance(loaded, list):
                values = loaded
            elif isinstance(loaded, str):
                values = re.split(r"[\n,;]+", loaded)
            else:
                values = re.split(r"[\n,;]+", text)
        except Exception:
            values = re.split(r"[\n,;]+", text)
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        normalized = _normalize_media_reference(value, public_base_url=public_base_url)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        out.append(normalized)
    return out

SPECIAL_TEMPLATE_MATCHERS = [
    # Online Trade standard template (long headers)
    (("индентификатор номенклатуры поставщика", "идентификатор номенклатуры поставщика"), "column", "supplier_article", "online_trade_standard", None),
    (("штрих-код товара", "штрихкод товара"), "column", "barcode", "online_trade_standard", None),
    (("артикул не менее",), "column", "article", "online_trade_standard", None),
    (("наименование (править", "наименование"), "column", "name", "online_trade_standard", None),
    (("производитель (название бренда)", "производитель"), "column", "brand", "online_trade_standard", None),
    (("вес брутто в кг",), "column", "gross_weight", "online_trade_standard", None),
    (("единица измерения",), "column", "uom", "online_trade_standard", None),
    (("длина/глубина упаковки",), "column", "package_length", "online_trade_standard", None),
    (("ширина упаковки",), "column", "package_width", "online_trade_standard", None),
    (("высота упаковки",), "column", "package_height", "online_trade_standard", None),
    (("описание полное текстовое",), "column", "description", "online_trade_standard", None),
    (("фото №1", "фото no1", "фото n1"), "column", "media_gallery", "online_trade_standard", "image_1"),
    (("фото №2", "фото no2", "фото n2"), "column", "media_gallery", "online_trade_standard", "image_2"),
    (("фото №3", "фото no3", "фото n3"), "column", "media_gallery", "online_trade_standard", "image_3"),
    (("фото №4", "фото no4", "фото n4"), "column", "media_gallery", "online_trade_standard", "image_4"),
    (("фото №5", "фото no5", "фото n5"), "column", "media_gallery", "online_trade_standard", "image_5"),
    # Detmir category template (bike sample)
    (("exact:id_sap", "exact:id sap"), "column", "internal_article", "detmir_standard", None),
    (("exact:title",), "column", "name", "detmir_standard", None),
    (("exact:bar_code", "exact:bar code"), "column", "barcode", "detmir_standard", None),
    (("exact:name_on_site", "exact:name on site"), "column", "name", "detmir_standard", None),
    (("exact:vendor_code", "exact:vendor code"), "column", "supplier_article", "detmir_standard", None),
    (("exact:packing_height",), "column", "package_height", "detmir_standard", None),
    (("exact:packing_width",), "column", "package_width", "detmir_standard", None),
    (("exact:package_length",), "column", "package_length", "detmir_standard", None),
    (("exact:package_weight",), "column", "gross_weight", "detmir_standard", None),
    (("exact:ves",), "column", "weight", "detmir_standard", None),
    (("exact:dlina",), "column", "length", "detmir_standard", None),
    (("exact:shirina",), "column", "width", "detmir_standard", None),
    (("exact:vysota",), "column", "height", "detmir_standard", None),
    (("exact:strana_proizvodstva",), "attribute", "country_of_origin", "detmir_standard", None),
    (("exact:pol",), "attribute", "gender", "detmir_standard", None),
    (("exact:material_igr_osn",), "attribute", "material", "detmir_standard", None),
    (("exact:tip_velosipedy",), "attribute", "child_bike_type", "detmir_standard", None),
    (("exact:tip_koles_kgt",), "attribute", "wheel_type", "detmir_standard", None),
    (("exact:tip_tormoza",), "attribute", "brake_type", "detmir_standard", None),
    (("exact:cvet_f",), "attribute", "color", "detmir_standard", None),
    (("exact:komplektaciya_velo",), "attribute", "equipment", "detmir_standard", None),
    (("exact:description",), "column", "description", "detmir_standard", None),
    (("exact:tnved",), "column", "tnved_code", "detmir_standard", None),
    (("exact:image_links", "exact:image links"), "column", "media_gallery", "detmir_standard", "join_images"),
    # Wildberries templates (category files)
    (("exact:артикул продавца",), "column", "article", "wb_standard", None),
    (("exact:артикул ozon",), "column", "article", "wb_standard", None),
    (("exact:наименование",), "column", "name", "wb_standard", None),
    (("exact:категория продавца",), "column", "category", "wb_standard", None),
    (("exact:бренд",), "column", "brand", "wb_standard", None),
    (("exact:описание",), "column", "description", "wb_standard", None),
    (("exact:фото",), "column", "media_gallery", "wb_standard", "join_images_semicolon"),
    (("exact:баркоды",), "column", "barcode", "wb_standard", None),
    (("exact:вес с упаковкой (кг)",), "column", "gross_weight", "wb_standard", None),
    (("exact:вес без упаковки (кг)",), "column", "weight", "wb_standard", None),
    (("exact:вес товара с упаковкой (г)",), "column", "gross_weight", "wb_standard", "kg_to_g"),
    (("exact:вес товара без упаковки (г)",), "column", "weight", "wb_standard", "kg_to_g"),
    (("exact:длина упаковки",), "column", "package_length", "wb_standard", None),
    (("exact:ширина упаковки",), "column", "package_width", "wb_standard", None),
    (("exact:высота упаковки",), "column", "package_height", "wb_standard", None),
    (("exact:длина предмета",), "column", "length", "wb_standard", None),
    (("exact:ширина предмета",), "column", "width", "wb_standard", None),
    (("exact:высота предмета",), "column", "height", "wb_standard", None),
    (("exact:глубина предмета",), "column", "length", "wb_standard", None),
    (("exact:тнвэд",), "column", "tnved_code", "wb_standard", None),
    # Sportmaster templates (shared layout)
    (("exact:артикул продавца",), "column", "article", "sportmaster_standard", None),
    (("exact:штрихкод",), "column", "barcode", "sportmaster_standard", None),
    (("exact:наименование",), "column", "name", "sportmaster_standard", None),
    (("exact:цена",), "attribute", "sportmaster_purchase_price", "sportmaster_standard", None),
    (("exact:цена со скидкой",), "attribute", "sportmaster_discount_price", "sportmaster_standard", None),
    (("exact:описание товара",), "column", "description", "sportmaster_standard", None),
    (("exact:тнвэд",), "column", "tnved_code", "sportmaster_standard", None),
    (("exact:страна производства",), "attribute", "country_of_origin", "sportmaster_standard", None),
    (("exact:высота в упаковке, см",), "column", "package_height", "sportmaster_standard", None),
    (("exact:длина в упаковке, см",), "column", "package_length", "sportmaster_standard", None),
    (("exact:ширина в упаковке, см",), "column", "package_width", "sportmaster_standard", None),
    (("exact:вес в упаковке, кг",), "column", "gross_weight", "sportmaster_standard", None),
    (("exact:ссылки на фото",), "column", "media_gallery", "sportmaster_standard", "join_images"),
]

PARTIAL_ALIAS_BLOCKLIST = {
    "name",
    "title",
    "товар",
    "товары",
    "product",
}


def normalize_key(value: str) -> str:
    text = str(value or "").strip().lower().replace("_", " ")
    text = text.replace("\n", " ")
    text = text.replace("*", "")
    text = re.sub(r"\.\d+$", "", text)
    return " ".join(text.split())


def detect_template_layout(template_bytes: bytes, sheet_name: str | None = None) -> dict[str, int | str]:
    safe_bytes = template_bytes
    try:
        workbook = load_workbook(BytesIO(safe_bytes), read_only=True, data_only=False)
    except Exception:
        safe_bytes = sanitize_template_xlsx_bytes(template_bytes)
        workbook = load_workbook(BytesIO(safe_bytes), read_only=True, data_only=False)

    try:
        chosen_sheet = sheet_name if sheet_name and sheet_name in workbook.sheetnames else workbook.sheetnames[0]
        ws = workbook[chosen_sheet]
        if "Коды характеристик" in workbook.sheetnames and chosen_sheet in {"Шаблон для заполнения", "Пример"}:
            row3 = [str(ws.cell(3, c).value or "").strip() for c in range(1, int(ws.max_column or 1) + 1)]
            row4 = [str(ws.cell(4, c).value or "").strip().lower() for c in range(1, int(ws.max_column or 1) + 1)]
            non_empty_row3 = [v for v in row3 if v]
            required_markers = sum(1 for v in row4 if "обязательный атрибут" in v)
            if len(non_empty_row3) >= 8 and required_markers >= 4:
                return {"header_row": 3, "data_start_row": 5, "template_kind": "sportmaster"}
    finally:
        workbook.close()

    try:
        xls = pd.ExcelFile(BytesIO(safe_bytes))
    except Exception:
        safe_bytes = sanitize_template_xlsx_bytes(template_bytes)
        xls = pd.ExcelFile(BytesIO(safe_bytes))
    chosen_sheet = sheet_name if sheet_name and sheet_name in xls.sheet_names else xls.sheet_names[0]
    probe = pd.read_excel(BytesIO(safe_bytes), sheet_name=chosen_sheet, header=None, nrows=120)

    header_tokens = {
        "артикул", "article", "sku", "name", "название", "наименование",
        "barcode", "штрихкод", "бренд", "brand", "описание", "description",
        "артикул продавца", "категория продавца", "артикул wb",
    }
    instruction_markers = (
        "введите",
        "выберите",
        "формат",
        "минимальное кол-во",
        "максимальное кол-во",
        "можно выбрать",
        "присваивается автоматически",
        "выпадающий список",
        "это номер или название",
        "список ссылок",
        "поставьте значение",
        "обязательный атрибут",
    )

    candidate_header_row = None
    best_score = -1
    for i in range(len(probe)):
        values = [str(v).strip().lower() for v in probe.iloc[i].tolist() if str(v).strip() and str(v).strip().lower() != "nan"]
        if not values:
            continue
        score = 0
        for v in values:
            if any(token in v for token in header_tokens):
                score += 3
            if len(v) <= 80:
                score += 1
        if score > best_score and len(values) >= 2:
            best_score = score
            candidate_header_row = i + 1

    if candidate_header_row is None:
        return {"header_row": 1, "data_start_row": 2, "template_kind": "generic"}

    data_start = int(candidate_header_row + 1)
    for r in range(data_start, min(data_start + 10, len(probe)) + 1):
        row_values = [str(v).strip().lower() for v in probe.iloc[r - 1].tolist() if str(v).strip() and str(v).strip().lower() != "nan"]
        if not row_values:
            continue
        if any(any(marker in val for marker in instruction_markers) for val in row_values):
            data_start = r + 1
            continue
        break
    return {"header_row": int(max(1, candidate_header_row)), "data_start_row": int(max(2, data_start)), "template_kind": "generic"}


def read_client_template_dataframe(template_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    safe_template_bytes = template_bytes
    try:
        pd.ExcelFile(BytesIO(safe_template_bytes))
    except Exception:
        safe_template_bytes = sanitize_template_xlsx_bytes(template_bytes)

    layout = detect_template_layout(safe_template_bytes, sheet_name=sheet_name)
    header_row = int(layout.get("header_row") or 1)
    chosen_sheet = sheet_name
    if not chosen_sheet:
        xls = pd.ExcelFile(BytesIO(safe_template_bytes))
        chosen_sheet = xls.sheet_names[0]
    return pd.read_excel(BytesIO(safe_template_bytes), sheet_name=chosen_sheet, header=max(0, header_row - 1))


def sanitize_template_xlsx_bytes(template_bytes: bytes) -> bytes:
    """
    Some client templates (notably some WB files) contain broken dataValidations
    that make openpyxl/pandas fail to parse. We remove those nodes in-memory.
    """
    try:
        src = BytesIO(template_bytes)
        with zipfile.ZipFile(src, "r") as zin:
            out = BytesIO()
            with zipfile.ZipFile(out, "w", compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                        try:
                            text = data.decode("utf-8")
                            text = re.sub(r"<dataValidations[^>]*>.*?</dataValidations>", "", text, flags=re.S)
                            data = text.encode("utf-8")
                        except Exception:
                            pass
                    zout.writestr(item, data)
            return out.getvalue()
    except Exception:
        return template_bytes


def build_candidate_map(conn) -> tuple[dict[str, tuple[str, str, str]], list[tuple[str, str, str, str]]]:
    result: dict[str, tuple[str, str, str]] = {}
    alias_rows: list[tuple[str, str, str, str]] = []

    for field_name, aliases in BASE_PRODUCT_FIELD_ALIASES.items():
        for alias in aliases:
            norm = normalize_key(alias)
            result[norm] = ("column", field_name, alias)
            alias_rows.append((norm, "column", field_name, alias))

    for attr in list_attribute_definitions(conn):
        result[normalize_key(attr["name"])] = ("attribute", attr["code"], attr["name"])
        result[normalize_key(attr["code"])] = ("attribute", attr["code"], attr["name"])

    return result, alias_rows


def auto_match_template_columns(conn, columns: list[str]) -> list[dict]:
    candidate_map, alias_rows = build_candidate_map(conn)
    matches: list[dict] = []

    for col in columns:
        norm = normalize_key(col)
        special = None
        for needles, source_type, source_name, matched_by, transform_rule in SPECIAL_TEMPLATE_MATCHERS:
            def _needle_match(needle_raw: str) -> bool:
                needle_raw = str(needle_raw or "")
                if needle_raw.startswith("exact:"):
                    return normalize_key(needle_raw.split(":", 1)[1]) == norm
                return normalize_key(needle_raw) in norm
            if any(_needle_match(n) for n in needles):
                special = (source_type, source_name, matched_by, transform_rule)
                break

        if special:
            source_type, source_name, matched_by, transform_rule = special
            resolved_transform = transform_rule or infer_transform_rule(col, source_type, source_name)
            matches.append(
                {
                    "template_column": col,
                    "status": "matched",
                    "source_type": source_type,
                    "source_name": source_name,
                    "matched_by": matched_by,
                    "transform_rule": resolved_transform,
                }
            )
            continue

        matched = candidate_map.get(norm)
        if not matched:
            # Partial alias match for verbose client headers.
            for alias_norm, source_type, source_name, matched_by in alias_rows:
                if alias_norm in PARTIAL_ALIAS_BLOCKLIST:
                    continue
                if len(alias_norm) >= 4 and alias_norm in norm:
                    matched = (source_type, source_name, f"alias_contains:{matched_by}")
                    break

        if matched:
            source_type, source_name, matched_by = matched
            resolved_transform = infer_transform_rule(col, source_type, source_name)
            matches.append(
                {
                    "template_column": col,
                    "status": "matched",
                    "source_type": source_type,
                    "source_name": source_name,
                    "matched_by": matched_by,
                    "transform_rule": resolved_transform,
                }
            )
            continue

        matches.append(
            {
                "template_column": col,
                "status": "unmatched",
                "source_type": None,
                "source_name": None,
                "matched_by": None,
                "transform_rule": None,
            }
        )

    return matches


def apply_saved_mapping_rules(conn, matches: list[dict], channel_code: str, category_code: str | None = None) -> list[dict]:
    rules = list_channel_mapping_rules(conn, channel_code=channel_code, category_code=category_code)
    if not rules:
        return matches

    rule_map = {r["target_field"]: r for r in rules}
    updated = []
    for match in matches:
        rule = rule_map.get(match["template_column"])
        if rule:
            saved_transform = rule.get("transform_rule")
            resolved_transform = saved_transform or infer_transform_rule(
                match["template_column"],
                rule.get("source_type"),
                rule.get("source_name"),
            )
            updated.append(
                {
                    "template_column": match["template_column"],
                    "status": "matched",
                    "source_type": rule["source_type"],
                    "source_name": rule["source_name"],
                    "matched_by": "saved_rule",
                    "transform_rule": resolved_transform,
                }
            )
        else:
            if "transform_rule" not in match:
                match["transform_rule"] = infer_transform_rule(
                    match.get("template_column"),
                    match.get("source_type"),
                    match.get("source_name"),
                )
            updated.append(match)
    return updated


def _resolve_transform_rule(match: dict) -> str | None:
    explicit = match.get("transform_rule")
    if explicit not in (None, ""):
        return str(explicit)
    return infer_transform_rule(
        match.get("template_column"),
        match.get("source_type"),
        match.get("source_name"),
    )


def build_product_value_map(conn, product_id: int) -> dict[str, object]:
    product_row = conn.execute("SELECT * FROM products WHERE id = ?", (int(product_id),)).fetchone()
    if not product_row:
        return {}
    # sqlite3.Row doesn't support .get(); normalize once to a plain dict.
    product = dict(product_row)
    public_media_base_url = _get_public_media_base_url(conn)

    value_map = dict(product)

    rows = conn.execute(
        """
        SELECT pav.attribute_code, pav.value_text, pav.value_number, pav.value_boolean, pav.value_json
        FROM product_attribute_values pav
        WHERE pav.product_id = ?
        """,
        (int(product_id),),
    ).fetchall()

    for row in rows:
        value = row["value_number"]
        if value is None:
            value = row["value_boolean"]
        if value is None:
            value = row["value_json"]
        if value is None:
            value = row["value_text"]
        value_map[row["attribute_code"]] = value

    media_rows = conn.execute(
        """
        SELECT value_text, value_json
        FROM product_attribute_values
        WHERE product_id = ?
          AND attribute_code IN ('gallery_images', 'main_image', 'generated_images')
        ORDER BY id
        """,
        (int(product_id),),
    ).fetchall()
    media_values = []
    for row in media_rows:
        if row["value_json"]:
            media_values.extend(_parse_media_value(row["value_json"], public_base_url=public_media_base_url))
        elif row["value_text"]:
            media_values.extend(_parse_media_value(row["value_text"], public_base_url=public_media_base_url))
    image_url = _normalize_media_reference(value_map.get("image_url"), public_base_url=public_media_base_url)
    if image_url:
        media_values.insert(0, image_url)
    deduped_media_values: list[str] = []
    seen_media: set[str] = set()
    for item in media_values:
        normalized = _normalize_media_reference(item, public_base_url=public_media_base_url)
        if not normalized or normalized in seen_media:
            continue
        seen_media.add(normalized)
        deduped_media_values.append(normalized)
        if len(deduped_media_values) >= 5:
            break
    if image_url:
        value_map["image_url"] = image_url
    value_map["media_gallery"] = deduped_media_values

    return value_map


def fill_template_dataframe(conn, template_df: pd.DataFrame, product_ids: list[int], matches: list[dict]) -> pd.DataFrame:
    rows: list[dict] = []
    for product_id in product_ids:
        value_map = build_product_value_map(conn, product_id)
        row_data = {}
        for match in matches:
            col = match["template_column"]
            status = str(match.get("status") or "").strip().lower()
            is_matched = status == "matched" or (
                not status
                and bool(str(match.get("template_column") or "").strip())
                and bool(str(match.get("source_type") or "").strip())
                and bool(str(match.get("source_name") or "").strip())
            )
            if not is_matched:
                row_data[col] = None
                continue
            transform_rule = _resolve_transform_rule(match)
            row_data[col] = apply_transform(value_map.get(match["source_name"]), transform_rule)
        rows.append(row_data)
    return pd.DataFrame(rows)


def apply_client_validated_values(conn, product_ids: list[int], matches: list[dict], channel_code: str | None = None) -> dict:
    applied = 0
    skipped = 0

    for product_id in product_ids:
        value_map = build_product_value_map(conn, product_id)
        for match in matches:
            if match.get("status") != "matched":
                continue
            field_name = match.get("source_name")
            source_type = match.get("source_type")
            if not field_name or not source_type:
                continue
            value = value_map.get(field_name)
            if value in (None, ""):
                continue

            if source_type == "column":
                if not can_overwrite_field(conn, product_id, field_name, "client_validated", force=False):
                    skipped += 1
                    continue
                conn.execute(
                    f"UPDATE products SET {field_name} = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
                    (value, int(product_id)),
                )
                save_field_source(
                    conn=conn,
                    product_id=product_id,
                    field_name=field_name,
                    source_type="client_validated",
                    source_value_raw=value,
                    source_url=channel_code,
                    confidence=0.95,
                    is_manual=False,
                )
                applied += 1
            elif source_type == "attribute":
                attr_field_name = f"attr:{field_name}"
                if not can_overwrite_field(conn, product_id, attr_field_name, "client_validated", force=False):
                    skipped += 1
                    continue
                set_product_attribute_value(conn, int(product_id), field_name, value, channel_code=channel_code)
                save_field_source(
                    conn=conn,
                    product_id=product_id,
                    field_name=attr_field_name,
                    source_type="client_validated",
                    source_value_raw=value,
                    source_url=channel_code,
                    confidence=0.95,
                    is_manual=False,
                )
                applied += 1

    conn.commit()
    return {"applied": applied, "skipped": skipped}


def dataframe_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "template") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return output.getvalue()


def detect_template_data_start_row(template_bytes: bytes, sheet_name: str | None = None) -> int:
    layout = detect_template_layout(template_bytes, sheet_name=sheet_name)
    return int(layout.get("data_start_row") or 2)


def fill_template_workbook_bytes(
    conn,
    template_bytes: bytes,
    product_ids: list[int],
    matches: list[dict],
    sheet_name: str | None = None,
    data_start_row: int | None = None,
) -> bytes:
    safe_template_bytes = template_bytes
    try:
        workbook = load_workbook(BytesIO(safe_template_bytes))
    except Exception:
        safe_template_bytes = sanitize_template_xlsx_bytes(template_bytes)
        try:
            workbook = load_workbook(BytesIO(safe_template_bytes))
        except Exception:
            # Fallback for malformed workbook structures (seen in some WB exports).
            chosen_sheet = sheet_name or "template"
            template_df = read_client_template_dataframe(safe_template_bytes, sheet_name=sheet_name)
            filled_df = fill_template_dataframe(conn, template_df, product_ids, matches)
            return dataframe_to_excel_bytes(filled_df, sheet_name=chosen_sheet)
    ws = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active

    layout = detect_template_layout(safe_template_bytes, sheet_name=ws.title)
    start_row = int(data_start_row or layout.get("data_start_row") or 2)
    if start_row < 2:
        start_row = 2
    header_row = int(layout.get("header_row") or max(1, start_row - 1))

    column_map: dict[str, int] = {}
    for col_idx in range(1, int(ws.max_column or 1) + 1):
        cell_value = ws.cell(row=header_row, column=col_idx).value
        if cell_value is None:
            continue
        key = str(cell_value).strip()
        if key and key not in column_map:
            column_map[key] = col_idx

    if not column_map:
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    # Clear previous values in the data area for mapped columns.
    for row_idx in range(start_row, int(ws.max_row or start_row) + 1):
        for match in matches:
            template_column = match.get("template_column")
            if not template_column or template_column not in column_map:
                continue
            ws.cell(row=row_idx, column=column_map[template_column]).value = None

    for row_offset, product_id in enumerate(product_ids):
        target_row = start_row + row_offset
        value_map = build_product_value_map(conn, int(product_id))
        for match in matches:
            status = str(match.get("status") or "").strip().lower()
            is_matched = status == "matched" or (
                not status
                and bool(str(match.get("template_column") or "").strip())
                and bool(str(match.get("source_type") or "").strip())
                and bool(str(match.get("source_name") or "").strip())
            )
            if not is_matched:
                continue
            template_column = match.get("template_column")
            source_name = match.get("source_name")
            if not template_column or template_column not in column_map or not source_name:
                continue
            transform_rule = _resolve_transform_rule(match)
            value = apply_transform(value_map.get(source_name), transform_rule)
            ws.cell(row=target_row, column=column_map[template_column]).value = value

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
