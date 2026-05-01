from __future__ import annotations

import sqlite3
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from services.client_registry import upsert_client_channel
from services.persistence_service import (
    get_uploaded_file_metadata,
    list_uploaded_files,
    persist_uploaded_file,
)
from services.template_matching import (
    apply_saved_mapping_rules,
    auto_match_template_columns,
    sanitize_template_xlsx_bytes,
)
from services.template_profiles import list_template_profiles, save_template_profile

SPORTMASTER_CHANNEL_CODE = "sportmaster"
SPORTMASTER_CLIENT_NAME = "Спортмастер"
SPORTMASTER_TEMPLATE_SHEET = "Шаблон для заполнения"
SPORTMASTER_SETTINGS_SHEET = "Настройки"
SPORTMASTER_CODES_SHEET = "Коды характеристик"
SPORTMASTER_REFERENCE_SHEET = "Справочные данные"

SPORTMASTER_GLOBAL_RULES: list[str] = [
    "Каждая категория Sportmaster живет в отдельном Excel-шаблоне из личного кабинета.",
    "Нельзя удалять листы, строки и столбцы шаблона, а также вставлять сторонние формулы или ссылки.",
    "После загрузки шаблон проходит первичную и вторичную модерацию; товары без критичных ошибок попадают в личный кабинет обычно в течение 2 часов.",
    "Атрибут `Цена` нужен для продажи товара; без него товар останется с критичной ошибкой и не будет продаваться.",
    "Sportmaster принимает штрихкоды EAN-13 и UPC; UPC нужно передавать в шаблоне как `0 + UPC`.",
    "На разных товарах должны быть уникальные штрихкоды.",
    "Фото через шаблон передаются только прямыми публичными ссылками `jpg/jpeg/png` в правильном порядке ракурсов.",
    "Первый ракурс фото обязателен; без фото №1 товар не может быть выведен на продажу.",
    "После появления товародвижения нельзя свободно менять штрихкод и весогабаритные характеристики через обычный flow.",
]

SPORTMASTER_SOURCE_LINKS: dict[str, str] = {
    "upload_flow": "https://seller-help.sportmaster.ru/pages/viewpage.action?pageId=5146264",
    "template_rules": "https://seller-help.sportmaster.ru/pages/viewpage.action?pageId=5146270",
    "photo_rules": "https://seller-help.sportmaster.ru/pages/viewpage.action?pageId=48365668",
    "movement_limits": "https://seller-help.sportmaster.ru/pages/viewpage.action?pageId=48365678",
}


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    return " ".join(text.replace("*", "").split())


def _clean_header_name(value: object) -> str:
    text = str(value or "").strip()
    return " ".join(text.replace("*", "").split())


def _to_attribute_code(name: str) -> str:
    clean = str(name or "").strip().lower()
    clean = "_".join("".join(ch if ch.isalnum() else " " for ch in clean).split())
    return clean[:120]


def _dedupe_headers_like_pandas(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    resolved: list[str] = []
    for raw in headers:
        base = str(raw or "").strip()
        if not base:
            resolved.append(base)
            continue
        count = seen.get(base, 0)
        resolved.append(base if count == 0 else f"{base}.{count}")
        seen[base] = count + 1
    return resolved


def _load_workbook_safe(template_bytes: bytes):
    payload = bytes(template_bytes or b"")
    try:
        return load_workbook(BytesIO(payload), read_only=True, data_only=False)
    except Exception:
        safe_payload = sanitize_template_xlsx_bytes(payload)
        return load_workbook(BytesIO(safe_payload), read_only=True, data_only=False)


def _extract_settings(workbook) -> dict[str, Any]:
    if SPORTMASTER_SETTINGS_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[SPORTMASTER_SETTINGS_SHEET]
    settings: dict[str, Any] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        key = str(row[0] or "").strip()
        if key:
            settings[key] = row[1]
    return settings


def _extract_characteristic_codes(workbook) -> dict[str, str]:
    if SPORTMASTER_CODES_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[SPORTMASTER_CODES_SHEET]
    code_map: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        internal_code = str(row[0] or "").strip()
        label = str(row[1] or "").strip()
        if internal_code and label:
            code_map[_normalize_text(label)] = internal_code
    return code_map


def _extract_reference_values(workbook) -> dict[str, list[str]]:
    if SPORTMASTER_REFERENCE_SHEET not in workbook.sheetnames:
        return {}
    ws = workbook[SPORTMASTER_REFERENCE_SHEET]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    headings = [str(value or "").strip() for value in rows[0]]
    refs: dict[str, list[str]] = {}
    for col_idx, heading in enumerate(headings):
        clean_heading = _clean_header_name(heading)
        if not clean_heading:
            continue
        values: list[str] = []
        seen: set[str] = set()
        for row in rows[1:]:
            raw = row[col_idx] if col_idx < len(row) else None
            text = str(raw or "").strip()
            if not text or text in seen:
                continue
            seen.add(text)
            values.append(text)
        refs[_normalize_text(clean_heading)] = values
    return refs


def _extract_special_setting(settings: dict[str, Any], marker: str) -> Any | None:
    marker_norm = _normalize_text(marker)
    for key, value in settings.items():
        if marker_norm in _normalize_text(key):
            return value
    return None


def extract_sportmaster_template_metadata(template_bytes: bytes) -> dict[str, Any]:
    workbook = _load_workbook_safe(template_bytes)
    try:
        settings = _extract_settings(workbook)
        characteristic_codes = _extract_characteristic_codes(workbook)
        reference_values = _extract_reference_values(workbook)

        tpl_sheet_name = SPORTMASTER_TEMPLATE_SHEET if SPORTMASTER_TEMPLATE_SHEET in workbook.sheetnames else workbook.sheetnames[0]
        ws = workbook[tpl_sheet_name]
        template_rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
        raw_headers: list[str] = []
        header_cells: list[tuple[int, str]] = []
        row1 = template_rows[0] if len(template_rows) >= 1 else ()
        row2 = template_rows[1] if len(template_rows) >= 2 else ()
        row3 = template_rows[2] if len(template_rows) >= 3 else ()
        row4 = template_rows[3] if len(template_rows) >= 4 else ()
        max_cols = max(len(row1), len(row2), len(row3), len(row4), int(ws.max_column or 0))
        for col_idx in range(max_cols):
            header_raw = row3[col_idx] if col_idx < len(row3) else None
            header = str(header_raw or "").strip()
            if not header:
                continue
            raw_headers.append(header)
            header_cells.append((col_idx, header))
        deduped_headers = _dedupe_headers_like_pandas(raw_headers)

        columns: list[dict[str, Any]] = []
        required_headers: list[str] = []
        for (col_idx, raw_header), header in zip(header_cells, deduped_headers):
            header_clean = _clean_header_name(raw_header)
            normalized_header = _normalize_text(header_clean)
            group_name = str((row1[col_idx] if col_idx < len(row1) else "") or "").strip() or None
            instruction = str((row2[col_idx] if col_idx < len(row2) else "") or "").strip() or None
            required_marker = str((row4[col_idx] if col_idx < len(row4) else "") or "").strip().lower()
            is_required = int("обязательный атрибут" in required_marker or "*" in raw_header)
            refs = reference_values.get(normalized_header, [])
            internal_code = characteristic_codes.get(normalized_header)
            item = {
                "column_index": int(col_idx + 1),
                "header": header,
                "header_raw": raw_header,
                "header_clean": header_clean,
                "normalized_header": normalized_header,
                "internal_code": internal_code,
                "group_name": group_name,
                "instruction": instruction,
                "required": int(is_required),
                "reference_values": refs,
                "reference_count": int(len(refs)),
            }
            columns.append(item)
            if is_required:
                required_headers.append(header)

        attr_class = str(settings.get("Атрибутный класс") or "").strip()
        attr_class_id = str(settings.get("Атрибутный класс id") or "").strip()
        attr_class_alias = str(settings.get("Атрибутный класс alias") or attr_class or "").strip()
        category_code = f"sportmaster:{attr_class_id}" if attr_class_id else f"sportmaster:{_to_attribute_code(attr_class_alias or attr_class or 'default')}"
        category_label = f"{attr_class or attr_class_alias or 'Sportmaster'} | class={attr_class_id or '-'}"

        return {
            "template_kind": "sportmaster",
            "sheet_name": tpl_sheet_name,
            "header_row": 3,
            "data_start_row": 5,
            "attr_class": attr_class or None,
            "attr_class_id": attr_class_id or None,
            "attr_class_alias": attr_class_alias or None,
            "category_code": category_code,
            "category_label": category_label,
            "critical_error_column_name": _extract_special_setting(settings, "критич"),
            "warning_column_name": _extract_special_setting(settings, "предупреж"),
            "global_rules": list(SPORTMASTER_GLOBAL_RULES),
            "source_links": dict(SPORTMASTER_SOURCE_LINKS),
            "columns": columns,
            "headers": [str(item["header"]) for item in columns],
            "required_headers": required_headers,
            "characteristic_codes": characteristic_codes,
            "reference_headers": sorted(reference_values.keys()),
        }
    finally:
        workbook.close()


def build_sportmaster_scope_labels(conn: sqlite3.Connection) -> dict[str, str]:
    labels: dict[str, str] = {}

    for row in list_uploaded_files(conn, storage_kind="client_template", channel_code=SPORTMASTER_CHANNEL_CODE, limit=500):
        category_code = str(row.get("category_code") or "").strip()
        if not category_code:
            continue
        metadata = get_uploaded_file_metadata(row)
        attr_class = str(metadata.get("attr_class") or metadata.get("attr_class_alias") or "").strip()
        attr_class_id = str(metadata.get("attr_class_id") or "").strip()
        if attr_class or attr_class_id:
            labels[category_code] = f"{attr_class or 'Sportmaster'} | class={attr_class_id or '-'}"
        else:
            labels.setdefault(category_code, category_code)

    for profile in list_template_profiles(conn, channel_code=SPORTMASTER_CHANNEL_CODE):
        category_code = str(profile.get("category_code") or "").strip()
        if not category_code:
            continue
        labels.setdefault(category_code, category_code)

    return labels


def _upsert_attribute_definition_fast(
    conn: sqlite3.Connection,
    *,
    code: str,
    name: str,
    description: str | None,
) -> bool:
    existing = conn.execute(
        "SELECT id FROM attribute_definitions WHERE code = ? LIMIT 1",
        (code,),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE attribute_definitions
            SET name = ?, description = COALESCE(NULLIF(?, ''), description), updated_at = CURRENT_TIMESTAMP
            WHERE code = ?
            """,
            (name, description, code),
        )
        return False
    conn.execute(
        """
        INSERT INTO attribute_definitions (
            code, name, data_type, scope, entity_type, is_required, is_multi_value, unit, description, created_at, updated_at
        )
        VALUES (?, ?, 'text', 'master', 'product', 0, 0, NULL, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (code, name, description),
    )
    return True


def _upsert_channel_requirement_fast(
    conn: sqlite3.Connection,
    *,
    channel_code: str,
    category_code: str,
    attribute_code: str,
    is_required: int,
    sort_order: int,
    notes: str | None,
) -> bool:
    existing = conn.execute(
        """
        SELECT id
        FROM channel_attribute_requirements
        WHERE channel_code = ?
          AND IFNULL(category_code, '') = IFNULL(?, '')
          AND attribute_code = ?
        LIMIT 1
        """,
        (channel_code, category_code, attribute_code),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE channel_attribute_requirements
            SET is_required = ?, sort_order = ?, notes = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (int(is_required), int(sort_order), notes, int(existing["id"])),
        )
        return False
    conn.execute(
        """
        INSERT INTO channel_attribute_requirements (
            channel_code, category_code, attribute_code, is_required, sort_order, notes, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (channel_code, category_code, attribute_code, int(is_required), int(sort_order), notes),
    )
    return True


def _upsert_mapping_rule_fast(
    conn: sqlite3.Connection,
    *,
    channel_code: str,
    category_code: str,
    target_field: str,
    source_type: str,
    source_name: str,
    transform_rule: str | None,
    is_required: int,
) -> bool:
    existing = conn.execute(
        """
        SELECT id
        FROM channel_mapping_rules
        WHERE channel_code = ?
          AND IFNULL(category_code, '') = IFNULL(?, '')
          AND target_field = ?
        LIMIT 1
        """,
        (channel_code, category_code, target_field),
    ).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE channel_mapping_rules
            SET source_type = ?, source_name = ?, transform_rule = ?, is_required = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            (source_type, source_name, transform_rule, int(is_required), int(existing["id"])),
        )
        return False
    conn.execute(
        """
        INSERT INTO channel_mapping_rules (
            channel_code, category_code, target_field, source_type, source_name, transform_rule, is_required, created_at, updated_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
        """,
        (channel_code, category_code, target_field, source_type, source_name, transform_rule, int(is_required)),
    )
    return True


def import_sportmaster_template(
    conn: sqlite3.Connection,
    template_bytes: bytes,
    original_file_name: str | None,
    *,
    profile_name: str | None = None,
) -> dict[str, Any]:
    metadata = extract_sportmaster_template_metadata(template_bytes)
    category_code = str(metadata.get("category_code") or "").strip()
    if not category_code:
        raise ValueError("Не удалось определить category_code шаблона Sportmaster")

    upsert_client_channel(
        conn,
        client_code=SPORTMASTER_CHANNEL_CODE,
        client_name=SPORTMASTER_CLIENT_NAME,
        notes="Категорийные Excel-шаблоны Sportmaster из личного кабинета",
    )

    upload_result = persist_uploaded_file(
        conn=conn,
        storage_kind="client_template",
        original_file_name=original_file_name,
        file_bytes=template_bytes,
        channel_code=SPORTMASTER_CHANNEL_CODE,
        category_code=category_code,
        metadata=metadata,
    )

    matches = auto_match_template_columns(conn, list(metadata.get("headers") or []))
    matches = apply_saved_mapping_rules(
        conn,
        matches,
        channel_code=SPORTMASTER_CHANNEL_CODE,
        category_code=category_code,
    )
    match_map = {str(item.get("template_column") or "").strip(): item for item in matches}

    created_attributes = 0
    created_requirements = 0
    created_rules = 0
    profile_columns: list[dict[str, Any]] = []

    for idx, column in enumerate(metadata.get("columns") or [], start=1):
        header = str(column.get("header") or "").strip()
        header_clean = str(column.get("header_clean") or header).strip()
        if not header:
            continue

        code = _to_attribute_code(header_clean or header)
        created_attr = _upsert_attribute_definition_fast(
            conn=conn,
            code=code,
            name=header_clean or header,
            description=f"Sportmaster template column: {header_clean or header}",
        )
        if created_attr:
            created_attributes += 1

        is_required = int(column.get("required") or 0)
        notes_parts = ["Импортировано из шаблона Sportmaster"]
        if column.get("group_name"):
            notes_parts.append(f"Группа: {column['group_name']}")
        if column.get("instruction"):
            notes_parts.append(str(column["instruction"]))
        if column.get("reference_count"):
            notes_parts.append(f"Справочных значений: {int(column['reference_count'])}")
        if column.get("internal_code"):
            notes_parts.append(f"Код характеристики: {column['internal_code']}")
        requirement_notes = " | ".join(part for part in notes_parts if str(part).strip())

        created_req = _upsert_channel_requirement_fast(
            conn=conn,
            channel_code=SPORTMASTER_CHANNEL_CODE,
            category_code=category_code,
            attribute_code=code,
            is_required=is_required,
            sort_order=1000 + idx,
            notes=requirement_notes,
        )
        if created_req:
            created_requirements += 1

        matched = match_map.get(header)
        if matched and str(matched.get("status") or "") == "matched" and str(matched.get("source_name") or "").strip():
            source_type = str(matched.get("source_type") or "attribute")
            source_name = str(matched.get("source_name") or code)
            matched_by = str(matched.get("matched_by") or "sportmaster_standard")
            transform_rule = matched.get("transform_rule")
        else:
            source_type = "attribute"
            source_name = code
            matched_by = "sportmaster_self"
            transform_rule = None

        created_rule = _upsert_mapping_rule_fast(
            conn=conn,
            channel_code=SPORTMASTER_CHANNEL_CODE,
            category_code=category_code,
            target_field=header,
            source_type=source_type,
            source_name=source_name,
            transform_rule=transform_rule,
            is_required=is_required,
        )
        if created_rule:
            created_rules += 1

        profile_columns.append(
            {
                "template_column": header,
                "source_type": source_type,
                "source_name": source_name,
                "matched_by": matched_by,
                "transform_rule": transform_rule,
            }
        )

    profile_suffix = str(metadata.get("attr_class_id") or "").strip() or _to_attribute_code(
        str(metadata.get("attr_class_alias") or metadata.get("attr_class") or "default")
    )
    profile_id = save_template_profile(
        conn=conn,
        profile_name=profile_name or f"sportmaster_{profile_suffix}",
        channel_code=SPORTMASTER_CHANNEL_CODE,
        category_code=category_code,
        file_name=original_file_name,
        columns=profile_columns,
    )

    return {
        "ok": True,
        "category_code": category_code,
        "category_label": metadata.get("category_label"),
        "attr_class": metadata.get("attr_class"),
        "attr_class_id": metadata.get("attr_class_id"),
        "uploaded_file_id": int(upload_result["id"]),
        "profile_id": int(profile_id),
        "columns_total": int(len(metadata.get("headers") or [])),
        "required_total": int(len(metadata.get("required_headers") or [])),
        "created_attributes": int(created_attributes),
        "created_requirements": int(created_requirements),
        "created_rules": int(created_rules),
    }
