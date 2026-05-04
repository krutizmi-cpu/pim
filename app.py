from __future__ import annotations

import json
import hashlib
from io import BytesIO
import math
import mimetypes
import ntpath
import os
import posixpath
import re
import zipfile
from pathlib import Path
import threading
from datetime import datetime, timezone
from urllib.parse import quote, urlparse

import pandas as pd
import sqlite3
import streamlit as st
import httpx
from openpyxl import load_workbook

from db import get_connection, init_db
try:
    from db import get_active_db_path as _get_active_db_path
except Exception:
    def _get_active_db_path():
        return None
from services.attribute_service import (
    get_product_attribute_values,
    list_attribute_definitions,
    list_channel_mapping_rules,
    list_channel_requirements,
    set_product_attribute_value,
    upsert_attribute_definition,
    upsert_channel_attribute_requirement,
    upsert_channel_mapping_rule,
)
from services.catalog_service import import_catalog_from_excel
from services.duplicate_service import refresh_duplicates_for_product
from services.source_tracking import get_field_sources, save_field_source, get_latest_field_source, field_is_manual
from services.source_priority import can_overwrite_field
try:
    from services import supplier_parser as _supplier_parser
except Exception:
    _supplier_parser = None


def parse_supplier_product_page(url: str, hints: list[str] | None = None, timeout: float = 8.0, max_hops: int = 1) -> dict:
    if _supplier_parser is not None and hasattr(_supplier_parser, "parse_supplier_product_page"):
        return _supplier_parser.parse_supplier_product_page(url, hints=hints, timeout=timeout, max_hops=max_hops)
    if _supplier_parser is not None and all(hasattr(_supplier_parser, x) for x in ("fetch_supplier_page", "extract_supplier_data", "normalize_supplier_data")):
        html = _supplier_parser.fetch_supplier_page(url, timeout=timeout)
        raw = _supplier_parser.extract_supplier_data(html, url)
        parsed = _supplier_parser.normalize_supplier_data(raw)
        parsed["resolved_url"] = url
        parsed["resolved_from_listing"] = False
        return parsed
    raise RuntimeError("Supplier parser module is unavailable")


def has_meaningful_supplier_data(parsed: dict) -> bool:
    if _supplier_parser is not None and hasattr(_supplier_parser, "has_meaningful_supplier_data"):
        return bool(_supplier_parser.has_meaningful_supplier_data(parsed))
    if not parsed:
        return False
    for key in ("description", "image_url", "weight", "length", "width", "height", "gross_weight"):
        if parsed.get(key) not in (None, "", 0, 0.0):
            return True
    attrs = parsed.get("attributes") or {}
    return bool(attrs)


def fallback_search_product_data(
    query: str,
    timeout: float = 8.0,
    max_results: int = 3,
    hints: list[str] | None = None,
    preferred_domain: str | None = None,
    blocked_source_domain: bool = False,
) -> dict:
    if _supplier_parser is not None and hasattr(_supplier_parser, "fallback_search_product_data"):
        return _supplier_parser.fallback_search_product_data(
            query,
            timeout=timeout,
            max_results=max_results,
            hints=hints,
            preferred_domain=preferred_domain,
            blocked_source_domain=blocked_source_domain,
        )
    return {}


def is_likely_blocked_supplier_domain(url: str | None) -> bool:
    if _supplier_parser is not None and hasattr(_supplier_parser, "is_likely_blocked_supplier_domain"):
        try:
            return bool(_supplier_parser.is_likely_blocked_supplier_domain(url))
        except Exception:
            return False
    return False
from services.template_matching import auto_match_template_columns, apply_saved_mapping_rules, fill_template_dataframe, apply_client_validated_values, fill_template_workbook_bytes, dataframe_to_excel_bytes, detect_template_data_start_row, sanitize_template_xlsx_bytes, read_client_template_dataframe, build_product_value_map
from services.template_profiles import save_template_profile, list_template_profiles, get_template_profile_columns
from services.client_registry import list_client_channels, upsert_client_channel
from services.readiness_service import analyze_template_readiness
from services.supplier_profiles import list_supplier_profiles, upsert_supplier_profile, ensure_default_supplier_profiles
from services.sportmaster_template_service import (
    SPORTMASTER_CHANNEL_CODE,
    SPORTMASTER_CLIENT_NAME,
    build_sportmaster_scope_labels,
    extract_sportmaster_template_metadata,
    import_sportmaster_template,
)
from services.persistence_service import (
    persist_uploaded_file,
    list_uploaded_files,
    get_uploaded_file_metadata,
    read_uploaded_file_bytes,
    record_catalog_import_history,
    list_catalog_import_history,
    save_ai_connection_profile,
    list_ai_connection_profiles,
    get_ai_connection_profile,
)
from services.backup_service import (
    backup_database_file,
    backup_ozon_snapshot_bytes,
    list_ozon_snapshot_backups,
    read_backup_bytes,
)
from services.certificate_registry import (
    search_fsa_registry_candidates,
    parse_fsa_document_resource,
    save_fsa_document,
    list_fsa_documents,
    delete_fsa_document,
)
from services.ozon_api_service import AUTO_MANUAL_ONLY_OZON_CODES, is_configured, sync_category_tree, list_cached_categories, list_cached_category_pairs, get_ozon_cache_stats, get_ozon_sync_coverage, sync_missing_category_attributes, sync_category_attributes, list_cached_attributes, sync_attribute_dictionary_values, sync_all_category_dictionary_values, list_cached_attribute_values, import_cached_attributes_to_pim, import_all_cached_attributes_to_pim, suggest_mappings_for_cached_attributes, save_suggested_mappings, analyze_product_ozon_coverage, ensure_ozon_master_attributes, build_product_ozon_payload, materialize_product_ozon_attributes, preview_product_ozon_dictionary_gaps, build_product_ozon_api_attributes, build_bulk_ozon_api_payloads, build_ozon_attributes_update_request, submit_ozon_attributes_update, list_ozon_update_jobs, get_ozon_update_job, retry_ozon_update_job, list_ozon_update_job_items, save_dictionary_override, list_dictionary_overrides, delete_dictionary_override, sync_all_categories_and_attributes, load_saved_credentials as load_saved_ozon_credentials, save_credentials as save_ozon_credentials, clear_saved_credentials as clear_saved_ozon_credentials, resolve_credentials as resolve_ozon_credentials, check_connection as check_ozon_connection
from services.ozon_category_match import bulk_assign_ozon_categories
from services.detmir_api_service import (
    load_detmir_settings,
    save_detmir_settings,
    is_configured as detmir_is_configured,
    check_connection as check_detmir_connection,
    sync_category_tree as sync_detmir_category_tree,
    sync_categories_with_attributes,
    sync_attribute_values as sync_detmir_attribute_values,
    sync_all_attribute_values as sync_all_detmir_attribute_values,
    sync_products as sync_detmir_products,
    list_cached_categories as list_detmir_cached_categories,
    list_cached_attributes as list_detmir_cached_attributes,
    list_cached_attribute_values as list_detmir_cached_attribute_values,
    list_cached_products as list_detmir_cached_products,
    get_detmir_cache_stats,
    import_category_requirements_to_pim as import_detmir_category_requirements_to_pim,
    get_cached_category as get_detmir_cached_category,
    suggest_categories_for_product as suggest_detmir_categories_for_product,
    detect_best_category_for_product as detect_best_detmir_category_for_product,
    analyze_product_detmir_readiness,
)
from services.wildberries_api_service import (
    load_settings as load_wb_settings,
    save_settings as save_wb_settings,
    clear_settings as clear_wb_settings,
    is_configured as wb_is_configured,
    check_connection as check_wb_connection,
    list_parent_categories as list_wb_parent_categories,
    search_subjects as search_wb_subjects,
    get_subject_characteristics as get_wb_subject_characteristics,
    build_card_draft as build_wb_card_draft,
    upload_product_cards as upload_wb_product_cards,
    list_failed_cards as list_wb_failed_cards,
)
from services.dimension_fallback import infer_category_fields, infer_dimensions_from_catalog, infer_dimensions_from_category_defaults, is_dimension_payload_suspicious
from services.ai_content_service import (
    PROVIDER_DEFAULTS,
    load_ai_settings,
    save_ai_settings,
    ai_is_configured,
    check_ai_connection,
    generate_selling_title_for_product,
    generate_seo_description_for_product,
    generate_product_copy_pack_for_product,
    generate_ai_attribute_suggestions_for_product,
    apply_ai_attribute_suggestions,
    run_ai_enrichment_for_product,
    build_marketing_image_prompts_for_product,
    build_image_gallery_plan_for_product,
    verify_parser_result_for_product,
    generate_images_from_prompts,
)

st.set_page_config(page_title="PIM", page_icon="📦", layout="wide")
OZON_OFFER_ID_OPTIONS = ["article", "internal_article", "supplier_article"]
TEMPLATE_TRANSFORM_OPTIONS = [
    "",
    "cm_to_mm",
    "mm_to_cm",
    "m_to_cm",
    "m_to_mm",
    "cm_to_m",
    "mm_to_m",
    "kg_to_g",
    "g_to_kg",
    "kg_to_lb",
    "lb_to_kg",
    "inch_to_cm",
    "cm_to_inch",
    "lower",
    "upper",
    "strip",
    "first_image",
    "join_images",
    "join_images_semicolon",
    "image_1",
    "image_2",
    "image_3",
    "image_4",
    "image_5",
]
OZON_CATEGORY_MIN_SCORE = 0.42
_OZON_SYNC_BG_LOCK = threading.Lock()
_OZON_SYNC_BG_THREAD: threading.Thread | None = None
_OZON_SYNC_BG_STATE: dict[str, object] = {
    "running": False,
    "started_at": None,
    "finished_at": None,
    "last_error": None,
    "result": None,
}


def get_db():
    conn = get_connection()
    init_db(conn)
    ensure_default_supplier_profiles(conn)
    return conn


def apply_app_theme() -> None:
    st.markdown(
        """
<style>
@import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=IBM+Plex+Mono:wght@400;500&display=swap');

:root {
  --pim-bg: #f5f5f7;
  --pim-bg-soft: #edf1f6;
  --pim-surface: rgba(255, 255, 255, 0.88);
  --pim-surface-strong: #ffffff;
  --pim-line: rgba(17, 24, 39, 0.08);
  --pim-text: #111111;
  --pim-sub: #6e7280;
  --pim-accent: #111111;
  --pim-accent-soft: #eef2ff;
  --pim-success: #0f766e;
  --pim-warn: #b45309;
  --pim-shadow: 0 22px 56px rgba(17, 24, 39, 0.07);
}

html, body, [class*="css"], [data-testid="stAppViewContainer"] {
  font-family: 'Manrope', sans-serif;
  color: var(--pim-text);
}

[data-testid="stAppViewContainer"] {
  background:
    radial-gradient(circle at top left, rgba(219, 234, 254, 0.7), transparent 28%),
    radial-gradient(circle at top right, rgba(229, 231, 235, 0.7), transparent 26%),
    linear-gradient(180deg, #fbfbfd 0%, #f5f5f7 52%, #eef1f5 100%);
}

[data-testid="stHeader"] {
  background: rgba(248, 250, 252, 0.78);
  backdrop-filter: blur(16px);
}

[data-testid="stSidebar"] {
  background: linear-gradient(180deg, rgba(255,255,255,0.96) 0%, rgba(244,246,248,0.98) 100%);
  border-right: 1px solid var(--pim-line);
}

[data-testid="stSidebar"] [data-testid="stMarkdownContainer"] p,
[data-testid="stSidebar"] label {
  color: var(--pim-text);
}

.pim-shell {
  padding-top: 0.2rem;
}

.pim-hero {
  background: linear-gradient(135deg, rgba(255,255,255,0.98) 0%, rgba(255,255,255,0.8) 100%);
  border: 1px solid rgba(17, 24, 39, 0.06);
  border-radius: 30px;
  padding: 1.45rem 1.55rem 1.2rem 1.55rem;
  box-shadow: var(--pim-shadow);
  margin-bottom: 1.1rem;
  backdrop-filter: blur(18px);
}

.pim-hero-title {
  font-size: 2.15rem;
  font-weight: 700;
  line-height: 1.05;
  letter-spacing: -0.03em;
  margin: 0 0 0.45rem 0;
}

.pim-hero-sub {
  color: var(--pim-sub);
  font-size: 0.98rem;
  margin-bottom: 0.9rem;
}

.pim-chip-row {
  display: flex;
  flex-wrap: wrap;
  gap: 0.55rem;
}

.pim-chip {
  display: inline-flex;
  align-items: center;
  gap: 0.35rem;
  padding: 0.45rem 0.72rem;
  border-radius: 999px;
  background: rgba(17, 24, 39, 0.035);
  border: 1px solid rgba(17, 24, 39, 0.05);
  color: var(--pim-text);
  font-size: 0.83rem;
}

.pim-chip strong {
  font-weight: 700;
}

.pim-side-card {
  background: rgba(255,255,255,0.84);
  border: 1px solid rgba(17, 24, 39, 0.06);
  border-radius: 22px;
  padding: 1rem 1rem 0.9rem 1rem;
  box-shadow: 0 16px 38px rgba(17, 24, 39, 0.04);
  margin-bottom: 0.9rem;
}

.pim-side-kicker {
  color: var(--pim-sub);
  font-size: 0.78rem;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  margin-bottom: 0.28rem;
}

.pim-side-title {
  font-size: 1.18rem;
  font-weight: 700;
  letter-spacing: -0.02em;
  margin-bottom: 0.35rem;
}

.pim-side-note {
  color: var(--pim-sub);
  font-size: 0.86rem;
}

div[data-testid="stMetric"] {
  background: rgba(255,255,255,0.82);
  border: 1px solid rgba(17, 24, 39, 0.06);
  border-radius: 20px;
  padding: 0.8rem 0.9rem;
  box-shadow: 0 12px 28px rgba(17, 24, 39, 0.04);
}

div[data-testid="stMetricLabel"] {
  color: var(--pim-sub);
}

div[data-testid="stMetricValue"] {
  font-weight: 700;
  letter-spacing: -0.03em;
}

div.stButton > button,
div.stDownloadButton > button {
  border-radius: 18px;
  border: 1px solid rgba(17, 24, 39, 0.08);
  min-height: 2.7rem;
  box-shadow: 0 10px 22px rgba(17, 24, 39, 0.04);
  transition: all 0.18s ease;
}

div.stButton > button:hover,
div.stDownloadButton > button:hover {
  border-color: rgba(17, 24, 39, 0.16);
  transform: translateY(-1px);
}

div.stButton > button[kind="primary"] {
  background: linear-gradient(180deg, #171717 0%, #111111 100%);
  color: #ffffff;
  border-color: rgba(17, 24, 39, 0.24);
}

div[data-testid="stExpander"] {
  border: 1px solid rgba(17, 24, 39, 0.07);
  border-radius: 22px;
  background: rgba(255,255,255,0.72);
  box-shadow: 0 10px 24px rgba(17, 24, 39, 0.03);
}

.stTabs [data-baseweb="tab-list"] {
  gap: 0.45rem;
  background: rgba(255,255,255,0.62);
  border: 1px solid rgba(17, 24, 39, 0.06);
  border-radius: 20px;
  padding: 0.3rem;
}

.stTabs [data-baseweb="tab"] {
  border-radius: 14px;
  padding: 0.55rem 0.9rem;
}

.stTabs [aria-selected="true"] {
  background: rgba(15, 23, 42, 0.08);
}

div[data-testid="stDataFrame"], div[data-testid="stTable"] {
  border-radius: 20px;
  overflow: hidden;
  border: 1px solid rgba(17, 24, 39, 0.07);
}

div[data-baseweb="input"],
div[data-baseweb="select"],
div[data-baseweb="textarea"] {
  border-radius: 18px !important;
}

div[data-baseweb="input"] > div,
div[data-baseweb="select"] > div,
div[data-baseweb="textarea"] > div {
  border-radius: 18px !important;
  border-color: rgba(17, 24, 39, 0.08) !important;
  background: rgba(255,255,255,0.82) !important;
  box-shadow: inset 0 1px 0 rgba(255,255,255,0.6);
}

.pim-section-note {
  color: var(--pim-sub);
  font-size: 0.92rem;
  margin: 0.15rem 0 0.85rem 0;
}

.pim-soft-card {
  background: rgba(255,255,255,0.78);
  border: 1px solid rgba(17, 24, 39, 0.06);
  border-radius: 24px;
  padding: 1rem 1.1rem;
  box-shadow: 0 16px 34px rgba(17, 24, 39, 0.035);
  margin-bottom: 0.85rem;
}

.pim-admin-card {
  background: rgba(255,255,255,0.92);
  border: 1px solid rgba(17, 24, 39, 0.07);
  border-radius: 24px;
  padding: 1rem 1.05rem;
  box-shadow: 0 14px 30px rgba(17, 24, 39, 0.04);
  margin-bottom: 0.85rem;
}

.pim-rail-card {
  background: rgba(255,255,255,0.96);
  border: 1px solid rgba(17, 24, 39, 0.07);
  border-radius: 22px;
  padding: 0.9rem 0.95rem;
  box-shadow: 0 12px 28px rgba(17, 24, 39, 0.04);
  margin-bottom: 0.8rem;
}

.pim-section-kicker {
  color: var(--pim-sub);
  font-size: 0.76rem;
  text-transform: uppercase;
  letter-spacing: 0.08em;
  margin-bottom: 0.2rem;
}

.pim-section-title {
  font-size: 1.02rem;
  font-weight: 700;
  letter-spacing: -0.02em;
  margin-bottom: 0.18rem;
}

.pim-compact-note {
  color: var(--pim-sub);
  font-size: 0.86rem;
  line-height: 1.42;
}

code, pre, .pim-mono {
  font-family: 'IBM Plex Mono', monospace;
}
</style>
        """,
        unsafe_allow_html=True,
    )


def build_workspace_summary(conn) -> dict[str, object]:
    stats = get_ozon_cache_stats(conn)
    detmir_stats = get_detmir_cache_stats(conn)
    products_total = int(conn.execute("SELECT COUNT(*) AS total FROM products").fetchone()["total"] or 0)
    imports_total = int(conn.execute("SELECT COUNT(*) AS total FROM catalog_import_history").fetchone()["total"] or 0)
    template_profiles_total = int(conn.execute("SELECT COUNT(*) AS total FROM template_profiles").fetchone()["total"] or 0)
    uploaded_templates_total = int(
        conn.execute(
            "SELECT COUNT(*) AS total FROM uploaded_files WHERE storage_kind = 'client_template'"
        ).fetchone()["total"]
        or 0
    )
    clients_total = int(len(list_client_channels(conn)))
    ozon_backups_total = int(len(list_ozon_snapshot_backups(limit=50)))
    return {
        "products_total": products_total,
        "imports_total": imports_total,
        "template_profiles_total": template_profiles_total,
        "uploaded_templates_total": uploaded_templates_total,
        "clients_total": clients_total,
        "ozon_pairs_total": int(stats.get("category_pairs") or 0),
        "ozon_attributes_total": int(stats.get("attributes_total") or 0),
        "ozon_backups_total": ozon_backups_total,
        "detmir_categories_total": int(detmir_stats.get("detmir_category_cache") or 0),
        "detmir_products_total": int(detmir_stats.get("detmir_product_cache") or 0),
        "active_db_path": str(_get_active_db_path() or Path("data/catalog.db")),
    }


WORKSPACE_NAV_OPTIONS: list[tuple[str, str]] = [
    ("⇣ Импорт", "import"),
    ("▦ Каталог", "catalog"),
    ("▣ Карточка", "product"),
    ("◪ Клиентский шаблон", "template"),
    ("◎ Ozon", "ozon"),
    ("✦ Атрибуты", "attributes"),
    ("⌁ Каналы", "channels"),
    ("⚙︎ Настройки", "settings"),
]

WORKSPACE_SECTION_META: dict[str, tuple[str, str]] = {
    "import": ("Поступление и фиксация памяти", "Загрузи прайс, зафиксируй поставщика, сохрани импорт в каталог и сразу закрепи память в БД."),
    "catalog": ("Каталог для ежедневной работы", "Фильтруй товары, запускай массовое наполнение и открывай карточку по артикулу без лишних технических шагов."),
    "product": ("Карточка товара", "Финальная доводка master-карточки: категории, overlay клиентов, фото, AI и контроль качества перед выгрузкой."),
    "template": ("Клиентские шаблоны", "Выбор клиента из базы, повторное открытие сохранённого шаблона и стабильный экспорт без новой ручной настройки."),
    "ozon": ("Эталон Ozon и кэш памяти", "Категории, атрибуты и справочники Ozon живут как локальный эталон и должны переживать синки и перезапуски."),
    "attributes": ("Справочник атрибутов", "Поддерживай мастер-слой и клиентские поля без путаницы и дублирования."),
    "channels": ("Каналы и overlay клиентов", "Здесь живут channel rules, требования клиентов и интеграции вроде Детского Мира, но не общие настройки PIM."),
    "settings": ("Настройки PIM", "Отдельное место для AI, парсинга, фото и рабочей конфигурации сервиса без засорения каталога и каналов."),
}

WORKSPACE_NAV_KEY_BY_LABEL = {label: key for label, key in WORKSPACE_NAV_OPTIONS}
WORKSPACE_NAV_LABEL_BY_KEY = {key: label for label, key in WORKSPACE_NAV_OPTIONS}
WORKSPACE_NAV_ALIASES = {
    "📥 Импорт": "import",
    "📚 Каталог": "catalog",
    "🧾 Карточка": "product",
    "🧠 Клиентский шаблон": "template",
    "🛒 Ozon": "ozon",
    "🧩 Атрибуты": "attributes",
    "🔌 Каналы": "channels",
    "🧭 Каналы": "channels",
    "⚙️ Настройки": "settings",
}


def _resolve_workspace_nav_key(value: object, default: str = "catalog") -> str:
    raw = str(value or "").strip()
    if not raw:
        return str(default)
    if raw in WORKSPACE_NAV_LABEL_BY_KEY:
        return raw
    if raw in WORKSPACE_NAV_KEY_BY_LABEL:
        return WORKSPACE_NAV_KEY_BY_LABEL[raw]
    if raw in WORKSPACE_NAV_ALIASES:
        return WORKSPACE_NAV_ALIASES[raw]
    return str(default)


def render_sidebar_navigation(summary: dict[str, object], selected_section: str) -> None:
    section_key = _resolve_workspace_nav_key(selected_section, "catalog")
    title, caption = WORKSPACE_SECTION_META.get(section_key, ("PIM", "Рабочая зона PIM."))
    with st.sidebar:
        st.markdown(
            f"""
<div class="pim-side-card">
  <div class="pim-side-kicker">PIM Workspace</div>
  <div class="pim-side-title">Galvanize Product Memory</div>
  <div class="pim-side-note">Товары, шаблоны клиентов, Ozon-структура и экспорт в одной постоянной рабочей зоне.</div>
</div>
<div class="pim-side-card">
  <div class="pim-side-kicker">Память</div>
  <div class="pim-side-note">Товаров: <strong>{int(summary.get('products_total') or 0)}</strong><br/>Клиентов: <strong>{int(summary.get('clients_total') or 0)}</strong><br/>Шаблонов: <strong>{int(summary.get('uploaded_templates_total') or 0)}</strong><br/>Ozon backup: <strong>{int(summary.get('ozon_backups_total') or 0)}</strong></div>
</div>
<div class="pim-side-card">
  <div class="pim-side-kicker">Текущий раздел</div>
  <div class="pim-side-title">{title}</div>
  <div class="pim-side-note">{caption}</div>
  <div class="pim-chip-row" style="margin-top:0.7rem;">
    <span class="pim-chip">Товаров <strong>{int(summary.get('products_total') or 0)}</strong></span>
    <span class="pim-chip">Импортов <strong>{int(summary.get('imports_total') or 0)}</strong></span>
    <span class="pim-chip">Клиентов <strong>{int(summary.get('clients_total') or 0)}</strong></span>
    <span class="pim-chip">Профилей <strong>{int(summary.get('template_profiles_total') or 0)}</strong></span>
    <span class="pim-chip">Ozon cat/type <strong>{int(summary.get('ozon_pairs_total') or 0)}</strong></span>
    <span class="pim-chip">Ozon атрибутов <strong>{int(summary.get('ozon_attributes_total') or 0)}</strong></span>
  </div>
  <div class="pim-side-note" style="margin-top:0.7rem;">Ориентир: 5–10 тыс. карточек в месяц силами одного менеджера.</div>
</div>
            """,
            unsafe_allow_html=True,
        )
        st.caption(f"База: `{summary.get('active_db_path')}`")


def render_workspace_top_navigation() -> str:
    nav_labels = [label for label, _ in WORKSPACE_NAV_OPTIONS]
    default_key = _resolve_workspace_nav_key(st.session_state.get("workspace_nav_section"), "catalog")
    default_label = WORKSPACE_NAV_LABEL_BY_KEY.get(default_key, nav_labels[1] if len(nav_labels) > 1 else nav_labels[0])
    if str(st.session_state.get("workspace_nav_header_label") or "") not in nav_labels:
        st.session_state["workspace_nav_header_label"] = default_label

    with st.container(border=True):
        st.caption("Разделы PIM")
        if hasattr(st, "segmented_control"):
            selected_label = st.segmented_control(
                "Разделы PIM",
                options=nav_labels,
                default=str(st.session_state.get("workspace_nav_header_label") or default_label),
                key="workspace_nav_header_label",
                label_visibility="collapsed",
            )
        else:
            selected_label = st.radio(
                "Разделы PIM",
                options=nav_labels,
                index=nav_labels.index(str(st.session_state.get("workspace_nav_header_label") or default_label)),
                key="workspace_nav_header_label",
                horizontal=True,
                label_visibility="collapsed",
            )
    selected_key = WORKSPACE_NAV_KEY_BY_LABEL.get(str(selected_label or default_label), default_key)
    st.session_state["workspace_nav_section"] = str(selected_key)
    return str(selected_key)


def request_workspace_navigation(nav_label: str) -> None:
    st.session_state["workspace_nav_target"] = _resolve_workspace_nav_key(nav_label, "catalog")
    st.rerun()


def render_workspace_hero(section_key: str, summary: dict[str, object]) -> None:
    title, caption = WORKSPACE_SECTION_META.get(section_key, ("PIM", "Рабочая зона PIM."))
    st.markdown(
        f"""
<div class="pim-shell">
  <div class="pim-hero">
    <div class="pim-hero-title">{title}</div>
    <div class="pim-hero-sub">{caption}</div>
    <div class="pim-chip-row">
      <span class="pim-chip">Товаров <strong>{int(summary.get('products_total') or 0)}</strong></span>
      <span class="pim-chip">Импортов <strong>{int(summary.get('imports_total') or 0)}</strong></span>
      <span class="pim-chip">Клиентов <strong>{int(summary.get('clients_total') or 0)}</strong></span>
      <span class="pim-chip">Профилей шаблонов <strong>{int(summary.get('template_profiles_total') or 0)}</strong></span>
      <span class="pim-chip">Ozon cat/type <strong>{int(summary.get('ozon_pairs_total') or 0)}</strong></span>
      <span class="pim-chip">Ozon атрибутов <strong>{int(summary.get('ozon_attributes_total') or 0)}</strong></span>
    </div>
  </div>
</div>
        """,
        unsafe_allow_html=True,
    )


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def _registry_kind_label(kind: str | None) -> str:
    value = str(kind or "").strip().lower()
    if value == "declaration":
        return "Декларация"
    if value == "certificate":
        return "Сертификат"
    return value or "-"


def _build_registry_product_kind(product_row) -> str:
    candidates = [
        str(product_row.get("subcategory") or "").strip(),
        "",
        str(product_row.get("base_category") or "").strip(),
        str(product_row.get("category") or "").strip(),
        str(product_row.get("name") or "").strip(),
    ]
    ozon_path = str(product_row.get("ozon_category_path") or "").strip()
    if ozon_path:
        parts = [p.strip() for p in ozon_path.split("/") if str(p).strip()]
        if parts:
            candidates[1] = parts[-1]
    for item in candidates:
        if item:
            return item
    return ""


PARSER_SETTINGS_DEFAULTS: dict[str, object] = {
    "timeout_seconds": 8.0,
    "max_hops": 1,
    "fallback_max_results": 4,
    "source_strategy": "auto_full",
    "extra_fallback_domains": "",
    "require_article_match": True,
    "min_name_overlap": 2,
    "min_fallback_score": 3.0,
    "enable_web_fallback": True,
    "enable_ozon_fallback": True,
    "enable_yandex_fallback": True,
    "enable_stats_fallback": True,
    "enable_defaults_fallback": True,
}

MEDIA_SETTINGS_DEFAULTS: dict[str, object] = {
    "public_base_url": "",
}


def _ensure_system_settings_table(conn) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at TEXT
        )
        """
    )
    conn.commit()


def _get_system_setting(conn, key: str, default: object = None) -> object:
    _ensure_system_settings_table(conn)
    row = conn.execute("SELECT value FROM system_settings WHERE key = ? LIMIT 1", (str(key),)).fetchone()
    if not row:
        return default
    return row["value"]


def _set_system_setting(conn, key: str, value: object) -> None:
    _ensure_system_settings_table(conn)
    conn.execute(
        """
        INSERT INTO system_settings (key, value, updated_at)
        VALUES (?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(key) DO UPDATE SET
            value = excluded.value,
            updated_at = CURRENT_TIMESTAMP
        """,
        (str(key), str(value) if value is not None else None),
    )
    conn.commit()


def _to_bool_setting(value: object, default: bool) -> bool:
    if value is None:
        return bool(default)
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return bool(default)


def load_parser_settings(conn) -> dict[str, object]:
    _ensure_system_settings_table(conn)
    timeout_raw = _get_system_setting(conn, "parser.timeout_seconds", PARSER_SETTINGS_DEFAULTS["timeout_seconds"])
    max_hops_raw = _get_system_setting(conn, "parser.max_hops", PARSER_SETTINGS_DEFAULTS["max_hops"])
    max_results_raw = _get_system_setting(conn, "parser.fallback_max_results", PARSER_SETTINGS_DEFAULTS["fallback_max_results"])
    min_name_overlap_raw = _get_system_setting(conn, "parser.min_name_overlap", PARSER_SETTINGS_DEFAULTS["min_name_overlap"])
    min_fallback_score_raw = _get_system_setting(conn, "parser.min_fallback_score", PARSER_SETTINGS_DEFAULTS["min_fallback_score"])
    settings = {
        "timeout_seconds": float(timeout_raw) if str(timeout_raw).strip() not in {"", "None"} else float(PARSER_SETTINGS_DEFAULTS["timeout_seconds"]),
        "max_hops": int(float(max_hops_raw)) if str(max_hops_raw).strip() not in {"", "None"} else int(PARSER_SETTINGS_DEFAULTS["max_hops"]),
        "fallback_max_results": int(float(max_results_raw)) if str(max_results_raw).strip() not in {"", "None"} else int(PARSER_SETTINGS_DEFAULTS["fallback_max_results"]),
        "source_strategy": str(_get_system_setting(conn, "parser.source_strategy", PARSER_SETTINGS_DEFAULTS["source_strategy"]) or PARSER_SETTINGS_DEFAULTS["source_strategy"]),
        "extra_fallback_domains": str(_get_system_setting(conn, "parser.extra_fallback_domains", PARSER_SETTINGS_DEFAULTS["extra_fallback_domains"]) or ""),
        "require_article_match": _to_bool_setting(_get_system_setting(conn, "parser.require_article_match", PARSER_SETTINGS_DEFAULTS["require_article_match"]), bool(PARSER_SETTINGS_DEFAULTS["require_article_match"])),
        "min_name_overlap": int(float(min_name_overlap_raw)) if str(min_name_overlap_raw).strip() not in {"", "None"} else int(PARSER_SETTINGS_DEFAULTS["min_name_overlap"]),
        "min_fallback_score": float(min_fallback_score_raw) if str(min_fallback_score_raw).strip() not in {"", "None"} else float(PARSER_SETTINGS_DEFAULTS["min_fallback_score"]),
        "enable_web_fallback": _to_bool_setting(_get_system_setting(conn, "parser.enable_web_fallback", PARSER_SETTINGS_DEFAULTS["enable_web_fallback"]), bool(PARSER_SETTINGS_DEFAULTS["enable_web_fallback"])),
        "enable_ozon_fallback": _to_bool_setting(_get_system_setting(conn, "parser.enable_ozon_fallback", PARSER_SETTINGS_DEFAULTS["enable_ozon_fallback"]), bool(PARSER_SETTINGS_DEFAULTS["enable_ozon_fallback"])),
        "enable_yandex_fallback": _to_bool_setting(_get_system_setting(conn, "parser.enable_yandex_fallback", PARSER_SETTINGS_DEFAULTS["enable_yandex_fallback"]), bool(PARSER_SETTINGS_DEFAULTS["enable_yandex_fallback"])),
        "enable_stats_fallback": _to_bool_setting(_get_system_setting(conn, "parser.enable_stats_fallback", PARSER_SETTINGS_DEFAULTS["enable_stats_fallback"]), bool(PARSER_SETTINGS_DEFAULTS["enable_stats_fallback"])),
        "enable_defaults_fallback": _to_bool_setting(_get_system_setting(conn, "parser.enable_defaults_fallback", PARSER_SETTINGS_DEFAULTS["enable_defaults_fallback"]), bool(PARSER_SETTINGS_DEFAULTS["enable_defaults_fallback"])),
    }
    settings["timeout_seconds"] = max(2.0, min(30.0, float(settings["timeout_seconds"])))
    settings["max_hops"] = max(1, min(3, int(settings["max_hops"])))
    settings["fallback_max_results"] = max(1, min(12, int(settings["fallback_max_results"])))
    valid_strategies = {"auto_full", "supplier_only", "supplier_plus_ozon", "supplier_plus_yandex", "web_only", "custom_domains"}
    if str(settings.get("source_strategy") or "") not in valid_strategies:
        settings["source_strategy"] = str(PARSER_SETTINGS_DEFAULTS["source_strategy"])
    settings["min_name_overlap"] = max(1, min(5, int(settings["min_name_overlap"])))
    settings["min_fallback_score"] = max(0.0, min(10.0, float(settings["min_fallback_score"])))
    return settings


def save_parser_settings(conn, settings: dict[str, object]) -> None:
    _set_system_setting(conn, "parser.timeout_seconds", float(settings.get("timeout_seconds", PARSER_SETTINGS_DEFAULTS["timeout_seconds"])))
    _set_system_setting(conn, "parser.max_hops", int(settings.get("max_hops", PARSER_SETTINGS_DEFAULTS["max_hops"])))
    _set_system_setting(conn, "parser.fallback_max_results", int(settings.get("fallback_max_results", PARSER_SETTINGS_DEFAULTS["fallback_max_results"])))
    _set_system_setting(conn, "parser.source_strategy", str(settings.get("source_strategy", PARSER_SETTINGS_DEFAULTS["source_strategy"])))
    _set_system_setting(conn, "parser.extra_fallback_domains", str(settings.get("extra_fallback_domains", "") or ""))
    _set_system_setting(conn, "parser.require_article_match", 1 if bool(settings.get("require_article_match", True)) else 0)
    _set_system_setting(conn, "parser.min_name_overlap", int(settings.get("min_name_overlap", PARSER_SETTINGS_DEFAULTS["min_name_overlap"])))
    _set_system_setting(conn, "parser.min_fallback_score", float(settings.get("min_fallback_score", PARSER_SETTINGS_DEFAULTS["min_fallback_score"])))
    _set_system_setting(conn, "parser.enable_web_fallback", 1 if bool(settings.get("enable_web_fallback", True)) else 0)
    _set_system_setting(conn, "parser.enable_ozon_fallback", 1 if bool(settings.get("enable_ozon_fallback", True)) else 0)
    _set_system_setting(conn, "parser.enable_yandex_fallback", 1 if bool(settings.get("enable_yandex_fallback", True)) else 0)
    _set_system_setting(conn, "parser.enable_stats_fallback", 1 if bool(settings.get("enable_stats_fallback", True)) else 0)
    _set_system_setting(conn, "parser.enable_defaults_fallback", 1 if bool(settings.get("enable_defaults_fallback", True)) else 0)


def load_media_settings(conn) -> dict[str, object]:
    _ensure_system_settings_table(conn)
    return {
        "public_base_url": str(_get_system_setting(conn, "media.public_base_url", MEDIA_SETTINGS_DEFAULTS["public_base_url"]) or "").strip(),
    }


def save_media_settings(conn, settings: dict[str, object]) -> None:
    _set_system_setting(conn, "media.public_base_url", str(settings.get("public_base_url", "") or "").strip())


def _coerce_state_number(key: str, default_value: object, caster):
    if key in st.session_state:
        try:
            st.session_state[key] = caster(st.session_state.get(key))
        except Exception:
            st.session_state[key] = caster(default_value)


def _current_ai_form_settings(saved_settings: dict[str, Any]) -> dict[str, Any]:
    provider = str(st.session_state.get("ai_cfg_provider", saved_settings.get("provider") or "openai")).strip().lower()
    defaults = PROVIDER_DEFAULTS.get(provider, PROVIDER_DEFAULTS["openai"])
    settings = {
        "enabled": bool(st.session_state.get("ai_cfg_enabled", saved_settings.get("enabled", True))),
        "provider": provider,
        "base_url": str(st.session_state.get("ai_cfg_base_url", saved_settings.get("base_url") or defaults.get("base_url") or "") or "").strip(),
        "chat_model": str(st.session_state.get("ai_cfg_chat_model", saved_settings.get("chat_model") or defaults.get("chat_model") or "") or "").strip(),
        "image_model": str(st.session_state.get("ai_cfg_image_model", saved_settings.get("image_model") or defaults.get("image_model") or "") or "").strip(),
        "api_key": str(st.session_state.get("ai_cfg_api_key", saved_settings.get("api_key") or "") or "").strip(),
        "use_env_api_key": bool(st.session_state.get("ai_cfg_use_env_key", saved_settings.get("use_env_api_key", True))),
        "temperature": float(st.session_state.get("ai_cfg_temperature", saved_settings.get("temperature", 0.3)) or 0.3),
        "max_tokens": int(float(st.session_state.get("ai_cfg_max_tokens", saved_settings.get("max_tokens", 1800)) or 1800)),
        "image_size": str(st.session_state.get("ai_cfg_image_size", saved_settings.get("image_size") or "1024x1024") or "1024x1024").strip(),
        "openrouter_referer": str(st.session_state.get("ai_cfg_or_referer", saved_settings.get("openrouter_referer") or "") or "").strip(),
        "openrouter_title": str(st.session_state.get("ai_cfg_or_title", saved_settings.get("openrouter_title") or "pim") or "pim").strip() or "pim",
    }
    known_provider_base_urls = {
        str(item.get("base_url") or "").strip()
        for item in PROVIDER_DEFAULTS.values()
        if str(item.get("base_url") or "").strip()
    }
    current_base_url = str(settings.get("base_url") or "").strip()
    default_base_url = str(defaults.get("base_url") or "").strip()
    if (not current_base_url) or (current_base_url in known_provider_base_urls and current_base_url != default_base_url):
        settings["base_url"] = default_base_url
    if not str(settings.get("chat_model") or "").strip():
        settings["chat_model"] = str(defaults.get("chat_model") or "").strip()
    if not str(settings.get("image_model") or "").strip():
        settings["image_model"] = str(defaults.get("image_model") or "").strip()
    return settings


def _apply_ai_provider_defaults_to_state(provider: str, *, force_provider_base_url: bool = False) -> None:
    key = str(provider or "openai").strip().lower()
    defaults = PROVIDER_DEFAULTS.get(key, PROVIDER_DEFAULTS["openai"])
    known_provider_base_urls = {
        str(item.get("base_url") or "").strip()
        for item in PROVIDER_DEFAULTS.values()
        if str(item.get("base_url") or "").strip()
    }
    current_base_url = str(st.session_state.get("ai_cfg_base_url") or "").strip()
    default_base_url = str(defaults.get("base_url") or "").strip()
    if (
        force_provider_base_url
        or not current_base_url
        or (current_base_url in known_provider_base_urls and current_base_url != default_base_url)
    ):
        st.session_state["ai_cfg_base_url"] = str(defaults.get("base_url") or "").strip()
    if not str(st.session_state.get("ai_cfg_chat_model") or "").strip():
        st.session_state["ai_cfg_chat_model"] = str(defaults.get("chat_model") or "").strip()
    if not str(st.session_state.get("ai_cfg_image_model") or "").strip():
        st.session_state["ai_cfg_image_model"] = str(defaults.get("image_model") or "").strip()
    if key == "openrouter" and not str(st.session_state.get("ai_cfg_or_title") or "").strip():
        st.session_state["ai_cfg_or_title"] = "pim"


def render_ai_settings_panel(conn) -> None:
    ai_settings = load_ai_settings(conn)
    ai_profiles = list_ai_connection_profiles(conn)
    for key, default in (
        ("ai_cfg_provider", str(ai_settings.get("provider") or "openai")),
        ("ai_cfg_chat_model", str(ai_settings.get("chat_model") or "")),
        ("ai_cfg_image_model", str(ai_settings.get("image_model") or "")),
        ("ai_cfg_base_url", str(ai_settings.get("base_url") or "")),
        ("ai_cfg_image_size", str(ai_settings.get("image_size") or "1024x1024")),
        ("ai_cfg_api_key", str(ai_settings.get("api_key") or "")),
        ("ai_cfg_or_referer", str(ai_settings.get("openrouter_referer") or "")),
        ("ai_cfg_or_title", str(ai_settings.get("openrouter_title") or "pim")),
    ):
        st.session_state.setdefault(key, default)
    _coerce_state_number("ai_cfg_temperature", ai_settings.get("temperature", 0.3), float)
    _coerce_state_number("ai_cfg_max_tokens", ai_settings.get("max_tokens", 1800), int)
    st.session_state.setdefault("ai_cfg_use_env_key", bool(ai_settings.get("use_env_api_key", True)))
    st.session_state.setdefault("ai_cfg_enabled", bool(ai_settings.get("enabled", True)))
    _apply_ai_provider_defaults_to_state(str(st.session_state.get("ai_cfg_provider") or ai_settings.get("provider") or "openai"))

    form_settings = _current_ai_form_settings(ai_settings)
    cfg_ok, cfg_msg = ai_is_configured(form_settings)
    active_ok, active_msg = ai_is_configured(ai_settings)

    st.markdown('<div class="pim-soft-card">', unsafe_allow_html=True)
    st.markdown("### AI-провайдер и профили")
    st.markdown(
        '<div class="pim-section-note">AI используется для продающего названия, описания, подсказок атрибутов и промптов для фото. Профиль можно сделать активным одним нажатием.</div>',
        unsafe_allow_html=True,
    )
    if ai_profiles:
        profile_options = [None] + [int(item["id"]) for item in ai_profiles]
        profile_map = {int(item["id"]): item for item in ai_profiles}
        ap1, ap2, ap3 = st.columns([2.2, 1.2, 1])
        with ap1:
            selected_ai_profile_id = st.selectbox(
                "Сохранённый AI-профиль",
                options=profile_options,
                format_func=lambda x: "-- выбрать --" if x is None else f"{profile_map[int(x)]['profile_name']} | {profile_map[int(x)].get('provider') or '-'} | {profile_map[int(x)].get('chat_model') or '-'}",
                key="ai_profile_select",
            )
        with ap2:
            if st.button("Активировать профиль", key="ai_profile_load_btn"):
                if selected_ai_profile_id is None:
                    st.warning("Сначала выбери профиль.")
                else:
                    loaded = get_ai_connection_profile(conn, int(selected_ai_profile_id))
                    if loaded:
                        st.session_state["ai_cfg_provider"] = str(loaded.get("provider") or "openai")
                        st.session_state["ai_cfg_chat_model"] = str(loaded.get("chat_model") or "")
                        st.session_state["ai_cfg_image_model"] = str(loaded.get("image_model") or "")
                        st.session_state["ai_cfg_base_url"] = str(loaded.get("base_url") or "")
                        st.session_state["ai_cfg_temperature"] = float(loaded.get("temperature") or 0.3)
                        st.session_state["ai_cfg_max_tokens"] = int(loaded.get("max_tokens") or 1800)
                        st.session_state["ai_cfg_image_size"] = str(loaded.get("image_size") or "1024x1024")
                        st.session_state["ai_cfg_use_env_key"] = bool(int(loaded.get("use_env_api_key") or 0))
                        st.session_state["ai_cfg_enabled"] = True
                        st.session_state["ai_cfg_api_key"] = str(loaded.get("api_key") or "")
                        st.session_state["ai_cfg_or_referer"] = str(loaded.get("openrouter_referer") or "")
                        st.session_state["ai_cfg_or_title"] = str(loaded.get("openrouter_title") or "pim")
                        save_ai_settings(conn, _current_ai_form_settings(ai_settings))
                        st.success("Профиль загружен и сделан активным для всего PIM.")
                        st.rerun()
        with ap3:
            st.caption(f"Профилей: {len(ai_profiles)}")
    if cfg_ok:
        st.success(cfg_msg)
    else:
        st.warning(cfg_msg)
    if form_settings != ai_settings:
        if active_ok:
            st.caption("Ниже сейчас редактируется новая конфигурация. Активная сохранённая конфигурация уже рабочая.")
        else:
            st.caption("Ниже есть несохранённые изменения. Для всей системы активна только сохранённая конфигурация.")
    else:
        st.caption(active_msg)
    st.markdown('</div>', unsafe_allow_html=True)

    provider_options = ["openai", "openrouter", "nvidia"]
    current_provider = str(form_settings.get("provider") or "openai").strip().lower()
    if current_provider not in provider_options:
        current_provider = "openai"
    previous_provider = str(st.session_state.get("ai_cfg_provider_last_applied") or current_provider).strip().lower()
    a1, a2, a3 = st.columns(3)
    with a1:
        ai_provider = st.selectbox(
            "Провайдер AI",
            options=provider_options,
            index=provider_options.index(current_provider),
            format_func=lambda x: {"openai": "OpenAI", "openrouter": "OpenRouter", "nvidia": "NVIDIA"}.get(x, x),
            key="ai_cfg_provider",
        )
        provider_changed = str(ai_provider).strip().lower() != previous_provider
        _apply_ai_provider_defaults_to_state(str(ai_provider), force_provider_base_url=provider_changed)
        st.session_state["ai_cfg_provider_last_applied"] = str(ai_provider).strip().lower()
    with a2:
        ai_chat_model = st.text_input(
            "Модель",
            value=str(_current_ai_form_settings(ai_settings).get("chat_model") or ""),
            key="ai_cfg_chat_model",
            placeholder="Например: openai/gpt-4o-mini или tencent/hunyuan-a13b-instruct:free",
        )
    with a3:
        ai_api_key = st.text_input(
            "API key",
            value=str(form_settings.get("api_key") or ""),
            type="password",
            key="ai_cfg_api_key",
            help="Для OpenRouter обычно достаточно только ключа и названия модели.",
        )

    quick1, quick2 = st.columns([1, 2])
    with quick1:
        ai_enabled = st.checkbox("AI включен", value=bool(form_settings.get("enabled", True)), key="ai_cfg_enabled")
    with quick2:
        provider_note = {
            "openrouter": "Для OpenRouter service URL и служебные заголовки подставляются автоматически. Обычно достаточно выбрать модель и вставить API key.",
            "nvidia": "Для NVIDIA базовый URL подставляется автоматически. Если модель медленная, лучше вынести её в отдельный профиль для deep repair.",
            "openai": "Для OpenAI базовый URL и image model подставляются автоматически. Обычно достаточно модели и ключа.",
        }.get(str(ai_provider), "Базовые настройки провайдера подставляются автоматически.")
        st.caption(provider_note)

    with st.expander("Расширенные настройки AI", expanded=False):
        advanced_form_settings = _current_ai_form_settings(ai_settings)
        b1, b2, b3 = st.columns(3)
        with b1:
            ai_base_url = st.text_input("Base URL", value=str(advanced_form_settings.get("base_url") or ""), key="ai_cfg_base_url")
        with b2:
            ai_temperature = st.number_input("Температура", min_value=0.0, max_value=1.5, value=float(advanced_form_settings.get("temperature") or 0.3), step=0.1, key="ai_cfg_temperature")
        with b3:
            ai_max_tokens = st.number_input("Max tokens", min_value=256, max_value=65536, value=int(advanced_form_settings.get("max_tokens") or 1800), step=64, key="ai_cfg_max_tokens")

        c1, c2, c3 = st.columns(3)
        image_size_options = ["1024x1024", "1536x1024", "1024x1536"]
        current_image_size = str(advanced_form_settings.get("image_size") or "1024x1024")
        if current_image_size not in image_size_options:
            current_image_size = "1024x1024"
        with c1:
            ai_image_model = st.text_input("Image model", value=str(advanced_form_settings.get("image_model") or ""), key="ai_cfg_image_model")
            ai_image_size = st.selectbox("Размер изображения", options=image_size_options, index=image_size_options.index(current_image_size), key="ai_cfg_image_size")
        with c2:
            ai_use_env_key = st.checkbox("Брать API key из env", value=bool(advanced_form_settings.get("use_env_api_key", True)), key="ai_cfg_use_env_key")
        with c3:
            if str(ai_provider) == "openrouter":
                ai_openrouter_referer = st.text_input("OpenRouter Referer", value=str(advanced_form_settings.get("openrouter_referer") or ""), key="ai_cfg_or_referer")
                ai_openrouter_title = st.text_input("OpenRouter App Title", value=str(advanced_form_settings.get("openrouter_title") or "pim"), key="ai_cfg_or_title")
            else:
                ai_openrouter_referer = str(advanced_form_settings.get("openrouter_referer") or "")
                ai_openrouter_title = str(advanced_form_settings.get("openrouter_title") or "pim")

    if "ai_temperature" not in locals():
        ai_temperature = float(_current_ai_form_settings(ai_settings).get("temperature") or 0.3)
    if "ai_max_tokens" not in locals():
        ai_max_tokens = int(_current_ai_form_settings(ai_settings).get("max_tokens") or 1800)
    if "ai_base_url" not in locals():
        ai_base_url = str(_current_ai_form_settings(ai_settings).get("base_url") or "")
    if "ai_image_model" not in locals():
        ai_image_model = str(_current_ai_form_settings(ai_settings).get("image_model") or "")
    if "ai_use_env_key" not in locals():
        ai_use_env_key = bool(_current_ai_form_settings(ai_settings).get("use_env_api_key", True))
    if "ai_openrouter_referer" not in locals():
        ai_openrouter_referer = str(_current_ai_form_settings(ai_settings).get("openrouter_referer") or "")
    if "ai_openrouter_title" not in locals():
        ai_openrouter_title = str(_current_ai_form_settings(ai_settings).get("openrouter_title") or "pim")
    if "ai_image_size" not in locals():
        ai_image_size = str(_current_ai_form_settings(ai_settings).get("image_size") or "1024x1024")

    ai_profile_name = st.text_input(
        "Имя профиля AI",
        value=str(st.session_state.get("ai_profile_name_input") or f"{str(ai_provider).strip()}_{str(ai_chat_model or 'default').strip() or 'default'}"),
        key="ai_profile_name_input",
        help="Профили нужны, чтобы быстро переключать провайдера, модель и токен без ручного переввода.",
    )
    draft_settings = {
        "enabled": bool(ai_enabled),
        "provider": str(ai_provider),
        "base_url": str(ai_base_url or "").strip(),
        "chat_model": str(ai_chat_model or "").strip(),
        "image_model": str(ai_image_model or "").strip(),
        "api_key": str(ai_api_key or "").strip(),
        "use_env_api_key": bool(ai_use_env_key),
        "temperature": float(ai_temperature),
        "max_tokens": int(ai_max_tokens),
        "image_size": str(ai_image_size),
        "openrouter_referer": str(ai_openrouter_referer or "").strip(),
        "openrouter_title": str(ai_openrouter_title or "pim").strip(),
    }

    s1, s2, s3 = st.columns([1, 1, 1.2])
    with s1:
        if st.button("Сохранить и сделать активным", key="ai_cfg_save_btn", type="primary"):
            save_ai_settings(conn, draft_settings)
            st.success("AI-настройки сохранены и стали активными для всего PIM.")
            st.rerun()
    with s2:
        if st.button("Проверить провайдера и модель", key="ai_cfg_test_btn"):
            check_result = check_ai_connection(draft_settings)
            if check_result.get("ok"):
                st.success(
                    f"AI доступен: provider={check_result.get('provider')}, model={check_result.get('model')}. "
                    f"{str(check_result.get('text') or '').strip()}"
                )
                if check_result.get("warning"):
                    st.info(f"Замечание: {check_result.get('warning')}")
            else:
                st.error(f"Ошибка AI-подключения: {check_result.get('error')}")
    with s3:
        if st.button("Сохранить как профиль", key="ai_profile_save_btn"):
            profile_id = save_ai_connection_profile(conn, ai_profile_name, draft_settings)
            st.success(f"AI-профиль сохранён: #{profile_id}")
            st.rerun()


def render_media_settings_panel(conn) -> None:
    media_settings = load_media_settings(conn)
    st.markdown("### Фото и медиа")
    st.markdown(
        '<div class="pim-section-note">Если в прайсах или 1С приходят локальные пути к картинкам, здесь задаётся единый публичный base URL для карточек, шаблонов и ZIP.</div>',
        unsafe_allow_html=True,
    )
    media_public_base_url = st.text_input(
        "Public base URL для фото",
        value=str(media_settings.get("public_base_url") or ""),
        placeholder="Например: https://cdn.example.ru/photos",
        key="media_cfg_public_base_url",
        help="Пример: `\\\\fs03\\1c_photo\\LinkPics\\153\\153246_0.jpg` -> `https://.../1c_photo/LinkPics/153/153246_0.jpg`.",
    )
    if st.button("Сохранить настройки фото", key="media_cfg_save_btn"):
        save_media_settings(conn, {"public_base_url": str(media_public_base_url or "").strip()})
        st.success("Настройки фото сохранены.")
        st.rerun()


def render_parser_settings_panel(conn) -> None:
    parser_settings = load_parser_settings(conn)
    st.markdown("### Парсинг и автообогащение")
    st.markdown(
        '<div class="pim-section-note">Глобальные настройки enrichment-пайплайна. В рабочем каталоге остаются только быстрые действия по товарам, а не технастройки.</div>',
        unsafe_allow_html=True,
    )
    _coerce_state_number("parser_cfg_timeout_seconds", parser_settings.get("timeout_seconds", 8.0), float)
    _coerce_state_number("parser_cfg_min_fallback_score", parser_settings.get("min_fallback_score", 3.0), float)
    _coerce_state_number("parser_cfg_max_hops", parser_settings.get("max_hops", 1), int)
    _coerce_state_number("parser_cfg_max_results", parser_settings.get("fallback_max_results", 4), int)
    _coerce_state_number("parser_cfg_min_name_overlap", parser_settings.get("min_name_overlap", 2), int)

    p1, p2, p3 = st.columns(3)
    with p1:
        ps_timeout = st.number_input("Таймаут запроса, сек", min_value=2.0, max_value=30.0, value=float(parser_settings.get("timeout_seconds", 8.0)), step=1.0, key="parser_cfg_timeout_seconds")
        ps_max_hops = st.number_input("Переходов с листинга в карточку", min_value=1, max_value=3, value=int(parser_settings.get("max_hops", 1)), step=1, key="parser_cfg_max_hops")
        ps_max_results = st.number_input("Лимит кандидатов fallback", min_value=1, max_value=12, value=int(parser_settings.get("fallback_max_results", 4)), step=1, key="parser_cfg_max_results")
        ps_min_score = st.number_input("Мин. score fallback", min_value=0.0, max_value=10.0, value=float(parser_settings.get("min_fallback_score", 3.0)), step=0.1, key="parser_cfg_min_fallback_score")
    with p2:
        strategy_options = [
            ("auto_full", "Авто: поставщик -> web-поиск -> AI"),
            ("supplier_only", "Только сайт поставщика"),
            ("web_only", "Только интернет-поиск + AI"),
            ("custom_domains", "Только выбранные домены"),
        ]
        strategy_values = [x[0] for x in strategy_options]
        current_strategy = str(parser_settings.get("source_strategy", "auto_full") or "auto_full")
        if current_strategy not in strategy_values:
            current_strategy = "auto_full"
        ps_source_strategy = st.selectbox(
            "Стратегия источников парсинга",
            options=strategy_values,
            index=strategy_values.index(current_strategy),
            format_func=lambda x: next((label for key, label in strategy_options if key == x), x),
            key="parser_cfg_source_strategy",
        )
        ps_require_article_match = st.checkbox(
            "Требовать совпадение артикула/кода",
            value=bool(parser_settings.get("require_article_match", True)),
            key="parser_cfg_require_article_match",
            help="Защищает от подстановки данных не того товара при fallback-поиске.",
        )
        ps_min_overlap = st.number_input(
            "Мин. пересечение токенов названия",
            min_value=1,
            max_value=5,
            value=int(parser_settings.get("min_name_overlap", 2)),
            step=1,
            key="parser_cfg_min_name_overlap",
            help="Минимум общих токенов в названии товара и найденной страницы.",
        )
    with p3:
        ps_extra_domains = st.text_area(
            "Доп. домены fallback (через запятую)",
            value=str(parser_settings.get("extra_fallback_domains", "") or ""),
            key="parser_cfg_extra_fallback_domains",
            help="Пример: sportmaster.ru, alltricks.com, chainreactioncycles.com. Для стратегии `Только выбранные домены` обязательно заполнить.",
        )
        ps_stats = st.checkbox("Fallback габаритов из каталога", value=bool(parser_settings.get("enable_stats_fallback", True)), key="parser_cfg_enable_stats_fallback")
        ps_defaults = st.checkbox("Fallback габаритов из defaults", value=bool(parser_settings.get("enable_defaults_fallback", True)), key="parser_cfg_enable_defaults_fallback")
    if st.button("Сохранить настройки парсинга", key="parser_cfg_save_button", type="primary"):
        strategy = str(ps_source_strategy or "auto_full")
        enable_web = strategy in {"auto_full", "web_only", "custom_domains"}
        enable_ozon = strategy == "auto_full"
        enable_yandex = strategy == "auto_full"
        new_settings = {
            "timeout_seconds": float(ps_timeout),
            "max_hops": int(ps_max_hops),
            "fallback_max_results": int(ps_max_results),
            "source_strategy": strategy,
            "extra_fallback_domains": str(ps_extra_domains or "").strip(),
            "require_article_match": bool(ps_require_article_match),
            "min_name_overlap": int(ps_min_overlap),
            "min_fallback_score": float(ps_min_score),
            "enable_web_fallback": bool(enable_web),
            "enable_ozon_fallback": bool(enable_ozon),
            "enable_yandex_fallback": bool(enable_yandex),
            "enable_stats_fallback": bool(ps_stats),
            "enable_defaults_fallback": bool(ps_defaults),
        }
        save_parser_settings(conn, new_settings)
        st.success("Настройки парсинга сохранены.")
        st.rerun()


def _set_ozon_bg_state(**kwargs) -> None:
    with _OZON_SYNC_BG_LOCK:
        _OZON_SYNC_BG_STATE.update(kwargs)


def _get_ozon_bg_state() -> dict:
    with _OZON_SYNC_BG_LOCK:
        state = dict(_OZON_SYNC_BG_STATE)
    state["thread_alive"] = bool(_OZON_SYNC_BG_THREAD and _OZON_SYNC_BG_THREAD.is_alive())
    return state


def _ozon_bg_worker(db_path: str, client_id: str, api_key: str) -> None:
    conn = None
    try:
        conn = get_connection(Path(db_path))
        init_db(conn)
        result = sync_all_categories_and_attributes(
            conn,
            client_id=client_id or None,
            api_key=api_key or None,
            max_pairs=None,
            import_to_pim=True,
            only_leaf=True,
            include_disabled=False,
        )
        _set_ozon_bg_state(
            running=False,
            finished_at=_now_iso(),
            result=result,
            last_error=None,
        )
    except Exception as e:
        _set_ozon_bg_state(
            running=False,
            finished_at=_now_iso(),
            result=None,
            last_error=str(e)[:1000],
        )
    finally:
        if conn is not None:
            conn.close()


def _start_ozon_bg_sync(client_id: str, api_key: str) -> tuple[bool, str]:
    global _OZON_SYNC_BG_THREAD
    state = _get_ozon_bg_state()
    if state.get("running") and state.get("thread_alive"):
        return False, "Фоновая синхронизация Ozon уже выполняется."
    active_db = _get_active_db_path() or str(Path("data/catalog.db"))
    _set_ozon_bg_state(
        running=True,
        started_at=_now_iso(),
        finished_at=None,
        result=None,
        last_error=None,
        db_path=active_db,
    )
    _OZON_SYNC_BG_THREAD = threading.Thread(
        target=_ozon_bg_worker,
        args=(str(active_db), str(client_id or ""), str(api_key or "")),
        daemon=True,
        name="ozon-full-sync-bg",
    )
    _OZON_SYNC_BG_THREAD.start()
    return True, f"Фоновая синхронизация Ozon запущена. База: {active_db}"


def to_attribute_code(name: str) -> str:
    clean = str(name or "").strip().lower()
    clean = "_".join("".join(ch if ch.isalnum() else " " for ch in clean).split())
    return clean[:120]


def list_distinct_values(conn, column_name: str) -> list[str]:
    rows = conn.execute(
        f"""
        SELECT DISTINCT TRIM({column_name}) AS value
        FROM products
        WHERE {column_name} IS NOT NULL
          AND TRIM({column_name}) <> ''
        ORDER BY value
        """
    ).fetchall()
    return [str(r["value"]) for r in rows if r["value"]]


def list_catalog_categories(conn) -> list[str]:
    rows = conn.execute(
        """
        SELECT category, base_category, subcategory, ozon_category_path
        FROM products
        """
    ).fetchall()
    if not rows:
        return []

    def split_ozon_path(path: str | None) -> list[str]:
        text = " ".join(str(path or "").strip().split())
        if not text:
            return []
        chunks = [x.strip() for x in re.split(r"\s*(?:/|>|»|→|\|)\s*", text) if str(x).strip()]
        return chunks if chunks else [text]

    ozon_values: set[str] = set()
    legacy_values: set[str] = set()
    for row in rows:
        category = str(row["category"] or "").strip()
        base_category = str(row["base_category"] or "").strip()
        subcategory = str(row["subcategory"] or "").strip()
        ozon_path = str(row["ozon_category_path"] or "").strip()

        if ozon_path:
            parts = split_ozon_path(ozon_path)
            ozon_values.add(ozon_path)
            if parts:
                ozon_values.add(parts[-1])
            if len(parts) >= 2:
                ozon_values.add(parts[-2])

        for value in (category, base_category, subcategory):
            if value:
                legacy_values.add(value)

    preferred = sorted([v for v in ozon_values if v], key=lambda x: x.lower())
    legacy_only = sorted([v for v in legacy_values if v and v not in ozon_values], key=lambda x: x.lower())
    # Ozon — эталон категорий. Legacy используем только если Ozon-категорий пока нет.
    return preferred if preferred else legacy_only


def _split_ozon_path_parts(path: str | None) -> list[str]:
    text = " ".join(str(path or "").strip().split())
    if not text:
        return []
    parts = [x.strip() for x in re.split(r"\s*(?:/|>|»|→|\|)\s*", text) if str(x).strip()]
    return parts if parts else [text]


def list_ozon_category_filters(conn) -> tuple[list[str], list[str]]:
    rows = conn.execute(
        """
        SELECT DISTINCT ozon_category_path
        FROM products
        WHERE ozon_category_path IS NOT NULL
          AND TRIM(ozon_category_path) <> ''
        """
    ).fetchall()
    categories: set[str] = set()
    subcategories: set[str] = set()
    for row in rows:
        path = str(row["ozon_category_path"] or "").strip()
        parts = _split_ozon_path_parts(path)
        if not parts:
            continue
        subcategories.add(parts[-1])
        categories.add(parts[-2] if len(parts) >= 2 else parts[-1])
    return (
        sorted([x for x in categories if x], key=lambda x: x.lower()),
        sorted([x for x in subcategories if x], key=lambda x: x.lower()),
    )


RU_COLUMN_MAP: dict[str, str] = {
    "id": "ID",
    "product_id": "ID товара",
    "article": "Артикул",
    "internal_article": "Внутренний артикул",
    "supplier_article": "Артикул поставщика",
    "name": "Название",
    "brand": "Бренд",
    "barcode": "Штрихкод",
    "category": "Категория",
    "base_category": "Базовая категория",
    "subcategory": "Подкатегория",
    "supplier_name": "Поставщик",
    "supplier_url": "Ссылка поставщика",
    "description": "Описание",
    "image_url": "Фото",
    "weight": "Вес, кг",
    "gross_weight": "Вес брутто, кг",
    "length": "Длина, см",
    "width": "Ширина, см",
    "height": "Высота, см",
    "package_length": "Длина упаковки, см",
    "package_width": "Ширина упаковки, см",
    "package_height": "Высота упаковки, см",
    "uom": "Ед. изм.",
    "tnved_code": "ТН ВЭД",
    "wheel_diameter_inch": "Диаметр колеса, inch",
    "updated_at": "Обновлено",
    "created_at": "Создано",
    "import_batch_id": "Партия импорта",
    "supplier_parse_status": "Статус парсинга",
    "supplier_parse_comment": "Комментарий парсинга",
    "ozon_description_category_id": "ID категории Ozon",
    "ozon_type_id": "ID типа Ozon",
    "ozon_category_path": "Ozon категория",
    "ozon_category_confidence": "Уверенность Ozon",
    "description_category_id": "ID категории Ozon",
    "type_id": "ID типа Ozon",
    "type_name": "Тип категории Ozon",
    "category_name": "Название категории Ozon",
    "full_path": "Путь категории Ozon",
    "children_count": "Дочерних категорий",
    "disabled": "Отключена",
    "fetched_at": "Загружено",
    "nodes": "Узлов",
    "attribute_id": "ID атрибута",
    "attribute_code": "Код атрибута",
    "code": "Код атрибута",
    "data_type": "Тип данных",
    "scope": "Область",
    "entity_type": "Сущность",
    "channel_code": "Канал",
    "locale": "Локаль",
    "is_required": "Обязательный",
    "is_required_for_category": "Обязательный для категории",
    "is_collection": "Множественный",
    "is_multi_value": "Множественный",
    "dictionary_id": "ID справочника",
    "group_name": "Группа",
    "max_value_count": "Макс. значений",
    "value": "Значение",
    "value_text": "Текстовое значение",
    "value_number": "Числовое значение",
    "value_boolean": "Булево значение",
    "value_json": "JSON значение",
    "value_id": "ID значения",
    "info": "Инфо",
    "picture": "Картинка",
    "field_name": "Поле",
    "source_type": "Источник",
    "source_url": "URL источника",
    "source_value_raw": "Сырое значение",
    "confidence": "Уверенность",
    "source_name": "Источник значения",
    "transform_rule": "Правило трансформации",
    "matched_by": "Метод сопоставления",
    "status": "Статус",
}

ATTRIBUTE_CODE_RU_OVERRIDES: dict[str, str] = {
    "main_image": "Главное изображение",
    "gallery_images": "Галерея изображений",
    "article": "Артикул",
    "supplier_article": "Артикул поставщика",
    "internal_article": "Внутренний артикул",
    "image_url": "Ссылка на фото",
}


def with_ru_columns(df: pd.DataFrame, extra_map: dict[str, str] | None = None) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    mapping = dict(RU_COLUMN_MAP)
    if extra_map:
        mapping.update(extra_map)
    return df.rename(columns={c: mapping.get(c, c) for c in df.columns})


def humanize_attribute_code(code: str | None) -> str:
    text = str(code or "").strip()
    if not text:
        return ""
    if text in ATTRIBUTE_CODE_RU_OVERRIDES:
        return ATTRIBUTE_CODE_RU_OVERRIDES[text]
    if text.startswith("ozon_attr_"):
        attr_id = text.replace("ozon_attr_", "", 1)
        return f"Ozon атрибут ID {attr_id}"
    return " ".join(text.replace("_", " ").split()).capitalize()


def _short_text(value: object, max_len: int = 88) -> str:
    text = str(value or "").strip()
    if len(text) <= int(max_len):
        return text
    return text[: max(0, int(max_len) - 1)].rstrip() + "…"


def _parse_status_label(status: object) -> str:
    value = str(status or "").strip().lower()
    if value == "success":
        return "Успех"
    if value == "error":
        return "Ошибка"
    return "Не запускался"


def _supplier_url_shape(url: object) -> str:
    text = str(url or "").strip()
    if not text:
        return "none"
    try:
        if _supplier_parser is not None and hasattr(_supplier_parser, "_is_root_like_url") and _supplier_parser._is_root_like_url(text):
            return "root"
        if _supplier_parser is not None and hasattr(_supplier_parser, "_is_listing_like_url") and _supplier_parser._is_listing_like_url(text):
            return "listing"
    except Exception:
        pass
    return "direct"


def _parse_source_type_from_comment(comment: object) -> str:
    text = str(comment or "").strip()
    if not text:
        return ""
    match = re.search(r"\bsource=([a-z0-9_]+)\b", text)
    return str(match.group(1) or "").strip().lower() if match else ""


def _parser_stage_info(product_row) -> tuple[str, str]:
    supplier_url = str(product_row.get("supplier_url") or "").strip() if hasattr(product_row, "get") else ""
    parse_status = str(product_row.get("supplier_parse_status") or "").strip().lower() if hasattr(product_row, "get") else ""
    parse_comment = str(product_row.get("supplier_parse_comment") or "").strip() if hasattr(product_row, "get") else ""
    url_shape = _supplier_url_shape(supplier_url)
    source_type = _parse_source_type_from_comment(parse_comment)

    if not supplier_url:
        return "Нет supplier-домена", "Нет supplier-домена"
    if parse_status == "success":
        if source_type in {"web_search_fallback", "web_search_fallback_domain", "domain_search_fallback", "yandex_search_fallback", "ozon_search_fallback"}:
            return "Найдено через интернет", "Успех: web fallback"
        if "listing->product resolved" in parse_comment or url_shape in {"root", "listing"}:
            return "Найдено от домена поставщика", "Успех: supplier domain"
        return "Прямая карточка поставщика", "Успех: supplier page"
    if parse_status == "error":
        low_comment = parse_comment.lower()
        if "access blocked" in low_comment:
            return "Сайт блокирует парсер", "Сайт блокирует доступ"
        if "fallback_rejected=" in low_comment:
            return "Нерелевантный кандидат", "Не найден релевантный товар"
        if "не удалось получить полезные данные" in low_comment:
            return "Карточка не найдена", "Не найден релевантный товар"
        return "Ошибка parser-flow", "Ошибка parser-flow"
    if url_shape in {"root", "listing"}:
        return "Ожидает поиск по домену", "Нужен parser-run"
    return "Не запускался", "Нужен parser-run"


def _barcode_status_label(product_row) -> str:
    barcode = str(product_row.get("barcode") or "").strip() if hasattr(product_row, "get") else ""
    return "Есть" if barcode else "Нет"


SERVICE_SIGNAL_CODES = (
    "service_ai_verdict",
    "service_ai_confidence",
    "service_ai_summary",
    "service_ai_mode",
    "service_image_stage",
    "service_gallery_count",
    "service_gallery_missing_slots",
    "service_image_queue",
)


def _decode_service_attr_value(row) -> object:
    if row is None:
        return None
    if row["value_number"] is not None:
        return row["value_number"]
    if row["value_boolean"] is not None:
        return bool(row["value_boolean"])
    if row["value_json"] is not None:
        try:
            return json.loads(row["value_json"])
        except Exception:
            return row["value_json"]
    return row["value_text"]


def _load_service_signal_map(conn, product_ids: list[int], codes: tuple[str, ...] = SERVICE_SIGNAL_CODES) -> dict[int, dict[str, object]]:
    ids = [int(x) for x in product_ids if int(x) > 0]
    if not ids:
        return {}
    placeholders_ids = ",".join(["?"] * len(ids))
    placeholders_codes = ",".join(["?"] * len(codes))
    rows = conn.execute(
        f"""
        SELECT product_id, attribute_code, value_text, value_number, value_boolean, value_json
        FROM product_attribute_values
        WHERE product_id IN ({placeholders_ids})
          AND attribute_code IN ({placeholders_codes})
        """,
        tuple(ids) + tuple(codes),
    ).fetchall()
    result: dict[int, dict[str, object]] = {}
    for row in rows:
        pid = int(row["product_id"])
        result.setdefault(pid, {})[str(row["attribute_code"])] = _decode_service_attr_value(row)
    return result


def _load_latest_field_source_type_map(conn, product_ids: list[int], field_names: tuple[str, ...] = ("name", "description")) -> dict[int, dict[str, str]]:
    ids = [int(x) for x in product_ids if int(x) > 0]
    if not ids:
        return {}
    placeholders_ids = ",".join(["?"] * len(ids))
    placeholders_fields = ",".join(["?"] * len(field_names))
    rows = conn.execute(
        f"""
        SELECT s.product_id, s.field_name, s.source_type
        FROM product_data_sources s
        JOIN (
            SELECT product_id, field_name, MAX(id) AS max_id
            FROM product_data_sources
            WHERE product_id IN ({placeholders_ids})
              AND field_name IN ({placeholders_fields})
            GROUP BY product_id, field_name
        ) latest
          ON latest.max_id = s.id
        """,
        tuple(ids) + tuple(field_names),
    ).fetchall()
    result: dict[int, dict[str, str]] = {}
    for row in rows:
        pid = int(row["product_id"])
        result.setdefault(pid, {})[str(row["field_name"])] = str(row["source_type"] or "").strip().lower()
    return result


def _photo_count_for_product(product_row, service_state: dict[str, object] | None = None) -> int:
    state = service_state or {}
    count = _safe_int_id(state.get("service_gallery_count"))
    if count > 0:
        return count
    return 1 if str(product_row.get("image_url") or "").strip() else 0


def _ai_stage_info(
    product_row,
    service_state: dict[str, object] | None = None,
    latest_sources: dict[str, str] | None = None,
) -> tuple[str, str]:
    parse_status = str(product_row.get("supplier_parse_status") or "").strip().lower() if hasattr(product_row, "get") else ""
    state = service_state or {}
    sources = latest_sources or {}
    verdict = str(state.get("service_ai_verdict") or "").strip().lower()
    confidence = _safe_float_value(state.get("service_ai_confidence"), default=0.0)
    title_ai = str(sources.get("name") or "").strip().lower() == "ai"
    desc_ai = str(sources.get("description") or "").strip().lower() == "ai"

    if parse_status != "success":
        return "Ожидает parser", "Сначала parser"
    if verdict == "reject":
        return "AI отклонил parser result", "AI отклонил parser result"
    if verdict == "review" or (verdict == "accept" and confidence < 0.62):
        return "Нужна быстрая AI-проверка", "Низкая уверенность parser/AI"
    if title_ai and desc_ai:
        return "AI rewrite ready", "AI rewrite ready"
    if verdict == "accept":
        return "AI verified", "Нужен AI-run"
    return "Нужен AI-run", "Нужен AI-run"


def _image_stage_info(product_row, service_state: dict[str, object] | None = None) -> tuple[str, str, int]:
    state = service_state or {}
    photo_count = _photo_count_for_product(product_row, state)
    main_image = bool(str(product_row.get("image_url") or "").strip()) if hasattr(product_row, "get") else False
    stage = str(state.get("service_image_stage") or "").strip().lower()
    queue = str(state.get("service_image_queue") or "").strip()
    if not stage:
        if not main_image:
            return "Нет главного фото", "Нет фото", photo_count
        if photo_count < 3:
            return "Фото меньше 3", "Фото меньше 3", photo_count
        if photo_count <= 5:
            return "Фото готовы", "Фото готовы", photo_count
        return "Расширенная галерея", "Фото готовы", photo_count
    if stage == "no_main_image":
        return "Нет главного фото", queue or "Нет фото", photo_count
    if stage == "under_min":
        return "Фото меньше 3", queue or "Фото меньше 3", photo_count
    if stage == "target_ready":
        return "Фото готовы", queue or "Фото готовы", photo_count
    if stage == "rich_gallery":
        return "Расширенная галерея", queue or "Фото готовы", photo_count
    return humanize_attribute_code(stage), queue or "Фото готовы", photo_count


def _operational_queue_label(
    product_row,
    service_state: dict[str, object] | None = None,
    latest_sources: dict[str, str] | None = None,
) -> str:
    has_ozon = bool(
        _safe_int_id(product_row.get("ozon_description_category_id")) > 0
        and _safe_int_id(product_row.get("ozon_type_id")) > 0
    ) if hasattr(product_row, "get") else False
    parse_status = str(product_row.get("supplier_parse_status") or "").strip().lower() if hasattr(product_row, "get") else ""
    has_barcode = bool(str(product_row.get("barcode") or "").strip()) if hasattr(product_row, "get") else False
    _, parser_queue = _parser_stage_info(product_row)
    _, ai_queue = _ai_stage_info(product_row, service_state, latest_sources)
    _, image_queue, photo_count = _image_stage_info(product_row, service_state)

    if not has_ozon:
        return "Нет Ozon-категории"
    if parse_status != "success":
        return parser_queue
    if photo_count <= 0:
        return "Нет фото"
    if photo_count < 3:
        return image_queue
    if not has_barcode:
        return "Нет штрихкода"
    if ai_queue in {"AI отклонил parser result", "Низкая уверенность parser/AI", "Нужен AI-run"}:
        return ai_queue
    return "Готово к AI/клиенту"


def _product_core_fill_stats(product_row) -> tuple[int, int]:
    fields = [
        "name",
        "brand",
        "description",
        "image_url",
        "weight",
        "length",
        "width",
        "height",
        "tnved_code",
    ]
    filled = 0
    for field_name in fields:
        value = product_row.get(field_name) if hasattr(product_row, "get") else None
        if value not in (None, "", 0, 0.0):
            filled += 1
    return filled, len(fields)


def _product_stage_label(product_row) -> str:
    has_ozon = bool(
        _safe_int_id(product_row.get("ozon_description_category_id")) > 0
        and _safe_int_id(product_row.get("ozon_type_id")) > 0
    ) if hasattr(product_row, "get") else False
    parse_status = str(product_row.get("supplier_parse_status") or "").strip().lower() if hasattr(product_row, "get") else ""
    filled, total = _product_core_fill_stats(product_row)
    if not has_ozon:
        return "Нужна Ozon-категория"
    if parse_status == "success" and filled >= max(6, total - 1):
        return "Почти готова"
    if parse_status == "success":
        return "Нужно добить атрибуты"
    if parse_status == "error":
        return "Нужна ручная проверка"
    return "Готова к заполнению"


def _readiness_value_present(value: object) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return value not in (0, 0.0)


def _safe_int_id(value: object) -> int:
    try:
        if value is None:
            return 0
        if pd.isna(value):
            return 0
    except Exception:
        pass
    try:
        return int(value or 0)
    except Exception:
        return 0


def _safe_float_value(value: object, default: float = 0.0) -> float:
    try:
        if value is None:
            return float(default)
        if pd.isna(value):
            return float(default)
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        return float(default)


def _clamp_unit_confidence(value: object) -> float:
    return max(0.0, min(1.0, _safe_float_value(value, default=0.0)))


def _detmir_confidence_from_match_score(score: object) -> float | None:
    normalized = _clamp_unit_confidence(_safe_float_value(score, default=0.0) / 10.0)
    return normalized if normalized > 0 else None


def compute_quick_ozon_readiness(conn, product_row) -> dict[str, object]:
    product_id = _safe_int_id(product_row.get("id")) if hasattr(product_row, "get") else 0
    desc_id = _safe_int_id(product_row.get("ozon_description_category_id")) if hasattr(product_row, "get") else 0
    type_id = _safe_int_id(product_row.get("ozon_type_id")) if hasattr(product_row, "get") else 0
    if product_id <= 0 or desc_id <= 0 or type_id <= 0:
        return {"status": "no_category", "required_total": 0, "required_filled": 0, "readiness_pct": 0, "missing": 0}

    category_code = f"ozon:{desc_id}:{type_id}"
    req_rows = conn.execute(
        """
        SELECT attribute_code, is_required
        FROM channel_attribute_requirements
        WHERE channel_code = 'ozon'
          AND category_code = ?
        """,
        (category_code,),
    ).fetchall()
    required_codes = [str(row["attribute_code"]) for row in req_rows if int(row["is_required"] or 0) == 1 and str(row["attribute_code"] or "").strip()]
    if not required_codes:
        return {"status": "no_requirements", "required_total": 0, "required_filled": 0, "readiness_pct": 0, "missing": 0}

    value_map = build_product_value_map(conn, product_id)
    filled = 0
    for code in required_codes:
        if _readiness_value_present(value_map.get(code)):
            filled += 1
    required_total = int(len(required_codes))
    readiness_pct = round((filled / required_total) * 100) if required_total else 0
    return {
        "status": "ok",
        "required_total": required_total,
        "required_filled": int(filled),
        "readiness_pct": int(readiness_pct),
        "missing": int(required_total - filled),
    }


def compute_quick_detmir_readiness(conn, product_row) -> dict[str, object]:
    product_id = _safe_int_id(product_row.get("id")) if hasattr(product_row, "get") else 0
    category_id = _safe_int_id(product_row.get("detmir_category_id")) if hasattr(product_row, "get") else 0
    if product_id <= 0 or category_id <= 0:
        return {
            "status": "no_category",
            "required_total": 0,
            "required_filled": 0,
            "readiness_pct": 0,
            "blockers": 1,
            "warnings": 0,
            "photos_count": 0,
            "dictionary_unmatched": 0,
        }
    try:
        result = analyze_product_detmir_readiness(conn, product_id=product_id, category_id=category_id)
    except Exception:
        return {
            "status": "error",
            "required_total": 0,
            "required_filled": 0,
            "readiness_pct": 0,
            "blockers": 1,
            "warnings": 0,
            "photos_count": 0,
            "dictionary_unmatched": 0,
        }
    summary = result.get("summary") or {}
    return {
        "status": str(summary.get("status") or "ok"),
        "required_total": int(summary.get("required_total") or 0),
        "required_filled": int(summary.get("required_filled") or 0),
        "readiness_pct": int(summary.get("readiness_pct") or 0),
        "blockers": int(summary.get("blockers") or 0),
        "warnings": int(summary.get("warnings") or 0),
        "photos_count": int(summary.get("photos_count") or 0),
        "dictionary_unmatched": int(summary.get("dictionary_unmatched") or 0),
    }


def compute_best_template_profile_readiness(conn, product_row, limit_profiles: int = 40) -> dict[str, object]:
    product_id = _safe_int_id(product_row.get("id")) if hasattr(product_row, "get") else 0
    if product_id <= 0:
        return {"profiles_total": 0, "best_readiness_pct": 0}

    category_code = ""
    desc_id = _safe_int_id(product_row.get("ozon_description_category_id")) if hasattr(product_row, "get") else 0
    type_id = _safe_int_id(product_row.get("ozon_type_id")) if hasattr(product_row, "get") else 0
    if desc_id > 0 and type_id > 0:
        category_code = f"ozon:{desc_id}:{type_id}"

    all_profiles = list_template_profiles(conn, channel_code=None)
    candidate_profiles = [
        profile
        for profile in all_profiles
        if str(profile.get("category_code") or "").strip() == str(category_code or "").strip()
    ] if category_code else all_profiles
    candidate_profiles = candidate_profiles[: max(1, int(limit_profiles))]
    if not candidate_profiles:
        return {"profiles_total": 0, "best_readiness_pct": 0}

    value_map = build_product_value_map(conn, product_id)
    best: dict[str, object] | None = None
    for profile in candidate_profiles:
        columns = get_template_profile_columns(conn, int(profile["id"]))
        matched_columns = [
            col
            for col in columns
            if str(col.get("source_type") or "").strip() in {"column", "attribute"}
            and str(col.get("source_name") or "").strip()
        ]
        if not matched_columns:
            continue
        filled = 0
        for col in matched_columns:
            if _readiness_value_present(value_map.get(str(col.get("source_name") or "").strip())):
                filled += 1
        matched_total = int(len(matched_columns))
        readiness_pct = round((filled / matched_total) * 100) if matched_total else 0
        row = {
            "profile_id": int(profile["id"]),
            "profile_name": str(profile.get("profile_name") or ""),
            "channel_code": str(profile.get("channel_code") or ""),
            "matched_columns": matched_total,
            "filled_columns": int(filled),
            "missing_columns": int(matched_total - filled),
            "readiness_pct": int(readiness_pct),
            "profiles_total": int(len(candidate_profiles)),
        }
        if best is None or (int(row["readiness_pct"]), int(row["filled_columns"])) > (int(best["readiness_pct"]), int(best["filled_columns"])):
            best = row

    if not best:
        return {"profiles_total": int(len(candidate_profiles)), "best_readiness_pct": 0}
    return best


def build_catalog_operational_view(
    conn,
    df: pd.DataFrame,
    service_signals: dict[int, dict[str, object]] | None = None,
    latest_sources: dict[int, dict[str, str]] | None = None,
) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    service_map = service_signals or {}
    source_map = latest_sources or {}
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        product_id = _safe_int_id(row_dict.get("id"))
        state = service_map.get(product_id, {})
        latest = source_map.get(product_id, {})
        filled, total = _product_core_fill_stats(row_dict)
        parser_stage, _ = _parser_stage_info(row_dict)
        ai_stage, _ = _ai_stage_info(row_dict, state, latest)
        image_stage, _, photo_count = _image_stage_info(row_dict, state)
        rows.append(
            {
                "article": row_dict.get("article") or row_dict.get("supplier_article") or row_dict.get("internal_article"),
                "name": _short_text(row_dict.get("name"), 72),
                "supplier_name": row_dict.get("supplier_name"),
                "ozon_category_path": _short_text(row_dict.get("ozon_category_path") or row_dict.get("category"), 68),
                "stage": _product_stage_label(row_dict),
                "parser_stage": parser_stage,
                "ai_stage": ai_stage,
                "image_stage": image_stage,
                "queue": _operational_queue_label(row_dict, state, latest),
                "supplier_parse_status": _parse_status_label(row_dict.get("supplier_parse_status")),
                "photo_status": "Есть" if photo_count > 0 else "Нет",
                "photo_count": int(photo_count),
                "barcode_status": _barcode_status_label(row_dict),
                "fill_score": f"{filled}/{total}",
                "updated_at": row_dict.get("updated_at"),
            }
        )
    return pd.DataFrame(rows)


def format_source_name_ui(
    source_name: str | None,
    source_type: str | None = None,
    attr_name_map: dict[str, str] | None = None,
) -> str:
    value = str(source_name or "").strip()
    if not value:
        return ""

    stype = str(source_type or "").strip().lower()
    if stype == "attribute":
        if attr_name_map and value in attr_name_map and str(attr_name_map[value]).strip():
            return str(attr_name_map[value]).strip()
        return humanize_attribute_code(value)

    if stype == "column":
        if value in RU_COLUMN_MAP:
            return str(RU_COLUMN_MAP[value])
        return humanize_attribute_code(value)

    if attr_name_map and value in attr_name_map and str(attr_name_map[value]).strip():
        return str(attr_name_map[value]).strip()
    return humanize_attribute_code(value)


_MATCH_STOPWORDS = {
    "велосипед", "bike", "товар", "product", "new", "новый", "новинка", "для", "and", "the", "with",
    "комплект", "set", "шт", "pcs", "item", "model", "модель",
}


def _compact_for_match(value: object) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def _name_tokens_for_match(value: object) -> set[str]:
    tokens = set(re.findall(r"[a-zа-я0-9]{3,}", str(value or "").lower()))
    return {t for t in tokens if t not in _MATCH_STOPWORDS}


def _collect_target_codes(product_row: dict) -> list[str]:
    out: list[str] = []
    for key in ("supplier_article", "article", "internal_article", "barcode"):
        raw = str(product_row.get(key) or "").strip()
        compact = _compact_for_match(raw)
        if len(compact) >= 4 and compact not in out:
            out.append(compact)
    return out


def _extract_parsed_article_candidates(parsed: dict) -> list[str]:
    values: list[str] = []
    attrs = parsed.get("attributes") or {}
    for k, v in attrs.items():
        key_l = str(k or "").lower()
        if any(x in key_l for x in ["артикул", "sku", "article", "код"]):
            values.append(str(v or ""))
    values.append(str(parsed.get("name") or ""))
    values.append(str(parsed.get("title") or ""))
    out: list[str] = []
    for val in values:
        compact = _compact_for_match(val)
        if len(compact) >= 4 and compact not in out:
            out.append(compact)
    return out


def _is_parsed_result_relevant(product_row: dict, parsed: dict, source_url: str, settings: dict[str, object]) -> tuple[bool, str]:
    target_codes = _collect_target_codes(product_row)
    parsed_codes = _extract_parsed_article_candidates(parsed)
    combined_text = " ".join(
        [
            str(source_url or ""),
            str(parsed.get("resolved_url") or ""),
            str(parsed.get("title") or ""),
            str(parsed.get("name") or ""),
            str(parsed.get("description") or "")[:500],
        ]
    )
    combined_compact = _compact_for_match(combined_text)

    code_match = False
    for code in target_codes:
        if code in combined_compact:
            code_match = True
            break
        if any(code in parsed_code or parsed_code in code for parsed_code in parsed_codes):
            code_match = True
            break

    product_brand = str(product_row.get("brand") or "").strip().lower()
    parsed_brand = str(parsed.get("brand") or "").strip().lower()
    brand_match = bool(product_brand and (product_brand in parsed_brand or parsed_brand in product_brand))

    source_tokens = _name_tokens_for_match(product_row.get("name"))
    parsed_tokens = _name_tokens_for_match(parsed.get("name"))
    overlap = len(source_tokens.intersection(parsed_tokens))

    fallback_score = float(parsed.get("fallback_score") or 0.0)
    min_overlap = int(settings.get("min_name_overlap", 2))
    min_fallback_score = float(settings.get("min_fallback_score", 3.0))
    require_article_match = bool(settings.get("require_article_match", True))

    if target_codes and code_match:
        return True, "article_or_code_match"
    if overlap >= min_overlap and (brand_match or fallback_score >= min_fallback_score):
        return True, f"name_overlap={overlap}"
    if (not require_article_match) and fallback_score >= (min_fallback_score + 1.0):
        return True, f"score={fallback_score:.2f}"
    return False, f"rejected: code_match={code_match}, overlap={overlap}, brand_match={brand_match}, score={fallback_score:.2f}"


_ATTR_MATCH_STOPWORDS = {
    "ozon",
    "товар",
    "товара",
    "характеристика",
    "характеристики",
    "параметр",
    "параметры",
    "значение",
    "для",
    "и",
    "в",
    "на",
    "с",
    "the",
    "with",
    "for",
    "item",
    "product",
}


def _is_empty_like(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, (list, tuple, dict, set)):
        return len(value) == 0
    text = str(value).strip().lower()
    return text in {"", "none", "null", "nan", "-", "—"}


def _normalize_attr_text(value: object) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"[\(\)\[\],;:]+", " ", text)
    return " ".join(text.split())


def _tokenize_attr_text(value: object) -> set[str]:
    tokens = set(re.findall(r"[a-zа-я0-9]{2,}", _normalize_attr_text(value)))
    return {t for t in tokens if t not in _ATTR_MATCH_STOPWORDS}


def _extract_number_like(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().lower().replace(",", ".")
    m = re.search(r"(-?\d+(?:\.\d+)?)", text)
    if not m:
        return None
    try:
        return float(m.group(1))
    except Exception:
        return None


def _coerce_value_for_attr_type(value: object, data_type: str | None) -> object:
    kind = str(data_type or "text").strip().lower()
    if kind == "number":
        num = _extract_number_like(value)
        return num if num is not None else value
    if kind == "boolean":
        text = str(value or "").strip().lower()
        if text in {"1", "true", "yes", "да", "есть"}:
            return True
        if text in {"0", "false", "no", "нет", "none"}:
            return False
    return value


def _parse_domain_list(value: object) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    out: list[str] = []
    for part in re.split(r"[,\n;]+", text):
        dom = str(part or "").strip().lower()
        if not dom:
            continue
        dom = re.sub(r"^https?://", "", dom)
        dom = dom.split("/")[0].strip()
        dom = dom.replace("www.", "")
        if not dom or "." not in dom:
            continue
        if dom not in out:
            out.append(dom)
    return out


def _row_value(row: object, key: str, default: object = None) -> object:
    if row is None:
        return default
    try:
        if hasattr(row, "get"):
            value = row.get(key, default)
            return default if value is None else value
    except Exception:
        pass
    try:
        if hasattr(row, "keys") and key in row.keys():
            value = row[key]
            return default if value is None else value
    except Exception:
        pass
    try:
        value = row[key]
        return default if value is None else value
    except Exception:
        return default


def _infer_attr_semantic_key(attr_name: str) -> str:
    key = _normalize_attr_text(attr_name)
    if any(x in key for x in ("вес брутто", "gross")):
        return "gross_weight"
    if "вес" in key:
        return "weight"
    if any(x in key for x in ("длина упаков", "длина короб", "package length")):
        return "package_length"
    if any(x in key for x in ("ширина упаков", "ширина короб", "package width")):
        return "package_width"
    if any(x in key for x in ("высота упаков", "высота короб", "package height")):
        return "package_height"
    if "длина" in key:
        return "length"
    if "ширина" in key:
        return "width"
    if any(x in key for x in ("высота", "глубина", "depth")):
        return "height"
    if any(x in key for x in ("бренд", "торговая марка", "manufacturer", "brand", "производитель")):
        return "brand"
    if any(x in key for x in ("описан", "description")):
        return "description"
    if any(x in key for x in ("артикул", "sku", "vendor code", "код товара", "модель")):
        return "article"
    if any(x in key for x in ("фото", "изображ", "image", "картин")):
        return "image_url"
    if any(x in key for x in ("цвет", "color")):
        return "color"
    if any(x in key for x in ("материал", "material")):
        return "material"
    return ""


def _build_parsed_candidates(parsed: dict, product_row: dict) -> tuple[list[dict], dict[str, object]]:
    candidates: list[dict] = []
    scalar_map: dict[str, object] = {
        "name": parsed.get("name") or product_row.get("name"),
        "brand": parsed.get("brand") or product_row.get("brand"),
        "description": parsed.get("description"),
        "image_url": parsed.get("image_url"),
        "weight": parsed.get("weight"),
        "gross_weight": parsed.get("gross_weight"),
        "length": parsed.get("length"),
        "width": parsed.get("width"),
        "height": parsed.get("height"),
        "package_length": parsed.get("package_length"),
        "package_width": parsed.get("package_width"),
        "package_height": parsed.get("package_height"),
        "article": product_row.get("supplier_article") or product_row.get("article") or product_row.get("internal_article"),
        "barcode": product_row.get("barcode"),
        "category": parsed.get("category") or product_row.get("category"),
        "color": None,
        "material": None,
    }
    for raw_key, raw_value in (parsed.get("attributes") or {}).items():
        key = str(raw_key or "").strip()
        if not key or _is_empty_like(raw_value):
            continue
        key_norm = _normalize_attr_text(key)
        key_tokens = _tokenize_attr_text(key_norm)
        value = raw_value
        candidates.append(
            {
                "key": key,
                "key_norm": key_norm,
                "key_tokens": key_tokens,
                "value": value,
            }
        )
        key_semantic = _infer_attr_semantic_key(key)
        if key_semantic and _is_empty_like(scalar_map.get(key_semantic)):
            scalar_map[key_semantic] = value

    for semantic_key, semantic_value in scalar_map.items():
        if _is_empty_like(semantic_value):
            continue
        candidates.append(
            {
                "key": f"scalar:{semantic_key}",
                "key_norm": _normalize_attr_text(str(semantic_key)),
                "key_tokens": _tokenize_attr_text(str(semantic_key)),
                "value": semantic_value,
            }
        )
    return candidates, scalar_map


def _build_product_state_candidates(conn, product_row: dict) -> tuple[list[dict], dict[str, object]]:
    product_id = _safe_int_id(product_row.get("id"))
    candidates: list[dict] = []
    scalar_map: dict[str, object] = {
        "name": product_row.get("name"),
        "brand": product_row.get("brand"),
        "description": product_row.get("description"),
        "image_url": product_row.get("image_url"),
        "weight": product_row.get("weight"),
        "gross_weight": product_row.get("gross_weight"),
        "length": product_row.get("length"),
        "width": product_row.get("width"),
        "height": product_row.get("height"),
        "package_length": product_row.get("package_length"),
        "package_width": product_row.get("package_width"),
        "package_height": product_row.get("package_height"),
        "article": product_row.get("supplier_article") or product_row.get("article") or product_row.get("internal_article"),
        "barcode": product_row.get("barcode"),
        "category": product_row.get("category"),
        "color": None,
        "material": None,
    }
    if product_id > 0:
        defs = list_attribute_definitions(conn)
        defs_map = {str(d.get("code") or ""): d for d in defs}
        for row in get_product_attribute_values(conn, product_id):
            code = str(row.get("attribute_code") or "").strip()
            if not code:
                continue
            value = row.get("value")
            if _is_empty_like(value):
                continue
            attr_def = defs_map.get(code, {})
            attr_name = str(attr_def.get("name") or humanize_attribute_code(code))
            key_norm = _normalize_attr_text(attr_name)
            key_tokens = _tokenize_attr_text(key_norm)
            candidates.append(
                {
                    "key": attr_name,
                    "key_norm": key_norm,
                    "key_tokens": key_tokens,
                    "value": value,
                }
            )
            code_norm = _normalize_attr_text(code)
            code_tokens = _tokenize_attr_text(code_norm)
            candidates.append(
                {
                    "key": code,
                    "key_norm": code_norm,
                    "key_tokens": code_tokens,
                    "value": value,
                }
            )
            semantic_key = _infer_attr_semantic_key(attr_name) or _infer_attr_semantic_key(code)
            if semantic_key and _is_empty_like(scalar_map.get(semantic_key)):
                scalar_map[semantic_key] = value

    for semantic_key, semantic_value in scalar_map.items():
        if _is_empty_like(semantic_value):
            continue
        candidates.append(
            {
                "key": f"scalar:{semantic_key}",
                "key_norm": _normalize_attr_text(str(semantic_key)),
                "key_tokens": _tokenize_attr_text(str(semantic_key)),
                "value": semantic_value,
            }
        )
    return candidates, scalar_map


def _resolve_best_value_for_attr(attr_name: str, candidates: list[dict], scalar_map: dict[str, object]) -> tuple[object, str, float] | None:
    attr_norm = _normalize_attr_text(attr_name)
    attr_tokens = _tokenize_attr_text(attr_norm)
    semantic = _infer_attr_semantic_key(attr_name)
    if semantic and not _is_empty_like(scalar_map.get(semantic)):
        return scalar_map.get(semantic), f"semantic:{semantic}", 10.0

    best: tuple[object, str, float] | None = None
    for c in candidates:
        key_norm = str(c.get("key_norm") or "")
        key_tokens = set(c.get("key_tokens") or set())
        value = c.get("value")
        if _is_empty_like(value):
            continue
        score = 0.0
        if key_norm == attr_norm:
            score += 7.0
        elif key_norm and (key_norm in attr_norm or attr_norm in key_norm):
            score += 4.0
        if attr_tokens and key_tokens:
            overlap = len(attr_tokens & key_tokens) / max(len(attr_tokens), 1)
            score += overlap * 4.0
        cand_semantic = _infer_attr_semantic_key(key_norm)
        if semantic and cand_semantic and semantic == cand_semantic:
            score += 3.0
        if best is None or score > best[2]:
            best = (value, str(c.get("key") or ""), float(score))
    if best and float(best[2]) >= 2.5:
        return best
    return None


def _load_target_attribute_rows(
    conn,
    channel_code: str,
    category_code: str | None = None,
) -> list[dict]:
    req_rows = list_channel_requirements(conn, channel_code=channel_code, category_code=category_code or None)
    rule_rows = list_channel_mapping_rules(conn, channel_code=channel_code, category_code=category_code or None)
    defs = list_attribute_definitions(conn)
    defs_map = {str(d.get("code") or ""): d for d in defs}
    target_codes: set[str] = set()
    for row in req_rows:
        code = str(row.get("attribute_code") or "").strip()
        if code:
            target_codes.add(code)
    for row in rule_rows:
        if str(row.get("source_type") or "") == "attribute":
            code = str(row.get("source_name") or "").strip()
            if code:
                target_codes.add(code)
    rows: list[dict] = []
    for code in sorted(target_codes):
        attr_def = defs_map.get(code, {})
        rows.append(
            {
                "attribute_code": code,
                "attribute_name": str(attr_def.get("name") or humanize_attribute_code(code)),
                "data_type": str(attr_def.get("data_type") or "text"),
                "is_required": int(next((int(r.get("is_required") or 0) for r in req_rows if str(r.get("attribute_code") or "") == code), 0)),
            }
        )
    return rows


def _fill_channel_attrs_from_product_state(
    conn,
    product_row: dict,
    channel_code: str,
    category_code: str | None = None,
    source_type: str = "derived_from_master",
    source_url: str = "product_state",
    force: bool = False,
    target_channel_code: str | None = None,
) -> dict[str, int]:
    product_id = _safe_int_id(product_row.get("id"))
    if product_id <= 0 or not str(channel_code or "").strip():
        return {"saved": 0, "skipped": 0, "targets": 0}
    target_rows = _load_target_attribute_rows(conn, str(channel_code), category_code=category_code)
    if not target_rows:
        return {"saved": 0, "skipped": 0, "targets": 0}
    candidates, scalar_map = _build_product_state_candidates(conn, product_row)
    saved = 0
    skipped = 0
    for row in target_rows:
        code = str(row.get("attribute_code") or "").strip()
        attr_name = str(row.get("attribute_name") or code)
        if not code:
            continue
        if str(channel_code).strip() == "ozon" and code in AUTO_MANUAL_ONLY_OZON_CODES:
            skipped += 1
            continue
        resolved = _resolve_best_value_for_attr(attr_name, candidates, scalar_map)
        if not resolved:
            continue
        value, matched_by, score = resolved
        if _is_empty_like(value):
            continue
        field_name = f"attr:{code}"
        effective_field_source = str(source_type or "derived_from_master")
        if not can_overwrite_field(conn, product_id, field_name, effective_field_source, force=force):
            skipped += 1
            continue
        coerced = _coerce_value_for_attr_type(value, str(_row_value(row, "data_type", "text") or "text"))
        try:
            set_product_attribute_value(
                conn,
                product_id,
                code,
                coerced,
                channel_code=target_channel_code,
            )
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name=field_name,
                source_type=effective_field_source,
                source_value_raw=coerced,
                source_url=f"{source_url} | match={matched_by}",
                confidence=min(0.95, max(0.45, float(score) / 10.0)),
            )
            saved += 1
        except Exception:
            skipped += 1
    return {"saved": int(saved), "skipped": int(skipped), "targets": int(len(target_rows))}


def _fill_ozon_attrs_from_parsed(
    conn,
    product_row: dict,
    parsed: dict,
    source_type: str,
    source_url: str,
    force: bool = False,
) -> dict[str, int]:
    product_id = _safe_int_id(product_row.get("id"))
    desc_id = _safe_int_id(product_row.get("ozon_description_category_id"))
    type_id = _safe_int_id(product_row.get("ozon_type_id"))
    if product_id <= 0 or desc_id <= 0 or type_id <= 0:
        return {"saved": 0, "skipped": 0}

    category_code = f"ozon:{desc_id}:{type_id}"
    req_rows = conn.execute(
        """
        SELECT
            car.attribute_code,
            ad.name AS attribute_name,
            ad.data_type
        FROM channel_attribute_requirements car
        JOIN attribute_definitions ad
          ON ad.code = car.attribute_code
        WHERE car.channel_code = 'ozon'
          AND car.category_code = ?
        ORDER BY car.is_required DESC, ad.name
        """,
        (category_code,),
    ).fetchall()
    if not req_rows:
        try:
            ensure_ozon_requirements_for_product_category(conn, desc_id, type_id)
        except Exception:
            pass
        req_rows = conn.execute(
            """
            SELECT
                car.attribute_code,
                ad.name AS attribute_name,
                ad.data_type
            FROM channel_attribute_requirements car
            JOIN attribute_definitions ad
              ON ad.code = car.attribute_code
            WHERE car.channel_code = 'ozon'
              AND car.category_code = ?
            ORDER BY car.is_required DESC, ad.name
            """,
            (category_code,),
        ).fetchall()
    if not req_rows:
        return {"saved": 0, "skipped": 0}

    candidates, scalar_map = _build_parsed_candidates(parsed, product_row)
    saved = 0
    skipped = 0
    for row in req_rows:
        code = str(row["attribute_code"] or "").strip()
        if not code:
            continue
        if code in AUTO_MANUAL_ONLY_OZON_CODES:
            skipped += 1
            continue
        resolved = _resolve_best_value_for_attr(str(row["attribute_name"] or code), candidates, scalar_map)
        if not resolved:
            continue
        value, matched_by, score = resolved
        if _is_empty_like(value):
            continue
        field_name = f"attr:{code}"
        if not can_overwrite_field(conn, product_id, field_name, source_type, force=force):
            skipped += 1
            continue
        coerced = _coerce_value_for_attr_type(value, str(_row_value(row, "data_type", "text") or "text"))
        try:
            set_product_attribute_value(conn, product_id, code, coerced)
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name=field_name,
                source_type=source_type,
                source_value_raw=coerced,
                source_url=f"{source_url} | match={matched_by}",
                confidence=min(0.95, max(0.45, float(score) / 10.0)),
            )
            saved += 1
        except Exception:
            skipped += 1
    return {"saved": int(saved), "skipped": int(skipped)}


AUTO_UNSAFE_OZON_SOURCE_TYPES = {
    "derived_from_master",
    "supplier_page",
    "web_search_fallback",
    "web_search_fallback_domain",
    "ozon_search_fallback",
    "yandex_search_fallback",
    "domain_search_fallback",
    "name_category_inference",
    "category_stats_fallback",
    "category_defaults_fallback",
}


def _cleanup_unsafe_ozon_autofill_values(conn, product_id: int) -> int:
    product_id = _safe_int_id(product_id)
    if product_id <= 0:
        return 0
    cleared = 0
    for code in sorted(AUTO_MANUAL_ONLY_OZON_CODES):
        field_name = f"attr:{code}"
        latest = get_latest_field_source(conn, product_id, field_name)
        if not latest:
            continue
        if int(latest.get("is_manual") or 0) == 1:
            continue
        latest_source_type = str(latest.get("source_type") or "").strip()
        if latest_source_type not in AUTO_UNSAFE_OZON_SOURCE_TYPES:
            continue
        current_rows = [
            row
            for row in get_product_attribute_values(conn, product_id)
            if str(row.get("attribute_code") or "").strip() == code
        ]
        if not current_rows:
            continue
        has_value = any(not _is_empty_like(row.get("value")) for row in current_rows)
        if not has_value:
            continue
        set_product_attribute_value(conn, product_id, code, None)
        save_field_source(
            conn=conn,
            product_id=product_id,
            field_name=field_name,
            source_type="auto_cleanup",
            source_value_raw=None,
            source_url=f"cleanup:{latest_source_type}",
            confidence=1.0,
            is_manual=False,
        )
        cleared += 1
    return int(cleared)


def _infer_dimension_heuristics(product_row: dict) -> dict[str, Any]:
    name = str(product_row.get("name") or "").lower()
    category_text = " ".join(
        [
            str(product_row.get("category") or ""),
            str(product_row.get("base_category") or ""),
            str(product_row.get("subcategory") or ""),
            str(product_row.get("ozon_category_path") or ""),
        ]
    ).lower()
    wheel = product_row.get("wheel_diameter_inch")
    if wheel in (None, "", 0, 0.0):
        try:
            inferred = infer_category_fields({"name": product_row.get("name")})
            wheel = inferred.get("wheel_diameter_inch")
        except Exception:
            wheel = None
    try:
        wheel_val = float(wheel) if wheel not in (None, "", 0, 0.0) else None
    except Exception:
        wheel_val = None

    is_bike = any(token in f"{name} {category_text}" for token in ["велосипед", "bike", "bicycle"])
    if is_bike:
        # Практические ориентиры для закрытия пустых логистических данных по велосипедам.
        if wheel_val is None:
            wheel_val = 27.5
        if wheel_val >= 29:
            length, width, height, weight = 138.0, 82.0, 22.0, 15.5
        elif wheel_val >= 27.5:
            length, width, height, weight = 130.0, 80.0, 20.0, 14.0
        elif wheel_val >= 26:
            length, width, height, weight = 126.0, 78.0, 20.0, 13.8
        elif wheel_val >= 24:
            length, width, height, weight = 118.0, 70.0, 18.0, 12.3
        elif wheel_val >= 20:
            length, width, height, weight = 108.0, 64.0, 18.0, 10.9
        else:
            length, width, height, weight = 92.0, 56.0, 16.0, 8.3
        return {
            "found": True,
            "scope": "bike_heuristics",
            "values": {
                "length": length,
                "width": width,
                "height": height,
                "weight": weight,
                "package_length": round(length + 4.0, 2),
                "package_width": round(width + 4.0, 2),
                "package_height": round(height + 3.0, 2),
                "gross_weight": round(weight + 1.8, 2),
            },
        }

    return {"found": False, "scope": None, "values": {}}


def _normalize_media_urls(values: list[object]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in values:
        url = str(raw or "").strip()
        if not url:
            continue
        url = normalize_media_reference(url) or ""
        if not url:
            continue
        if url in seen:
            continue
        seen.add(url)
        out.append(url)
    return out


def _looks_like_local_media_path(value: str | None) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    low = text.lower()
    if text.startswith("\\\\") and re.search(r"\.(jpg|jpeg|png|webp|gif)$", low):
        return True
    if re.match(r"^[a-zA-Z]:\\", text) and re.search(r"\.(jpg|jpeg|png|webp|gif)$", low):
        return True
    return False


def _normalize_local_media_path_to_public(value: str | None, public_base_url: str | None = None) -> str | None:
    text = str(value or "").strip()
    base = str(public_base_url or "").strip().rstrip("/")
    if not text or not base or not _looks_like_local_media_path(text):
        return None
    normalized = text.replace("/", "\\")
    if normalized.startswith("\\\\"):
        parts = [p for p in normalized.split("\\") if p]
        rel_parts = parts[1:] if len(parts) >= 2 else parts
    elif re.match(r"^[a-zA-Z]:\\", normalized):
        rel_parts = [p for p in normalized[3:].split("\\") if p]
    else:
        rel_parts = [p for p in normalized.split("\\") if p]
    if not rel_parts:
        return None
    rel_path = posixpath.join(*rel_parts)
    return f"{base}/{rel_path}"


def normalize_media_reference(value: str | None, public_base_url: str | None = None) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.startswith("//"):
        return "https:" + text
    if text.lower().startswith(("http://", "https://")):
        return text
    if re.match(r"^[a-z0-9][a-z0-9\.\-]+\.[a-z]{2,}.*$", text.lower()) and " " not in text:
        return f"https://{text}"
    return _normalize_local_media_path_to_public(text, public_base_url=public_base_url)


def _parse_gallery_value(raw_value: object, public_base_url: str | None = None) -> list[str]:
    if raw_value is None:
        return []
    if isinstance(raw_value, (list, tuple, set)):
        return _normalize_media_urls([normalize_media_reference(v, public_base_url=public_base_url) or v for v in list(raw_value)])
    text = str(raw_value).strip()
    if not text:
        return []
    try:
        loaded = json.loads(text)
        if isinstance(loaded, list):
            return _normalize_media_urls([normalize_media_reference(v, public_base_url=public_base_url) or v for v in loaded])
        if isinstance(loaded, str):
            text = loaded
    except Exception:
        pass
    parts = re.split(r"[\n,;]+", text)
    return _normalize_media_urls([normalize_media_reference(v, public_base_url=public_base_url) or v for v in parts])


def _collect_product_gallery_urls(conn, product_id: int, fallback_image_url: str | None = None, public_base_url: str | None = None) -> list[str]:
    rows = conn.execute(
        """
        SELECT attribute_code, value_text, value_json
        FROM product_attribute_values
        WHERE product_id = ?
          AND attribute_code IN ('main_image', 'gallery_images')
        ORDER BY id DESC
        """,
        (int(product_id),),
    ).fetchall()
    values: list[object] = []
    for row in rows:
        if row["attribute_code"] == "gallery_images":
            if row["value_json"] not in (None, ""):
                values.extend(_parse_gallery_value(row["value_json"], public_base_url=public_base_url))
            elif row["value_text"] not in (None, ""):
                values.extend(_parse_gallery_value(row["value_text"], public_base_url=public_base_url))
        else:
            if row["value_text"] not in (None, ""):
                values.append(normalize_media_reference(row["value_text"], public_base_url=public_base_url) or row["value_text"])
            elif row["value_json"] not in (None, ""):
                values.extend(_parse_gallery_value(row["value_json"], public_base_url=public_base_url))
    if str(fallback_image_url or "").strip():
        values.insert(0, normalize_media_reference(str(fallback_image_url).strip(), public_base_url=public_base_url) or str(fallback_image_url).strip())
    return _normalize_media_urls(values)


def _reset_card_filters() -> None:
    st.session_state["card_product_search"] = ""
    st.session_state["card_product_category_filter"] = "Все"
    st.session_state["card_product_subcategory_filter"] = "Все"
    st.session_state["card_product_supplier_filter"] = "Все"


def _sanitize_filename_part(value: str | None) -> str:
    text = str(value or "").strip()
    if not text:
        return "ITEM"
    text = re.sub(r'[<>:"/\\|?*\x00-\x1F]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return text[:120] or "ITEM"


def _download_binary_resource(source: str, timeout: float = 25.0) -> bytes | None:
    src = str(source or "").strip()
    if not src:
        return None
    if src.lower().startswith(("http://", "https://")):
        try:
            with httpx.Client(follow_redirects=True, timeout=timeout) as client:
                response = client.get(src)
                response.raise_for_status()
                return response.content
        except Exception:
            return None
    if _looks_like_local_media_path(src):
        try:
            return Path(src).read_bytes()
        except Exception:
            return None
    return None


def build_product_images_zip(
    conn,
    product_ids: list[int],
    public_base_url: str | None = None,
) -> tuple[bytes, dict[str, int]]:
    ids = [int(x) for x in product_ids if str(x).strip()]
    output = BytesIO()
    stats = {"products": 0, "images_written": 0, "images_skipped": 0}
    with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for product_id in ids:
            row = conn.execute(
                """
                SELECT id, article, supplier_article, image_url
                FROM products
                WHERE id = ?
                LIMIT 1
                """,
                (int(product_id),),
            ).fetchone()
            if not row:
                continue
            stats["products"] += 1
            article = str(row["article"] or row["supplier_article"] or f"product_{int(product_id)}")
            base_name = _sanitize_filename_part(article)
            urls = _collect_product_gallery_urls(
                conn,
                int(product_id),
                fallback_image_url=str(row["image_url"] or ""),
                public_base_url=public_base_url,
            )
            if not urls:
                stats["images_skipped"] += 1
                continue
            for idx, url in enumerate(urls, start=1):
                binary = _download_binary_resource(url)
                if not binary:
                    stats["images_skipped"] += 1
                    continue
                ext = os.path.splitext(urlparse(url).path)[1].lower()
                if ext not in {".jpg", ".jpeg", ".png", ".webp", ".gif"}:
                    guessed = mimetypes.guess_extension(
                        mimetypes.guess_type(url)[0] or "image/jpeg"
                    ) or ".jpg"
                    ext = guessed if guessed in {".jpg", ".jpeg", ".png", ".webp", ".gif"} else ".jpg"
                zf.writestr(f"{base_name}_{idx}{ext}", binary)
                stats["images_written"] += 1
    output.seek(0)
    return output.getvalue(), stats


def estimate_dimensions_for_product(
    conn,
    product_id: int,
    force: bool = False,
    min_samples: int = 4,
) -> dict[str, Any]:
    row = get_product(conn, int(product_id))
    if not row:
        return {"ok": False, "message": "Товар не найден", "updated_fields": 0}
    product_row = dict(row)

    stats_res = infer_dimensions_from_catalog(conn, product_row, min_samples=max(1, int(min_samples)))
    defaults_res = infer_dimensions_from_category_defaults(conn, product_row)
    heuristic_res = _infer_dimension_heuristics(product_row)

    sources: list[tuple[str, dict[str, Any], str]] = [
        ("category_stats_fallback", stats_res, str(stats_res.get("scope") or "catalog_stats")),
        ("category_defaults_fallback", defaults_res, str(defaults_res.get("scope") or "category_defaults")),
        ("type_heuristic_estimate", heuristic_res, str(heuristic_res.get("scope") or "heuristic")),
    ]

    target_fields = [
        "weight",
        "gross_weight",
        "length",
        "width",
        "height",
        "package_length",
        "package_width",
        "package_height",
    ]
    updates: dict[str, Any] = {}
    source_by_field: dict[str, str] = {}
    detail_by_field: dict[str, str] = {}

    for source_type, result, detail in sources:
        if not bool(result.get("found")):
            continue
        values = dict(result.get("values") or {})
        for field in target_fields:
            if field in updates:
                continue
            value = values.get(field)
            if value in (None, "", 0, 0.0):
                continue
            current = product_row.get(field)
            if (current not in (None, "", 0, 0.0)) and not force:
                continue
            if not can_overwrite_field(conn, int(product_id), field, source_type, force=bool(force)):
                continue
            updates[field] = float(value)
            source_by_field[field] = source_type
            detail_by_field[field] = detail

    if not updates:
        return {
            "ok": True,
            "message": "Нет пустых полей для расчёта габаритов/веса",
            "updated_fields": 0,
            "updates": {},
            "used_sources": [],
        }

    set_clause = ", ".join([f"{k} = ?" for k in updates.keys()])
    params = [updates[k] for k in updates.keys()] + [int(product_id)]
    conn.execute(
        f"""
        UPDATE products
        SET {set_clause},
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        params,
    )

    for field_name, value in updates.items():
        src_type = source_by_field.get(field_name) or "category_stats_fallback"
        src_detail = detail_by_field.get(field_name) or ""
        save_field_source(
            conn=conn,
            product_id=int(product_id),
            field_name=field_name,
            source_type=src_type,
            source_value_raw=value,
            source_url=f"dimension_estimation:{src_detail}",
            confidence=0.42 if src_type == "category_stats_fallback" else 0.36 if src_type == "category_defaults_fallback" else 0.31,
            is_manual=False,
        )

    conn.commit()
    used_sources = sorted(set(source_by_field.values()))
    return {
        "ok": True,
        "message": f"Рассчитано полей: {len(updates)}",
        "updated_fields": int(len(updates)),
        "updates": updates,
        "used_sources": used_sources,
    }


def _build_ozon_scope_labels(conn) -> dict[str, str]:
    rows = conn.execute(
        """
        SELECT DISTINCT category_code
        FROM channel_attribute_requirements
        WHERE channel_code = 'ozon'
          AND category_code IS NOT NULL
          AND TRIM(category_code) <> ''
        ORDER BY category_code
        """
    ).fetchall()
    labels: dict[str, str] = {}
    for row in rows:
        code = str(row["category_code"])
        labels[code] = code
        if not code.startswith("ozon:"):
            continue
        parts = code.split(":")
        if len(parts) != 3:
            continue
        try:
            desc_id = int(parts[1])
            type_id = int(parts[2])
        except Exception:
            continue
        cat = conn.execute(
            """
            SELECT MAX(full_path) AS full_path, MAX(type_name) AS type_name
            FROM ozon_category_cache
            WHERE description_category_id = ? AND type_id = ?
            """,
            (desc_id, type_id),
        ).fetchone()
        full_path = str(cat["full_path"] or "").strip() if cat else ""
        type_name = str(cat["type_name"] or "").strip() if cat else ""
        if full_path or type_name:
            labels[code] = f"{full_path or '-'} | {type_name or '-'} | cat={desc_id}, type={type_id}"
        else:
            fallback = conn.execute(
                """
                SELECT MAX(category_name) AS category_name
                FROM ozon_category_cache
                WHERE description_category_id = ?
                """,
                (desc_id,),
            ).fetchone()
            category_name = str(fallback["category_name"] or "").strip() if fallback else ""
            if category_name:
                labels[code] = f"{category_name} | тип={type_id} | cat={desc_id}"
            else:
                labels[code] = f"Ozon категория {desc_id} | тип {type_id}"
    return labels


def _build_detmir_scope_labels(conn) -> dict[str, str]:
    rows = conn.execute(
        """
        SELECT DISTINCT category_code
        FROM channel_attribute_requirements
        WHERE channel_code = 'detmir'
          AND category_code IS NOT NULL
          AND TRIM(category_code) <> ''
        ORDER BY category_code
        """
    ).fetchall()
    labels: dict[str, str] = {}
    for row in rows:
        code = str(row["category_code"] or "")
        labels[code] = code
        if not code.startswith("detmir:"):
            continue
        parts = code.split(":")
        if len(parts) != 2:
            continue
        try:
            category_id = int(parts[1])
        except Exception:
            continue
        cat = get_detmir_cached_category(conn, category_id)
        if cat:
            labels[code] = f"{cat.get('full_path') or cat.get('name') or '-'} | cat={category_id}"
        else:
            labels[code] = f"Detmir категория {category_id}"
    return labels


def ensure_ozon_requirements_for_product_category(
    conn,
    description_category_id: int,
    type_id: int,
) -> dict[str, int | bool]:
    desc_id = int(description_category_id or 0)
    typ_id = int(type_id or 0)
    if desc_id <= 0 or typ_id <= 0:
        return {"ok": False, "imported": 0, "required": 0, "existing": 0, "cached": 0}

    category_code = f"ozon:{desc_id}:{typ_id}"
    existing = int(
        conn.execute(
            """
            SELECT COUNT(*)
            FROM channel_attribute_requirements
            WHERE channel_code = 'ozon'
              AND category_code = ?
            """,
            (category_code,),
        ).fetchone()[0]
        or 0
    )

    cached = int(
        conn.execute(
            """
            SELECT COUNT(*)
            FROM ozon_attribute_cache
            WHERE description_category_id = ?
              AND type_id = ?
            """,
            (desc_id, typ_id),
        ).fetchone()[0]
        or 0
    )
    if cached <= 0:
        return {"ok": False, "imported": 0, "required": 0, "existing": existing, "cached": cached}
    if existing >= cached:
        return {"ok": True, "imported": 0, "required": 0, "existing": existing, "cached": cached}

    result = import_cached_attributes_to_pim(conn, description_category_id=desc_id, type_id=typ_id)
    final_existing = int(
        conn.execute(
            """
            SELECT COUNT(*)
            FROM channel_attribute_requirements
            WHERE channel_code = 'ozon'
              AND category_code = ?
            """,
            (category_code,),
        ).fetchone()[0]
        or 0
    )
    return {
        "ok": True,
        "imported": int(result.get("imported") or 0),
        "required": int(result.get("required") or 0),
        "existing": final_existing,
        "cached": cached,
    }


def ensure_ozon_requirements_for_products(
    conn,
    product_ids: list[int],
) -> dict[str, int]:
    ids = [int(x) for x in product_ids if str(x).strip()]
    if not ids:
        return {
            "products_total": 0,
            "products_with_ozon_category": 0,
            "category_pairs": 0,
            "category_pairs_missing_cache": 0,
            "imported_attributes": 0,
            "required_attributes": 0,
        }

    pair_set: set[tuple[int, int]] = set()
    products_with_ozon_category = 0
    chunk_size = 900
    for start in range(0, len(ids), chunk_size):
        chunk = ids[start : start + chunk_size]
        placeholders = ", ".join(["?"] * len(chunk))
        rows = conn.execute(
            f"""
            SELECT id, ozon_description_category_id, ozon_type_id
            FROM products
            WHERE id IN ({placeholders})
            """,
            tuple(chunk),
        ).fetchall()
        for row in rows:
            desc_id = int(row["ozon_description_category_id"] or 0)
            type_id = int(row["ozon_type_id"] or 0)
            if desc_id > 0 and type_id > 0:
                products_with_ozon_category += 1
                pair_set.add((desc_id, type_id))

    imported_attributes = 0
    required_attributes = 0
    missing_cache = 0
    for desc_id, type_id in sorted(pair_set):
        result = ensure_ozon_requirements_for_product_category(
            conn,
            description_category_id=int(desc_id),
            type_id=int(type_id),
        )
        imported_attributes += int(result.get("imported") or 0)
        required_attributes += int(result.get("required") or 0)
        if int(result.get("cached") or 0) <= 0:
            missing_cache += 1

    return {
        "products_total": int(len(ids)),
        "products_with_ozon_category": int(products_with_ozon_category),
        "category_pairs": int(len(pair_set)),
        "category_pairs_missing_cache": int(missing_cache),
        "imported_attributes": int(imported_attributes),
        "required_attributes": int(required_attributes),
    }


def materialize_ozon_attribute_slots_for_product(
    conn,
    product_id: int,
    description_category_id: int,
    type_id: int,
    required_only: bool = False,
) -> dict[str, int | bool]:
    pid = int(product_id or 0)
    desc_id = int(description_category_id or 0)
    typ_id = int(type_id or 0)
    if pid <= 0 or desc_id <= 0 or typ_id <= 0:
        return {"ok": False, "created": 0, "existing": 0, "requirements": 0}

    ensure_ozon_requirements_for_product_category(conn, desc_id, typ_id)
    category_code = f"ozon:{desc_id}:{typ_id}"
    req_rows = list_channel_requirements(conn, channel_code="ozon", category_code=category_code)
    if bool(required_only):
        req_rows = [row for row in req_rows if int(row.get("is_required") or 0) == 1]
    attr_codes = [str(row.get("attribute_code") or "").strip() for row in req_rows if str(row.get("attribute_code") or "").strip()]
    if not attr_codes:
        return {"ok": True, "created": 0, "existing": 0, "requirements": 0}

    placeholders = ", ".join(["?"] * len(attr_codes))
    existing_rows = conn.execute(
        f"""
        SELECT DISTINCT attribute_code
        FROM product_attribute_values
        WHERE product_id = ?
          AND attribute_code IN ({placeholders})
        """,
        [pid] + attr_codes,
    ).fetchall()
    existing_codes = {str(row["attribute_code"] or "").strip() for row in existing_rows}

    created = 0
    for code in attr_codes:
        if code in existing_codes:
            continue
        set_product_attribute_value(
            conn=conn,
            product_id=pid,
            attribute_code=code,
            value=None,
            channel_code=None,
        )
        created += 1

    return {
        "ok": True,
        "created": int(created),
        "existing": int(len(existing_codes)),
        "requirements": int(len(attr_codes)),
    }


def materialize_ozon_attribute_slots_for_products(
    conn,
    product_ids: list[int],
    required_only: bool = False,
) -> dict[str, int]:
    ids = [int(x) for x in product_ids if str(x).strip()]
    if not ids:
        return {"products_total": 0, "products_with_ozon_category": 0, "slots_created": 0, "requirements_total": 0}

    chunk_size = 900
    rows: list[sqlite3.Row] = []
    for start in range(0, len(ids), chunk_size):
        chunk = ids[start : start + chunk_size]
        placeholders = ", ".join(["?"] * len(chunk))
        rows.extend(
            conn.execute(
                f"""
                SELECT id, ozon_description_category_id, ozon_type_id
                FROM products
                WHERE id IN ({placeholders})
                """,
                tuple(chunk),
            ).fetchall()
        )

    total_with_ozon = 0
    total_created = 0
    total_requirements = 0
    for row in rows:
        desc_id = int(row["ozon_description_category_id"] or 0)
        type_id = int(row["ozon_type_id"] or 0)
        if desc_id <= 0 or type_id <= 0:
            continue
        total_with_ozon += 1
        result = materialize_ozon_attribute_slots_for_product(
            conn,
            product_id=int(row["id"]),
            description_category_id=desc_id,
            type_id=type_id,
            required_only=bool(required_only),
        )
        total_created += int(result.get("created") or 0)
        total_requirements += int(result.get("requirements") or 0)

    return {
        "products_total": int(len(ids)),
        "products_with_ozon_category": int(total_with_ozon),
        "slots_created": int(total_created),
        "requirements_total": int(total_requirements),
    }


def _build_ozon_template_category_options(
    conn,
    channel_code: str | None = None,
    limit: int = 5000,
) -> tuple[list[str], dict[str, str]]:
    options: list[str] = [""]
    labels: dict[str, str] = {"": "-- без категории --"}
    seen: set[str] = {""}

    if str(channel_code or "").strip() == SPORTMASTER_CHANNEL_CODE:
        scope_labels = build_sportmaster_scope_labels(conn)
        for code in sorted(scope_labels.keys(), key=lambda x: scope_labels.get(x, x).lower()):
            if not code or code in seen:
                continue
            labels[code] = scope_labels.get(code, code)
            options.append(code)
            seen.add(code)
        return options, labels

    if str(channel_code or "").strip() == "detmir":
        scope_labels = _build_detmir_scope_labels(conn)
        for code in sorted(scope_labels.keys(), key=lambda x: scope_labels.get(x, x).lower()):
            if not code or code in seen:
                continue
            labels[code] = scope_labels.get(code, code)
            options.append(code)
            seen.add(code)
        return options, labels

    try:
        pairs = list_cached_category_pairs(conn, limit=max(200, int(limit)))
    except Exception:
        pairs = []

    for row in pairs:
        desc_id = row.get("description_category_id")
        type_id = row.get("type_id")
        if desc_id is None or type_id is None:
            continue
        code = f"ozon:{int(desc_id)}:{int(type_id)}"
        if code in seen:
            continue
        full_path = str(row.get("full_path") or row.get("category_name") or "").strip()
        type_name = str(row.get("type_name") or "").strip()
        labels[code] = f"{full_path or '-'} | {type_name or '-'} | {code}"
        options.append(code)
        seen.add(code)

    # Добавляем legacy-категории из сохранённых профилей канала, чтобы не потерять совместимость.
    profile_categories: set[str] = set()
    if channel_code:
        for profile in list_template_profiles(conn, channel_code=channel_code):
            raw = str(profile.get("category_code") or "").strip()
            if raw:
                profile_categories.add(raw)
    if profile_categories:
        scope_labels = _build_ozon_scope_labels(conn)
        for code in sorted(profile_categories):
            if code in seen:
                continue
            labels[code] = scope_labels.get(code, code)
            options.append(code)
            seen.add(code)

    return options, labels


def _is_blank_value(value: object) -> bool:
    if value is None:
        return True
    text = str(value).strip()
    return text == "" or text.lower() == "none"


def render_supplier_url(url_template: str, row: dict) -> str | None:
    if not url_template:
        return None
    article = str(row.get("article") or "").strip()
    supplier_article = str(row.get("supplier_article") or "").strip()
    name = str(row.get("name") or "").strip()
    category = str(row.get("category") or "").strip()
    code = str(row.get("article") or row.get("supplier_article") or "").strip()
    rendered = str(url_template)
    rendered = rendered.replace("{article}", article)
    rendered = rendered.replace("{article_q}", quote(article, safe=""))
    rendered = rendered.replace("{supplier_article}", supplier_article)
    rendered = rendered.replace("{supplier_article_q}", quote(supplier_article, safe=""))
    rendered = rendered.replace("{name}", name)
    rendered = rendered.replace("{name_q}", quote(name, safe=""))
    rendered = rendered.replace("{category}", category)
    rendered = rendered.replace("{category_q}", quote(category, safe=""))
    rendered = rendered.replace("{code}", code)
    rendered = rendered.replace("{code_q}", quote(code, safe=""))
    rendered = rendered.strip()
    if not rendered:
        return None
    if rendered.startswith("http://") or rendered.startswith("https://"):
        return rendered
    if "." in rendered and " " not in rendered:
        return f"https://{rendered}"
    return None


def load_product_ids(
    conn,
    search: str = "",
    category: str = "",
    supplier: str = "",
    import_batch_id: str = "",
    parse_filter: str = "Все",
    limit: int | None = None,
    offset: int = 0,
) -> list[int]:
    where, params = _build_product_filters(
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=import_batch_id,
        parse_filter=parse_filter,
    )
    sql = "SELECT id FROM products"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    if limit is not None:
        sql += " LIMIT ? OFFSET ?"
        params.extend([int(limit), int(offset)])
    rows = conn.execute(sql, params).fetchall()
    return [int(r["id"]) for r in rows]


def load_product_ids_with_supplier_url(
    conn,
    search: str = "",
    category: str = "",
    supplier: str = "",
    import_batch_id: str = "",
    parse_filter: str = "Все",
    limit: int | None = None,
    offset: int = 0,
) -> list[int]:
    where, params = _build_product_filters(
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=import_batch_id,
        parse_filter=parse_filter,
    )
    where.append("supplier_url IS NOT NULL AND TRIM(supplier_url) <> ''")
    sql = "SELECT id FROM products"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC"
    if limit is not None:
        sql += " LIMIT ? OFFSET ?"
        params.extend([int(limit), int(offset)])
    rows = conn.execute(sql, params).fetchall()
    return [int(r["id"]) for r in rows]


def apply_mass_product_updates(
    conn,
    product_ids: list[int],
    updates: dict[str, str],
    supplier_url_template: str | None = None,
    only_empty: bool = False,
) -> dict:
    if not product_ids:
        return {"updated_products": 0, "updated_fields": 0}
    tracked_fields = ["supplier_name", "supplier_url", "category", "base_category", "subcategory", "brand"]
    updated_products = 0
    updated_fields = 0
    for pid in product_ids:
        row = conn.execute(
            """
            SELECT id, article, supplier_article, name, category, supplier_name, supplier_url, base_category, subcategory, brand
                 , ozon_description_category_id, ozon_type_id
            FROM products
            WHERE id = ?
            LIMIT 1
            """,
            (int(pid),),
        ).fetchone()
        if not row:
            continue
        current = dict(row)
        row_updates: dict[str, str] = {}
        for field, value in updates.items():
            if value is None:
                continue
            if only_empty and not _is_blank_value(current.get(field)):
                continue
            row_updates[field] = str(value).strip()

        if supplier_url_template:
            generated = render_supplier_url(supplier_url_template, current)
            if generated:
                if (not only_empty) or _is_blank_value(current.get("supplier_url")):
                    row_updates["supplier_url"] = generated

        if not row_updates:
            continue

        set_clause = ", ".join([f"{k} = ?" for k in row_updates.keys()])
        params = list(row_updates.values()) + [int(pid)]
        conn.execute(
            f"UPDATE products SET {set_clause}, updated_at = CURRENT_TIMESTAMP WHERE id = ?",
            params,
        )
        for field_name, value in row_updates.items():
            if field_name in tracked_fields:
                save_field_source(
                    conn=conn,
                    product_id=int(pid),
                    field_name=field_name,
                    source_type="manual",
                    source_value_raw=value,
                    source_url=None,
                    confidence=1.0,
                    is_manual=True,
                )
                updated_fields += 1
        updated_products += 1
    conn.commit()
    return {"updated_products": updated_products, "updated_fields": updated_fields}


def render_section_help() -> None:
    with st.expander("Быстрый контур работы", expanded=False):
        st.markdown(
            """
1. `Импорт`: загрузи прайс поставщика и сохрани его в мастер-каталог.
2. `Каталог`: отфильтруй нужные товары и запусти массовое наполнение.
3. `Карточка`: доведи спорные товары руками, проверь Ozon/Detmir и фото.
4. `Клиентский шаблон`: выбери клиента, подтяни shortlist и выгрузи готовую пачку.
5. `Настройки`: здесь живут AI, парсинг и фото. `Каналы` оставлены только для клиентских правил и интеграций.
            """
        )


def _search_terms(value: str) -> list[str]:
    return [token for token in re.split(r"\s+", str(value or "").strip().lower()) if token]


def _matches_search_tokens(values: list[object], query: str) -> bool:
    terms = _search_terms(query)
    if not terms:
        return True
    haystacks = [str(value or "").strip().lower() for value in values if str(value or "").strip()]
    if not haystacks:
        return False
    return all(any(term in haystack for haystack in haystacks) for term in terms)


def _build_product_filters(
    search: str = "",
    category: str = "",
    supplier: str = "",
    import_batch_id: str = "",
    parse_filter: str = "Все",
) -> tuple[list[str], list[object]]:
    where = []
    params: list[object] = []

    if search:
        search_terms = _search_terms(search)
        for term in search_terms:
            where.append(
                "("
                "article LIKE ? OR supplier_article LIKE ? OR internal_article LIKE ? OR barcode LIKE ? "
                "OR name LIKE ? OR brand LIKE ? OR category LIKE ? OR base_category LIKE ? OR subcategory LIKE ?"
                ")"
            )
            s = f"%{term}%"
            params.extend([s, s, s, s, s, s, s, s, s])

    if category:
        where.append(
            "("
            "LOWER(TRIM(category)) = LOWER(TRIM(?)) "
            "OR LOWER(TRIM(base_category)) = LOWER(TRIM(?)) "
            "OR LOWER(TRIM(subcategory)) = LOWER(TRIM(?)) "
            "OR LOWER(TRIM(IFNULL(ozon_category_path, ''))) = LOWER(TRIM(?)) "
            "OR LOWER(IFNULL(ozon_category_path, '')) LIKE LOWER(?)"
            ")"
        )
        params.extend([category, category, category, category, f"%{category}%"])

    if supplier:
        where.append("LOWER(TRIM(supplier_name)) = LOWER(TRIM(?))")
        params.append(supplier)

    if import_batch_id:
        where.append("import_batch_id = ?")
        params.append(import_batch_id)

    if parse_filter == "Есть supplier_url":
        where.append("supplier_url IS NOT NULL AND TRIM(supplier_url) <> ''")
    elif parse_filter == "Не парсено":
        where.append("(supplier_parse_status IS NULL OR TRIM(supplier_parse_status) = '')")
    elif parse_filter == "Ошибка":
        where.append("supplier_parse_status = 'error'")
    elif parse_filter == "Успех":
        where.append("supplier_parse_status = 'success'")

    return where, params


def count_products(
    conn,
    search: str = "",
    category: str = "",
    supplier: str = "",
    import_batch_id: str = "",
    parse_filter: str = "Все",
) -> int:
    where, params = _build_product_filters(
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=import_batch_id,
        parse_filter=parse_filter,
    )
    sql = "SELECT COUNT(*) AS total FROM products"
    if where:
        sql += " WHERE " + " AND ".join(where)
    row = conn.execute(sql, params).fetchone()
    return int(row["total"]) if row and row["total"] is not None else 0


def load_products(
    conn,
    search: str = "",
    category: str = "",
    supplier: str = "",
    limit: int = 200,
    import_batch_id: str = "",
    parse_filter: str = "Все",
    offset: int = 0,
) -> pd.DataFrame:
    where, params = _build_product_filters(
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=import_batch_id,
        parse_filter=parse_filter,
    )

    sql = """
        SELECT
            id,
            article,
            internal_article,
            supplier_article,
            name,
            brand,
            supplier_name,
            supplier_url,
            barcode,
            category,
            base_category,
            subcategory,
            wheel_diameter_inch,
            weight,
            length,
            width,
            height,
            package_length,
            package_width,
            package_height,
            gross_weight,
            enrichment_status,
            enrichment_comment,
            supplier_parse_status,
            duplicate_status,
            ozon_description_category_id,
            ozon_type_id,
            ozon_category_path,
            ozon_category_confidence,
            import_batch_id,
            updated_at
        FROM products
    """

    if where:
        sql += " WHERE " + " AND ".join(where)

    sql += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params.append(int(limit))
    params.append(int(offset))

    rows = conn.execute(sql, params).fetchall()
    return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()


def get_product(conn, product_id: int):
    return conn.execute("SELECT * FROM products WHERE id = ?", (product_id,)).fetchone()


def find_products_for_card(
    conn,
    search: str = "",
    ozon_category: str = "",
    ozon_subcategory: str = "",
    supplier: str = "",
    limit: int = 5000,
) -> list[dict]:
    where: list[str] = []
    params: list[object] = []
    if supplier and supplier != "Все":
        where.append("LOWER(TRIM(supplier_name)) = LOWER(TRIM(?))")
        params.append(supplier)

    sql = """
        SELECT
            id, article, internal_article, supplier_article, name, brand, barcode,
            category, base_category, subcategory, supplier_name, ozon_category_path,
            ozon_description_category_id, ozon_type_id, supplier_parse_status, updated_at
        FROM products
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY id DESC LIMIT ?"
    params.append(int(limit))
    rows = conn.execute(sql, params).fetchall()
    out: list[dict] = []
    category_filter = str(ozon_category or "").strip().lower()
    subcategory_filter = str(ozon_subcategory or "").strip().lower()
    for row in rows:
        item = dict(row)
        parts = _split_ozon_path_parts(item.get("ozon_category_path"))
        item["ozon_subcategory"] = parts[-1] if parts else ""
        item["ozon_category"] = parts[-2] if len(parts) >= 2 else item["ozon_subcategory"]
        if search and not _matches_search_tokens(
            [
                item.get("article"),
                item.get("internal_article"),
                item.get("supplier_article"),
                item.get("name"),
                item.get("brand"),
                item.get("barcode"),
                item.get("supplier_name"),
                item.get("category"),
                item.get("base_category"),
                item.get("subcategory"),
                item.get("ozon_category"),
                item.get("ozon_subcategory"),
                item.get("ozon_category_path"),
            ],
            search,
        ):
            continue
        if category_filter and category_filter != "все":
            if str(item.get("ozon_category") or "").strip().lower() != category_filter:
                continue
        if subcategory_filter and subcategory_filter != "все":
            if str(item.get("ozon_subcategory") or "").strip().lower() != subcategory_filter:
                continue
        out.append(item)
        if len(out) >= int(limit):
            break
    return out


def _card_product_sort_key(item: dict, search: str = "") -> tuple:
    query = str(search or "").strip().lower()
    terms = _search_terms(query)
    article = str(item.get("article") or "").strip().lower()
    internal_article = str(item.get("internal_article") or "").strip().lower()
    supplier_article = str(item.get("supplier_article") or "").strip().lower()
    name = str(item.get("name") or "").strip().lower()
    brand = str(item.get("brand") or "").strip().lower()
    exact = 0 if query and query in {article, internal_article, supplier_article} else 1
    starts = 0 if query and (
        article.startswith(query) or internal_article.startswith(query) or supplier_article.startswith(query)
    ) else 1
    token_hit = 0 if terms and any(
        (article.startswith(term) or internal_article.startswith(term) or supplier_article.startswith(term))
        for term in terms
    ) else 1
    contains = 0 if query and (
        _matches_search_tokens([article, internal_article, supplier_article, name, brand], query)
    ) else 1
    return (
        exact,
        starts,
        token_hit,
        contains,
        article or supplier_article or internal_article or name,
        -int(item.get("id") or 0),
    )


def list_channel_codes(conn) -> list[str]:
    rows = conn.execute(
        """
        SELECT DISTINCT channel_code
        FROM (
            SELECT channel_code FROM channel_profiles
            UNION ALL
            SELECT channel_code FROM channel_attribute_requirements
            UNION ALL
            SELECT channel_code FROM channel_mapping_rules
        )
        WHERE channel_code IS NOT NULL
          AND TRIM(channel_code) <> ''
        ORDER BY channel_code
        """
    ).fetchall()
    return [str(r["channel_code"]) for r in rows if r["channel_code"]]


def list_channel_category_codes(conn, channel_code: str) -> list[str]:
    rows = conn.execute(
        """
        SELECT DISTINCT category_code
        FROM (
            SELECT category_code
            FROM channel_attribute_requirements
            WHERE channel_code = ?
            UNION ALL
            SELECT category_code
            FROM channel_mapping_rules
            WHERE channel_code = ?
        )
        WHERE category_code IS NOT NULL
          AND TRIM(category_code) <> ''
        ORDER BY category_code
        """,
        (channel_code, channel_code),
    ).fetchall()
    return [str(r["category_code"]) for r in rows if r["category_code"]]


def ensure_template_columns_registered(
    conn,
    channel_code: str,
    category_code: str | None,
    template_columns: list[object],
    required_by_column: dict[str, int] | None = None,
    mapping_by_column: dict[str, dict[str, object]] | None = None,
) -> dict[str, int]:
    if not channel_code:
        return {"attributes": 0, "requirements": 0, "rules": 0}

    created_attributes = 0
    created_requirements = 0
    created_rules = 0

    for idx, col in enumerate(template_columns):
        col_name = str(col or "").strip()
        if not col_name:
            continue
        code = to_attribute_code(col_name)
        if not code:
            continue
        is_required = int((required_by_column or {}).get(col_name, 0) or 0)
        saved_match = (mapping_by_column or {}).get(col_name) or {}
        source_type = str(saved_match.get("source_type") or "attribute").strip() or "attribute"
        source_name = str(saved_match.get("source_name") or code).strip() or code
        transform_rule = saved_match.get("transform_rule")
        requirement_notes = str(saved_match.get("notes") or "").strip() or "Автодобавлено из клиентского шаблона"

        existed_attr = conn.execute(
            "SELECT 1 FROM attribute_definitions WHERE code = ?",
            (code,),
        ).fetchone()
        upsert_attribute_definition(
            conn=conn,
            code=code,
            name=col_name,
            data_type="text",
            scope="master",
            unit=None,
            description=f"Автосоздано из клиентского шаблона: {col_name}",
        )
        if not existed_attr:
            created_attributes += 1

        existed_req = conn.execute(
            """
            SELECT 1
            FROM channel_attribute_requirements
            WHERE channel_code = ?
              AND IFNULL(category_code, '') = IFNULL(?, '')
              AND attribute_code = ?
            """,
            (channel_code, category_code, code),
        ).fetchone()
        upsert_channel_attribute_requirement(
            conn=conn,
            channel_code=channel_code,
            category_code=category_code or None,
            attribute_code=code,
            is_required=is_required,
            sort_order=1000 + int(idx),
            notes=requirement_notes,
        )
        if not existed_req:
            created_requirements += 1

        existed_rule = conn.execute(
            """
            SELECT 1
            FROM channel_mapping_rules
            WHERE channel_code = ?
              AND IFNULL(category_code, '') = IFNULL(?, '')
              AND target_field = ?
            """,
            (channel_code, category_code, col_name),
        ).fetchone()
        if not existed_rule:
            upsert_channel_mapping_rule(
                conn=conn,
                channel_code=channel_code,
                category_code=category_code or None,
                target_field=col_name,
                source_type=source_type,
                source_name=source_name,
                transform_rule=transform_rule,
                is_required=is_required,
            )
            created_rules += 1

    return {
        "attributes": created_attributes,
        "requirements": created_requirements,
        "rules": created_rules,
    }


def save_product(conn, product_id: int, payload: dict):
    current = get_product(conn, int(product_id))
    if not current:
        raise ValueError(f"Товар не найден: {product_id}")
    merged = dict(current)
    merged.update(payload or {})
    conn.execute(
        """
        UPDATE products
        SET
            article = ?,
            internal_article = ?,
            supplier_article = ?,
            name = ?,
            brand = ?,
            supplier_name = ?,
            barcode = ?,
            barcode_source = ?,
            category = ?,
            base_category = ?,
            subcategory = ?,
            wheel_diameter_inch = ?,
            supplier_url = ?,
            ozon_description_category_id = ?,
            ozon_type_id = ?,
            ozon_category_path = ?,
            ozon_category_confidence = ?,
            detmir_category_id = ?,
            detmir_category_path = ?,
            detmir_category_confidence = ?,
            uom = ?,
            weight = ?,
            length = ?,
            width = ?,
            height = ?,
            package_length = ?,
            package_width = ?,
            package_height = ?,
            gross_weight = ?,
            image_url = ?,
            description = ?,
            tnved_code = ?,
            updated_at = CURRENT_TIMESTAMP
        WHERE id = ?
        """,
        (
            merged.get("article"),
            merged.get("internal_article"),
            merged.get("supplier_article"),
            merged.get("name"),
            merged.get("brand"),
            merged.get("supplier_name"),
            merged.get("barcode"),
            merged.get("barcode_source"),
            merged.get("category"),
            merged.get("base_category"),
            merged.get("subcategory"),
            merged.get("wheel_diameter_inch"),
            merged.get("supplier_url"),
            merged.get("ozon_description_category_id"),
            merged.get("ozon_type_id"),
            merged.get("ozon_category_path"),
            merged.get("ozon_category_confidence"),
            merged.get("detmir_category_id"),
            merged.get("detmir_category_path"),
            merged.get("detmir_category_confidence"),
            merged.get("uom"),
            merged.get("weight"),
            merged.get("length"),
            merged.get("width"),
            merged.get("height"),
            merged.get("package_length"),
            merged.get("package_width"),
            merged.get("package_height"),
            merged.get("gross_weight"),
            merged.get("image_url"),
            merged.get("description"),
            merged.get("tnved_code"),
            product_id,
        ),
    )
    tracked_fields = [
        "article", "internal_article", "supplier_article", "name", "brand", "supplier_name", "barcode",
        "category", "base_category", "subcategory", "wheel_diameter_inch", "supplier_url",
        "ozon_description_category_id", "ozon_type_id", "ozon_category_path", "ozon_category_confidence",
        "detmir_category_id", "detmir_category_path", "detmir_category_confidence", "uom",
        "weight", "length", "width", "height", "package_length", "package_width", "package_height",
        "gross_weight", "image_url", "description", "tnved_code"
    ]
    for field_name in tracked_fields:
        value = payload.get(field_name)
        if value not in (None, ""):
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name=field_name,
                source_type="manual",
                source_value_raw=value,
                source_url=None,
                confidence=1.0,
                is_manual=True,
            )
    conn.commit()


def export_current_df(df: pd.DataFrame):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="products")
    return output.getvalue()


def dataframes_to_excel_bytes(sheets: dict[str, pd.DataFrame]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = str(sheet_name)[:31] if sheet_name else "sheet"
            frame = df if df is not None else pd.DataFrame()
            frame.to_excel(writer, index=False, sheet_name=safe_name)
    return output.getvalue()


def _write_excel_sheet_chunked(writer: pd.ExcelWriter, sheet_name: str, df: pd.DataFrame, chunk_size: int = 900000) -> None:
    frame = df if df is not None else pd.DataFrame()
    if frame.empty:
        frame.to_excel(writer, index=False, sheet_name=str(sheet_name)[:31])
        return

    total = len(frame)
    if total <= int(chunk_size):
        frame.to_excel(writer, index=False, sheet_name=str(sheet_name)[:31])
        return

    parts = math.ceil(total / int(chunk_size))
    for idx in range(parts):
        start = idx * int(chunk_size)
        end = min((idx + 1) * int(chunk_size), total)
        suffix = f"_{idx + 1}"
        base = str(sheet_name)
        safe_name = f"{base[: max(1, 31 - len(suffix))]}{suffix}"[:31]
        frame.iloc[start:end].to_excel(writer, index=False, sheet_name=safe_name)


def _read_excel_sheet_group(xls: pd.ExcelFile, base_sheet: str) -> pd.DataFrame | None:
    names: list[str] = []
    for sheet in xls.sheet_names:
        if sheet == base_sheet or sheet.startswith(f"{base_sheet}_"):
            names.append(sheet)
    if not names:
        return None

    def _sort_key(name: str) -> tuple[int, int]:
        if name == base_sheet:
            return (0, 0)
        suffix = name[len(base_sheet):]
        if suffix.startswith("_"):
            suffix = suffix[1:]
        try:
            return (1, int(suffix))
        except Exception:
            return (1, 999999)

    names = sorted(names, key=_sort_key)
    parts = [pd.read_excel(xls, sheet_name=name) for name in names]
    if not parts:
        return pd.DataFrame()
    if len(parts) == 1:
        return parts[0]
    return pd.concat(parts, ignore_index=True)


def _select_ozon_snapshot_df(conn, table_name: str) -> pd.DataFrame:
    allowed_columns: dict[str, list[str]] = {
        "ozon_category_cache": [
            "description_category_id",
            "category_name",
            "full_path",
            "type_id",
            "type_name",
            "disabled",
            "children_count",
            "fetched_at",
        ],
        "ozon_attribute_cache": [
            "description_category_id",
            "type_id",
            "attribute_id",
            "name",
            "description",
            "type",
            "group_id",
            "group_name",
            "dictionary_id",
            "is_required",
            "is_collection",
            "max_value_count",
            "category_dependent",
            "fetched_at",
        ],
        "ozon_attribute_value_cache": [
            "description_category_id",
            "type_id",
            "attribute_id",
            "dictionary_id",
            "value_id",
            "value",
            "info",
            "picture",
            "fetched_at",
        ],
    }
    columns = allowed_columns.get(table_name) or []
    if not columns:
        return pd.DataFrame()
    query = f"SELECT {', '.join(columns)} FROM {table_name}"
    return pd.read_sql_query(query, conn)


def build_ozon_cache_snapshot_excel(conn, include_value_cache: bool = False) -> bytes:
    meta_df = pd.DataFrame(
        [
            {
                "generated_at": _now_iso(),
                "db_path": _get_active_db_path() or "",
                "include_value_cache": 1 if bool(include_value_cache) else 0,
            }
        ]
    )
    category_df = _select_ozon_snapshot_df(conn, "ozon_category_cache")
    attribute_df = _select_ozon_snapshot_df(conn, "ozon_attribute_cache")
    value_df = _select_ozon_snapshot_df(conn, "ozon_attribute_value_cache") if bool(include_value_cache) else pd.DataFrame()

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _write_excel_sheet_chunked(writer, "meta", meta_df)
        _write_excel_sheet_chunked(writer, "ozon_category_cache", category_df)
        _write_excel_sheet_chunked(writer, "ozon_attribute_cache", attribute_df)
        if bool(include_value_cache):
            _write_excel_sheet_chunked(writer, "ozon_attribute_value_cache", value_df)
    return output.getvalue()


def _restore_snapshot_table(conn, table_name: str, df: pd.DataFrame | None) -> int:
    if df is None:
        return 0
    if df.empty:
        return 0
    table_cols = [str(r["name"]) for r in conn.execute(f"PRAGMA table_info({table_name})").fetchall()]
    if not table_cols:
        return 0
    allowed_cols = [c for c in df.columns if str(c) in table_cols and str(c) != "id"]
    if not allowed_cols:
        return 0
    frame = df[allowed_cols].copy()
    frame = frame.where(pd.notna(frame), None)
    frame.to_sql(table_name, conn, if_exists="append", index=False, method="multi", chunksize=3000)
    return int(len(frame))


def restore_ozon_cache_snapshot_excel(conn, snapshot_bytes: bytes) -> dict:
    try:
        xls = pd.ExcelFile(BytesIO(snapshot_bytes))
    except Exception as e:
        return {"ok": False, "message": f"Не удалось открыть snapshot Excel: {e}"}

    category_df = _read_excel_sheet_group(xls, "ozon_category_cache")
    attribute_df = _read_excel_sheet_group(xls, "ozon_attribute_cache")
    value_df = _read_excel_sheet_group(xls, "ozon_attribute_value_cache")

    if category_df is None or attribute_df is None:
        return {"ok": False, "message": "В snapshot нет обязательных листов `ozon_category_cache` и/или `ozon_attribute_cache`."}

    try:
        conn.execute("BEGIN")
        conn.execute("DELETE FROM ozon_category_cache")
        conn.execute("DELETE FROM ozon_attribute_cache")
        if value_df is not None:
            conn.execute("DELETE FROM ozon_attribute_value_cache")

        restored_categories = _restore_snapshot_table(conn, "ozon_category_cache", category_df)
        restored_attributes = _restore_snapshot_table(conn, "ozon_attribute_cache", attribute_df)
        restored_values = 0
        if value_df is not None:
            restored_values = _restore_snapshot_table(conn, "ozon_attribute_value_cache", value_df)

        conn.commit()
        return {
            "ok": True,
            "categories": int(restored_categories),
            "attributes": int(restored_attributes),
            "values": int(restored_values),
            "has_values": bool(value_df is not None),
        }
    except Exception as e:
        conn.rollback()
        return {"ok": False, "message": f"Ошибка восстановления snapshot: {e}"}


def build_supplier_catalog_template_excel() -> bytes:
    df = pd.DataFrame(
        {
            "Артикул": ["SUP-001", "SUP-002"],
            "Артикул поставщика": ["RB_0001", "RB_0002"],
            "Номенклатура": ["Велофара Rockbros", "Насос SKS"],
            "Штрихкод": ["4600000000011", "4600000000012"],
            "Категория": ["Вело", "Вело"],
            "Поставщик": ["Rockbros", "SKS"],
            "Ссылка на товар": ["https://example.com/item1", "https://example.com/item2"],
            "Вес": [0.35, 0.42],
            "Длина": [15, 21],
            "Ширина": [9, 6],
            "Высота": [5, 4],
            "Длина упаковки": [17, 23],
            "Ширина упаковки": [10, 7],
            "Высота упаковки": [6, 5],
            "Вес брутто": [0.41, 0.48],
            "Фото": ["https://example.com/img1.jpg", "https://example.com/img2.jpg"],
            "Описание": ["Яркая велофара", "Легкий насос"],
        }
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Каталог")
    return output.getvalue()


def inspect_excel_sheets(file_bytes: bytes) -> dict:
    try:
        xls = pd.ExcelFile(BytesIO(file_bytes))
    except Exception as e:
        return {"ok": False, "message": str(e)}

    sheets = xls.sheet_names
    header_keywords = {
        "артикул",
        "номенклатура",
        "наименование",
        "название",
        "категория",
        "бренд",
        "штрихкод",
        "код",
        "поставщик",
        "ссылка",
    }

    def _score_preview_header(values: list[object]) -> float:
        tokens = [str(v).strip() for v in values if str(v).strip() and str(v).strip().lower() != "nan"]
        if not tokens:
            return -1.0
        lower_tokens = [v.lower() for v in tokens]
        exact_hits = sum(1 for v in lower_tokens if v in header_keywords)
        contains_hits = 0
        for value in lower_tokens:
            if value in header_keywords:
                continue
            if any(keyword in value or value in keyword for keyword in header_keywords if len(keyword) >= 4):
                contains_hits += 1
        text_like = sum(1 for v in tokens if any(ch.isalpha() for ch in v))
        numeric_like = sum(1 for v in tokens if re.fullmatch(r"[\d\.,\-_/]+", v) is not None)
        url_like = sum(1 for v in lower_tokens if "http" in v or "www." in v)
        score = (exact_hits * 100.0) + (contains_hits * 30.0) + (text_like * 2.0)
        score -= (numeric_like * 4.0) + (url_like * 8.0)
        if exact_hits == 0 and contains_hits <= 1:
            score -= 40.0
        return score

    preview_rows = []
    for sheet in sheets[:10]:
        try:
            probe = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, header=None, nrows=12)
            row_scores = []
            for i in range(len(probe)):
                values = [str(v).strip() for v in probe.iloc[i].tolist() if str(v).strip() and str(v).strip().lower() != "nan"]
                lower_values = [v.lower() for v in values]
                keyword_hits = sum(1 for v in lower_values if v in header_keywords)
                text_like = sum(1 for v in values if any(ch.isalpha() for ch in v))
                score = _score_preview_header(probe.iloc[i].tolist())
                row_scores.append({"row": i + 1, "non_empty": len(values), "keywords": keyword_hits, "sample": ", ".join(values[:6]), "score": score})
            recommended = 1
            if row_scores:
                best = max(row_scores, key=lambda r: r["score"])
                recommended = int(best["row"])
            preview_rows.append({"sheet": sheet, "rows": row_scores, "recommended_header_row": recommended})
        except Exception:
            preview_rows.append({"sheet": sheet, "rows": [], "recommended_header_row": 1})
    return {"ok": True, "sheets": sheets, "preview": preview_rows}


def build_ozon_product_list_template_excel() -> bytes:
    df = pd.DataFrame(
        {
            "article": ["ART-001", "ART-002"],
            "internal_article": ["INT-001", "INT-002"],
            "supplier_article": ["SUP-001", "SUP-002"],
            "id": [1, 2],
        }
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="products")
    return output.getvalue()


def build_ozon_dictionary_overrides_template_excel() -> bytes:
    df = pd.DataFrame(
        {
            "attribute_id": [85, 8229],
            "raw_value": ["stels", "bike"],
            "value_id": [123456, 654321],
            "value": ["Stels", "Велосипед"],
            "comment": ["Бренд нормализован", "Тип товара"],
        }
    )
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="overrides")
    return output.getvalue()


def build_ozon_retry_jobs_template_excel() -> bytes:
    df = pd.DataFrame({"job_id": [101, 102, 103]})
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="jobs")
    return output.getvalue()


def _cell_to_lookup_text(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def resolve_product_ids_from_excel(
    conn,
    file_bytes: bytes,
    lookup_field: str,
    sheet_name: str | None = None,
    column_name: str | None = None,
) -> dict:
    try:
        if sheet_name:
            df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
        else:
            df = pd.read_excel(BytesIO(file_bytes))
    except Exception as e:
        return {"ok": False, "message": f"Не удалось прочитать Excel: {e}"}

    if df.empty:
        return {"ok": False, "message": "Excel пустой"}

    lookup_aliases = {
        "id": ["id", "product_id", "товар id", "id товара"],
        "article": ["article", "артикул", "sku", "vendor code", "код товара"],
        "internal_article": ["internal_article", "внутренний артикул", "артикул 1с"],
        "supplier_article": ["supplier_article", "артикул поставщика"],
    }
    field = str(lookup_field or "article").strip().lower()
    aliases = lookup_aliases.get(field, lookup_aliases["article"])

    normalized_columns = {str(c).strip().lower(): str(c) for c in df.columns}
    selected_column = None
    manual_column = (column_name or "").strip()
    if manual_column:
        manual_norm = manual_column.lower()
        selected_column = normalized_columns.get(manual_norm)
        if not selected_column:
            return {"ok": False, "message": f"Колонка '{manual_column}' не найдена в Excel."}
    for alias in aliases:
        if selected_column:
            break
        if alias in normalized_columns:
            selected_column = normalized_columns[alias]
            break
    if not selected_column:
        selected_column = str(df.columns[0])

    raw_values = [_cell_to_lookup_text(v) for v in df[selected_column].tolist()]
    values = []
    seen = set()
    for v in raw_values:
        if not v:
            continue
        key = v.lower() if field != "id" else v
        if key in seen:
            continue
        seen.add(key)
        values.append(v)

    resolved_ids = []
    not_found = []
    for v in values:
        row = None
        if field == "id":
            try:
                row = conn.execute("SELECT id FROM products WHERE id = ? LIMIT 1", (int(float(v)),)).fetchone()
            except Exception:
                row = None
        elif field in {"article", "internal_article", "supplier_article"}:
            row = conn.execute(f"SELECT id FROM products WHERE lower(IFNULL({field}, '')) = lower(?) LIMIT 1", (v,)).fetchone()
        if row:
            resolved_ids.append(int(row["id"]))
        else:
            not_found.append(v)

    return {
        "ok": True,
        "lookup_field": field,
        "used_column": selected_column,
        "input_values": int(len(values)),
        "resolved_ids": sorted(list(set(resolved_ids))),
        "resolved_count": int(len(set(resolved_ids))),
        "not_found": not_found,
        "not_found_count": int(len(not_found)),
    }


def import_dictionary_overrides_from_excel(
    conn,
    file_bytes: bytes,
    description_category_id: int,
    type_id: int,
) -> dict:
    try:
        df = pd.read_excel(BytesIO(file_bytes))
    except Exception as e:
        return {"ok": False, "message": f"Не удалось прочитать Excel: {e}"}

    required_cols = {"attribute_id", "raw_value", "value_id"}
    actual_cols = {str(c).strip().lower(): str(c) for c in df.columns}
    missing = [c for c in required_cols if c not in actual_cols]
    if missing:
        return {"ok": False, "message": f"В Excel нет обязательных колонок: {', '.join(missing)}"}

    applied = 0
    skipped = 0
    errors = []
    for idx, row in df.iterrows():
        try:
            attribute_id_raw = row[actual_cols["attribute_id"]]
            raw_value = _cell_to_lookup_text(row[actual_cols["raw_value"]])
            value_id_raw = row[actual_cols["value_id"]]
            value = row[actual_cols["value"]] if "value" in actual_cols else None
            comment = row[actual_cols["comment"]] if "comment" in actual_cols else None

            if not raw_value:
                skipped += 1
                continue

            attribute_id = int(float(attribute_id_raw))
            value_id = int(float(value_id_raw))
            save_dictionary_override(
                conn=conn,
                description_category_id=int(description_category_id),
                type_id=int(type_id),
                attribute_id=attribute_id,
                raw_value=raw_value,
                value_id=value_id,
                value=_cell_to_lookup_text(value) if value is not None else None,
                comment=_cell_to_lookup_text(comment) if comment is not None else None,
            )
            applied += 1
        except Exception as e:
            skipped += 1
            errors.append({"row": int(idx) + 2, "error": str(e)})

    return {
        "ok": True,
        "applied": int(applied),
        "skipped": int(skipped),
        "errors": errors[:100],
    }


def resolve_job_ids_from_excel(file_bytes: bytes, column_name: str | None = None) -> dict:
    try:
        df = pd.read_excel(BytesIO(file_bytes))
    except Exception as e:
        return {"ok": False, "message": f"Не удалось прочитать Excel: {e}"}

    if df.empty:
        return {"ok": False, "message": "Excel пустой"}

    columns = {str(c).strip().lower(): str(c) for c in df.columns}
    selected_column = None
    if column_name:
        selected_column = columns.get(str(column_name).strip().lower())
        if not selected_column:
            return {"ok": False, "message": f"Колонка '{column_name}' не найдена"}
    if not selected_column:
        for alias in ["job_id", "id", "job"]:
            if alias in columns:
                selected_column = columns[alias]
                break
    if not selected_column:
        selected_column = str(df.columns[0])

    job_ids = []
    errors = []
    seen = set()
    for idx, value in enumerate(df[selected_column].tolist(), start=2):
        text = _cell_to_lookup_text(value)
        if not text:
            continue
        try:
            job_id = int(float(text))
            if job_id in seen:
                continue
            seen.add(job_id)
            job_ids.append(job_id)
        except Exception:
            errors.append({"row": idx, "value": text, "error": "Не удалось распознать job_id"})

    return {
        "ok": True,
        "used_column": selected_column,
        "job_ids": job_ids,
        "count": len(job_ids),
        "errors": errors[:100],
    }


def render_template_readiness(filled_df: pd.DataFrame, manual_rows: list[dict]) -> None:
    readiness = analyze_template_readiness(filled_df, manual_rows)
    summary = readiness["summary"]
    avg_readiness = int(summary["avg_readiness"])
    unmatched_columns = int(summary["unmatched_columns"])
    blocked_rows = int(summary["blocked_rows"])
    rows_to_fix = int(summary["partial_rows"] + summary["blocked_rows"])

    st.markdown("### Готовность шаблона")

    if avg_readiness >= 95 and unmatched_columns == 0 and blocked_rows == 0:
        st.success("Шаблон выглядит почти готовым, критичных дыр не видно.")
    elif avg_readiness >= 80:
        st.warning("Шаблон уже в рабочем состоянии, но часть полей и строк ещё нужно добить.")
    else:
        st.error("Шаблон пока сырой, сначала лучше закрыть пробелы в матчинге и данных.")

    st.progress(max(0, min(avg_readiness, 100)) / 100)

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    c1.metric("Средняя готовность", f"{avg_readiness}%")
    c2.metric("Matched", summary["matched_columns"])
    c3.metric("Unmatched", unmatched_columns)
    c4.metric("Готовых строк", summary["ready_rows"])
    c5.metric("Частично готовы", summary["partial_rows"])
    c6.metric("Блокеры", blocked_rows)

    coverage_df = pd.DataFrame(readiness["column_coverage"])
    row_df = pd.DataFrame(readiness["row_readiness"])

    if not coverage_df.empty:
        weak_columns = coverage_df[coverage_df["Покрытие, %"] < 100]
        if not weak_columns.empty:
            st.caption(f"Колонки, которые требуют внимания: {len(weak_columns)}")
            st.dataframe(weak_columns.head(15), use_container_width=True, hide_index=True)

    if rows_to_fix > 0:
        st.caption(f"Строки, которые ещё нужно добить: {rows_to_fix}")

    tab_cov, tab_rows, tab_all = st.tabs(["Проблемные колонки", "Проблемные строки", "Все колонки"])
    with tab_cov:
        if not coverage_df.empty:
            problem_df = coverage_df[coverage_df["Покрытие, %"] < 100]
            if problem_df.empty:
                st.success("Все колонки шаблона заполнены на 100%.")
            else:
                st.dataframe(problem_df, use_container_width=True, hide_index=True)
        else:
            st.info("Пока нет данных для оценки колонок.")
    with tab_rows:
        if not row_df.empty:
            st.dataframe(row_df.head(200), use_container_width=True, hide_index=True)
        else:
            st.success("Проблемных строк не найдено.")
    with tab_all:
        if not coverage_df.empty:
            st.dataframe(coverage_df, use_container_width=True, hide_index=True)
        else:
            st.info("Пока нет данных для полной сводки.")


def show_import_tab():
    st.subheader("Импорт каталога")
    st.caption("Логика этапа Импорт: 1) загрузка прайса 2) автопривязка к эталонным Ozon категориям 3) переход в Каталог для enrichment.")
    with st.expander("Инструкция по кнопкам раздела Импорт", expanded=False):
        st.markdown(
            """
1. `Скачать шаблон импорта поставщика (Excel)`: эталонный формат для поставщиков.
2. `Сохранить профиль`: записывает профиль поставщика (имя, сайт, URL template) в БД.
3. `Импортировать`: запускает импорт файла в мастер-каталог.
4. `После импорта автоматически привязывать товары к Ozon категориям`: сразу фиксирует эталонную категорию/подкатегорию Ozon.
5. `После импорта автоматически подтягивать Ozon-атрибуты категорий`: подготовит атрибуты категории для карточек товаров.
6. Далее переходи в `Каталог` и запускай supplier enrichment уже по Ozon-структуре.
            """
        )
    st.download_button(
        "Скачать шаблон импорта поставщика (Excel)",
        data=build_supplier_catalog_template_excel(),
        file_name="supplier_catalog_import_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="supplier_import_template",
    )
    profiles_conn = get_db()
    profiles = list_supplier_profiles(profiles_conn, only_active=True)
    existing_suppliers = list_distinct_values(profiles_conn, "supplier_name")
    recent_imports = list_catalog_import_history(profiles_conn, limit=12)
    profiles_conn.close()
    profile_map = {p["supplier_name"]: p for p in profiles}
    selected_profile_name = st.selectbox(
        "Профиль поставщика",
        options=[""] + sorted(profile_map.keys()),
        key="import_supplier_profile_name",
        help="Выбери профиль, чтобы автоматически подставить поставщика и URL-шаблон.",
    )
    if selected_profile_name:
        profile = profile_map[selected_profile_name]
        st.session_state["import_default_supplier_name"] = profile.get("supplier_name") or ""
        st.session_state["import_default_supplier_url_template"] = profile.get("url_template") or profile.get("base_url") or ""
        if profile.get("base_url"):
            st.caption(f"Базовый сайт поставщика: {profile.get('base_url')}")
        if profile.get("legal_entity_name"):
            st.caption(f"Юрлицо поставщика: {profile.get('legal_entity_name')}")
    uploaded = st.file_uploader("Excel файл", type=["xlsx", "xls"])
    s1, s2 = st.columns(2)
    with s1:
        supplier_options = [""] + sorted(set(existing_suppliers + list(profile_map.keys())))
        session_supplier = st.session_state.get("import_default_supplier_name", "")
        supplier_index = supplier_options.index(session_supplier) if session_supplier in supplier_options else 0
        default_supplier_name = st.selectbox(
            "Поставщик по умолчанию (из базы)",
            options=supplier_options,
            index=supplier_index,
            help="Если в файле нет колонки Поставщик, будет выбран этот поставщик из базы.",
        )
        st.session_state["import_default_supplier_name"] = default_supplier_name
    with s2:
        default_supplier_url_template = st.text_input(
            "Шаблон URL поставщика (опционально)",
            value=st.session_state.get("import_default_supplier_url_template", ""),
            placeholder="https://site.ru/product/{supplier_article}",
            help="Поддерживает {article}, {supplier_article}, {code}, {name}, а также *_q для URL-encoding.",
        )
        st.session_state["import_default_supplier_url_template"] = default_supplier_url_template
    with st.expander("Профили поставщиков", expanded=False):
        editing_profile = profile_map.get(selected_profile_name or "", {})
        sp1, sp2, sp3, sp4 = st.columns([2, 2, 2, 1])
        with sp1:
            profile_name_input = st.text_input("Имя профиля", value=default_supplier_name or "", key="supplier_profile_name_input")
        with sp2:
            profile_base_url = st.text_input("Базовый URL", value=str(editing_profile.get("base_url") or ""), key="supplier_profile_base_url")
        with sp3:
            profile_legal_entity = st.text_input("Юрлицо поставщика", value=str(editing_profile.get("legal_entity_name") or ""), key="supplier_profile_legal_entity")
        with sp4:
            save_profile_btn = st.button("Сохранить профиль", help="Сохранить/обновить профиль поставщика в базе")
        profile_url_template = st.text_input(
            "URL template профиля",
            value=str(editing_profile.get("url_template") or default_supplier_url_template or ""),
            key="supplier_profile_url_template",
        )
        if save_profile_btn and profile_name_input.strip():
            conn = get_db()
            profile_id = upsert_supplier_profile(
                conn=conn,
                supplier_name=profile_name_input.strip(),
                legal_entity_name=profile_legal_entity.strip() or None,
                base_url=profile_base_url.strip() or None,
                url_template=profile_url_template.strip() or None,
                notes="Сохранено из вкладки Импорт",
                is_active=1,
            )
            conn.close()
            st.success(f"Профиль поставщика сохранён: #{profile_id}")
    auto_match_ozon_after_import = st.checkbox(
        "После импорта автоматически привязывать товары к Ozon категориям (эталон)",
        value=True,
        help="Работает, если кэш категорий Ozon уже синхронизирован во вкладке Ozon.",
        key="import_auto_ozon_match",
    )
    auto_seed_ozon_attrs_after_import = st.checkbox(
        "После импорта автоматически подтягивать Ozon-атрибуты категорий для карточек",
        value=True,
        help="После Ozon-автопривязки добавит category requirements, чтобы атрибуты сразу были доступны в карточках.",
        key="import_auto_seed_ozon_attrs",
    )
    if recent_imports:
        with st.expander("История импортов каталога", expanded=False):
            history_df = pd.DataFrame(recent_imports)
            history_cols = [
                c
                for c in [
                    "created_at",
                    "original_file_name",
                    "supplier_name",
                    "selected_sheet",
                    "header_row",
                    "imported_count",
                    "created_count",
                    "updated_count",
                    "duplicates_count",
                    "stored_rel_path",
                    "batch_id",
                ]
                if c in history_df.columns
            ]
            st.dataframe(with_ru_columns(history_df[history_cols] if history_cols else history_df), use_container_width=True, hide_index=True)

    if uploaded is not None:
        uploaded_bytes = uploaded.getvalue() if hasattr(uploaded, "getvalue") else uploaded.read()
        if not uploaded_bytes:
            st.error("Файл прочитан пустым. Перезагрузи файл и повтори импорт.")
            return
        temp_path = Path("data/_import_temp.xlsx")
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path.write_bytes(uploaded_bytes)

        excel_info = inspect_excel_sheets(uploaded_bytes)
        import_mode = st.radio(
            "Режим определения структуры Excel",
            options=["Автоопределение", "Ручной выбор листа и строки заголовка"],
            horizontal=True,
            key="import_mode",
        )

        selected_sheet = None
        selected_header_row_zero = None
        if excel_info.get("ok"):
            sheets = excel_info.get("sheets") or []
            preview = excel_info.get("preview") or []
            recommended_by_sheet = {item.get("sheet"): int(item.get("recommended_header_row") or 1) for item in preview}
            if sheets:
                preview_df_rows = []
                for item in preview:
                    rows = item.get("rows") or []
                    recommended = int(item.get("recommended_header_row") or 1)
                    row_obj = next((r for r in rows if int(r.get("row", 0)) == recommended), {})
                    preview_df_rows.append(
                        {
                            "Лист": item.get("sheet"),
                            "Строки-превью": len(rows),
                            "Рекоменд. строка заголовка": recommended,
                            "Sample рекоменд. строки": row_obj.get("sample", ""),
                        }
                    )
                if preview_df_rows:
                    st.dataframe(pd.DataFrame(preview_df_rows), use_container_width=True, hide_index=True)

            if import_mode == "Ручной выбор листа и строки заголовка" and sheets:
                c1, c2 = st.columns(2)
                with c1:
                    selected_sheet = st.selectbox("Лист для импорта", options=sheets, index=0, key="manual_import_sheet")
                with c2:
                    default_header_row = int(recommended_by_sheet.get(selected_sheet, 2))
                    header_row_human = st.number_input(
                        "Строка заголовка (1 = первая строка)",
                        min_value=1,
                        max_value=50,
                        value=default_header_row,
                        step=1,
                        key=f"manual_import_header_row_{selected_sheet}",
                    )
                    selected_header_row_zero = int(header_row_human) - 1
        else:
            st.warning(f"Не удалось прочитать структуру Excel: {excel_info.get('message')}")

        if st.button("Импортировать", type="primary", help="Импортировать текущий Excel в мастер-каталог"):
            conn = get_db()
            try:
                if import_mode == "Ручной выбор листа и строки заголовка":
                    result = import_catalog_from_excel(
                        conn,
                        temp_path,
                        sheet_name=selected_sheet,
                        header_row=selected_header_row_zero,
                        default_supplier_name=default_supplier_name or None,
                        default_supplier_url_template=default_supplier_url_template or None,
                    )
                else:
                    result = import_catalog_from_excel(
                        conn,
                        temp_path,
                        default_supplier_name=default_supplier_name or None,
                        default_supplier_url_template=default_supplier_url_template or None,
                    )
                uploaded_record = persist_uploaded_file(
                    conn=conn,
                    storage_kind="supplier_catalog",
                    original_file_name=getattr(uploaded, "name", None),
                    file_bytes=uploaded_bytes,
                    batch_id=result.batch_id,
                    metadata={
                        "import_mode": import_mode,
                        "selected_sheet": selected_sheet,
                        "header_row_zero_based": selected_header_row_zero,
                        "auto_match_ozon_after_import": bool(auto_match_ozon_after_import),
                        "auto_seed_ozon_attrs_after_import": bool(auto_seed_ozon_attrs_after_import),
                    },
                )
                record_catalog_import_history(
                    conn=conn,
                    batch_id=result.batch_id,
                    uploaded_file_id=int(uploaded_record["id"]),
                    original_file_name=getattr(uploaded, "name", None),
                    supplier_name=default_supplier_name or None,
                    supplier_url_template=default_supplier_url_template or None,
                    selected_sheet=selected_sheet,
                    header_row=(int(selected_header_row_zero) + 1) if selected_header_row_zero is not None else None,
                    imported_count=int(result.imported),
                    created_count=int(result.created),
                    updated_count=int(result.updated),
                    duplicates_count=int(len(result.duplicates)),
                    notes="Импорт из вкладки Импорт",
                )
                if auto_match_ozon_after_import and result.batch_id:
                    batch_rows = conn.execute(
                        "SELECT id FROM products WHERE import_batch_id = ? ORDER BY id DESC LIMIT 20000",
                        (result.batch_id,),
                    ).fetchall()
                    batch_ids = [int(r["id"]) for r in batch_rows]
                    if batch_ids:
                        ozon_match_result = bulk_assign_ozon_categories(conn, batch_ids, min_score=OZON_CATEGORY_MIN_SCORE, force=False)
                        if ozon_match_result.get("message"):
                            st.info(str(ozon_match_result["message"]))
                        else:
                            st.caption(
                                f"Ozon автопривязка: обработано {ozon_match_result['processed']}, "
                                f"привязано {ozon_match_result['assigned']}, пропущено {ozon_match_result['skipped']}"
                            )
                        if auto_seed_ozon_attrs_after_import:
                            seeded = ensure_ozon_requirements_for_products(conn, batch_ids)
                            st.caption(
                                "Ozon-атрибуты категорий подтянуты: "
                                f"товаров с Ozon-категорией {int(seeded.get('products_with_ozon_category') or 0)} из {int(seeded.get('products_total') or 0)}, "
                                f"пар категорий {int(seeded.get('category_pairs') or 0)}, "
                                f"импортировано атрибутов {int(seeded.get('imported_attributes') or 0)}."
                            )
                batch_df = load_products(conn, limit=1000, import_batch_id=result.batch_id)
                missing_supplier_count = conn.execute(
                    """
                    SELECT COUNT(*) AS c
                    FROM products
                    WHERE import_batch_id = ?
                      AND (supplier_name IS NULL OR TRIM(supplier_name) = '')
                    """,
                    (result.batch_id,),
                ).fetchone()["c"]
            except sqlite3.OperationalError as e:
                conn.close()
                st.error(f"Ошибка базы при импорте: {e}")
                st.info("Попробуй автоопределение или выбери другой лист/строку заголовка. Если ошибка повторяется, база требует миграции.")
                return
            except Exception as e:
                conn.close()
                st.error(f"Ошибка импорта: {e}")
                return
            conn.close()
            backup_result = backup_database_file(reason="catalog_import")
            st.session_state["last_import_batch_id"] = result.batch_id
            st.success(
                f"Импорт завершён. Всего: {result.imported}, создано: {result.created}, обновлено: {result.updated}, дублей: {len(result.duplicates)}"
            )
            if backup_result.get("ok"):
                st.caption(f"Память каталога зафиксирована в backup: `{Path(str(backup_result['path'])).name}`")
            if int(missing_supplier_count or 0) > 0:
                st.warning(
                    f"У {int(missing_supplier_count)} товаров в этой партии не назначен поставщик. "
                    "Назначь его массово во вкладке Каталог."
                )
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Импортировано", int(result.imported))
            c2.metric("Создано", int(result.created))
            c3.metric("Обновлено", int(result.updated))
            c4.metric("Дублей", int(len(result.duplicates)))

            st.markdown("### Последняя загруженная партия")
            if not batch_df.empty:
                st.dataframe(with_ru_columns(batch_df), use_container_width=True, hide_index=True)
            else:
                st.info("В текущей партии нет отображаемых записей. Попробуй ручной выбор листа и строки заголовка.")

            if result.duplicates:
                st.dataframe(with_ru_columns(pd.DataFrame(result.duplicates)), use_container_width=True)


def show_catalog_tab():
    conn = get_db()
    parser_settings = load_parser_settings(conn)
    ai_settings = load_ai_settings(conn)
    st.subheader("Каталог")
    st.caption("Реестр товаров для массовой работы: фильтр, shortlist, переход в карточку и запуск пайплайна по выборке.")

    with st.container(border=True):
        st.markdown("### Реестр товаров")
        st.caption("Категория опирается прежде всего на Ozon-эталон, а затем на категории из каталога.")

        category_values = list_catalog_categories(conn)
        supplier_values = list_distinct_values(conn, "supplier_name")
        supplier_profile_values = [str(p["supplier_name"]) for p in list_supplier_profiles(conn, only_active=True)]
        supplier_values = sorted(set(supplier_values + supplier_profile_values))
        c1, c2, c3, c4, c5, c6 = st.columns([2.3, 1.2, 1.1, 1, 1, 1])
        with c1:
            search = st.text_input("Поиск", placeholder="Название / артикул / штрихкод")
        with c2:
            category_option = st.selectbox("Категория", options=["Все"] + category_values, index=0)
        with c3:
            supplier_option = st.selectbox("Поставщик", options=["Все"] + supplier_values, index=0)
        with c4:
            page_size = st.selectbox("На странице", options=[50, 100, 200, 500], index=1)
        with c5:
            only_last_batch = st.checkbox("Последняя загрузка", value=False)
        with c6:
            parse_filter = st.selectbox("Парсинг", ["Все", "Есть supplier_url", "Не парсено", "Ошибка", "Успех"], index=0)

        pset1, pset2, pset3 = st.columns([1.25, 1.25, 3.5])
        with pset1:
            ps_include_without_url = st.checkbox(
                "Без supplier_url тоже",
                value=bool(st.session_state.get("catalog_enrich_include_without_url", True)),
                key="catalog_enrich_include_without_url",
                help="Если включено, товары без URL поставщика тоже пойдут в fallback-поиск.",
            )
        with pset2:
            pre_ozon_before_enrich = st.checkbox(
                "Сначала Ozon-match",
                value=bool(st.session_state.get("catalog_pre_ozon_before_enrich", True)),
                key="catalog_pre_ozon_before_enrich",
                help="Рекомендуется держать включённым: сначала Ozon-категория, потом парсинг и AI.",
            )
        with pset3:
            st.caption(
                f"Стратегия парсинга сейчас: `{str(parser_settings.get('source_strategy') or 'auto_full')}`. "
                "Постоянные настройки живут в разделе `Настройки`."
            )
    category = "" if category_option == "Все" else category_option
    supplier = "" if supplier_option == "Все" else supplier_option

    batch_id = st.session_state.get("last_import_batch_id") if only_last_batch else ""
    total_rows = count_products(
        conn,
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=batch_id or "",
        parse_filter=parse_filter,
    )
    total_pages = max(1, int(math.ceil(total_rows / max(int(page_size), 1))))
    page_options = list(range(1, total_pages + 1))
    current_page = int(st.session_state.get("catalog_page_current", st.session_state.get("catalog_page", 1)))
    if current_page > total_pages:
        current_page = 1
    tb1, tb2, tb3, tb4 = st.columns([1, 1, 1, 2.2])
    with tb1:
        page = st.selectbox("Страница", options=page_options, index=page_options.index(current_page), key="catalog_page_widget")
        st.session_state["catalog_page_current"] = int(page)
    with tb2:
        st.metric("Страниц", total_pages)
    with tb3:
        st.metric("Всего товаров", int(total_rows))
    with tb4:
        st.caption(f"Показана выборка для страницы {int(page)}. Ниже таблица, shortlist и быстрый переход в карточку.")
    offset = (int(page) - 1) * int(page_size)
    df = load_products(
        conn,
        search=search,
        category=category,
        supplier=supplier,
        limit=int(page_size),
        offset=int(offset),
        import_batch_id=batch_id or "",
        parse_filter=parse_filter,
    )

    if df.empty:
        st.info("Нет товаров")
        conn.close()
        return

    if batch_id:
        st.caption("Показана только последняя загруженная партия")

    page_product_ids = [int(x) for x in df["id"].tolist()]
    page_service_signals = _load_service_signal_map(conn, page_product_ids)
    page_latest_sources = _load_latest_field_source_type_map(conn, page_product_ids)

    summary1, summary2, summary3, summary4, summary5 = st.columns([1, 1, 1, 1, 1.2])
    summary1.metric("На странице", int(len(df)))
    summary2.metric("С supplier_url", int((df["supplier_url"].fillna("").astype(str).str.strip() != "").sum()) if "supplier_url" in df.columns else 0)
    summary3.metric("Парсинг ок", int((df["supplier_parse_status"] == "success").sum()) if "supplier_parse_status" in df.columns else 0)
    summary4.metric("С Ozon", int((df["ozon_description_category_id"].notna()).sum()) if "ozon_description_category_id" in df.columns else 0)
    with summary5:
        st.markdown("<br>", unsafe_allow_html=True)
        st.download_button(
            "Excel страницы",
            data=export_current_df(df),
            file_name=f"pim_products_page_{int(page)}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    operational_df = build_catalog_operational_view(
        conn,
        df,
        service_signals=page_service_signals,
        latest_sources=page_latest_sources,
    )
    operational_df_display = operational_df.copy() if operational_df is not None else pd.DataFrame()
    if operational_df is not None and not operational_df.empty:
        queue_counts = operational_df["queue"].fillna("Без очереди").astype(str).value_counts().to_dict()
        oq1, oq2, oq3, oq4 = st.columns(4)
        oq1.metric("Нужен parser-run", int(queue_counts.get("Нужен parser-run", 0)))
        oq2.metric("Нужен AI-run", int(queue_counts.get("Нужен AI-run", 0)) + int(queue_counts.get("Низкая уверенность parser/AI", 0)))
        oq3.metric("Фото < 3", int(queue_counts.get("Фото меньше 3", 0)) + int(queue_counts.get("Нет фото", 0)))
        oq4.metric("Готово к AI/клиенту", int(queue_counts.get("Готово к AI/клиенту", 0)))
        queue_priority = [
            "Нужен parser-run",
            "Не найден релевантный товар",
            "Сайт блокирует доступ",
            "Ошибка parser-flow",
            "Нет supplier-домена",
            "Нет Ozon-категории",
            "Нужен AI-run",
            "Низкая уверенность parser/AI",
            "AI отклонил parser result",
            "Нет фото",
            "Фото меньше 3",
            "Нет штрихкода",
            "Готово к AI/клиенту",
        ]
        queue_options = ["Все на странице"] + [q for q in queue_priority if q in queue_counts]
        selected_queue = st.selectbox(
            "Операционная очередь страницы",
            options=queue_options,
            index=0,
            key="catalog_operational_queue_filter",
            help="Это рабочий фокус по текущей странице каталога: сначала добивай проблемные очереди, а не все товары подряд.",
        )
        if selected_queue != "Все на странице":
            operational_df_display = operational_df[operational_df["queue"] == selected_queue].copy()
            st.caption(f"Рабочий фокус страницы: `{selected_queue}`.")
        else:
            st.caption("Рабочий фокус страницы: показаны все товары текущей страницы.")
    media_settings = load_media_settings(conn)
    media_public_base_url = str(media_settings.get("public_base_url") or "").strip()

    ids = df["id"].tolist()
    page_selector_options = [int(x) for x in ids]
    selected_page_ids = st.multiselect(
        "Shortlist на этой странице",
        options=page_selector_options,
        default=[int(st.session_state.get("selected_product_id"))] if int(st.session_state.get("selected_product_id") or 0) in page_selector_options else [],
        format_func=lambda x: next(
            (
                f"{str(row.get('article') or row.get('supplier_article') or row.get('internal_article') or '-')} | "
                f"{_short_text(row.get('name'), 54)}"
                for _, row in df.iterrows()
                if int(row["id"]) == int(x)
            ),
            f"ID {x}",
        ),
        key="catalog_selected_page_ids",
        help="Используй этот список, если нужно запустить массовое действие только по shortlist, а не по всей странице или всему фильтру.",
    )
    st.session_state["template_selected_ids_from_catalog"] = [int(x) for x in selected_page_ids]
    with st.container(border=True):
        st.markdown("### Быстрый переход и фокус")
        focus1, focus2 = st.columns([3, 2])
        with focus1:
            selected_id = st.selectbox(
                "Открыть карточку товара",
                ids,
                format_func=lambda x: next(
                    (
                        f"{str(row.get('article') or row.get('supplier_article') or row.get('internal_article') or '-')} | "
                        f"{str(row.get('name') or '-')} | "
                        f"{str(row.get('supplier_name') or '-')} | "
                        f"{str(row.get('ozon_category_path') or row.get('category') or '-')}"
                        for _, row in df.iterrows()
                        if int(row["id"]) == int(x)
                    ),
                    f"ID {x}",
                ),
            )
        with focus2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Открыть выбранный товар в Карточке", type="primary", key="catalog_open_selected_in_card"):
                st.session_state["selected_product_id"] = int(selected_id)
                request_workspace_navigation("product")

        selected_row = next((row for _, row in df.iterrows() if int(row["id"]) == int(selected_id)), None)
        if selected_row is not None:
            selected_row_dict = selected_row.to_dict()
            selected_service_state = page_service_signals.get(int(selected_row_dict.get("id") or 0), {})
            selected_latest_sources = page_latest_sources.get(int(selected_row_dict.get("id") or 0), {})
            quick_ozon_ready = compute_quick_ozon_readiness(conn, selected_row_dict)
            quick_template_ready = compute_best_template_profile_readiness(conn, selected_row_dict)
            filled_core, total_core = _product_core_fill_stats(selected_row_dict)
            parser_stage_label, parser_queue_label = _parser_stage_info(selected_row_dict)
            ai_stage_label, ai_queue_label = _ai_stage_info(selected_row_dict, selected_service_state, selected_latest_sources)
            image_stage_label, image_queue_label, image_photo_count = _image_stage_info(selected_row_dict, selected_service_state)
            sf1, sf2, sf3, sf4, sf5, sf6, sf7, sf8 = st.columns(8)
            sf1.metric("Артикул", str(selected_row_dict.get("article") or selected_row_dict.get("supplier_article") or selected_row_dict.get("internal_article") or "-"))
            sf2.metric("Статус карточки", _product_stage_label(selected_row_dict))
            sf3.metric("Парсинг", _parse_status_label(selected_row_dict.get("supplier_parse_status")))
            sf4.metric("Заполнено ядро", f"{filled_core}/{total_core}")
            sf5.metric("Ozon ready", f"{int(quick_ozon_ready.get('readiness_pct') or 0)}%")
            sf6.metric("Шаблон ready", f"{int(quick_template_ready.get('readiness_pct') or 0)}%")
            sf7.metric("AI stage", ai_stage_label)
            sf8.metric("Фото", int(image_photo_count))
            st.caption(
                f"Поставщик: {selected_row_dict.get('supplier_name') or '-'} | "
                f"Ozon: {selected_row_dict.get('ozon_category_path') or selected_row_dict.get('category') or '-'} | "
                f"Фото stage: {image_stage_label} | "
                f"Штрихкод: {_barcode_status_label(selected_row_dict)}"
            )
            st.caption(
                f"Parser stage: {parser_stage_label} | AI очередь: {ai_queue_label} | "
                f"Фото очередь: {image_queue_label} | Рабочая очередь: {_operational_queue_label(selected_row_dict, selected_service_state, selected_latest_sources)}"
            )
            parse_comment = str(selected_row_dict.get("supplier_parse_comment") or "").strip()
            if parse_comment and str(selected_row_dict.get("supplier_parse_status") or "").strip().lower() == "error":
                st.error(f"Причина ошибки парсинга: {parse_comment}")
            elif parse_comment and str(selected_row_dict.get("supplier_parse_status") or "").strip().lower() == "success":
                st.caption(f"Комментарий парсинга: {parse_comment}")
            if int(quick_ozon_ready.get("required_total") or 0) > 0:
                st.caption(
                    f"Ozon: заполнено {int(quick_ozon_ready.get('required_filled') or 0)} из {int(quick_ozon_ready.get('required_total') or 0)} обязательных."
                )
            if int(quick_template_ready.get("profiles_total") or 0) > 0:
                st.caption(
                    f"Лучший клиентский профиль: {quick_template_ready.get('channel_code') or '-'} / "
                    f"{quick_template_ready.get('profile_name') or '-'} "
                    f"({int(quick_template_ready.get('filled_columns') or 0)}/{int(quick_template_ready.get('matched_columns') or 0)})."
                )
            selected_gallery_urls = _collect_product_gallery_urls(
                conn,
                int(selected_row_dict["id"]),
                fallback_image_url=str(selected_row_dict.get("image_url") or ""),
                public_base_url=media_public_base_url,
            )
            if selected_gallery_urls:
                gallery_preview = selected_gallery_urls[:4]
                preview_cols = st.columns(len(gallery_preview))
                for idx, img_url in enumerate(gallery_preview):
                    with preview_cols[idx]:
                        st.image(str(img_url), caption=f"Фото {idx + 1}", use_container_width=True)
    supplier_candidate_ids = [
        int(row["id"])
        for _, row in df.iterrows()
        if str(row.get("supplier_url") or "").strip()
    ]
    filtered_supplier_candidate_ids = load_product_ids_with_supplier_url(
        conn,
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=batch_id or "",
        parse_filter=parse_filter,
        limit=None,
        offset=0,
    )
    all_filtered_candidate_ids = load_product_ids(
        conn,
        search=search,
        category=category,
        supplier=supplier,
        import_batch_id=batch_id or "",
        parse_filter=parse_filter,
        limit=None,
        offset=0,
    )
    include_without_url = bool(st.session_state.get("catalog_enrich_include_without_url", True))
    effective_filtered_candidate_ids = (
        [int(x) for x in all_filtered_candidate_ids]
        if include_without_url
        else [int(x) for x in filtered_supplier_candidate_ids]
    )
    selected_shortlist_ids = [int(x) for x in selected_page_ids]
    effective_selected_shortlist_ids = list(selected_shortlist_ids)
    if not effective_selected_shortlist_ids and int(selected_id or 0) > 0:
        # If the operator picked a single product in the focus block but did not
        # add anything to the shortlist, treat that focused product as the
        # working selection for "Selected products" actions.
        effective_selected_shortlist_ids = [int(selected_id)]
    page_zip_signature = f"page:{int(page)}:{hashlib.sha1(','.join([str(int(x)) for x in ids]).encode('utf-8')).hexdigest()}"
    filtered_zip_signature = f"filtered:{hashlib.sha1(','.join([str(int(x)) for x in all_filtered_candidate_ids]).encode('utf-8')).hexdigest()}"
    selected_zip_signature = f"selected:{hashlib.sha1(','.join([str(int(x)) for x in effective_selected_shortlist_ids]).encode('utf-8')).hexdigest()}" if effective_selected_shortlist_ids else "selected:none"

    catalog_scope_options = ["Выбранные товары", "Текущая страница", "Вся выборка по фильтру"]

    def resolve_catalog_scope_ids(scope_value: str) -> list[int]:
        if scope_value == "Выбранные товары":
            return [int(x) for x in effective_selected_shortlist_ids]
        if scope_value == "Текущая страница":
            return [int(x) for x in ids]
        return [int(x) for x in effective_filtered_candidate_ids]

    def load_catalog_brief_rows(product_ids: list[int]) -> list[dict]:
        ids_local = [int(x) for x in product_ids if str(x).strip()]
        if not ids_local:
            return []
        rows: list[dict] = []
        chunk_size = 900
        for start in range(0, len(ids_local), chunk_size):
            chunk = ids_local[start : start + chunk_size]
            placeholders = ", ".join(["?"] * len(chunk))
            fetched = conn.execute(
                f"""
                SELECT
                    id,
                    article,
                    supplier_article,
                    name,
                    supplier_name,
                    ozon_description_category_id,
                    ozon_type_id,
                    ozon_category_path,
                    detmir_category_id,
                    detmir_category_path,
                    supplier_parse_status,
                    image_url
                FROM products
                WHERE id IN ({placeholders})
                """,
                tuple(chunk),
            ).fetchall()
            rows.extend([dict(row) for row in fetched])
        return rows

    def infer_catalog_template_category(product_rows: list[dict]) -> str:
        pair_counter: dict[str, int] = {}
        for row in product_rows:
            desc_id = _safe_int_id(row.get("ozon_description_category_id"))
            type_id = _safe_int_id(row.get("ozon_type_id"))
            if desc_id <= 0 or type_id <= 0:
                continue
            code = f"ozon:{desc_id}:{type_id}"
            pair_counter[code] = int(pair_counter.get(code, 0)) + 1
        if not pair_counter:
            return ""
        return max(sorted(pair_counter.items()), key=lambda item: (int(item[1]), str(item[0])))[0]

    bc1, bc2, bc3 = st.columns(3)
    with bc1:
        max_bulk_enrich_page = st.number_input(
            "Лимит (текущая страница)",
            min_value=1,
            max_value=1000,
            value=min(20, max(1, len(supplier_candidate_ids))),
            step=1,
            help="Сколько товаров обогащать за один запуск по текущей странице.",
            key="catalog_max_bulk_enrich_page",
        )
    with bc2:
        max_bulk_enrich_filtered = st.number_input(
            "Лимит (вся выборка фильтра)",
            min_value=1,
            max_value=10000,
            value=min(300, max(1, len(filtered_supplier_candidate_ids))),
            step=10,
            help="Сколько товаров обогащать за один запуск по всей выборке фильтров.",
            key="catalog_max_bulk_enrich_filtered",
        )
    with bc3:
        supplier_timeout_seconds = st.number_input(
            "Таймаут supplier_url, сек",
            min_value=2,
            max_value=30,
            value=int(float(parser_settings.get("timeout_seconds", 8.0))),
            step=1,
            help="Максимальное время ожидания ответа от сайта поставщика для одного товара.",
            key="catalog_supplier_timeout_seconds",
        )
    enrich_force = st.checkbox(
        "Перезаписывать значения (force, кроме manual)",
        value=False,
        help="Если включено, enrichment сможет перезаписывать не пустые значения, но manual-поля останутся защищены.",
        key="catalog_enrich_force",
    )
    st.caption(
        f"Кандидатов: страница {len(supplier_candidate_ids)}, вся выборка {len(filtered_supplier_candidate_ids)}. "
        f"выбрано вручную: {len(selected_shortlist_ids)}. "
        f"включая без supplier_url: {len(all_filtered_candidate_ids)}. "
        f"Лимиты: страница {int(max_bulk_enrich_page)}, выборка {int(max_bulk_enrich_filtered)}. "
        f"Ozon перед обогащением: {'вкл' if bool(pre_ozon_before_enrich) else 'выкл'}."
    )
    if not selected_shortlist_ids and effective_selected_shortlist_ids:
        st.caption(
            "Shortlist пуст, поэтому для режима `Выбранные товары` будет использован товар из блока `Быстрый переход и фокус`."
        )
    ai_cfg_ok, ai_cfg_msg = ai_is_configured(ai_settings)
    with st.expander("Тонкая настройка массового наполнения", expanded=False):
        if ai_cfg_ok:
            st.success(ai_cfg_msg)
        else:
            st.warning(ai_cfg_msg)
        ai_mode = st.radio(
            "AI-режим конвейера",
            options=["fast_batch", "deep_repair"],
            index=0,
            horizontal=True,
            key="catalog_ai_mode",
            format_func=lambda x: "Fast batch" if x == "fast_batch" else "Deep repair",
            help="Fast batch — обязательный массовый verifier + rewrite после parser. Deep repair — более тяжёлый режим для спорных и сложных SKU.",
        )
        aic1, aic2, aic3, aic4 = st.columns(4)
        with aic1:
            ai_batch_include_supplier = st.checkbox(
                "Сначала найти данные на сайте и в интернете",
                value=True,
                key="catalog_ai_include_supplier",
                help="Сначала попробует найти карточку товара по домену поставщика и через web fallback, затем передаст результат в AI-слой.",
            )
        with aic2:
            ai_batch_include_title = st.checkbox(
                "AI-рерайт названия",
                value=True,
                key="catalog_ai_include_title",
            )
        with aic3:
            ai_batch_include_description = st.checkbox(
                "AI-рерайт описания",
                value=True,
                key="catalog_ai_include_description",
            )
        with aic4:
            ai_batch_include_attributes = st.checkbox(
                "AI-дозаполнение Ozon атрибутов",
                value=True,
                key="catalog_ai_include_attributes",
            )
        st.caption(
            "Этот режим уже ближе к целевому conveyor: домен поставщика -> поиск карточки товара -> AI verifier -> массовый рерайт -> Ozon/client-ready данные. "
            "Fast batch нужен для основного потока, Deep repair оставляй для тяжёлых SKU и ручного восстановления."
        )
    dcfg1, dcfg2 = st.columns([1, 2])
    with dcfg1:
        dim_min_samples = st.number_input(
            "Мин. выборка для статистики",
            min_value=1,
            max_value=50,
            value=4,
            step=1,
            key="catalog_dim_min_samples",
            help="Сколько похожих товаров нужно минимум, чтобы использовать статистический расчет.",
        )
    with dcfg2:
        dim_force = st.checkbox(
            "Перезаписывать существующие габариты/вес",
            value=False,
            key="catalog_dim_force",
            help="Если выключено, расчет заполняет только пустые логистические поля.",
        )

    def run_supplier_enrichment_batch(candidate_ids: list[int], run_limit: int, run_label: str) -> None:
        if not candidate_ids:
            st.info(f"Для режима `{run_label}` нет товаров для обогащения.")
            return
        target_ids = candidate_ids[: int(run_limit)]
        ozon_processed = 0
        ozon_assigned = 0
        ozon_attr_imported = 0
        if bool(pre_ozon_before_enrich) and target_ids:
            pre_res = bulk_assign_ozon_categories(
                conn,
                [int(x) for x in target_ids],
                min_score=OZON_CATEGORY_MIN_SCORE,
                force=False,
            )
            ozon_processed = int(pre_res.get("processed") or 0)
            ozon_assigned = int(pre_res.get("assigned") or 0)
            seeded = ensure_ozon_requirements_for_products(conn, [int(x) for x in target_ids])
            ozon_attr_imported = int(seeded.get("imported_attributes") or 0)
        progress = st.progress(0)
        processed = 0
        success = 0
        failed = 0
        used_fallback = 0
        resolved_from_listing = 0
        failed_details: list[str] = []
        for i, pid in enumerate(target_ids, start=1):
            current_row = get_product(conn, int(pid))
            current_supplier_url = str(current_row["supplier_url"] or "").strip() if current_row and "supplier_url" in current_row.keys() else ""
            current_article = (
                str(current_row["article"] or current_row["supplier_article"] or current_row["internal_article"] or f"ID {int(pid)}")
                if current_row and hasattr(current_row, "keys")
                else f"ID {int(pid)}"
            )
            try:
                result = enrich_product_from_supplier(
                    conn,
                    int(pid),
                    force=bool(enrich_force),
                    timeout_seconds=float(supplier_timeout_seconds),
                    parser_settings=parser_settings,
                )
                if result.get("ok"):
                    success += 1
                    if "fallback" in str(result.get("source_type") or ""):
                        used_fallback += 1
                    if str(result.get("source_url") or "").strip() and str(result.get("source_url") or "").strip() != current_supplier_url:
                        resolved_from_listing += 1
                else:
                    failed += 1
                    if len(failed_details) < 5:
                        failed_details.append(f"{current_article}: {str(result.get('message') or 'ошибка без текста')[:220]}")
            except Exception:
                failed += 1
                if len(failed_details) < 5:
                    failed_details.append(f"{current_article}: внутренняя ошибка запуска enrichment")
            processed += 1
            progress.progress(i / len(target_ids))
        skipped_by_limit = max(0, len(candidate_ids) - len(target_ids))
        st.success(
            f"[{run_label}] Обогащение завершено: обработано {processed}, успешно {success}, ошибок {failed}, "
            f"fallback {used_fallback}, listing->product {resolved_from_listing}, "
            f"Ozon автопривязка до парсинга: обработано {ozon_processed}, назначено {ozon_assigned}, "
            f"подтянуто Ozon-атрибутов {ozon_attr_imported}, "
            f"отложено по лимиту {skipped_by_limit}."
        )
        if failed_details:
            st.warning("Причины последних ошибок parser-flow:\n\n- " + "\n- ".join(failed_details))

    def run_ai_enrichment_batch(candidate_ids: list[int], run_limit: int, run_label: str, mode: str = "fast_batch") -> None:
        if not candidate_ids:
            st.info(f"Для режима `{run_label}` нет товаров для AI-дозаполнения.")
            return
        if not ai_cfg_ok:
            st.warning(f"AI не настроен: {ai_cfg_msg}")
            return
        target_ids = [int(x) for x in candidate_ids[: int(run_limit)]]
        ozon_processed = 0
        ozon_assigned = 0
        ozon_attr_imported = 0
        if target_ids:
            pre_res = bulk_assign_ozon_categories(
                conn,
                target_ids,
                min_score=OZON_CATEGORY_MIN_SCORE,
                force=False,
            )
            ozon_processed = int(pre_res.get("processed") or 0)
            ozon_assigned = int(pre_res.get("assigned") or 0)
            seeded = ensure_ozon_requirements_for_products(conn, target_ids)
            ozon_attr_imported = int(seeded.get("imported_attributes") or 0)
        progress = st.progress(0)
        processed = 0
        supplier_success = 0
        ai_success = 0
        ai_errors = 0
        title_applied = 0
        description_applied = 0
        attributes_saved = 0
        photos_found = 0
        ai_verified = 0
        ai_rejected = 0
        ai_rewrite_ready = 0
        image_ready = 0
        image_under_min = 0
        for i, pid in enumerate(target_ids, start=1):
            try:
                if bool(ai_batch_include_supplier):
                    supplier_res = enrich_product_from_supplier(
                        conn,
                        int(pid),
                        force=bool(enrich_force),
                        timeout_seconds=float(supplier_timeout_seconds),
                        parser_settings=parser_settings,
                    )
                    if supplier_res.get("ok"):
                        supplier_success += 1
                        photos_found += len(supplier_res.get("image_urls") or [])
                ai_res = run_ai_enrichment_for_product(
                    conn=conn,
                    product_id=int(pid),
                    settings=ai_settings,
                    include_title=bool(ai_batch_include_title),
                    include_description=bool(ai_batch_include_description),
                    include_attributes=bool(ai_batch_include_attributes),
                    force=bool(enrich_force),
                    mode=str(mode or "fast_batch"),
                )
                if ai_res.get("ok"):
                    ai_success += 1
                    verdict = str(ai_res.get("verification_verdict") or "").strip().lower()
                    if verdict == "accept":
                        ai_verified += 1
                    elif verdict == "reject":
                        ai_rejected += 1
                    if bool(ai_res.get("title_applied")) and bool(ai_res.get("description_applied")):
                        ai_rewrite_ready += 1
                    if bool(ai_res.get("title_applied")):
                        title_applied += 1
                    if bool(ai_res.get("description_applied")):
                        description_applied += 1
                    attributes_saved += int(ai_res.get("attributes_saved") or 0)
                    if str(ai_res.get("image_stage") or "").strip() in {"target_ready", "rich_gallery"}:
                        image_ready += 1
                    elif str(ai_res.get("image_stage") or "").strip() in {"no_main_image", "under_min"}:
                        image_under_min += 1
                    if ai_res.get("errors"):
                        ai_errors += 1
                else:
                    ai_errors += 1
            except Exception:
                ai_errors += 1
            processed += 1
            progress.progress(i / len(target_ids))
        skipped_by_limit = max(0, len(candidate_ids) - len(target_ids))
        st.success(
            f"[{run_label}] AI-пакет ({str(mode or 'fast_batch')}) завершён: обработано {processed}, "
            f"supplier/web ok {supplier_success}, AI ok {ai_success}, AI ошибок {ai_errors}, "
            f"AI verified {ai_verified}, AI rejected {ai_rejected}, rewrite ready {ai_rewrite_ready}, "
            f"новых названий {title_applied}, описаний {description_applied}, "
            f"сохранено AI-атрибутов {attributes_saved}, найдено фото {photos_found}, "
            f"image ready {image_ready}, image gaps {image_under_min}, "
            f"Ozon категорий назначено {ozon_assigned} из {ozon_processed}, "
            f"подтянуто Ozon-атрибутов {ozon_attr_imported}, "
            f"отложено по лимиту {skipped_by_limit}."
        )

    def run_dimension_estimation_batch(candidate_ids: list[int], run_limit: int, run_label: str) -> None:
        if not candidate_ids:
            st.info(f"Для режима `{run_label}` нет товаров для расчета.")
            return
        target_ids = [int(x) for x in candidate_ids[: int(run_limit)]]
        progress = st.progress(0)
        processed = 0
        updated_products = 0
        updated_fields_total = 0
        source_counter: dict[str, int] = {}
        for i, pid in enumerate(target_ids, start=1):
            try:
                result = estimate_dimensions_for_product(
                    conn=conn,
                    product_id=int(pid),
                    force=bool(dim_force),
                    min_samples=int(dim_min_samples),
                )
                if bool(result.get("ok")) and int(result.get("updated_fields") or 0) > 0:
                    updated_products += 1
                    updated_fields_total += int(result.get("updated_fields") or 0)
                    for src in (result.get("used_sources") or []):
                        source_counter[str(src)] = int(source_counter.get(str(src), 0)) + 1
            except Exception:
                pass
            processed += 1
            progress.progress(i / len(target_ids))
        skipped_by_limit = max(0, len(candidate_ids) - len(target_ids))
        source_text = ", ".join([f"{k}:{v}" for k, v in sorted(source_counter.items())]) if source_counter else "-"
        st.success(
            f"[{run_label}] Расчёт завершён: обработано {processed}, "
            f"обновлено товаров {updated_products}, полей {updated_fields_total}, "
            f"источники {source_text}, отложено по лимиту {skipped_by_limit}."
        )

    def run_detmir_overlay_batch(candidate_ids: list[int], run_limit: int, run_label: str) -> None:
        if not candidate_ids:
            st.info(f"Для режима `{run_label}` нет товаров для подготовки Detmir.")
            return
        target_ids = [int(x) for x in candidate_ids[: int(run_limit)]]
        progress = st.progress(0)
        processed = 0
        matched_categories = 0
        missing_categories = 0
        imported_requirements = 0
        overlay_saved = 0
        overlay_skipped = 0
        errors = 0
        for i, pid in enumerate(target_ids, start=1):
            try:
                current_product = get_product(conn, int(pid))
                if not current_product:
                    errors += 1
                    continue
                product_row = dict(current_product)
                detmir_category_id = _safe_int_id(product_row.get("detmir_category_id"))
                if detmir_category_id <= 0:
                    detmir_match = detect_best_detmir_category_for_product(conn, product_row)
                    if detmir_match.get("ok"):
                        matched = detmir_match.get("category") or {}
                        detmir_category_id = _safe_int_id(matched.get("category_id"))
                        payload = {
                            "detmir_category_id": detmir_category_id or None,
                            "detmir_category_path": str(matched.get("full_path") or matched.get("name") or "").strip() or None,
                            "detmir_category_confidence": _detmir_confidence_from_match_score(matched.get("match_score")),
                        }
                        save_product(conn, int(pid), payload)
                        save_field_source(
                            conn=conn,
                            product_id=int(pid),
                            field_name="detmir_category_id",
                            source_type="detmir_category_match",
                            source_value_raw=detmir_category_id,
                            source_url=str(matched.get("full_path") or matched.get("name") or "detmir_match"),
                            confidence=min(0.99, max(0.35, _clamp_unit_confidence(payload.get("detmir_category_confidence")))),
                            is_manual=False,
                        )
                        matched_categories += 1
                        product_row.update(payload)
                    else:
                        missing_categories += 1
                        processed += 1
                        progress.progress(i / len(target_ids))
                        continue

                detmir_scope = f"detmir:{detmir_category_id}" if detmir_category_id > 0 else ""
                if not detmir_scope:
                    missing_categories += 1
                    processed += 1
                    progress.progress(i / len(target_ids))
                    continue

                existing_requirements = list_channel_requirements(conn, channel_code="detmir", category_code=detmir_scope)
                if not existing_requirements:
                    import_result = import_detmir_category_requirements_to_pim(conn, category_id=int(detmir_category_id))
                    imported_requirements += int(import_result.get("imported") or 0)

                gapfill_result = _fill_channel_attrs_from_product_state(
                    conn=conn,
                    product_row=product_row,
                    channel_code="detmir",
                    category_code=detmir_scope,
                    source_type="derived_from_master",
                    source_url="catalog_detmir_overlay_gapfill",
                    force=False,
                    target_channel_code="detmir",
                )
                overlay_saved += int(gapfill_result.get("saved") or 0)
                overlay_skipped += int(gapfill_result.get("skipped") or 0)
            except Exception:
                errors += 1
            processed += 1
            progress.progress(i / len(target_ids))

        skipped_by_limit = max(0, len(candidate_ids) - len(target_ids))
        st.success(
            f"[{run_label}] Detmir overlay готов: обработано {processed}, "
            f"категорий сматчено {matched_categories}, без категории {missing_categories}, "
            f"импортировано требований {imported_requirements}, заполнено overlay-полей {overlay_saved}, "
            f"пропущено {overlay_skipped}, ошибок {errors}, отложено по лимиту {skipped_by_limit}."
        )

    workflow_scope = st.radio(
        "Рабочая область каталога",
        options=catalog_scope_options,
        index=catalog_scope_options.index(str(st.session_state.get("catalog_workflow_scope") or "Выбранные товары")) if str(st.session_state.get("catalog_workflow_scope") or "Выбранные товары") in catalog_scope_options else 0,
        horizontal=True,
        key="catalog_workflow_scope",
        help="Каталог теперь работает как основной конвейер: выбери область, подготовь категории и overlay, затем массово заполни и выгрузи в шаблон клиента.",
    )
    workflow_target_ids = resolve_catalog_scope_ids(workflow_scope)
    workflow_brief_rows = load_catalog_brief_rows(workflow_target_ids)
    workflow_default_category_code = infer_catalog_template_category(workflow_brief_rows)
    workflow_ozon_ready = sum(
        1
        for row in workflow_brief_rows
        if _safe_int_id(row.get("ozon_description_category_id")) > 0 and _safe_int_id(row.get("ozon_type_id")) > 0
    )
    workflow_detmir_ready = sum(1 for row in workflow_brief_rows if _safe_int_id(row.get("detmir_category_id")) > 0)
    workflow_with_images = sum(1 for row in workflow_brief_rows if str(row.get("image_url") or "").strip())
    workflow_parse_success = sum(1 for row in workflow_brief_rows if str(row.get("supplier_parse_status") or "").strip().lower() == "success")

    with st.container(border=True):
        st.markdown("### Основной конвейер каталога")
        st.caption(
            "Здесь теперь главный путь работы: сначала готовим Ozon-ядро, затем client overlays, потом массово наполняем карточки и сразу готовим выгрузку под клиента."
        )
        ws1, ws2, ws3, ws4 = st.columns(4)
        ws1.metric("В работе", int(len(workflow_target_ids)))
        ws2.metric("С Ozon", int(workflow_ozon_ready))
        ws3.metric("С Detmir", int(workflow_detmir_ready))
        ws4.metric("С фото", int(workflow_with_images))
        st.caption(
            f"Область: {workflow_scope}. "
            f"Парсинг успех: {workflow_parse_success}. "
            "Для Excel-клиентов отдельный overlay обычно не нужен: шаблоны читают master-memory напрямую. "
            "Отдельно готовим только API/overlay-каналы вроде Детского Мира."
        )

        flow1, flow2, flow3 = st.columns([1.05, 1.05, 1.35], gap="large")

        with flow1:
            with st.container(border=True):
                st.markdown("#### 1. Категории и схемы")
                st.caption("Подготовить эталон Ozon и client overlay перед массовым заполнением.")
                if st.button("Подготовить Ozon-ядро", type="primary", use_container_width=True, key="catalog_workflow_prepare_ozon"):
                    if not workflow_target_ids:
                        st.warning("В выбранной области нет товаров.")
                    else:
                        prep_res = bulk_assign_ozon_categories(
                            conn,
                            [int(x) for x in workflow_target_ids],
                            min_score=OZON_CATEGORY_MIN_SCORE,
                            force=False,
                        )
                        seeded = ensure_ozon_requirements_for_products(conn, [int(x) for x in workflow_target_ids])
                        materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in workflow_target_ids])
                        st.success(
                            f"Ozon готов: обработано {int(prep_res.get('processed') or 0)}, привязано {int(prep_res.get('assigned') or 0)}, "
                            f"импортировано атрибутов {int(seeded.get('imported_attributes') or 0)}, создано слотов {int(materialized.get('slots_created') or 0)}."
                        )
                        st.rerun()
                if st.button("Подготовить Detmir overlay", use_container_width=True, key="catalog_workflow_prepare_detmir"):
                    run_detmir_overlay_batch(
                        candidate_ids=[int(x) for x in workflow_target_ids],
                        run_limit=max(len(workflow_target_ids), 1),
                        run_label=f"Detmir / {workflow_scope}",
                    )
                if st.button("Открыть Каналы", use_container_width=True, key="catalog_workflow_open_channels"):
                    request_workspace_navigation("channels")
                st.caption(
                    "Если Detmir-категории и атрибуты ещё не синхронизированы, сначала зайди в `Каналы -> Детский Мир API`, затем вернись сюда."
                )

        with flow2:
            with st.container(border=True):
                st.markdown("#### 2. Массовое наполнение")
                st.caption("Массово найти данные по товарам, переписать контент и подготовить master-карточки под клиентов.")
                fill_detmir_after_ai = st.checkbox(
                    "После AI сразу подготовить Detmir overlay",
                    value=True,
                    key="catalog_fill_detmir_after_ai",
                    help="После supplier/web + AI система сразу перенесёт уже найденные значения в Detmir overlay.",
                )
                if st.button("Запустить массовое заполнение для области", type="primary", use_container_width=True, key="catalog_workflow_fill_all"):
                    run_ai_enrichment_batch(
                        candidate_ids=[int(x) for x in workflow_target_ids],
                        run_limit=max(len(workflow_target_ids), 1),
                        run_label=f"AI / {workflow_scope}",
                        mode=str(ai_mode or "fast_batch"),
                    )
                    if bool(fill_detmir_after_ai) and workflow_target_ids:
                        run_detmir_overlay_batch(
                            candidate_ids=[int(x) for x in workflow_target_ids],
                            run_limit=max(len(workflow_target_ids), 1),
                            run_label=f"Detmir-after-AI / {workflow_scope}",
                        )
                st.caption(
                    "Этот шаг должен быть базовым для 90% товаров. `Карточка` остаётся для редких ручных правок, если массовый проход что-то не добрал."
                )

        with flow3:
            with st.container(border=True):
                st.markdown("#### 3. Клиент и выгрузка")
                st.caption("Выбери клиента и категорию, проверь готовность пачки и выгрузи исходный клиентский Excel прямо отсюда.")
                catalog_client_entries = list_client_channels(conn)
                catalog_client_map = {
                    str(item.get("client_code") or "").strip(): item
                    for item in catalog_client_entries
                    if str(item.get("client_code") or "").strip()
                }
                catalog_client_codes = sorted(
                    catalog_client_map.keys(),
                    key=lambda x: (str(catalog_client_map.get(x, {}).get("client_name") or x).lower(), x.lower()),
                )

                def _catalog_client_label(code: str) -> str:
                    item = catalog_client_map.get(str(code or "").strip(), {})
                    client_name = str(item.get("client_name") or "").strip()
                    return f"{client_name} ({code})" if client_name else str(code)

                export_client_code = st.selectbox(
                    "Клиент",
                    options=[""] + catalog_client_codes,
                    index=0,
                    format_func=lambda value: "-- выбери клиента --" if not value else _catalog_client_label(str(value)),
                    key="catalog_export_client_code",
                )
                export_category_options, export_category_labels = _build_ozon_template_category_options(
                    conn,
                    channel_code=export_client_code or None,
                    limit=5000,
                )
                if workflow_default_category_code and workflow_default_category_code in export_category_options and not st.session_state.get("catalog_export_category_code"):
                    st.session_state["catalog_export_category_code"] = workflow_default_category_code
                export_category_code = st.selectbox(
                    "Ozon-категория для профиля/шаблона",
                    options=export_category_options,
                    index=(
                        export_category_options.index(str(st.session_state.get("catalog_export_category_code") or ""))
                        if str(st.session_state.get("catalog_export_category_code") or "") in export_category_options
                        else 0
                    ),
                    format_func=lambda value: export_category_labels.get(str(value), str(value)),
                    key="catalog_export_category_code",
                )

                export_profiles = [
                    p for p in list_template_profiles(conn, channel_code=export_client_code or None)
                    if str(p.get("category_code") or "").strip() == str(export_category_code or "").strip()
                ] if export_client_code else []
                export_profile_options = [None] + [int(p["id"]) for p in export_profiles]
                export_profile_id = st.selectbox(
                    "Профиль шаблона",
                    options=export_profile_options,
                    format_func=lambda value: "-- нет --" if value is None else next(
                        (f"{p['profile_name']} (#{p['id']})" for p in export_profiles if int(p["id"]) == int(value)),
                        str(value),
                    ),
                    key="catalog_export_profile_id",
                )

                export_template_files = list_uploaded_files(
                    conn,
                    storage_kind="client_template",
                    channel_code=export_client_code or None,
                    category_code=export_category_code or None,
                    limit=30,
                ) if export_client_code else []
                export_template_options = [None] + [int(row["id"]) for row in export_template_files]
                export_template_file_id = st.selectbox(
                    "Сохранённый Excel-шаблон",
                    options=export_template_options,
                    format_func=lambda value: "-- нет --" if value is None else next(
                        (
                            f"{row.get('original_file_name') or Path(str(row.get('stored_rel_path') or '')).name} | #{int(row['id'])}"
                            for row in export_template_files
                            if int(row["id"]) == int(value)
                        ),
                        str(value),
                    ),
                    key="catalog_export_template_file_id",
                )

                export_scope_text = f"Пачка для выгрузки: {len(workflow_target_ids)} товаров."
                if workflow_default_category_code:
                    export_scope_text += f" Доминирующая категория области: {workflow_default_category_code}."
                st.caption(export_scope_text)

                export_profile_columns = get_template_profile_columns(conn, int(export_profile_id)) if export_profile_id else []
                export_template_row = next(
                    (row for row in export_template_files if int(row["id"]) == int(export_template_file_id)),
                    None,
                ) if export_template_file_id else None
                export_template_bytes = read_uploaded_file_bytes(conn, int(export_template_file_id)) if export_template_file_id else None
                export_template_metadata = get_uploaded_file_metadata(export_template_row)
                export_sheet_name = str(export_template_metadata.get("sheet_name") or "").strip() or "Товары"
                export_data_start_row = int(export_template_metadata.get("data_start_row") or 2)

                ex1, ex2 = st.columns(2)
                with ex1:
                    if st.button("Открыть пачку в Клиентский шаблон", use_container_width=True, key="catalog_open_template_workspace"):
                        if export_client_code:
                            st.session_state["template_client_code"] = export_client_code
                            st.session_state["template_client_selector"] = export_client_code
                        if export_category_code:
                            st.session_state["template_category_select"] = export_category_code
                        if export_template_file_id:
                            st.session_state["template_saved_file_id"] = int(export_template_file_id)
                        st.session_state["template_selected_ids_from_catalog"] = [int(x) for x in workflow_target_ids]
                        st.session_state["template_selected_ids"] = [int(x) for x in workflow_target_ids]
                        request_workspace_navigation("template")
                with ex2:
                    if st.button("Проверить readiness пачки", use_container_width=True, key="catalog_export_check_readiness"):
                        if not workflow_target_ids:
                            st.warning("В выбранной области нет товаров.")
                        elif not export_profile_columns:
                            st.warning("Сначала выбери профиль шаблона клиента.")
                        elif not export_template_bytes:
                            st.warning("Сначала выбери сохранённый Excel-шаблон клиента.")
                        else:
                            safe_catalog_template_bytes = sanitize_template_xlsx_bytes(export_template_bytes)
                            catalog_template_df = read_client_template_dataframe(safe_catalog_template_bytes, sheet_name=export_sheet_name)
                            catalog_filled_df = fill_template_dataframe(conn, catalog_template_df, workflow_target_ids, export_profile_columns)
                            catalog_batch_readiness = analyze_template_readiness(catalog_filled_df, export_profile_columns)
                            summary = catalog_batch_readiness.get("summary") or {}
                            cr1, cr2, cr3, cr4 = st.columns(4)
                            cr1.metric("Средняя готовность", f"{int(summary.get('avg_readiness') or 0)}%")
                            cr2.metric("Готовых строк", int(summary.get("ready_rows") or 0))
                            cr3.metric("Частично готовы", int(summary.get("partial_rows") or 0))
                            cr4.metric("Блокеры", int(summary.get("blocked_rows") or 0))
                            if int(summary.get("blocked_rows") or 0) > 0:
                                st.warning("Есть блокеры. Лучше открыть `Клиентский шаблон` и добить gaps перед финальной выгрузкой.")
                            else:
                                st.success("Пачка выглядит готовой к клиентской выгрузке.")

                export_signature = f"{export_client_code}|{export_category_code}|{export_profile_id}|{export_template_file_id}|{workflow_scope}|{hashlib.sha1(','.join([str(int(x)) for x in workflow_target_ids]).encode('utf-8')).hexdigest() if workflow_target_ids else 'none'}"
                if st.button("Собрать готовый Excel клиента", type="primary", use_container_width=True, key="catalog_export_build_button"):
                    if not workflow_target_ids:
                        st.warning("В выбранной области нет товаров.")
                    elif not export_profile_columns:
                        st.warning("Сначала выбери профиль шаблона клиента.")
                    elif not export_template_bytes:
                        st.warning("Сначала выбери сохранённый Excel-шаблон клиента.")
                    else:
                        safe_catalog_template_bytes = sanitize_template_xlsx_bytes(export_template_bytes)
                        ready_export_bytes = fill_template_workbook_bytes(
                            conn,
                            safe_catalog_template_bytes,
                            workflow_target_ids,
                            export_profile_columns,
                            sheet_name=export_sheet_name,
                            data_start_row=int(export_data_start_row),
                        )
                        st.session_state["catalog_client_export_bytes"] = ready_export_bytes
                        st.session_state["catalog_client_export_signature"] = export_signature
                        st.session_state["catalog_client_export_name"] = (
                            export_template_row.get("original_file_name")
                            if export_template_row
                            else "client_template.xlsx"
                        )
                        st.success("Клиентский Excel собран. Ниже можно скачать готовый файл.")
                if (
                    st.session_state.get("catalog_client_export_signature") == export_signature
                    and st.session_state.get("catalog_client_export_bytes")
                ):
                    st.download_button(
                        "Скачать готовый Excel клиента",
                        data=st.session_state["catalog_client_export_bytes"],
                        file_name=f"filled_{Path(str(st.session_state.get('catalog_client_export_name') or 'client_template.xlsx')).name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="catalog_export_download_ready_file",
                    )
                st.caption(
                    "Если нужен тонкий gap-анализ, ручной remap колонок или client_validated, переходи в `Клиентский шаблон`. "
                    "Но обычную рабочую выгрузку теперь можно собирать прямо из `Каталога`."
                )

    with st.expander("Проверка парсинга и сервис", expanded=False):
        st.caption("Для чистой проверки parser-flow используй вкладку `Supplier/Web parser` и не запускай `Заполнить всё для области`.")
        service_tab_ozon, service_tab_parser, service_tab_logistics, service_tab_media = st.tabs(
            ["Ozon категории", "Supplier/Web parser", "Логистика", "Фото и сервис"]
        )

        with service_tab_ozon:
            cextra1, cextra2, cextra3 = st.columns(3)
            with cextra1:
                if st.button("Автопривязать Ozon категории\nтекущая страница", help="Сначала назначить эталонную Ozon категорию для товаров этой страницы"):
                    res = bulk_assign_ozon_categories(conn, [int(x) for x in ids], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in ids])
                    if res.get("message"):
                        st.info(str(res["message"]))
                    else:
                        st.success(
                            f"Ozon автопривязка: обработано {res['processed']}, привязано {res['assigned']}, пропущено {res['skipped']}. "
                            f"Слотов Ozon-атрибутов создано: {int(materialized.get('slots_created') or 0)}."
                        )
                    st.rerun()
            with cextra2:
                if st.button("Перепривязать Ozon категории\nforce, текущая страница", help="Повторный подбор Ozon категории с возможной перезаписью текущей привязки"):
                    res = bulk_assign_ozon_categories(conn, [int(x) for x in ids], min_score=OZON_CATEGORY_MIN_SCORE, force=True)
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in ids])
                    if res.get("message"):
                        st.info(str(res["message"]))
                    else:
                        st.success(
                            f"Ozon force-привязка: обработано {res['processed']}, привязано {res['assigned']}, пропущено {res['skipped']}. "
                            f"Слотов Ozon-атрибутов создано: {int(materialized.get('slots_created') or 0)}."
                        )
                    st.rerun()
            with cextra3:
                if st.button("Автопривязать Ozon категории\nвыбранные", help="Назначить Ozon категорию только отмеченным товарам"):
                    res = bulk_assign_ozon_categories(conn, [int(x) for x in effective_selected_shortlist_ids], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in effective_selected_shortlist_ids])
                    if res.get("message"):
                        st.info(str(res["message"]))
                    else:
                        st.success(
                            f"Ozon автопривязка по выбранным: обработано {res['processed']}, привязано {res['assigned']}, пропущено {res['skipped']}. "
                            f"Слотов создано: {int(materialized.get('slots_created') or 0)}."
                        )
                    st.rerun()

            oextra1, oextra2, oextra3 = st.columns(3)
            with oextra1:
                if st.button("Подтянуть Ozon-атрибуты\nтекущая страница", help="Подготовить category requirements Ozon для товаров текущей страницы, чтобы атрибуты появились в карточках"):
                    seeded = ensure_ozon_requirements_for_products(conn, [int(x) for x in ids])
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in ids])
                    st.success(
                        "Готово по текущей странице: "
                        f"товаров с Ozon-категорией {int(seeded.get('products_with_ozon_category') or 0)} из {int(seeded.get('products_total') or 0)}, "
                        f"пар категорий {int(seeded.get('category_pairs') or 0)}, "
                        f"импортировано атрибутов {int(seeded.get('imported_attributes') or 0)}, "
                        f"создано слотов атрибутов у товаров {int(materialized.get('slots_created') or 0)}."
                    )
            with oextra2:
                if st.button("Подтянуть Ozon-атрибуты\nвся выборка фильтра", help="Подготовить category requirements Ozon для всей текущей выборки"):
                    seeded = ensure_ozon_requirements_for_products(conn, [int(x) for x in all_filtered_candidate_ids])
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in all_filtered_candidate_ids])
                    st.success(
                        "Готово по всей выборке: "
                        f"товаров с Ozon-категорией {int(seeded.get('products_with_ozon_category') or 0)} из {int(seeded.get('products_total') or 0)}, "
                        f"пар категорий {int(seeded.get('category_pairs') or 0)}, "
                        f"импортировано атрибутов {int(seeded.get('imported_attributes') or 0)}, "
                        f"создано слотов атрибутов у товаров {int(materialized.get('slots_created') or 0)}."
                    )
            with oextra3:
                if st.button("Подтянуть Ozon-атрибуты\nвыбранные", help="Подготовить Ozon-атрибуты только для отмеченных товаров"):
                    seeded = ensure_ozon_requirements_for_products(conn, [int(x) for x in effective_selected_shortlist_ids])
                    materialized = materialize_ozon_attribute_slots_for_products(conn, [int(x) for x in effective_selected_shortlist_ids])
                    st.success(
                        "Готово по выбранным: "
                        f"товаров с Ozon-категорией {int(seeded.get('products_with_ozon_category') or 0)} из {int(seeded.get('products_total') or 0)}, "
                        f"пар категорий {int(seeded.get('category_pairs') or 0)}, "
                        f"импортировано атрибутов {int(seeded.get('imported_attributes') or 0)}, "
                        f"создано слотов {int(materialized.get('slots_created') or 0)}."
                    )

        with service_tab_parser:
            st.info(
                "Чистая проверка парсера: выбери один товар и нажми `Только supplier/web enrichment (выбранные)`. "
                "Этот режим не запускает AI и после выполнения покажет точные причины последних ошибок parser-flow."
            )
            b1, b2, b3, b4 = st.columns(4)
            with b1:
                if st.button("Только supplier/web enrichment\nстраница", help="Запустить supplier/web enrichment без AI по текущей странице"):
                    run_supplier_enrichment_batch(
                        candidate_ids=supplier_candidate_ids,
                        run_limit=int(max_bulk_enrich_page),
                        run_label="Текущая страница",
                    )
            with b2:
                if st.button("Только supplier/web enrichment\nфильтр", help="Запустить supplier/web enrichment без AI по всей выборке фильтров"):
                    run_supplier_enrichment_batch(
                        candidate_ids=effective_filtered_candidate_ids,
                        run_limit=int(max_bulk_enrich_filtered),
                        run_label="Вся выборка фильтра",
                    )
            with b3:
                if st.button("Обновить дубли\nтекущая выборка", help="Пересчитать кандидатов дублей только для товаров на текущей странице"):
                    total = 0
                    progress = st.progress(0)
                    for i, pid in enumerate(ids, start=1):
                        refresh_duplicates_for_product(conn, int(pid))
                        total += 1
                        progress.progress(i / len(ids))
                    st.success(f"Проверка дублей завершена: {total} товаров")
            with b4:
                if st.button("Только supplier/web enrichment\nвыбранные", help="Запустить supplier/web enrichment без AI только по отмеченным товарам"):
                    run_supplier_enrichment_batch(
                        candidate_ids=[int(x) for x in effective_selected_shortlist_ids],
                        run_limit=max(len(effective_selected_shortlist_ids), 1),
                        run_label="Выбранные товары",
                    )

        with service_tab_logistics:
            d1, d2, d3 = st.columns(3)
            with d1:
                if st.button("Рассчитать габариты/вес\nтекущая страница", help="Заполнить пустые логистические поля товара статистикой похожих товаров и типовыми значениями."):
                    run_dimension_estimation_batch(
                        candidate_ids=[int(x) for x in ids],
                        run_limit=int(max_bulk_enrich_page),
                        run_label="Текущая страница",
                    )
            with d2:
                if st.button("Рассчитать габариты/вес\nвся выборка фильтра", help="Массово заполнить пустые логистические поля по всей текущей фильтрации."):
                    run_dimension_estimation_batch(
                        candidate_ids=[int(x) for x in all_filtered_candidate_ids],
                        run_limit=int(max_bulk_enrich_filtered),
                        run_label="Вся выборка фильтра",
                    )
            with d3:
                if st.button("Рассчитать габариты/вес\nвыбранные", help="Рассчитать логистику только по отмеченным товарам"):
                    run_dimension_estimation_batch(
                        candidate_ids=[int(x) for x in effective_selected_shortlist_ids],
                        run_limit=max(len(effective_selected_shortlist_ids), 1),
                        run_label="Выбранные товары",
                    )

        with service_tab_media:
            z1, z2, z3 = st.columns(3)
            with z1:
                if st.button("Подготовить ZIP фото текущей страницы", key="catalog_prepare_images_page_zip"):
                    zip_page_bytes, zip_page_stats = build_product_images_zip(
                        conn,
                        [int(x) for x in ids],
                        public_base_url=media_public_base_url,
                    )
                    st.session_state["catalog_page_zip_bytes"] = zip_page_bytes
                    st.session_state["catalog_page_zip_stats"] = zip_page_stats
                    st.session_state["catalog_page_zip_signature"] = page_zip_signature
                    if int(zip_page_stats.get("images_written") or 0) > 0:
                        st.success(
                            f"ZIP по текущей странице подготовлен. Фото: {int(zip_page_stats.get('images_written') or 0)}, "
                            f"пропущено: {int(zip_page_stats.get('images_skipped') or 0)}."
                        )
                    else:
                        st.warning("Для текущей страницы не удалось собрать публичные фото в ZIP.")
                if (
                    st.session_state.get("catalog_page_zip_signature") == page_zip_signature
                    and st.session_state.get("catalog_page_zip_bytes")
                ):
                    st.download_button(
                        "Скачать фото текущей страницы ZIP",
                        data=st.session_state["catalog_page_zip_bytes"],
                        file_name=f"pim_product_images_page_{int(page)}.zip",
                        mime="application/zip",
                        key="catalog_export_images_page_zip",
                    )
            with z2:
                if st.button("Подготовить ZIP фото всей выборки", key="catalog_prepare_images_filtered_zip"):
                    zip_filtered_bytes, zip_filtered_stats = build_product_images_zip(
                        conn,
                        [int(x) for x in all_filtered_candidate_ids],
                        public_base_url=media_public_base_url,
                    )
                    st.session_state["catalog_filtered_zip_bytes"] = zip_filtered_bytes
                    st.session_state["catalog_filtered_zip_stats"] = zip_filtered_stats
                    st.session_state["catalog_filtered_zip_signature"] = filtered_zip_signature
                    if int(zip_filtered_stats.get("images_written") or 0) > 0:
                        st.success(
                            f"ZIP по всей выборке подготовлен. Фото: {int(zip_filtered_stats.get('images_written') or 0)}, "
                            f"пропущено: {int(zip_filtered_stats.get('images_skipped') or 0)}."
                        )
                    else:
                        st.warning("Для текущей фильтрации не удалось собрать публичные фото в ZIP.")
                if (
                    st.session_state.get("catalog_filtered_zip_signature") == filtered_zip_signature
                    and st.session_state.get("catalog_filtered_zip_bytes")
                ):
                    st.download_button(
                        "Скачать фото всей выборки ZIP",
                        data=st.session_state["catalog_filtered_zip_bytes"],
                        file_name="pim_product_images_filtered.zip",
                        mime="application/zip",
                        key="catalog_export_images_filtered_zip",
                    )
            with z3:
                if st.button("Подготовить ZIP фото выбранных", key="catalog_prepare_images_selected_zip"):
                    zip_selected_bytes, zip_selected_stats = build_product_images_zip(
                        conn,
                        [int(x) for x in effective_selected_shortlist_ids],
                        public_base_url=media_public_base_url,
                    )
                    st.session_state["catalog_selected_zip_bytes"] = zip_selected_bytes
                    st.session_state["catalog_selected_zip_stats"] = zip_selected_stats
                    st.session_state["catalog_selected_zip_signature"] = selected_zip_signature
                    if int(zip_selected_stats.get("images_written") or 0) > 0:
                        st.success(
                            f"ZIP по выбранным подготовлен. Фото: {int(zip_selected_stats.get('images_written') or 0)}, "
                            f"пропущено: {int(zip_selected_stats.get('images_skipped') or 0)}."
                        )
                    else:
                        st.warning("Для выбранных товаров не удалось собрать публичные фото в ZIP.")
                if (
                    st.session_state.get("catalog_selected_zip_signature") == selected_zip_signature
                    and st.session_state.get("catalog_selected_zip_bytes")
                ):
                    st.download_button(
                        "Скачать фото выбранных ZIP",
                        data=st.session_state["catalog_selected_zip_bytes"],
                        file_name="pim_product_images_selected.zip",
                        mime="application/zip",
                        key="catalog_export_images_selected_zip",
                    )

    if operational_df_display is not None and not operational_df_display.empty:
        st.markdown("### Таблица товаров")
        st.caption("Главный рабочий список: артикул, этап, parser-stage, AI-stage, image-stage, очередь, фото, штрихкод и последнее обновление.")
        op_cols = [
            c
            for c in [
                "article",
                "name",
                "supplier_name",
                "ozon_category_path",
                "stage",
                "parser_stage",
                "ai_stage",
                "image_stage",
                "queue",
                "supplier_parse_status",
                "photo_status",
                "photo_count",
                "barcode_status",
                "fill_score",
                "updated_at",
            ]
            if c in operational_df_display.columns
        ]
        st.dataframe(
            with_ru_columns(
                operational_df_display[op_cols],
                extra_map={
                    "stage": "Этап карточки",
                    "parser_stage": "Parser stage",
                    "ai_stage": "AI stage",
                    "image_stage": "Image stage",
                    "queue": "Очередь",
                    "photo_status": "Фото",
                    "photo_count": "Кол-во фото",
                    "barcode_status": "Штрихкод",
                    "fill_score": "Ядро",
                },
            ),
            use_container_width=True,
            hide_index=True,
        )

    if selected_id:
        st.session_state["selected_product_id"] = int(selected_id)

    with st.expander("Массовое изменение данных", expanded=False):
        st.caption("Здесь можно массово назначить поставщика, URL, категории и бренд.")
        mm1, mm2, mm3 = st.columns(3)
        with mm1:
            scope = st.selectbox(
                "Область применения",
                options=["Выбранные товары", "Текущая страница", "Вся выборка по фильтру"],
                key="mass_edit_scope",
                help="Выбранные товары: только shortlist на этой странице. Текущая страница: только видимые товары. Вся выборка: все товары по текущим фильтрам.",
            )
            only_empty = st.checkbox(
                "Заполнять только пустые поля",
                value=True,
                key="mass_edit_only_empty",
                help="Если включено, заполнит только пустые значения.",
            )
        with mm2:
            mass_supplier = st.selectbox(
                "Поставщик",
                options=[""] + supplier_values,
                index=0,
                key="mass_edit_supplier",
                help="Назначить поставщика выбранным товарам.",
            )
            mass_category = st.selectbox(
                "Категория",
                options=[""] + category_values,
                index=0,
                key="mass_edit_category",
            )
            mass_base_category = st.selectbox(
                "Базовая категория",
                options=[""] + category_values,
                index=0,
                key="mass_edit_base_category",
            )
            mass_subcategory = st.selectbox(
                "Подкатегория",
                options=[""] + category_values,
                index=0,
                key="mass_edit_subcategory",
            )
        with mm3:
            mass_brand = st.text_input("Бренд", value="", key="mass_edit_brand")
            profile_for_url = st.selectbox(
                "Профиль URL поставщика",
                options=[""] + supplier_profile_values,
                index=0,
                key="mass_edit_profile_for_url",
                help="Можно выбрать профиль, чтобы автоматически подставить URL template.",
            )
            profile_map = {p["supplier_name"]: p for p in list_supplier_profiles(conn, only_active=True)}
            mass_supplier_url_template = st.text_input(
                "URL template",
                value=st.session_state.get("mass_edit_supplier_url_template", ""),
                key="mass_edit_supplier_url_template",
                help="Поддержка плейсхолдеров: {article}, {supplier_article}, {code}, {name} и *_q.",
            )
            if st.button("Подставить URL template из профиля", key="mass_edit_apply_profile_template"):
                if profile_for_url and profile_for_url in profile_map:
                    st.session_state["mass_edit_supplier_url_template"] = profile_map[profile_for_url].get("url_template") or ""
                    st.rerun()

        apply_mass = st.button(
            "Применить массовые изменения",
            type="primary",
            help="Применить изменения к выбранной области товаров.",
            key="mass_edit_apply_btn",
        )
        if apply_mass:
            if scope == "Выбранные товары":
                target_ids = [int(x) for x in effective_selected_shortlist_ids]
            elif scope == "Текущая страница":
                target_ids = [int(x) for x in ids]
            else:
                target_ids = load_product_ids(
                    conn,
                    search=search,
                    category=category,
                    supplier=supplier,
                    import_batch_id=batch_id or "",
                    parse_filter=parse_filter,
                    limit=None,
                    offset=0,
                )
            updates = {
                "supplier_name": mass_supplier.strip() if mass_supplier else None,
                "category": mass_category.strip() if mass_category else None,
                "base_category": mass_base_category.strip() if mass_base_category else None,
                "subcategory": mass_subcategory.strip() if mass_subcategory else None,
                "brand": mass_brand.strip() if mass_brand else None,
            }
            result = apply_mass_product_updates(
                conn=conn,
                product_ids=target_ids,
                updates=updates,
                supplier_url_template=mass_supplier_url_template.strip() or None,
                only_empty=bool(only_empty),
            )
            st.success(
                f"Обновлено товаров: {result['updated_products']}, обновлено полей: {result['updated_fields']}"
            )
            st.rerun()

    nav1, nav2, nav3 = st.columns([1, 1, 4])
    with nav1:
        if st.button("◀ Назад", disabled=int(page) <= 1, help="Перейти на предыдущую страницу каталога", key="catalog_nav_prev_bottom"):
            st.session_state["catalog_page_current"] = int(page) - 1
            st.rerun()
    with nav2:
        if st.button("Вперед ▶", disabled=int(page) >= total_pages, help="Перейти на следующую страницу каталога", key="catalog_nav_next_bottom"):
            st.session_state["catalog_page_current"] = int(page) + 1
            st.rerun()
    with nav3:
        st.caption(f"Навигация по каталогу: страница {int(page)} из {int(total_pages)}")

    conn.close()


def enrich_product_from_supplier(
    conn,
    product_id: int,
    force: bool = False,
    timeout_seconds: float | None = None,
    parser_settings: dict[str, object] | None = None,
) -> dict:
    product = get_product(conn, product_id)
    if not product:
        return {"ok": False, "message": "Товар не найден"}
    existing_ozon_desc = int(product["ozon_description_category_id"] or 0) if "ozon_description_category_id" in product.keys() else 0
    existing_ozon_type = int(product["ozon_type_id"] or 0) if "ozon_type_id" in product.keys() else 0
    if not (existing_ozon_desc > 0 and existing_ozon_type > 0):
        # Эталонный порядок: сначала пытаемся привязать Ozon категорию, потом парсинг/обогащение.
        try:
            bulk_assign_ozon_categories(conn, [int(product_id)], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
            refreshed = get_product(conn, product_id)
            if refreshed:
                product = refreshed
        except Exception:
            # Не блокируем enrichment, если Ozon-кэш временно недоступен.
            pass
    product_row = dict(product)
    settings = dict(parser_settings or load_parser_settings(conn))
    effective_timeout = float(timeout_seconds if timeout_seconds is not None else settings.get("timeout_seconds", 8.0))
    effective_timeout = max(2.0, min(30.0, float(effective_timeout)))
    max_hops = max(1, min(3, int(settings.get("max_hops", 1))))
    fallback_max_results = max(1, min(12, int(settings.get("fallback_max_results", 4))))
    source_strategy = str(settings.get("source_strategy", "auto_full") or "auto_full")
    extra_fallback_domains = _parse_domain_list(settings.get("extra_fallback_domains", ""))
    enable_web_fallback = bool(settings.get("enable_web_fallback", True))
    enable_ozon_fallback = bool(settings.get("enable_ozon_fallback", True))
    enable_yandex_fallback = bool(settings.get("enable_yandex_fallback", True))
    enable_stats_fallback = bool(settings.get("enable_stats_fallback", True))
    enable_defaults_fallback = bool(settings.get("enable_defaults_fallback", True))

    if source_strategy == "supplier_only":
        enable_web_fallback = False
        enable_ozon_fallback = False
        enable_yandex_fallback = False
    elif source_strategy == "supplier_plus_ozon":
        enable_yandex_fallback = False
        enable_web_fallback = False
    elif source_strategy == "supplier_plus_yandex":
        enable_ozon_fallback = False
        enable_web_fallback = False
    elif source_strategy == "web_only":
        enable_ozon_fallback = False
        enable_yandex_fallback = False
        enable_web_fallback = True
    elif source_strategy == "custom_domains":
        enable_ozon_fallback = False
        enable_yandex_fallback = False
        enable_web_fallback = bool(extra_fallback_domains)

    supplier_url = (product["supplier_url"] or "").strip() if product["supplier_url"] else ""

    try:
        parse_hints = [
            str(product["article"] or ""),
            str(product["supplier_article"] or ""),
            str(product["name"] or ""),
            str(product["brand"] or ""),
            str(product["category"] or ""),
            str(product["subcategory"] or ""),
            str(product["base_category"] or ""),
            str(product["ozon_category_path"] or "") if "ozon_category_path" in product.keys() else "",
        ]
        parsed: dict = {}
        source_url = supplier_url
        source_type = "supplier_page"
        used_fallback = False
        fallback_rejected_reason = ""
        parse_error_text = ""
        used_stats_fallback = False
        used_category_defaults = False
        field_source_types: dict[str, str] = {}

        if supplier_url:
            # Support supplier search pages like https://velocitygroup.ru/catalog/?q=
            effective_supplier_url = supplier_url
            low_url = effective_supplier_url.lower()
            if ("?q=" in low_url) and low_url.rstrip().endswith("?q="):
                query_candidate = str(product["supplier_article"] or product["article"] or product["name"] or "").strip()
                if query_candidate:
                    effective_supplier_url = f"{effective_supplier_url}{quote(query_candidate, safe='')}"
            try:
                parsed = parse_supplier_product_page(
                    effective_supplier_url,
                    hints=parse_hints,
                    timeout=float(effective_timeout),
                    max_hops=max_hops,
                )
                source_url = parsed.get("resolved_url") or effective_supplier_url
                listing_like_source = False
                root_like_source = False
                if _supplier_parser is not None:
                    try:
                        if hasattr(_supplier_parser, "_is_listing_like_url"):
                            listing_like_source = bool(_supplier_parser._is_listing_like_url(effective_supplier_url))
                        if hasattr(_supplier_parser, "_is_root_like_url"):
                            root_like_source = bool(_supplier_parser._is_root_like_url(effective_supplier_url))
                    except Exception:
                        listing_like_source = False
                        root_like_source = False
                if parsed and has_meaningful_supplier_data(parsed) and (
                    bool(parsed.get("resolved_from_listing"))
                    or listing_like_source
                    or root_like_source
                ):
                    strict_relevance_settings = dict(settings or {})
                    strict_relevance_settings["require_article_match"] = True
                    is_relevant, relevance_reason = _is_parsed_result_relevant(
                        product_row=product_row,
                        parsed=parsed,
                        source_url=str(source_url or effective_supplier_url or ""),
                        settings=strict_relevance_settings,
                    )
                    if not is_relevant:
                        parsed = {}
                        source_type = "web_search_fallback"
                        source_url = effective_supplier_url
                        fallback_rejected_reason = relevance_reason
            except Exception as parse_error:
                parsed = {}
                source_type = "web_search_fallback"
                source_url = effective_supplier_url
                parse_error_text = str(parse_error)[:220]
                # Keep flow alive: fallback to internet search below.

        dim_fields = [
            "weight",
            "gross_weight",
            "length",
            "width",
            "height",
            "package_length",
            "package_width",
            "package_height",
        ]
        has_dims = any(parsed.get(k) not in (None, "", 0, 0.0) for k in dim_fields)
        need_fallback = (not has_meaningful_supplier_data(parsed)) or bool(parsed.get("listing_only")) or (not has_dims) or is_dimension_payload_suspicious(parsed) or (not supplier_url)

        if need_fallback and (enable_web_fallback or enable_ozon_fallback or enable_yandex_fallback):
            preferred_domain = ""
            try:
                preferred_domain = (urlparse(str(source_url or supplier_url)).netloc or "").lower().replace("www.", "")
            except Exception:
                preferred_domain = ""
            blocked_supplier_domain = is_likely_blocked_supplier_domain(source_url or supplier_url)
            if blocked_supplier_domain:
                preferred_domain = ""
            fallback_query_parts = [
                str(product["name"] or "").strip(),
                str(product["article"] or "").strip(),
                str(product["supplier_article"] or "").strip(),
                str(product["brand"] or "").strip(),
                str(product["subcategory"] or "").strip(),
                str(product["category"] or "").strip(),
                "габариты",
            ]
            fallback_query = " ".join([p for p in fallback_query_parts if p])
            fallback = {}
            fallback_stage = "generic_web"
            fallback_targets: list[tuple[str, str | None]] = []
            if enable_ozon_fallback:
                fallback_targets.append(("ozon_search_fallback", "ozon.ru"))
            if enable_yandex_fallback:
                fallback_targets.append(("yandex_search_fallback", "market.yandex.ru"))
            for dom in extra_fallback_domains:
                fallback_targets.append(("domain_search_fallback", dom))
            if enable_web_fallback:
                if preferred_domain:
                    fallback_targets.append(("web_search_fallback_domain", preferred_domain))
                fallback_targets.append(("web_search_fallback", None))

            for stage_name, domain in fallback_targets:
                fallback = fallback_search_product_data(
                    fallback_query,
                    timeout=float(effective_timeout),
                    max_results=int(fallback_max_results),
                    hints=parse_hints,
                    preferred_domain=domain,
                    blocked_source_domain=bool(blocked_supplier_domain),
                )
                if fallback:
                    fallback_stage = stage_name
                    break
            if fallback and has_meaningful_supplier_data(fallback):
                is_relevant, relevance_reason = _is_parsed_result_relevant(
                    product_row=product_row,
                    parsed=fallback,
                    source_url=str(fallback.get("fallback_url") or source_url or ""),
                    settings=settings,
                )
                if not is_relevant:
                    fallback = {}
                    fallback_rejected_reason = relevance_reason
            if fallback and has_meaningful_supplier_data(fallback):
                for key in [
                    "name", "brand", "category", "description", "image_url", "weight", "gross_weight",
                    "length", "width", "height", "package_length", "package_width", "package_height"
                ]:
                    if parsed.get(key) in (None, "", 0, 0.0):
                        parsed[key] = fallback.get(key)
                merged_attrs = dict(parsed.get("attributes") or {})
                for k, v in (fallback.get("attributes") or {}).items():
                    if k not in merged_attrs:
                        merged_attrs[k] = v
                parsed["attributes"] = merged_attrs
                if not parsed.get("image_urls"):
                    parsed["image_urls"] = fallback.get("image_urls") or []
                source_url = fallback.get("fallback_url") or source_url
                source_type = str(fallback_stage)
                used_fallback = True
                for key in [
                    "name",
                    "brand",
                    "category",
                    "description",
                    "image_url",
                    "weight",
                    "gross_weight",
                    "length",
                    "width",
                    "height",
                    "package_length",
                    "package_width",
                    "package_height",
                ]:
                    if fallback.get(key) not in (None, "", 0, 0.0):
                        field_source_types[key] = str(fallback_stage)
            elif not supplier_url:
                source_type = "web_search_fallback"
                source_url = source_url or ""

        category_inferred = infer_category_fields(
            {
                "name": parsed.get("name") or product_row.get("name"),
                "category": parsed.get("category") or product_row.get("category"),
                "base_category": product_row.get("base_category"),
                "subcategory": product_row.get("subcategory"),
            }
        )
        weak_categories = {"товары", "каталог", "продукция", "все товары", "catalog", "products", "shop"}
        if (
            category_inferred.get("category")
            and (
                parsed.get("category") in (None, "")
                or str(parsed.get("category") or "").strip().lower() in weak_categories
            )
        ):
            parsed["category"] = category_inferred.get("category")
            field_source_types["category"] = "name_category_inference"
        if category_inferred.get("base_category"):
            parsed["base_category"] = category_inferred.get("base_category")
            field_source_types["base_category"] = "name_category_inference"
        if category_inferred.get("subcategory"):
            parsed["subcategory"] = category_inferred.get("subcategory")
            field_source_types["subcategory"] = "name_category_inference"
        if category_inferred.get("wheel_diameter_inch") is not None:
            parsed["wheel_diameter_inch"] = category_inferred.get("wheel_diameter_inch")
            field_source_types["wheel_diameter_inch"] = "name_category_inference"

        has_dims_after_web = any(parsed.get(k) not in (None, "", 0, 0.0) for k in dim_fields)
        if (not has_dims_after_web) or is_dimension_payload_suspicious(parsed):
            if enable_stats_fallback:
                stats_fallback = infer_dimensions_from_catalog(conn, product_row, min_samples=4)
                if stats_fallback.get("found"):
                    for key, value in (stats_fallback.get("values") or {}).items():
                        if parsed.get(key) in (None, "", 0, 0.0):
                            parsed[key] = value
                            field_source_types[key] = "category_stats_fallback"
                    used_stats_fallback = True
            if ((not used_stats_fallback) or is_dimension_payload_suspicious(parsed)) and enable_defaults_fallback:
                defaults_fallback = infer_dimensions_from_category_defaults(
                    conn,
                    {
                        "category": parsed.get("category") or product_row.get("category"),
                        "base_category": parsed.get("base_category") or product_row.get("base_category"),
                        "subcategory": parsed.get("subcategory") or product_row.get("subcategory"),
                        "wheel_diameter_inch": parsed.get("wheel_diameter_inch") or product_row.get("wheel_diameter_inch"),
                    },
                )
                if defaults_fallback.get("found"):
                    for key, value in (defaults_fallback.get("values") or {}).items():
                        if parsed.get(key) in (None, "", 0, 0.0):
                            parsed[key] = value
                            field_source_types[key] = "category_defaults_fallback"
                    used_category_defaults = True

        if not has_meaningful_supplier_data(parsed):
            has_any_dims_after_fallback = any(parsed.get(k) not in (None, "", 0, 0.0) for k in dim_fields)
            if not (used_stats_fallback or used_category_defaults or has_any_dims_after_fallback):
                fail_comment = "Не удалось получить полезные данные"
                if parse_error_text:
                    fail_comment += f"; supplier_parse_error={parse_error_text}"
                if fallback_rejected_reason:
                    fail_comment += f"; fallback_rejected={fallback_rejected_reason[:180]}"
                conn.execute(
                    """
                    UPDATE products
                    SET supplier_parse_status = ?,
                        supplier_parse_comment = ?,
                        supplier_last_parsed_at = CURRENT_TIMESTAMP,
                        updated_at = CURRENT_TIMESTAMP
                    WHERE id = ?
                    """,
                    ("error", fail_comment[:500], product_id),
                )
                conn.commit()
                return {"ok": False, "message": fail_comment}

        updates = {}
        skipped_manual_fields = []
        has_ozon_priority = bool(
            int(product["ozon_description_category_id"] or 0) > 0
            and int(product["ozon_type_id"] or 0) > 0
        )
        fields = [
            "name",
            "brand",
            "category",
            "base_category",
            "subcategory",
            "wheel_diameter_inch",
            "description",
            "image_url",
            "weight",
            "length",
            "width",
            "height",
            "package_length",
            "package_width",
            "package_height",
            "gross_weight",
        ]
        for field in fields:
            new_value = parsed.get(field)
            old_value = product[field] if field in product.keys() else None
            if new_value is None:
                continue
            if has_ozon_priority and field in {"category", "base_category", "subcategory"} and not force:
                skipped_manual_fields.append(f"{field}:ozon_priority")
                continue
            if field == "category" and str(new_value).strip().lower() in weak_categories and not force:
                continue
            if field_is_manual(conn, product_id, field) and not force:
                skipped_manual_fields.append(field)
                continue
            row_source_type = field_source_types.get(field, source_type)
            if not can_overwrite_field(conn, product_id, field, row_source_type, force=force):
                skipped_manual_fields.append(field)
                continue
            if old_value not in (None, "", 0, 0.0) and not force:
                continue
            updates[field] = new_value

        attributes_saved = 0
        skipped_attribute_fields = []
        existing_defs = list_attribute_definitions(conn)
        defs_by_code = {str(d.get("code")): d for d in existing_defs if str(d.get("code") or "").strip()}
        defs_by_name_norm = {
            _normalize_attr_text(str(d.get("name") or "")): str(d.get("code") or "")
            for d in existing_defs
            if str(d.get("name") or "").strip() and str(d.get("code") or "").strip()
        }
        for attr_name, attr_value in (parsed.get("attributes") or {}).items():
            clean_code = str(attr_name).strip().lower()
            clean_code = "_".join("".join(ch if ch.isalnum() else " " for ch in clean_code).split())
            if not clean_code:
                continue
            name_norm = _normalize_attr_text(attr_name)
            target_code = defs_by_name_norm.get(name_norm) or clean_code
            attr_field_name = f"attr:{target_code}"
            if not can_overwrite_field(conn, product_id, attr_field_name, source_type, force=force):
                skipped_attribute_fields.append(target_code)
                continue
            existing_def = defs_by_code.get(target_code)
            if not existing_def:
                conn.execute(
                    """
                    INSERT OR IGNORE INTO attribute_definitions
                    (code, name, data_type, scope, entity_type, is_required, is_multi_value, unit, description, created_at, updated_at)
                    VALUES (?, ?, 'text', 'master', 'product', 0, 0, NULL, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                    """,
                    (target_code, str(attr_name).strip(), f"Автосоздано из source: {source_url}"),
                )
            set_product_attribute_value(conn, product_id, target_code, str(attr_value))
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name=attr_field_name,
                source_type=source_type,
                source_value_raw=attr_value,
                source_url=source_url,
                confidence=0.6 if source_type == "supplier_page" else 0.45,
            )
            attributes_saved += 1

        ozon_attr_fill = _fill_ozon_attrs_from_parsed(
            conn=conn,
            product_row=product_row,
            parsed=parsed,
            source_type=source_type,
            source_url=source_url,
            force=force,
        )
        post_state_backfill = {"saved": 0, "skipped": 0, "targets": 0}

        image_urls = _normalize_media_urls([str(x).strip() for x in (parsed.get("image_urls") or []) if str(x).strip()])
        if not image_urls and str(parsed.get("image_url") or "").strip():
            image_urls = _normalize_media_urls([str(parsed.get("image_url") or "").strip()])
        if image_urls:
            parsed["image_urls"] = image_urls
            parsed["image_url"] = image_urls[0]
        if image_urls:
            set_product_attribute_value(conn, product_id, "main_image", image_urls[0])
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name="attr:main_image",
                source_type=source_type,
                source_value_raw=image_urls[0],
                source_url=source_url,
                confidence=0.75 if source_type == "supplier_page" else 0.5,
            )
            # Keep gallery_images even when there is only one image.
            set_product_attribute_value(conn, product_id, "gallery_images", image_urls)
            save_field_source(
                conn=conn,
                product_id=product_id,
                field_name="attr:gallery_images",
                source_type=source_type,
                source_value_raw=json.dumps(image_urls, ensure_ascii=False),
                source_url=source_url,
                confidence=0.7 if source_type == "supplier_page" else 0.45,
            )

        parse_comment = f"source={source_type}; strategy={source_strategy}; url={source_url}"
        if parsed.get("resolved_from_listing"):
            parse_comment += "; listing->product resolved"
        if used_fallback:
            parse_comment += "; web_fallback=1"
        if parse_error_text:
            parse_comment += f"; supplier_parse_error={parse_error_text}"
        if fallback_rejected_reason:
            parse_comment += f"; fallback_rejected={fallback_rejected_reason[:180]}"
        if used_stats_fallback:
            parse_comment += "; category_stats_fallback=1"
        if used_category_defaults:
            parse_comment += "; category_defaults_fallback=1"

        if updates:
            set_clause = ", ".join([f"{k} = ?" for k in updates.keys()])
            params = list(updates.values()) + ["success", parse_comment[:500], product_id]
            conn.execute(
                f"""
                UPDATE products
                SET {set_clause},
                    supplier_parse_status = ?,
                    supplier_parse_comment = ?,
                    supplier_last_parsed_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                params,
            )
            for field_name, value in updates.items():
                row_source_type = field_source_types.get(field_name, source_type)
                save_field_source(
                    conn=conn,
                    product_id=product_id,
                    field_name=field_name,
                    source_type=row_source_type,
                    source_value_raw=value,
                    source_url=source_url,
                    confidence=(
                        0.72
                        if row_source_type == "supplier_page"
                        else 0.58
                        if row_source_type == "web_search_fallback"
                        else 0.38
                        if row_source_type == "category_stats_fallback"
                        else 0.33
                        if row_source_type == "category_defaults_fallback"
                        else 0.65
                        if row_source_type == "name_category_inference"
                        else 0.45
                    ),
                )
        else:
            conn.execute(
                """
                UPDATE products
                SET supplier_parse_status = ?,
                    supplier_parse_comment = ?,
                    supplier_last_parsed_at = CURRENT_TIMESTAMP,
                    updated_at = CURRENT_TIMESTAMP
                WHERE id = ?
                """,
                ("success", f"Новых данных для записи не найдено; {parse_comment}"[:500], product_id),
            )

        conn.commit()
        refreshed_row = get_product(conn, int(product_id))
        if refreshed_row:
            product_row = dict(refreshed_row)
        if _safe_int_id(product_row.get("ozon_description_category_id")) > 0 and _safe_int_id(product_row.get("ozon_type_id")) > 0:
            post_state_backfill = _fill_channel_attrs_from_product_state(
                conn=conn,
                product_row=product_row,
                channel_code="ozon",
                category_code=f"ozon:{_safe_int_id(product_row.get('ozon_description_category_id'))}:{_safe_int_id(product_row.get('ozon_type_id'))}",
                source_type=source_type,
                source_url=f"{source_url} | product_state",
                force=force,
                target_channel_code=None,
            )
            conn.commit()
        ozon_match = bulk_assign_ozon_categories(conn, [int(product_id)], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
        skipped_msg = f", пропущено ручных полей: {len(skipped_manual_fields)}" if skipped_manual_fields else ""
        skipped_attr_msg = f", пропущено атрибутов по приоритету: {len(skipped_attribute_fields)}" if skipped_attribute_fields else ""
        ozon_attr_msg = (
            f", Ozon-атрибутов заполнено: {int(ozon_attr_fill.get('saved') or 0)}"
            if int(ozon_attr_fill.get("saved") or 0) > 0
            else ""
        )
        post_state_msg = (
            f", из карточки/мастера дозаполнено: {int(post_state_backfill.get('saved') or 0)}"
            if int(post_state_backfill.get("saved") or 0) > 0
            else ""
        )
        ozon_msg = f", Ozon category match: {ozon_match.get('assigned', 0)}" if ozon_match.get("processed") else ""
        fallback_msg = f", использован fallback: {source_type}" if used_fallback else ""
        stats_msg = ", использован category-stats fallback" if used_stats_fallback else ""
        defaults_msg = ", использованы category defaults" if used_category_defaults else ""
        photo_msg = f", фото найдено: {len(image_urls)}" if image_urls else ""
        return {
            "ok": True,
            "message": f"Обогащение завершено, обновлено полей: {len(updates)}, атрибутов сохранено: {attributes_saved}{photo_msg}{ozon_attr_msg}{post_state_msg}{fallback_msg}{stats_msg}{defaults_msg}{skipped_msg}{skipped_attr_msg}{ozon_msg}",
            "updates": updates,
            "attributes": parsed.get("attributes", {}),
            "image_urls": parsed.get("image_urls", []),
            "skipped_manual_fields": skipped_manual_fields,
            "skipped_attribute_fields": skipped_attribute_fields,
            "ozon_attrs_filled": int(ozon_attr_fill.get("saved") or 0),
            "state_attr_backfill": int(post_state_backfill.get("saved") or 0),
            "source_url": source_url,
            "source_type": source_type,
        }
    except Exception as e:
        conn.execute(
            """
            UPDATE products
            SET supplier_parse_status = ?,
                supplier_parse_comment = ?,
                supplier_last_parsed_at = CURRENT_TIMESTAMP,
                updated_at = CURRENT_TIMESTAMP
            WHERE id = ?
            """,
            ("error", str(e)[:500], product_id),
        )
        conn.commit()
        return {"ok": False, "message": str(e)}


def show_product_tab(summary: dict[str, object] | None = None):
    conn = get_db()
    parser_settings = load_parser_settings(conn)
    ai_settings = load_ai_settings(conn)
    media_settings = load_media_settings(conn)
    media_public_base_url = str(media_settings.get("public_base_url") or "").strip()
    st.subheader("Поиск и выбор товара для редактирования")
    st.caption("Сначала найди нужный товар, а уже потом спокойно переходи к доводке карточки ниже.")
    with st.container(border=True):
        fs1, fs2, fs3, fs4, fs5 = st.columns([3, 2, 2, 2, 1])
        with fs1:
            card_search = st.text_input(
                "Поиск товара",
                value=st.session_state.get("card_product_search", ""),
                placeholder="Артикул или наименование товара",
                key="card_product_search",
            )
        ozon_category_values, ozon_subcategory_values = list_ozon_category_filters(conn)
        supplier_values = list_distinct_values(conn, "supplier_name")
        with fs2:
            card_category = st.selectbox(
                "Категория Ozon",
                options=["Все"] + ozon_category_values,
                index=0,
                key="card_product_category_filter",
            )
        with fs3:
            card_subcategory = st.selectbox(
                "Подкатегория Ozon",
                options=["Все"] + ozon_subcategory_values,
                index=0,
                key="card_product_subcategory_filter",
            )
        with fs4:
            card_supplier = st.selectbox(
                "Поставщик",
                options=["Все"] + supplier_values,
                index=0,
                key="card_product_supplier_filter",
            )
        with fs5:
            st.markdown("<br>", unsafe_allow_html=True)
            st.button("Сброс", key="card_filters_reset", on_click=_reset_card_filters)

    filtered_products = find_products_for_card(
        conn,
        search=card_search or "",
        ozon_category=card_category or "",
        ozon_subcategory=card_subcategory or "",
        supplier=card_supplier or "",
        limit=5000,
    )
    if not filtered_products:
        st.warning("По фильтрам не найдено товаров. Измени фильтр или очисти поиск.")
        conn.close()
        return

    filtered_products = sorted(filtered_products, key=lambda item: _card_product_sort_key(item, card_search or ""))
    st.caption(f"Найдено товаров по фильтру: {len(filtered_products)}")
    preview_rows = [
        {
            "article": row.get("article") or row.get("supplier_article") or row.get("internal_article"),
            "name": _short_text(row.get("name"), 68),
            "supplier_name": row.get("supplier_name"),
            "ozon_category": row.get("ozon_category") or row.get("category"),
            "ozon_subcategory": row.get("ozon_subcategory") or row.get("subcategory"),
            "stage": _product_stage_label(row),
            "supplier_parse_status": _parse_status_label(row.get("supplier_parse_status")),
            "photo_status": "Есть" if str(row.get("image_url") or "").strip() else "Нет",
            "updated_at": row.get("updated_at"),
            "id": row.get("id"),
        }
        for row in filtered_products[:80]
    ]
    if preview_rows:
        st.dataframe(with_ru_columns(pd.DataFrame(preview_rows)), use_container_width=True, hide_index=True)
    product_options = [int(r["id"]) for r in filtered_products]
    current_product_id = int(st.session_state.get("selected_product_id") or 0)
    default_product_id = current_product_id if current_product_id in product_options else int(product_options[0])
    nav_prev_disabled = product_options.index(default_product_id) <= 0 if default_product_id in product_options else True
    nav_next_disabled = product_options.index(default_product_id) >= (len(product_options) - 1) if default_product_id in product_options else True
    n1, n2, n3 = st.columns([1, 1, 4])
    with n1:
        if st.button("◀ Предыдущий", disabled=nav_prev_disabled, key="card_prev_product_btn"):
            idx = product_options.index(default_product_id)
            st.session_state["selected_product_id"] = int(product_options[max(0, idx - 1)])
            st.rerun()
    with n2:
        if st.button("Следующий ▶", disabled=nav_next_disabled, key="card_next_product_btn"):
            idx = product_options.index(default_product_id)
            st.session_state["selected_product_id"] = int(product_options[min(len(product_options) - 1, idx + 1)])
            st.rerun()
    with n3:
        st.caption("Список выше нужен для быстрого выбора. Поиск работает по нескольким словам и ключевым полям товара.")
    selected_product_id = st.selectbox(
        "Товар по артикулу / названию",
        options=product_options,
        index=product_options.index(default_product_id),
        format_func=lambda x: next(
            (
                f"{str(row.get('article') or row.get('supplier_article') or row.get('internal_article') or '-')} | "
                f"{str(row.get('name') or '-')} | "
                f"{str(row.get('ozon_category') or row.get('category') or row.get('base_category') or '-')} / "
                f"{str(row.get('ozon_subcategory') or row.get('subcategory') or '-')} | "
                f"{str(row.get('supplier_name') or '-')} | ID {int(row['id'])}"
                for row in filtered_products
                if int(row["id"]) == int(x)
            ),
            f"ID {x}",
        ),
        key="card_selected_product_id",
    )
    st.session_state["selected_product_id"] = int(selected_product_id)
    product_id = int(selected_product_id)
    product = get_product(conn, product_id)

    if not product:
        st.warning("Товар не найден")
        conn.close()
        return

    cleanup_cleared = _cleanup_unsafe_ozon_autofill_values(conn, int(product_id))
    if cleanup_cleared > 0:
        product = get_product(conn, product_id)

    current_gallery_urls = _collect_product_gallery_urls(
        conn,
        int(product_id),
        fallback_image_url=str(product["image_url"] or ""),
        public_base_url=media_public_base_url,
    )
    product_core_filled, product_core_total = _product_core_fill_stats(dict(product))
    supplier_profiles = list_supplier_profiles(conn, only_active=True)
    supplier_profile_map = {str(p["supplier_name"]): p for p in supplier_profiles}
    strategy_options = [
        ("auto_full", "Авто: поставщик -> web-поиск -> AI"),
        ("supplier_only", "Только сайт поставщика"),
        ("web_only", "Только интернет-поиск + AI"),
        ("custom_domains", "Только выбранные домены"),
    ]
    strategy_values = [x[0] for x in strategy_options]
    default_strategy = str(parser_settings.get("source_strategy", "auto_full") or "auto_full")
    if default_strategy not in strategy_values:
        default_strategy = "auto_full"
    supplier_profile_options = sorted(supplier_profile_map.keys())
    profile_name = st.session_state.get(
        f"product_supplier_profile_{int(product_id)}",
        str(product["supplier_name"]) if (product["supplier_name"] and str(product["supplier_name"]) in supplier_profile_map) else "",
    )
    if profile_name not in supplier_profile_map:
        profile_name = ""
    selected_profile_template = supplier_profile_map.get(profile_name, {}).get("url_template") if profile_name else ""
    supplier_url_template = str(
        st.session_state.get(f"product_supplier_url_template_{int(product_id)}", selected_profile_template or "")
    )
    product_source_strategy = str(
        st.session_state.get(f"product_source_strategy_{int(product_id)}", default_strategy)
    )
    if product_source_strategy not in strategy_values:
        product_source_strategy = default_strategy
    product_extra_domains = str(
        st.session_state.get(
            f"product_extra_domains_{int(product_id)}",
            str(parser_settings.get("extra_fallback_domains", "") or ""),
        )
    )

    runtime_parser_settings = dict(parser_settings)
    runtime_parser_settings["source_strategy"] = str(locals().get("product_source_strategy") or parser_settings.get("source_strategy", "auto_full"))
    runtime_parser_settings["extra_fallback_domains"] = str(locals().get("product_extra_domains") or parser_settings.get("extra_fallback_domains", ""))

    if cleanup_cleared > 0:
        st.info(
            f"Из карточки автоматически убраны {int(cleanup_cleared)} служебных Ozon-атрибутов, "
            "которые были заполнены автоэвристикой слишком агрессивно. Их можно заполнить только вручную при необходимости."
        )

    top1, top2, top3, top4 = st.columns(4)
    top1.metric("Артикул", product["article"] or "-")
    top2.metric("Бренд", product["brand"] or "-")
    top3.metric("Категория", product["ozon_category_path"] or product["base_category"] or product["category"] or "-")
    top4.metric("Поставщик", product["supplier_name"] or "-")

    exact_ozon_ready = compute_quick_ozon_readiness(conn, dict(product))
    exact_detmir_ready = compute_quick_detmir_readiness(conn, dict(product))
    best_template_ready = compute_best_template_profile_readiness(conn, dict(product))
    detmir_category_id_current = int(product["detmir_category_id"] or 0)
    detmir_category_scope = f"detmir:{detmir_category_id_current}" if detmir_category_id_current > 0 else ""
    detmir_category_current = get_detmir_cached_category(conn, detmir_category_id_current) if detmir_category_id_current > 0 else None
    detmir_readiness_details = analyze_product_detmir_readiness(
        conn,
        product_id=int(product_id),
        category_id=detmir_category_id_current if detmir_category_id_current > 0 else None,
    )
    try:
        detmir_category_suggestions = suggest_detmir_categories_for_product(conn, dict(product), limit=8)
    except Exception:
        detmir_category_suggestions = []
    ready1, ready2, ready3, ready4, ready5 = st.columns(5)
    ready1.metric("Ozon ready", f"{int(exact_ozon_ready.get('readiness_pct') or 0)}%")
    ready2.metric("Ozon обязательные", f"{int(exact_ozon_ready.get('required_filled') or 0)}/{int(exact_ozon_ready.get('required_total') or 0)}")
    ready3.metric("Лучший шаблон", f"{int(best_template_ready.get('readiness_pct') or 0)}%")
    ready4.metric("Профилей по категории", int(best_template_ready.get("profiles_total") or 0))
    ready5.metric("Detmir ready", f"{int(exact_detmir_ready.get('readiness_pct') or 0)}%")
    if int(best_template_ready.get("profiles_total") or 0) > 0:
        st.caption(
            f"Лучший клиентский профиль сейчас: {best_template_ready.get('channel_code') or '-'} / "
            f"{best_template_ready.get('profile_name') or '-'} "
            f"({int(best_template_ready.get('filled_columns') or 0)}/{int(best_template_ready.get('matched_columns') or 0)})."
        )
    if str(exact_detmir_ready.get("status") or "") == "ok":
        st.caption(
            f"Detmir: обязательных заполнено {int(exact_detmir_ready.get('required_filled') or 0)}/"
            f"{int(exact_detmir_ready.get('required_total') or 0)}, "
            f"блокеров {int(exact_detmir_ready.get('blockers') or 0)}, "
            f"предупреждений {int(exact_detmir_ready.get('warnings') or 0)}, "
            f"фото {int(exact_detmir_ready.get('photos_count') or 0)}."
        )

    parse_status = product["supplier_parse_status"] if "supplier_parse_status" in product.keys() else None
    parse_comment = product["supplier_parse_comment"] if "supplier_parse_comment" in product.keys() else None
    parsed_at = product["supplier_last_parsed_at"] if "supplier_last_parsed_at" in product.keys() else None

    with st.container(border=True):
        h1, h2 = st.columns([1.3, 1])
        with h1:
            st.markdown(f"### Карточка товара #{product['id']}")
            st.caption(
                f"{str(product['article'] or product['supplier_article'] or product['internal_article'] or '-')} | "
                f"{str(product['name'] or '-')} | "
                f"этап: {_product_stage_label(dict(product))} | "
                f"парсинг: {_parse_status_label(parse_status)}"
            )
        with h2:
            headb1, headb2 = st.columns(2)
            with headb1:
                if st.button("← В каталог", use_container_width=True, key=f"product_back_to_catalog_{int(product_id)}"):
                    request_workspace_navigation("catalog")
            with headb2:
                if st.button(
                    "Заполнить карточку автоматически",
                    type="primary",
                    use_container_width=True,
                    help="Система последовательно попробует: Ozon-категорию, сайт поставщика / web-поиск, затем AI для названия, описания и атрибутов.",
                ):
                    bulk_assign_ozon_categories(conn, [int(product_id)], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
                    enrich_product_from_supplier(conn, int(product_id), force=False, parser_settings=runtime_parser_settings)
                    ai_fill_result = run_ai_enrichment_for_product(
                        conn=conn,
                        product_id=int(product_id),
                        settings=ai_settings,
                        include_title=True,
                        include_description=True,
                        include_attributes=True,
                        force=False,
                    )
                    if ai_fill_result.get("ok"):
                        st.success(
                            f"Автозаполнение завершено: название={'да' if ai_fill_result.get('title_applied') else 'нет'}, "
                            f"описание={'да' if ai_fill_result.get('description_applied') else 'нет'}, "
                            f"AI-атрибутов сохранено {int(ai_fill_result.get('attributes_saved') or 0)}."
                        )
                        st.rerun()
                    else:
                        st.error(str(ai_fill_result.get("error") or "Не удалось выполнить автоматическое заполнение карточки."))

        hm1, hm2, hm3, hm4, hm5 = st.columns(5)
        hm1.metric("Артикул", str(product["article"] or product["supplier_article"] or product["internal_article"] or "-"))
        hm2.metric("Категория", _short_text(product["ozon_category_path"] or product["base_category"] or product["category"] or "-", 34))
        hm3.metric("Фото", int(len(current_gallery_urls)))
        hm4.metric("Ozon ready", f"{int(exact_ozon_ready.get('readiness_pct') or 0)}%")
        hm5.metric("Detmir ready", f"{int(exact_detmir_ready.get('readiness_pct') or 0)}%")
        st.caption(
            f"Ядро: {int(product_core_filled)}/{int(product_core_total)} | "
            f"Поставщик: {product['supplier_name'] or '-'} | "
            f"Лучший шаблон: {int(best_template_ready.get('readiness_pct') or 0)}%"
        )

    card_main_col, card_rail_col = st.columns([2.15, 0.95], gap="large")

    with card_rail_col:
        st.markdown('<div class="pim-rail-card">', unsafe_allow_html=True)
        st.markdown('<div class="pim-section-kicker">Контекст</div><div class="pim-section-title">Состояние товара</div>', unsafe_allow_html=True)
        st.markdown(
            f'<div class="pim-compact-note">Парсинг: <strong>{_parse_status_label(parse_status)}</strong><br>'
            f'Обязательных Ozon: {int(exact_ozon_ready.get("required_filled") or 0)}/{int(exact_ozon_ready.get("required_total") or 0)}<br>'
            f'Обязательных Detmir: {int(exact_detmir_ready.get("required_filled") or 0)}/{int(exact_detmir_ready.get("required_total") or 0)}</div>',
            unsafe_allow_html=True,
        )
        if parse_comment:
            st.caption(_short_text(str(parse_comment), 180))
        if parsed_at:
            st.caption(f"Последний запуск: {parsed_at}")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="pim-rail-card">', unsafe_allow_html=True)
        st.markdown('<div class="pim-section-kicker">Медиа</div><div class="pim-section-title">Фото и галерея</div>', unsafe_allow_html=True)
        if current_gallery_urls:
            st.image(str(current_gallery_urls[0]), use_container_width=True)
            st.caption(f"Фото в карточке: {len(current_gallery_urls)}")
        else:
            st.caption("Фото в карточке пока нет.")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="pim-rail-card">', unsafe_allow_html=True)
        st.markdown('<div class="pim-section-kicker">Сохранение</div><div class="pim-section-title">Ручная правка</div>', unsafe_allow_html=True)
        st.markdown(
            '<div class="pim-compact-note">Сначала правь основные поля слева, затем сохраняй карточку. '
            'Редкие supplier/Ozon/Detmir-операции собраны ниже в одном блоке `Дополнительно`.</div>',
            unsafe_allow_html=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)

    with card_rail_col:
        with st.container(border=True):
            st.markdown("#### Дополнительно")
            st.caption("Здесь живут ручные операции по поставщику, Ozon, Detmir и сервисные режимы.")
        ctop1, ctop2, ctop3 = st.columns(3)
        with ctop1:
            if st.button("Спарсить поставщика", use_container_width=True, help="Обогатить карточку с сайта поставщика без жесткой перезаписи ручных значений"):
                result = enrich_product_from_supplier(conn, int(product_id), force=False, parser_settings=runtime_parser_settings)
                if result["ok"]:
                    st.success(result["message"])
                    if result.get("updates"):
                        st.json(result["updates"])
                    st.rerun()
                else:
                    st.error(result["message"])
        with ctop2:
            if st.button("Подобрать Ozon категорию", use_container_width=True, help="Подобрать эталонную Ozon категорию автоматически"):
                res = bulk_assign_ozon_categories(conn, [int(product_id)], min_score=OZON_CATEGORY_MIN_SCORE, force=False)
                refreshed_product = get_product(conn, int(product_id))
                materialized = {"slots_created": 0}
                if refreshed_product:
                    materialized = materialize_ozon_attribute_slots_for_product(
                        conn,
                        product_id=int(product_id),
                        description_category_id=int(refreshed_product["ozon_description_category_id"] or 0),
                        type_id=int(refreshed_product["ozon_type_id"] or 0),
                    )
                if res.get("message"):
                    st.info(str(res["message"]))
                else:
                    st.success(
                        f"Ozon автопривязка: обработано {res['processed']}, привязано {res['assigned']}, пропущено {res['skipped']}. "
                        f"Создано слотов Ozon-атрибутов: {int(materialized.get('slots_created') or 0)}."
                    )
                st.rerun()
        with ctop3:
            if st.button("Подобрать Detmir категорию", use_container_width=True, help="Подобрать клиентскую категорию Детского Мира по мастер-карточке и Ozon-ядру"):
                detmir_match = detect_best_detmir_category_for_product(conn, dict(product))
                if detmir_match.get("ok"):
                    matched = detmir_match.get("category") or {}
                    category_id = int(matched.get("category_id") or 0)
                    payload = {
                        "detmir_category_id": category_id or None,
                        "detmir_category_path": str(matched.get("full_path") or matched.get("name") or "").strip() or None,
                        "detmir_category_confidence": _detmir_confidence_from_match_score(matched.get("match_score")),
                    }
                    save_product(conn, int(product_id), payload)
                    save_field_source(
                        conn=conn,
                        product_id=int(product_id),
                        field_name="detmir_category_id",
                        source_type="detmir_category_match",
                        source_value_raw=category_id,
                        source_url=str(matched.get("full_path") or matched.get("name") or "detmir_match"),
                        confidence=min(0.99, max(0.35, _clamp_unit_confidence(payload.get("detmir_category_confidence")))),
                        is_manual=False,
                    )
                    st.success(
                        f"Detmir-категория подобрана: {payload['detmir_category_path'] or category_id}. "
                        "Следом можно импортировать требования категории и дозаполнить gaps."
                    )
                    st.rerun()
                else:
                    st.warning(str(detmir_match.get("message") or "Не удалось уверенно подобрать Detmir-категорию. Ниже доступны варианты."))
        with st.expander("Сервисные настройки и force-режимы", expanded=False):
            st.caption("Шаблон URL поставщика, force-перезаполнение, Detmir-gapfill и статистический расчёт габаритов.")
            sp1, sp2, sp3 = st.columns([2, 3, 1])
            with sp1:
                profile_name = st.selectbox(
                    "Профиль поставщика",
                    options=[""] + supplier_profile_options,
                    index=((supplier_profile_options.index(profile_name) + 1) if profile_name in supplier_profile_options else 0),
                    key=f"product_supplier_profile_{int(product_id)}",
                )
            with sp2:
                supplier_url_template = st.text_input(
                    "URL template для товара",
                    value=selected_profile_template or supplier_url_template or "",
                    placeholder="https://site.ru/catalog/?q={supplier_article_q}",
                    key=f"product_supplier_url_template_{int(product_id)}",
                    help="Поддерживаются плейсхолдеры: {article}, {supplier_article}, {name}, {code} и *_q.",
                )
            with sp3:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("Подставить URL", key=f"product_apply_supplier_url_{int(product_id)}", use_container_width=True):
                    render_payload = {
                        "article": product["article"],
                        "supplier_article": product["supplier_article"],
                        "name": product["name"],
                        "category": product["category"],
                        "code": product["supplier_article"] or product["article"],
                    }
                    generated_url = render_supplier_url(supplier_url_template, render_payload) if supplier_url_template else None
                    if generated_url:
                        save_product(
                            conn,
                            int(product_id),
                            {
                                "supplier_name": profile_name or product["supplier_name"] or None,
                                "supplier_url": generated_url,
                            },
                        )
                        save_field_source(
                            conn=conn,
                            product_id=int(product_id),
                            field_name="supplier_url",
                            source_type="manual",
                            source_value_raw=generated_url,
                            source_url=None,
                            confidence=1.0,
                            is_manual=True,
                        )
                        st.success("URL поставщика подставлен из шаблона профиля.")
                        st.rerun()
                    else:
                        st.warning("Не удалось собрать URL. Проверь шаблон и поля товара.")

            sp4, sp5 = st.columns([2, 3])
            with sp4:
                product_source_strategy = st.selectbox(
                    "Стратегия парсинга для этого товара",
                    options=strategy_values,
                    index=strategy_values.index(product_source_strategy),
                    format_func=lambda x: next((label for key, label in strategy_options if key == x), x),
                    key=f"product_source_strategy_{int(product_id)}",
                )
            with sp5:
                product_extra_domains = st.text_input(
                    "Доп. домены (override для товара)",
                    value=product_extra_domains,
                    key=f"product_extra_domains_{int(product_id)}",
                    help="Через запятую: домены, где искать карточку товара для fallback.",
                )

            adv1, adv2, adv3 = st.columns(3)
            with adv1:
                if st.button("Перезаполнить из поставщика", use_container_width=True, help="Жесткая перезапись значений из supplier page"):
                    result = enrich_product_from_supplier(conn, int(product_id), force=True, parser_settings=runtime_parser_settings)
                    if result["ok"]:
                        st.success(result["message"])
                        if result.get("updates"):
                            st.json(result["updates"])
                        st.rerun()
                    else:
                        st.error(result["message"])
            with adv2:
                if st.button("Перепривязать Ozon категорию", use_container_width=True, help="Повторно назначить Ozon категорию с перезаписью текущей привязки"):
                    res = bulk_assign_ozon_categories(conn, [int(product_id)], min_score=OZON_CATEGORY_MIN_SCORE, force=True)
                    refreshed_product = get_product(conn, int(product_id))
                    materialized = {"slots_created": 0}
                    if refreshed_product:
                        materialized = materialize_ozon_attribute_slots_for_product(
                            conn,
                            product_id=int(product_id),
                            description_category_id=int(refreshed_product["ozon_description_category_id"] or 0),
                            type_id=int(refreshed_product["ozon_type_id"] or 0),
                        )
                    if res.get("message"):
                        st.info(str(res["message"]))
                    else:
                        st.success(
                            f"Ozon force-привязка: обработано {res['processed']}, привязано {res['assigned']}, пропущено {res['skipped']}. "
                            f"Создано слотов Ozon-атрибутов: {int(materialized.get('slots_created') or 0)}."
                        )
                    st.rerun()
            with adv3:
                if st.button("Очистить Ozon-привязку", use_container_width=True, help="Убрать текущую Ozon категорию, если матчинг был неверным"):
                    save_product(
                        conn,
                        int(product_id),
                        {
                            "ozon_description_category_id": None,
                            "ozon_type_id": None,
                            "ozon_category_path": None,
                            "ozon_category_confidence": None,
                        },
                    )
                    st.success("Ozon-привязка очищена. Ниже можно вручную поправить category/base/subcategory и сохранить карточку.")
                    st.rerun()

            det1, det2, det3 = st.columns(3)
            with det1:
                if st.button("Импортировать требования Detmir", use_container_width=True, help="Подтянуть overlay-атрибуты выбранной Detmir-категории в PIM"):
                    if detmir_category_id_current <= 0:
                        st.warning("Сначала назначь Detmir-категорию.")
                    else:
                        import_result = import_detmir_category_requirements_to_pim(conn, category_id=detmir_category_id_current)
                        st.success(
                            f"Detmir overlay импортирован: атрибутов {int(import_result.get('imported') or 0)}, "
                            f"обязательных {int(import_result.get('required') or 0)}, "
                            f"mapping rules {int(import_result.get('mapping_saved') or 0)}."
                        )
                        st.rerun()
            with det2:
                if st.button("Заполнить gaps под Detmir", use_container_width=True, help="Перенести уже найденные master/Ozon-данные в Detmir overlay для текущей категории"):
                    if detmir_category_id_current <= 0:
                        st.warning("Сначала назначь Detmir-категорию.")
                    else:
                        existing_detmir_requirements = list_channel_requirements(conn, channel_code="detmir", category_code=detmir_category_scope)
                        if not existing_detmir_requirements:
                            import_detmir_category_requirements_to_pim(conn, category_id=detmir_category_id_current)
                        gapfill_result = _fill_channel_attrs_from_product_state(
                            conn=conn,
                            product_row=dict(product),
                            channel_code="detmir",
                            category_code=detmir_category_scope,
                            source_type="derived_from_master",
                            source_url="detmir_overlay_gapfill",
                            force=False,
                            target_channel_code="detmir",
                        )
                        st.success(
                            f"Detmir gaps обработаны: заполнено {int(gapfill_result.get('saved') or 0)}, "
                            f"пропущено {int(gapfill_result.get('skipped') or 0)}, "
                            f"всего целей {int(gapfill_result.get('targets') or 0)}."
                        )
                        st.rerun()
            with det3:
                if st.button("Очистить Detmir-привязку", use_container_width=True, help="Убрать текущую категорию Детского Мира, если матчинг был неверным"):
                    save_product(
                        conn,
                        int(product_id),
                        {
                            "detmir_category_id": None,
                            "detmir_category_path": None,
                            "detmir_category_confidence": None,
                        },
                    )
                    st.success("Detmir-привязка очищена. Ниже можно вручную выбрать или скорректировать категорию и сохранить карточку.")
                    st.rerun()

            ctop5, ctop6, ctop7 = st.columns([1, 1, 2])
            with ctop5:
                card_dim_min_samples = st.number_input(
                    "Мин. выборка",
                    min_value=1,
                    max_value=50,
                    value=4,
                    step=1,
                    key=f"card_dim_min_samples_{int(product_id)}",
                )
            with ctop6:
                card_dim_force = st.checkbox(
                    "Перезаписать существующие",
                    value=False,
                    key=f"card_dim_force_{int(product_id)}",
                )
            with ctop7:
                st.markdown("<br>", unsafe_allow_html=True)
                if st.button("Рассчитать габариты/вес (статистика)", key=f"card_estimate_dims_{int(product_id)}", use_container_width=True):
                    dim_result = estimate_dimensions_for_product(
                        conn=conn,
                        product_id=int(product_id),
                        force=bool(card_dim_force),
                        min_samples=int(card_dim_min_samples),
                    )
                    if dim_result.get("ok"):
                        st.success(
                            f"{dim_result.get('message')}. Источники: {', '.join(dim_result.get('used_sources') or []) or '-'}"
                        )
                        if dim_result.get("updates"):
                            st.json(dim_result.get("updates"))
                        st.rerun()
                    else:
                        st.error(str(dim_result.get("message") or "Не удалось рассчитать габариты/вес"))

    registry_legal_key = f"registry_legal_entity_{int(product_id)}"
    registry_kind_key = f"registry_product_kind_{int(product_id)}"
    registry_tnved_key = f"registry_tnved_{int(product_id)}"
    registry_url_key = f"registry_manual_url_{int(product_id)}"
    registry_candidates_key = f"registry_candidates_{int(product_id)}"
    registry_queries_key = f"registry_queries_{int(product_id)}"
    registry_errors_key = f"registry_errors_{int(product_id)}"
    registry_timeout_key = f"registry_timeout_{int(product_id)}"
    registry_limit_key = f"registry_limit_{int(product_id)}"

    profile_legal_entity = ""
    if product["supplier_name"]:
        profile_legal_entity = str(supplier_profile_map.get(str(product["supplier_name"]), {}).get("legal_entity_name") or "").strip()
    if registry_legal_key not in st.session_state:
        st.session_state[registry_legal_key] = profile_legal_entity
    if registry_kind_key not in st.session_state:
        st.session_state[registry_kind_key] = _build_registry_product_kind(dict(product))
    if registry_tnved_key not in st.session_state:
        st.session_state[registry_tnved_key] = str(product["tnved_code"] or "").strip()
    if registry_url_key not in st.session_state:
        st.session_state[registry_url_key] = ""
    if registry_timeout_key not in st.session_state:
        st.session_state[registry_timeout_key] = 35
    if registry_limit_key not in st.session_state:
        st.session_state[registry_limit_key] = 8

    with st.expander("Сертификаты / декларации ФСА", expanded=False):
        st.caption(
            "Поиск идёт по юрлицу поставщика, виду товара и коду ТН ВЭД. "
            "Найденный документ сохраняется к товару вместе с PDF, если файл доступен."
        )
        if not profile_legal_entity:
            st.info("У профиля поставщика пока не задано юрлицо. Его можно заполнить во вкладке Импорт -> Профили поставщиков.")
        rg1, rg2, rg3 = st.columns([2, 2, 1.2])
        with rg1:
            registry_legal_entity = st.text_input("Юрлицо поставщика", key=registry_legal_key)
        with rg2:
            registry_product_kind = st.text_input("Вид товара", key=registry_kind_key, help="Например: беговел, велосипед, велофонарь, насос.")
        with rg3:
            registry_tnved = st.text_input("Код ТН ВЭД", key=registry_tnved_key)
        rg4, rg5, rg6 = st.columns([3, 1, 1])
        with rg4:
            registry_manual_url = st.text_input(
                "Ручной URL документа/карточки ФСА",
                key=registry_url_key,
                placeholder="https://pub.fsa.gov.ru/rds/declaration/view/...",
            )
        with rg5:
            registry_limit = st.number_input("Кандидатов", min_value=1, max_value=20, step=1, key=registry_limit_key)
        with rg6:
            registry_timeout = st.number_input("Таймаут, сек", min_value=5, max_value=90, step=1, key=registry_timeout_key)

        rb1, rb2 = st.columns([1, 1])
        with rb1:
            if st.button("Найти в реестре ФСА", key=f"registry_search_btn_{int(product_id)}", type="primary"):
                search_result = search_fsa_registry_candidates(
                    legal_entity=registry_legal_entity,
                    product_name=str(product["name"] or "").strip(),
                    product_kind=registry_product_kind,
                    tnved_code=registry_tnved,
                    max_results=int(registry_limit),
                    timeout=float(registry_timeout),
                )
                st.session_state[registry_candidates_key] = search_result.get("items") or []
                st.session_state[registry_queries_key] = search_result.get("queries") or []
                st.session_state[registry_errors_key] = search_result.get("errors") or []
                if search_result.get("ok"):
                    st.success(f"Найдено кандидатов: {len(search_result.get('items') or [])}")
                else:
                    st.warning(str(search_result.get("error") or "Кандидаты не найдены"))
        with rb2:
            if st.button("Разобрать ручной URL", key=f"registry_parse_manual_btn_{int(product_id)}"):
                if not str(registry_manual_url or "").strip():
                    st.warning("Сначала вставь URL документа или карточки ФСА.")
                else:
                    try:
                        parsed_registry_doc = parse_fsa_document_resource(
                            str(registry_manual_url).strip(),
                            timeout=float(registry_timeout),
                        )
                        save_result = save_fsa_document(
                            conn,
                            int(product_id),
                            parsed_registry_doc,
                            pdf_bytes=parsed_registry_doc.get("pdf_bytes"),
                        )
                        st.success(
                            f"Документ сохранён: {_registry_kind_label(parsed_registry_doc.get('kind'))}, "
                            f"№ {parsed_registry_doc.get('doc_number') or '-'}"
                        )
                        if save_result.get("file_path"):
                            st.caption(f"PDF сохранён: {save_result.get('file_path')}")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Не удалось разобрать URL ФСА: {e}")

        search_errors = st.session_state.get(registry_errors_key) or []
        search_queries = st.session_state.get(registry_queries_key) or []
        if search_queries:
            st.caption("Поисковые запросы: " + " | ".join([str(x) for x in search_queries if x]))
        if search_errors:
            for err in search_errors[:4]:
                st.warning(str(err))

        candidate_rows = st.session_state.get(registry_candidates_key) or []
        if candidate_rows:
            candidate_options = list(range(len(candidate_rows)))
            selected_registry_idx = st.selectbox(
                "Найденные кандидаты",
                options=candidate_options,
                index=0,
                format_func=lambda idx: (
                    f"{_registry_kind_label(candidate_rows[idx].get('kind'))} | "
                    f"{candidate_rows[idx].get('title') or '-'} | "
                    f"{candidate_rows[idx].get('link') or '-'}"
                ),
                key=f"registry_candidate_idx_{int(product_id)}",
            )
            chosen_candidate = candidate_rows[int(selected_registry_idx)]
            st.caption(str(chosen_candidate.get("description") or ""))
            if st.button("Разобрать и сохранить выбранный документ", key=f"registry_parse_candidate_btn_{int(product_id)}"):
                try:
                    parsed_registry_doc = parse_fsa_document_resource(
                        str(chosen_candidate.get("link") or "").strip(),
                        timeout=float(registry_timeout),
                    )
                    save_result = save_fsa_document(
                        conn,
                        int(product_id),
                        parsed_registry_doc,
                        pdf_bytes=parsed_registry_doc.get("pdf_bytes"),
                    )
                    st.success(
                        f"Документ сохранён: {_registry_kind_label(parsed_registry_doc.get('kind'))}, "
                        f"№ {parsed_registry_doc.get('doc_number') or '-'}"
                    )
                    if save_result.get("file_path"):
                        st.caption(f"PDF сохранён: {save_result.get('file_path')}")
                    st.rerun()
                except Exception as e:
                    st.error(f"Не удалось разобрать выбранный документ: {e}")

        registry_docs = list_fsa_documents(conn, int(product_id))
        if registry_docs:
            st.markdown("#### Сохранённые документы ФСА")
            latest_doc = registry_docs[0]
            m1, m2, m3, m4, m5 = st.columns(5)
            m1.metric("Тип", _registry_kind_label(latest_doc.get("doc_kind")))
            m2.metric("Номер", str(latest_doc.get("doc_number") or "-"))
            m3.metric("Начало", str(latest_doc.get("valid_from") or "-"))
            m4.metric("Окончание", str(latest_doc.get("valid_to") or "-"))
            m5.metric("Файлов", str(sum(1 for row in registry_docs if row.get("local_file_path") or row.get("pdf_url"))))
            for doc in registry_docs:
                with st.container(border=True):
                    d1, d2, d3 = st.columns([3, 2, 1])
                    with d1:
                        st.markdown(
                            f"**{_registry_kind_label(doc.get('doc_kind'))}**  \n"
                            f"Номер: `{doc.get('doc_number') or '-'}`  \n"
                            f"Срок: {doc.get('valid_from') or '-'} -> {doc.get('valid_to') or '-'}  \n"
                            f"Орган: {doc.get('authority_name') or '-'}  \n"
                            f"Заявитель: {doc.get('applicant_name') or '-'}"
                        )
                        if doc.get("source_url"):
                            st.caption(f"Источник: {doc.get('source_url')}")
                    with d2:
                        st.caption(f"ТН ВЭД: {doc.get('tnved_code') or '-'}")
                        if doc.get("pdf_url"):
                            st.caption(f"PDF URL: {doc.get('pdf_url')}")
                    with d3:
                        local_file = str(doc.get("local_file_path") or "").strip()
                        if local_file and Path(local_file).exists():
                            try:
                                pdf_bytes = Path(local_file).read_bytes()
                                st.download_button(
                                    "Скачать PDF",
                                    data=pdf_bytes,
                                    file_name=Path(local_file).name,
                                    mime="application/pdf",
                                    key=f"registry_pdf_download_{int(doc['id'])}",
                                )
                            except Exception as e:
                                st.caption(f"PDF недоступен: {e}")
                        if st.button("Удалить", key=f"registry_delete_{int(doc['id'])}"):
                            delete_fsa_document(conn, int(doc["id"]))
                            st.success("Документ удалён")
                            st.rerun()
        else:
            st.info("По этому товару ещё нет сохранённых сертификатов или деклараций ФСА.")

    with st.expander("AI-контент: описание, атрибуты, фото", expanded=False):
        ai_ok, ai_msg = ai_is_configured(ai_settings)
        if ai_ok:
            st.success(ai_msg)
        else:
            st.warning(ai_msg)
            st.caption("Проверь настройки в разделе `Настройки` -> `AI`.")

        ai_desc_key = f"ai_desc_candidate_{int(product_id)}"
        ai_title_key = f"ai_title_candidate_{int(product_id)}"
        ai_attr_key = f"ai_attr_candidate_{int(product_id)}"
        ai_img_key = f"ai_img_prompts_{int(product_id)}"
        ai_img_result_key = f"ai_img_results_{int(product_id)}"
        ai_desc_widget_key = f"ai_desc_text_{int(product_id)}"
        ai_title_widget_key = f"ai_title_text_{int(product_id)}"
        ai_prompt1_key = f"ai_img_prompt_1_{int(product_id)}"
        ai_prompt2_key = f"ai_img_prompt_2_{int(product_id)}"
        ai_source_label = f"{str(ai_settings.get('provider') or '-')}/{str(ai_settings.get('chat_model') or '-')}"
        ai_mode_key = f"ai_mode_{int(product_id)}"
        ai_verify_key = f"ai_verify_result_{int(product_id)}"
        ai_image_plan_key = f"ai_image_plan_{int(product_id)}"

        ai_mode = st.radio(
            "Режим AI",
            options=["fast_batch", "deep_repair"],
            index=0,
            horizontal=True,
            key=ai_mode_key,
            format_func=lambda x: "Fast batch" if x == "fast_batch" else "Deep repair",
            help="Fast batch — verifier + массовый рерайт. Deep repair — более тяжёлый режим для спорных SKU и редких ручных вмешательств.",
        )

        current_verify_state = st.session_state.get(ai_verify_key) or {}
        if current_verify_state:
            st.caption(
                f"AI verifier: {str(current_verify_state.get('verdict') or '-')} | "
                f"confidence {int(_safe_float_value(current_verify_state.get('confidence'), 0.0) * 100)}% | "
                f"{str(current_verify_state.get('summary') or '-')}"
            )
        current_image_plan = st.session_state.get(ai_image_plan_key) or {}
        if current_image_plan:
            st.caption(
                f"Image plan: {str(current_image_plan.get('queue') or '-')} | "
                f"фото {int(current_image_plan.get('gallery_count') or 0)} | "
                f"не хватает слотов {int(current_image_plan.get('missing_slots') or 0)}."
            )

        d1, d2, d3, d4 = st.columns([1, 1, 1, 1])
        with d1:
            if st.button("AI: Проверить parser result", key=f"btn_ai_verify_{int(product_id)}"):
                verify_result = verify_parser_result_for_product(conn, int(product_id), ai_settings, mode=str(ai_mode))
                if verify_result.get("ok"):
                    st.session_state[ai_verify_key] = verify_result
                    st.success(
                        f"Verifier: {verify_result.get('verdict')} "
                        f"({int(_safe_float_value(verify_result.get('confidence'), 0.0) * 100)}%)."
                    )
                else:
                    st.error(f"Не удалось проверить parser result: {verify_result.get('error')}")
        with d2:
            if st.button("AI: Быстрый рерайт названия", key=f"btn_ai_title_{int(product_id)}"):
                title_result = generate_selling_title_for_product(conn, int(product_id), ai_settings)
                if title_result.get("ok"):
                    generated_title = str(title_result.get("title") or "").strip()
                    st.session_state[ai_title_key] = generated_title
                    st.session_state[ai_title_widget_key] = generated_title
                    st.success(
                        f"AI-черновик названия готов "
                        f"(модель: {title_result.get('model')})."
                    )
                else:
                    st.error(f"Не удалось сгенерировать название: {title_result.get('error')}")
        with d3:
            if st.button("AI: Чистое описание + SEO поля", key=f"btn_ai_desc_{int(product_id)}"):
                if str(ai_mode) == "deep_repair":
                    desc_result = generate_seo_description_for_product(conn, int(product_id), ai_settings)
                    generated_desc = str(desc_result.get("text") or "").strip() if desc_result.get("ok") else ""
                else:
                    desc_result = generate_product_copy_pack_for_product(conn, int(product_id), ai_settings, mode=str(ai_mode))
                    generated_desc = str(desc_result.get("description") or "").strip() if desc_result.get("ok") else ""
                if desc_result.get("ok"):
                    st.session_state[ai_desc_key] = generated_desc
                    st.session_state[ai_desc_widget_key] = generated_desc
                    st.success(
                        f"AI-черновик описания готов "
                        f"(модель: {desc_result.get('model')})."
                    )
                    if str(ai_mode) != "deep_repair":
                        st.caption(
                            f"SEO вынесено отдельно: meta title `{str(desc_result.get('meta_title') or '-')[:64]}`, "
                            f"meta description `{str(desc_result.get('meta_description') or '-')[:96]}`."
                        )
                else:
                    st.error(f"Не удалось сгенерировать описание: {desc_result.get('error')}")
        with d4:
            if st.button("AI: Найти пустые Ozon-атрибуты", key=f"btn_ai_attrs_{int(product_id)}"):
                attr_result = generate_ai_attribute_suggestions_for_product(conn, int(product_id), ai_settings, limit=20)
                if attr_result.get("ok"):
                    st.session_state[ai_attr_key] = attr_result.get("suggestions") or []
                    st.success(
                        f"AI подготовил подсказки: {len(st.session_state[ai_attr_key])} "
                        f"из {int(attr_result.get('missing_total') or 0)} пустых атрибутов."
                    )
                else:
                    st.error(f"Не удалось получить AI-подсказки атрибутов: {attr_result.get('error')}")

        imgc1, imgc2 = st.columns([1, 1])
        with imgc1:
            if st.button("AI: План 3–5 фото", key=f"btn_ai_photo_plan_{int(product_id)}"):
                image_plan = build_image_gallery_plan_for_product(conn, int(product_id))
                st.session_state[ai_image_plan_key] = image_plan
                prompt_result = {
                    "context_prompt": str(image_plan.get("context_prompt") or ""),
                    "color_prompt": str(image_plan.get("color_prompt") or ""),
                }
                st.session_state[ai_img_key] = prompt_result
                st.session_state[ai_prompt1_key] = str(prompt_result.get("context_prompt") or "")
                st.session_state[ai_prompt2_key] = str(prompt_result.get("color_prompt") or "")
                st.success(
                    f"План фото готов: {image_plan.get('queue')} | "
                    f"фото {int(image_plan.get('gallery_count') or 0)} | "
                    f"не хватает {int(image_plan.get('missing_slots') or 0)}."
                )
        with imgc2:
            if st.button("AI: Подготовить 2 промпта для фото", key=f"btn_ai_photo_prompts_{int(product_id)}"):
                prompt_result = build_marketing_image_prompts_for_product(conn, int(product_id))
                st.session_state[ai_img_key] = prompt_result
                st.session_state[ai_prompt1_key] = str(prompt_result.get("context_prompt") or "")
                st.session_state[ai_prompt2_key] = str(prompt_result.get("color_prompt") or "")
                st.success("Промпты для генерации изображений подготовлены.")

        ai_title_value = str(st.session_state.get(ai_title_key) or "").strip()
        if ai_title_widget_key not in st.session_state:
            st.session_state[ai_title_widget_key] = ai_title_value
        ai_title_text = st.text_area(
            "Черновик AI-названия",
            height=90,
            key=ai_title_widget_key,
            placeholder="Сначала нажми `AI: Продающее название`.",
        )
        st.session_state[ai_title_key] = ai_title_text

        tt1, tt2 = st.columns([1, 1])
        with tt1:
            if st.button("Применить AI-название в карточку", key=f"btn_ai_title_apply_save_{int(product_id)}"):
                text_to_apply = str(st.session_state.get(ai_title_key) or "").strip()
                if not text_to_apply:
                    st.warning("Нет AI-названия для применения.")
                else:
                    if can_overwrite_field(conn, int(product_id), "name", "ai", force=False):
                        save_product(conn, int(product_id), {"name": text_to_apply})
                        save_field_source(
                            conn=conn,
                            product_id=int(product_id),
                            field_name="name",
                            source_type="ai",
                            source_value_raw=text_to_apply,
                            source_url=ai_source_label,
                            confidence=0.72,
                            is_manual=False,
                        )
                        st.success("AI-название сохранено в поле `Название`.")
                        st.rerun()
                    else:
                        st.warning("Поле `Название` защищено более приоритетным источником.")
        with tt2:
            st.caption("Название собирается под карточку товара: тип товара + бренд/модель + полезные характеристики без пустого маркетинга.")

        ai_desc_value = str(st.session_state.get(ai_desc_key) or "").strip()
        if ai_desc_widget_key not in st.session_state:
            st.session_state[ai_desc_widget_key] = ai_desc_value
        ai_desc_text = st.text_area(
            "Черновик AI-описания",
            height=220,
            key=ai_desc_widget_key,
            placeholder="Сначала нажми `AI: Чистое описание + SEO поля`.",
        )
        st.session_state[ai_desc_key] = ai_desc_text

        dd1, dd2 = st.columns([1, 1])
        with dd1:
            if st.button("Применить AI-описание в карточку", key=f"btn_ai_desc_apply_save_{int(product_id)}"):
                text_to_apply = str(st.session_state.get(ai_desc_key) or "").strip()
                if not text_to_apply:
                    st.warning("Нет AI-описания для применения.")
                else:
                    if can_overwrite_field(conn, int(product_id), "description", "ai", force=False):
                        save_product(conn, int(product_id), {"description": text_to_apply})
                        save_field_source(
                            conn=conn,
                            product_id=int(product_id),
                            field_name="description",
                            source_type="ai",
                            source_value_raw=text_to_apply,
                            source_url=ai_source_label,
                            confidence=0.7,
                            is_manual=False,
                        )
                        st.success("AI-описание сохранено в поле `Описание`.")
                        st.rerun()
                    else:
                        st.warning("Поле `Описание` защищено более приоритетным источником (например manual).")
        with dd2:
            if st.button("Подставить AI-описание в форму ниже", key=f"btn_ai_desc_prefill_{int(product_id)}"):
                text_to_prefill = str(st.session_state.get(ai_desc_key) or "").strip()
                if not text_to_prefill:
                    st.warning("Нет AI-описания для подстановки.")
                else:
                    st.session_state[f"ai_description_prefill_{int(product_id)}"] = text_to_prefill
                    st.success("Описание подставлено в форму карточки.")
                    st.rerun()

        ai_suggestions = st.session_state.get(ai_attr_key) or []
        if ai_suggestions:
            st.markdown("#### AI-подсказки для пустых Ozon-атрибутов")
            ai_attr_df = pd.DataFrame(ai_suggestions)
            st.dataframe(
                with_ru_columns(ai_attr_df, extra_map={"reason": "Обоснование"}),
                use_container_width=True,
                hide_index=True,
            )
            if st.button("Применить AI-подсказки атрибутов", key=f"btn_ai_attrs_apply_{int(product_id)}"):
                apply_res = apply_ai_attribute_suggestions(
                    conn=conn,
                    product_id=int(product_id),
                    suggestions=ai_suggestions,
                    channel_code=None,
                    source_url=ai_source_label,
                )
                st.success(
                    f"AI-атрибуты применены: сохранено {int(apply_res.get('saved') or 0)}, "
                    f"пропущено {int(apply_res.get('skipped') or 0)}, ошибок {int(apply_res.get('errors') or 0)}."
                )
                st.rerun()

        prompts_obj = st.session_state.get(ai_img_key) or {}
        context_prompt_default = str(prompts_obj.get("context_prompt") or "")
        color_prompt_default = str(prompts_obj.get("color_prompt") or "")
        if ai_prompt1_key not in st.session_state:
            st.session_state[ai_prompt1_key] = context_prompt_default
        if ai_prompt2_key not in st.session_state:
            st.session_state[ai_prompt2_key] = color_prompt_default
        img_p1 = st.text_area(
            "Промпт 1: контекстный фон + инфографика",
            height=130,
            key=ai_prompt1_key,
        )
        img_p2 = st.text_area(
            "Промпт 2: цветной фон + инфографика",
            height=130,
            key=ai_prompt2_key,
        )
        if st.button("AI: Сгенерировать 2 изображения", key=f"btn_ai_generate_images_{int(product_id)}"):
            generation_results = generate_images_from_prompts(
                ai_settings,
                prompts=[img_p1, img_p2],
                size=str(ai_settings.get("image_size") or "1024x1024"),
            )
            st.session_state[ai_img_result_key] = generation_results
            if generation_results and any(bool(x.get("ok")) for x in generation_results):
                st.success("Генерация изображений завершена.")
            else:
                first_error = generation_results[0].get("error") if generation_results else "Пустой ответ image API."
                st.error(f"Не удалось сгенерировать изображения: {first_error}")

        generated_images = st.session_state.get(ai_img_result_key) or []
        if generated_images:
            st.markdown("#### AI-результат по изображениям")
            for idx, item in enumerate(generated_images, start=1):
                if not bool(item.get("ok")):
                    st.error(f"Изображение {idx}: {item.get('error')}")
                    continue
                st.caption(f"Вариант {idx}. Модель: {item.get('model')}")
                if item.get("image_bytes"):
                    img_bytes = item.get("image_bytes")
                    st.image(img_bytes, caption=f"AI-вариант {idx}")
                    st.download_button(
                        f"Скачать AI-вариант {idx}",
                        data=img_bytes,
                        file_name=f"product_{int(product_id)}_ai_variant_{idx}.png",
                        mime="image/png",
                        key=f"ai_img_download_{int(product_id)}_{idx}",
                    )
                elif item.get("image_url"):
                    st.image(str(item.get("image_url")), caption=f"AI-вариант {idx}")
                    st.code(str(item.get("image_url")), language="text")

    ozon_desc_id = int(product["ozon_description_category_id"] or 0)
    ozon_type_id = int(product["ozon_type_id"] or 0)
    if ozon_desc_id > 0 and ozon_type_id > 0:
        auto_req = ensure_ozon_requirements_for_product_category(conn, ozon_desc_id, ozon_type_id)
        auto_slots = materialize_ozon_attribute_slots_for_product(conn, int(product_id), ozon_desc_id, ozon_type_id)
        auto_backfill = _fill_channel_attrs_from_product_state(
            conn=conn,
            product_row=dict(product),
            channel_code="ozon",
            category_code=f"ozon:{ozon_desc_id}:{ozon_type_id}",
            source_type="derived_from_master",
            source_url="product_state",
            force=False,
            target_channel_code=None,
        )
        if int(auto_req.get("imported") or 0) > 0:
            st.success(
                "Ozon-атрибуты этой категории автоматически добавлены в карточку: "
                f"{int(auto_req.get('imported') or 0)} (обязательных: {int(auto_req.get('required') or 0)})."
            )
        if int(auto_slots.get("created") or 0) > 0:
            st.info(
                "Для товара автоматически созданы пустые Ozon-атрибуты для заполнения: "
                f"{int(auto_slots.get('created') or 0)} из {int(auto_slots.get('requirements') or 0)}."
            )
        if int(auto_backfill.get("saved") or 0) > 0:
            st.info(
                "Для товара автоматически дозаполнены Ozon-атрибуты из уже накопленных полей карточки: "
                f"{int(auto_backfill.get('saved') or 0)}."
            )

    st.markdown("### Детский Мир: overlay и готовность")
    detmir_summary = detmir_readiness_details.get("summary") or {}
    detmir_current_path = str(product["detmir_category_path"] or "") or str((detmir_category_current or {}).get("full_path") or "")
    dm1, dm2, dm3, dm4 = st.columns([1.2, 1.2, 1.2, 1.6])
    dm1.metric("Detmir ready", f"{int(detmir_summary.get('readiness_pct') or 0)}%")
    dm2.metric("Detmir обязательные", f"{int(detmir_summary.get('required_filled') or 0)}/{int(detmir_summary.get('required_total') or 0)}")
    dm3.metric("Detmir блокеры", int(detmir_summary.get("blockers") or 0))
    dm4.metric("Detmir фото", int(detmir_summary.get("photos_count") or 0))
    if detmir_current_path:
        st.caption(f"Текущая Detmir-категория: {detmir_current_path}")
    if not str(product["barcode"] or "").strip():
        st.warning("Для Детского Мира штрихкод критичен: без него карточка и фото-flow будут с высоким риском блокировки.")
    photo_count = int(detmir_summary.get("photos_count") or 0)
    if photo_count < 3:
        st.info("Для Детского Мира лучше держать минимум 3 фото. Если у поставщика фото мало, добивай галерею web/AI-пайплайном или вручную.")
    if current_gallery_urls:
        detmir_non_public = [url for url in current_gallery_urls if not str(url).strip().startswith(("http://", "https://"))]
        if detmir_non_public:
            st.warning("В галерее есть непубличные ссылки/локальные пути. Для Detmir лучше оставить только публичные http/https ссылки.")

    suggestion_options = [int(row.get("category_id") or 0) for row in detmir_category_suggestions if int(row.get("category_id") or 0) > 0]
    suggestion_map = {int(row.get("category_id") or 0): row for row in detmir_category_suggestions if int(row.get("category_id") or 0) > 0}
    if suggestion_options:
        sm1, sm2 = st.columns([2, 1])
        with sm1:
            selected_detmir_suggestion = st.selectbox(
                "Подсказки по Detmir-категории",
                options=suggestion_options,
                index=0,
                format_func=lambda cid: (
                    f"{suggestion_map[int(cid)].get('full_path') or suggestion_map[int(cid)].get('name')} "
                    f"| score={float(suggestion_map[int(cid)].get('match_score') or 0):.2f}"
                ),
                key=f"product_detmir_suggestion_{int(product_id)}",
            )
        with sm2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Применить выбранную Detmir-категорию", key=f"apply_detmir_suggestion_{int(product_id)}"):
                chosen = suggestion_map.get(int(selected_detmir_suggestion)) or {}
                chosen_id = int(chosen.get("category_id") or 0)
                save_product(
                    conn,
                    int(product_id),
                    {
                        "detmir_category_id": chosen_id or None,
                        "detmir_category_path": str(chosen.get("full_path") or chosen.get("name") or "").strip() or None,
                        "detmir_category_confidence": _detmir_confidence_from_match_score(chosen.get("match_score")),
                    },
                )
                save_field_source(
                    conn=conn,
                    product_id=int(product_id),
                    field_name="detmir_category_id",
                    source_type="manual",
                    source_value_raw=chosen_id,
                    source_url=str(chosen.get("full_path") or chosen.get("name") or "detmir_manual_pick"),
                    confidence=1.0,
                    is_manual=True,
                )
                st.success("Detmir-категория применена. Теперь импортируй требования и дозаполни gaps.")
                st.rerun()

    with st.expander("Detmir readiness: обязательные поля, справочники и payload", expanded=False):
        readiness_rows = detmir_readiness_details.get("rows") or []
        if readiness_rows:
            readiness_df = pd.DataFrame(readiness_rows)
            st.dataframe(
                with_ru_columns(
                    readiness_df,
                    extra_map={
                        "target": "Цель Detmir",
                        "attribute_code": "Код атрибута PIM",
                        "attribute_name": "Название атрибута",
                        "required": "Обязательный",
                        "status": "Статус",
                        "value": "Текущее значение",
                        "resolved_value": "Значение для API",
                        "notes": "Комментарий",
                        "data_type": "Тип Detmir",
                        "is_variant_attribute": "Вариантный",
                    },
                ),
                use_container_width=True,
                hide_index=True,
            )
            blocked_rows = [row for row in readiness_rows if int(row.get("required") or 0) == 1 and str(row.get("status") or "") != "ok"]
            if blocked_rows:
                st.warning(
                    "Есть блокеры под Детский Мир. Сначала добей обязательные поля и проверь справочники, "
                    "потом уже готовь payload/отправку."
                )
        else:
            st.info("Readiness пока нечего показать: сначала назначь Detmir-категорию и подтяни её schema в память.")
        detmir_payload_json = json.dumps(detmir_readiness_details.get("payload") or {}, ensure_ascii=False, indent=2)
        st.code(detmir_payload_json, language="json")
        st.download_button(
            "Скачать Detmir payload JSON",
            data=detmir_payload_json.encode("utf-8"),
            file_name=f"detmir_payload_{_sanitize_filename_part(str(product['article'] or product_id))}.json",
            mime="application/json",
            key=f"detmir_payload_download_{int(product_id)}",
        )

    st.markdown("### Атрибуты для заполнения (Ozon и клиентские шаблоны)")
    channel_codes = list_channel_codes(conn)
    if channel_codes:
        product_ozon_scope = f"ozon:{ozon_desc_id}:{ozon_type_id}" if ozon_desc_id > 0 and ozon_type_id > 0 else ""
        product_detmir_scope = detmir_category_scope
        channel_widget_key = f"card_attr_channel_{int(product_id)}"
        if product_ozon_scope and "ozon" in channel_codes and channel_widget_key not in st.session_state:
            st.session_state[channel_widget_key] = "ozon"
        elif product_detmir_scope and "detmir" in channel_codes and channel_widget_key not in st.session_state:
            st.session_state[channel_widget_key] = "detmir"
        default_channel = str(st.session_state.get(channel_widget_key) or st.session_state.get("card_attr_channel") or "")
        if product_ozon_scope and "ozon" in channel_codes:
            default_channel = "ozon"
        elif product_detmir_scope and "detmir" in channel_codes and not default_channel:
            default_channel = "detmir"
        if default_channel not in channel_codes:
            if product_ozon_scope and "ozon" in channel_codes:
                default_channel = "ozon"
            elif product_detmir_scope and "detmir" in channel_codes:
                default_channel = "detmir"
            elif "onlinetrade" in channel_codes:
                default_channel = "onlinetrade"
            else:
                default_channel = str(channel_codes[0])
        ch1, ch2, ch3 = st.columns([2, 2, 1])
        with ch1:
            selected_channel = st.selectbox(
                "Канал атрибутов",
                options=channel_codes,
                index=channel_codes.index(default_channel),
                key=channel_widget_key,
            )
        st.session_state["card_attr_channel"] = selected_channel

        category_scopes = list_channel_category_codes(conn, selected_channel)
        if selected_channel == "ozon" and product_ozon_scope and product_ozon_scope not in category_scopes:
            category_scopes = [product_ozon_scope] + category_scopes
        if selected_channel == "detmir" and product_detmir_scope and product_detmir_scope not in category_scopes:
            category_scopes = [product_detmir_scope] + category_scopes
        category_scopes = list(dict.fromkeys(category_scopes))
        scope_options = [""] + category_scopes
        if selected_channel == "ozon":
            scope_labels = _build_ozon_scope_labels(conn)
        elif selected_channel == "detmir":
            scope_labels = _build_detmir_scope_labels(conn)
        else:
            scope_labels = {}
        if selected_channel not in {"ozon", "detmir"}:
            for code in category_scopes:
                scope_labels.setdefault(code, str(code))

        default_scope = str(st.session_state.get("card_attr_category_scope") or "")
        if default_scope not in scope_options:
            default_scope = ""
        if selected_channel == "ozon" and product_ozon_scope:
            default_scope = product_ozon_scope
        elif selected_channel == "detmir" and product_detmir_scope:
            default_scope = product_detmir_scope
        elif not default_scope:
            product_scope_candidates = [
                str(product["subcategory"] or "").strip(),
                str(product["category"] or "").strip(),
                str(product["base_category"] or "").strip(),
            ]
            options_lc = {str(opt).strip().lower(): opt for opt in scope_options if str(opt).strip()}
            for cand in product_scope_candidates:
                if cand and cand.lower() in options_lc:
                    default_scope = str(options_lc[cand.lower()])
                    break
        scope_widget_key = f"card_attr_scope_{int(product_id)}_{selected_channel}"
        if selected_channel == "ozon" and product_ozon_scope and scope_widget_key not in st.session_state:
            st.session_state[scope_widget_key] = product_ozon_scope
        if selected_channel == "detmir" and product_detmir_scope and scope_widget_key not in st.session_state:
            st.session_state[scope_widget_key] = product_detmir_scope
        with ch2:
            selected_scope = st.selectbox(
                "Категория атрибутов",
                options=scope_options,
                index=(scope_options.index(default_scope) if default_scope in scope_options else 0),
                format_func=lambda x: "Все категории канала" if x == "" else scope_labels.get(x, str(x)),
                key=scope_widget_key,
            )
        st.session_state["card_attr_category_scope"] = selected_scope
        with ch3:
            save_as_channel = st.checkbox(
                "Сохранять как канал",
                value=False,
                key=f"card_attr_save_as_channel_{int(product_id)}_{selected_channel}_{selected_scope or 'all'}",
                help="Если выключено, значения сохраняются в мастер-карточку.",
            )

        req_rows = list_channel_requirements(
            conn,
            channel_code=selected_channel,
            category_code=selected_scope or None,
        )
        if selected_channel == "ozon" and selected_scope and selected_scope.startswith("ozon:") and not req_rows:
            try:
                _, raw_desc, raw_type = selected_scope.split(":", 2)
                ensure_ozon_requirements_for_product_category(conn, int(raw_desc), int(raw_type))
                req_rows = list_channel_requirements(
                    conn,
                    channel_code=selected_channel,
                    category_code=selected_scope or None,
                )
            except Exception:
                pass
        rule_rows = list_channel_mapping_rules(
            conn,
            channel_code=selected_channel,
            category_code=selected_scope or None,
        )
        required_map = {str(r["attribute_code"]): int(r.get("is_required") or 0) for r in req_rows}
        attribute_codes = set(required_map.keys())
        for rule in rule_rows:
            if str(rule.get("source_type") or "") == "attribute" and rule.get("source_name"):
                attribute_codes.add(str(rule["source_name"]))

        defs = list_attribute_definitions(conn)
        defs_map = {str(d["code"]): d for d in defs}
        type_map_ru = {"text": "Текст", "number": "Число", "boolean": "Да/Нет", "json": "JSON"}

        attr_values = get_product_attribute_values(conn, int(product_id), channel_code=selected_channel)
        value_by_code: dict[str, dict] = {}
        for row in attr_values:
            code = str(row.get("attribute_code") or "")
            if not code:
                continue
            priority = 2 if str(row.get("channel_code") or "").strip() == str(selected_channel).strip() else 1
            existing = value_by_code.get(code)
            if (not existing) or priority > int(existing.get("_priority") or 0):
                value_by_code[code] = {"value": row.get("value"), "_priority": priority}

        editor_rows = []
        for code in sorted(attribute_codes, key=lambda x: (0 if required_map.get(x, 0) else 1, humanize_attribute_code(x).lower())):
            attr_def = defs_map.get(code, {})
            current_value = value_by_code.get(code, {}).get("value")
            current_text = ""
            if current_value is not None:
                if isinstance(current_value, (dict, list)):
                    current_text = json.dumps(current_value, ensure_ascii=False)
                else:
                    current_text = str(current_value)
            editor_rows.append(
                {
                    "attribute_code": code,
                    "attribute_code_ru": humanize_attribute_code(code),
                    "name": str(attr_def.get("name") or humanize_attribute_code(code)),
                    "data_type": type_map_ru.get(str(attr_def.get("data_type") or "text"), str(attr_def.get("data_type") or "text")),
                    "is_required": int(required_map.get(code, 0)),
                    "current_value": current_text,
                    "new_value": current_text,
                }
            )

        if editor_rows:
            ed_df = pd.DataFrame(editor_rows)
            st.caption(
                "Здесь собраны атрибуты категории канала и атрибуты из mapping rules (source_type=attribute). "
                "Для Ozon и Detmir по умолчанию открывается scope текущей категории товара."
            )
            edited_df = st.data_editor(
                ed_df,
                use_container_width=True,
                hide_index=True,
                key=f"card_attr_editor_{int(product_id)}_{selected_channel}_{selected_scope or 'all'}",
                disabled=["attribute_code", "attribute_code_ru", "name", "data_type", "is_required", "current_value"],
                column_order=["name", "data_type", "is_required", "current_value", "new_value"],
                column_config={
                    "attribute_code_ru": None,
                    "attribute_code": None,
                    "name": st.column_config.TextColumn("Название"),
                    "data_type": st.column_config.TextColumn("Тип данных"),
                    "is_required": st.column_config.NumberColumn("Обязательный", format="%d"),
                    "current_value": st.column_config.TextColumn("Текущее значение"),
                    "new_value": st.column_config.TextColumn("Новое значение"),
                },
            )
            if st.button("Сохранить атрибуты этого блока", type="primary", key=f"card_attr_save_btn_{int(product_id)}_{selected_channel}_{selected_scope or 'all'}"):
                updated_count = 0
                for _, row in edited_df.iterrows():
                    code = str(row.get("attribute_code") or "").strip()
                    if not code:
                        continue
                    old_text = str(row.get("current_value") or "").strip()
                    new_text = str(row.get("new_value") or "").strip()
                    if old_text == new_text:
                        continue
                    value_to_save = new_text if new_text else None
                    try:
                        set_product_attribute_value(
                            conn=conn,
                            product_id=int(product_id),
                            attribute_code=code,
                            value=value_to_save,
                            channel_code=(selected_channel if save_as_channel else None),
                        )
                        save_field_source(
                            conn=conn,
                            product_id=int(product_id),
                            field_name=f"attr:{code}",
                            source_type="manual",
                            source_value_raw=value_to_save,
                            source_url=None,
                            confidence=1.0,
                            is_manual=True,
                        )
                        updated_count += 1
                    except Exception as e:
                        st.error(f"Не удалось сохранить атрибут `{code}`: {e}")
                st.success(f"Сохранено атрибутов: {updated_count}")
                st.rerun()
        else:
            st.info("Для выбранного канала и категории пока нет атрибутов. Загрузите клиентский шаблон и сохраните mapping rules.")
    else:
        st.info("Каналы пока не настроены. Добавь канал во вкладке Каналы, затем загрузи клиентский шаблон.")

    prefill_description = st.session_state.pop(f"ai_description_prefill_{int(product_id)}", None)
    description_initial_value = (
        str(prefill_description).strip()
        if prefill_description not in (None, "")
        else (product["description"] or "")
    )

    with card_main_col:
        with st.form("product_form"):
            supplier_options = [""] + sorted(set(supplier_values + ([str(product["supplier_name"])] if product["supplier_name"] else [])))
            supplier_default = str(product["supplier_name"] or "")
            supplier_idx = supplier_options.index(supplier_default) if supplier_default in supplier_options else 0

            fhead1, fhead2 = st.columns([2.2, 1])
            with fhead1:
                st.markdown("### Редактирование карточки")
                st.caption("Главные поля товара и ручная корректировка категории, логистики, фото и описания.")
            with fhead2:
                submitted_top = st.form_submit_button(
                    "Сохранить карточку",
                    type="primary",
                    use_container_width=True,
                    key=f"product_form_submit_top_{int(product_id)}",
                )

            with st.expander("1. Идентификация товара", expanded=True):
                id1, id2 = st.columns(2)
                with id1:
                    article = st.text_input("Артикул", value=product["article"] or "")
                    internal_article = st.text_input("Внутренний артикул", value=product["internal_article"] or "")
                    supplier_article = st.text_input("Артикул поставщика", value=product["supplier_article"] or "")
                    name = st.text_input("Название", value=product["name"] or "")
                with id2:
                    brand = st.text_input("Бренд", value=product["brand"] or "")
                    supplier_name = st.selectbox("Поставщик (из базы)", options=supplier_options, index=supplier_idx)
                    barcode = st.text_input("Штрихкод", value=product["barcode"] or "")
                    barcode_source = st.text_input("Источник штрихкода", value=product["barcode_source"] or "")
                    uom = st.text_input("Ед. изм.", value=product["uom"] or "")

            with st.expander("2. Категория и Ozon-эталон", expanded=True):
                st.caption(
                    "Если Ozon подобрал каталог неточно, эти поля можно скорректировать вручную. "
                    "Ручное значение сохраняется как `manual` и не должно затираться обычной автопривязкой без force."
                )
                if product["ozon_category_path"]:
                    st.caption(f"Ozon path сейчас: {product['ozon_category_path']}")
                cat1, cat2 = st.columns(2)
                with cat1:
                    category = st.text_input(
                        "Категория (ручная корректировка)",
                        value=str(product["category"] or ""),
                        help="Можно ввести своё значение, даже если его ещё нет в списках PIM.",
                    )
                    base_category = st.text_input("Базовая категория (ручная корректировка)", value=str(product["base_category"] or ""))
                    subcategory = st.text_input("Подкатегория (ручная корректировка)", value=str(product["subcategory"] or ""))
                    wheel_diameter_inch = st.number_input("Диаметр колеса, inch", value=float(product["wheel_diameter_inch"] or 0.0), step=0.5)
                    tnved_code = st.text_input("ТН ВЭД", value=product["tnved_code"] or "")
                with cat2:
                    supplier_url = st.text_input("URL поставщика", value=product["supplier_url"] or "")
                    ozon_description_category_id = st.number_input("Ozon description_category_id", min_value=0, value=int(product["ozon_description_category_id"] or 0), step=1)
                    ozon_type_id = st.number_input("Ozon type_id", min_value=0, value=int(product["ozon_type_id"] or 0), step=1)
                    ozon_category_path = st.text_input("Ozon категория (path)", value=product["ozon_category_path"] or "")
                    ozon_category_confidence = st.number_input("Уверенность Ozon категории (0..1)", min_value=0.0, max_value=1.0, value=float(product["ozon_category_confidence"] or 0.0), step=0.01)
                det1, det2 = st.columns(2)
                with det1:
                    detmir_category_id = st.number_input("Detmir category_id", min_value=0, value=int(product["detmir_category_id"] or 0), step=1, help="Можно выбрать вручную, если автоподбор Детского Мира ошибся.")
                    detmir_category_path = st.text_input("Detmir категория (path)", value=str(product["detmir_category_path"] or ""))
                with det2:
                    detmir_category_confidence = st.number_input("Уверенность Detmir категории (0..1)", min_value=0.0, max_value=1.0, value=_clamp_unit_confidence(product["detmir_category_confidence"]), step=0.01)
                    st.caption("Для Детского Мира сначала привяжи категорию, потом импортируй её требования и только после этого добивай gaps / готовь payload.")

            with st.expander("3. Логистика и упаковка", expanded=False):
                lg1, lg2 = st.columns(2)
                with lg1:
                    weight = st.number_input("Вес, кг", value=float(product["weight"] or 0.0), step=0.1)
                    length = st.number_input("Длина, см", value=float(product["length"] or 0.0), step=1.0)
                    width = st.number_input("Ширина, см", value=float(product["width"] or 0.0), step=1.0)
                    height = st.number_input("Высота, см", value=float(product["height"] or 0.0), step=1.0)
                with lg2:
                    package_length = st.number_input("Длина упаковки", value=float(product["package_length"] or 0.0), step=1.0)
                    package_width = st.number_input("Ширина упаковки", value=float(product["package_width"] or 0.0), step=1.0)
                    package_height = st.number_input("Высота упаковки", value=float(product["package_height"] or 0.0), step=1.0)
                    gross_weight = st.number_input("Вес брутто", value=float(product["gross_weight"] or 0.0), step=0.1)

            with st.expander("4. Фото и описание", expanded=True):
                image_url = st.text_input(
                    "Фото (основное)",
                    value=normalize_media_reference(product["image_url"] or "", public_base_url=media_public_base_url) or (product["image_url"] or ""),
                )
                gallery_text = st.text_area(
                    "Галерея фото (по одной ссылке в строке)",
                    value="\n".join(current_gallery_urls),
                    height=130,
                    help="Можно указать 1+ ссылок. Лучше использовать публичные http/https ссылки на .jpg/.jpeg/.png. Если настроен public media URL, локальные пути будут конвертированы в веб-ссылки.",
                )
                description = st.text_area("Описание", value=description_initial_value, height=180)

            submitted_bottom = st.form_submit_button(
                "Сохранить карточку",
                type="primary",
                key=f"product_form_submit_bottom_{int(product_id)}",
            )
            submitted = bool(submitted_top or submitted_bottom)

            if submitted:
                gallery_urls = _parse_gallery_value(gallery_text, public_base_url=media_public_base_url)
                primary_image = normalize_media_reference(str(image_url or "").strip(), public_base_url=media_public_base_url) or str(image_url or "").strip()
                if primary_image:
                    gallery_urls = _normalize_media_urls([primary_image] + gallery_urls)
                elif gallery_urls:
                    primary_image = str(gallery_urls[0]).strip()

                payload = {
                    "article": article or None,
                    "internal_article": internal_article or None,
                    "supplier_article": supplier_article or None,
                    "name": name or None,
                    "brand": brand or None,
                    "supplier_name": supplier_name or None,
                    "barcode": barcode or None,
                    "barcode_source": barcode_source or None,
                    "category": category or None,
                    "base_category": base_category or None,
                    "subcategory": subcategory or None,
                    "wheel_diameter_inch": wheel_diameter_inch or None,
                    "supplier_url": supplier_url or None,
                    "ozon_description_category_id": int(ozon_description_category_id) if int(ozon_description_category_id) > 0 else None,
                    "ozon_type_id": int(ozon_type_id) if int(ozon_type_id) > 0 else None,
                    "ozon_category_path": ozon_category_path or None,
                    "ozon_category_confidence": float(ozon_category_confidence) if float(ozon_category_confidence) > 0 else None,
                    "detmir_category_id": int(detmir_category_id) if int(detmir_category_id) > 0 else None,
                    "detmir_category_path": detmir_category_path or None,
                    "detmir_category_confidence": _clamp_unit_confidence(detmir_category_confidence) if _clamp_unit_confidence(detmir_category_confidence) > 0 else None,
                    "uom": uom or None,
                    "weight": weight or None,
                    "length": length or None,
                    "width": width or None,
                    "height": height or None,
                    "package_length": package_length or None,
                    "package_width": package_width or None,
                    "package_height": package_height or None,
                    "gross_weight": gross_weight or None,
                    "image_url": primary_image or None,
                    "description": description or None,
                    "tnved_code": tnved_code or None,
                }
                save_product(conn, int(product_id), payload)
                set_product_attribute_value(conn, int(product_id), "main_image", primary_image or None)
                set_product_attribute_value(conn, int(product_id), "gallery_images", gallery_urls if gallery_urls else None)
                if primary_image:
                    save_field_source(
                        conn=conn,
                        product_id=int(product_id),
                        field_name="attr:main_image",
                        source_type="manual",
                        source_value_raw=primary_image,
                        source_url=None,
                        confidence=1.0,
                        is_manual=True,
                    )
                save_field_source(
                    conn=conn,
                    product_id=int(product_id),
                    field_name="attr:gallery_images",
                    source_type="manual",
                    source_value_raw=json.dumps(gallery_urls, ensure_ascii=False) if gallery_urls else "[]",
                    source_url=None,
                    confidence=1.0,
                    is_manual=True,
                )
                refresh_duplicates_for_product(conn, int(product_id))
                backup_result = backup_database_file(reason="product_card_manual_save")
                st.success("Сохранено")
                if backup_result.get("ok"):
                    st.caption(f"Карточка зафиксирована в backup: `{Path(str(backup_result['path'])).name}`")
                st.rerun()

    if current_gallery_urls:
        zip_bytes, zip_stats = build_product_images_zip(conn, [int(product_id)], public_base_url=media_public_base_url)
        if int(zip_stats.get("images_written") or 0) > 0:
            st.download_button(
                "Скачать все фото товара ZIP",
                data=zip_bytes,
                file_name=f"{_sanitize_filename_part(str(product['article'] or product['supplier_article'] or product_id))}_images.zip",
                mime="application/zip",
                key=f"product_images_zip_{int(product_id)}",
            )
        else:
            st.caption("Фото у товара есть, но сейчас не удалось собрать ZIP. Проверь, что ссылки публичные или настроен public media URL.")

    with st.expander("Источники и аудит", expanded=False):
        st.caption("Здесь видно, что пришло руками, что от поставщика, а что создано AI или служебными пайплайнами.")
        key_fields = ["name", "brand", "description", "weight", "length", "width", "height", "package_length", "package_width", "package_height", "gross_weight", "image_url"]
        source_summary = []
        for field_name in key_fields:
            src = get_latest_field_source(conn, int(product_id), field_name)
            source_summary.append({
                "field_name": field_name,
                "source_type": src.get("source_type") if src else None,
                "is_manual": bool(src.get("is_manual")) if src else False,
                "confidence": src.get("confidence") if src else None,
                "created_at": src.get("created_at") if src else None,
            })
        st.dataframe(with_ru_columns(pd.DataFrame(source_summary)), use_container_width=True, hide_index=True)

        sources = get_field_sources(conn, int(product_id))
        if sources:
            src_df = pd.DataFrame(sources)
            if not src_df.empty and "field_name" in src_df.columns:
                src_df["field_name_ru"] = src_df["field_name"].map(humanize_attribute_code)
            st.dataframe(
                with_ru_columns(
                    src_df,
                    extra_map={"field_name_ru": "Поле (рус.)"},
                ),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.caption("Источники данных пока не записаны")

    conn.close()


def show_attributes_tab():
    conn = get_db()
    product_id = st.session_state.get("selected_product_id")
    def _reset_attrs_filters():
        st.session_state["attrs_search"] = ""
        st.session_state["attrs_source_filter"] = "Все"
        st.session_state["attrs_category_scope"] = "Все"
        st.session_state["attrs_only_required"] = False
        st.session_state["attrs_rows_limit"] = 300

    st.subheader("Атрибуты")
    st.caption("Рабочее место по атрибутам: фильтруй справочник, проверяй значения у выбранного товара и добавляй новые поля без путаницы.")
    with st.expander("Инструкция по разделу Атрибуты", expanded=False):
        st.markdown(
            """
1. В `Фильтры атрибутов` выбери источник (`Ozon`/`Кастомные`) и при необходимости `Категория Ozon`.
2. Вкладка `Справочник атрибутов` показывает доступные поля по текущим фильтрам.
3. Вкладка `Атрибуты выбранного товара` показывает заполненные значения текущего товара и позволяет быстро изменить поле.
4. Вкладка `Добавить/обновить атрибут` используется, когда появилось новое поле из клиентского шаблона.

`Только обязательные` работает в связке с выбранной Ozon-категорией.
`Лимит строк` нужен для удобной навигации по большим справочникам.
            """
        )

    selected_product = None
    if product_id:
        selected_product = conn.execute(
            """
            SELECT id, article, name, supplier_name, ozon_category_path, ozon_description_category_id, ozon_type_id
            FROM products
            WHERE id = ?
            LIMIT 1
            """,
            (int(product_id),),
        ).fetchone()

    scope_labels = _build_ozon_scope_labels(conn)
    category_scope_options = ["Все"] + sorted(scope_labels.keys())
    if selected_product:
        scoped_code = None
        try:
            desc_id = int(selected_product["ozon_description_category_id"] or 0)
            type_id = int(selected_product["ozon_type_id"] or 0)
            if desc_id > 0 and type_id > 0:
                scoped_code = f"ozon:{desc_id}:{type_id}"
        except Exception:
            scoped_code = None
        marker = st.session_state.get("attrs_scope_product_marker")
        if scoped_code and scoped_code in category_scope_options and marker != int(selected_product["id"]):
            current_scope = st.session_state.get("attrs_category_scope")
            if current_scope in (None, "", "Все"):
                st.session_state["attrs_category_scope"] = scoped_code
        st.session_state["attrs_scope_product_marker"] = int(selected_product["id"])

    current_scope_value = st.session_state.get("attrs_category_scope")
    if current_scope_value not in category_scope_options:
        st.session_state["attrs_category_scope"] = "Все"
        current_scope_value = "Все"
    scope_index = category_scope_options.index(current_scope_value) if current_scope_value in category_scope_options else 0

    with st.container(border=True):
        st.markdown("### Фильтры атрибутов")
        f1, f2, f3, f4, f5, f6 = st.columns([2.2, 1.2, 2.2, 1.2, 1.1, 0.7])
        with f1:
            attr_search = st.text_input(
                "Поиск",
                value=st.session_state.get("attrs_search", ""),
                placeholder="Название или описание атрибута",
                key="attrs_search",
            )
        with f2:
            source_options = ["Все", "Ozon", "Кастомные"]
            source_value = st.session_state.get("attrs_source_filter", "Все")
            source_index = source_options.index(source_value) if source_value in source_options else 0
            attr_source_filter = st.selectbox(
                "Источник",
                options=source_options,
                index=source_index,
                key="attrs_source_filter",
            )
        with f3:
            category_scope = st.selectbox(
                "Категория Ozon (область)",
                options=category_scope_options,
                index=scope_index,
                key="attrs_category_scope",
                format_func=lambda x: "Все" if x == "Все" else scope_labels.get(x, x),
            )
        with f4:
            only_required = st.checkbox(
                "Только обязательные",
                value=bool(st.session_state.get("attrs_only_required", False)),
                key="attrs_only_required",
            )
        with f5:
            limit_options = [100, 300, 1000, 3000]
            limit_value = int(st.session_state.get("attrs_rows_limit", 300))
            limit_index = limit_options.index(limit_value) if limit_value in limit_options else 1
            rows_limit = st.selectbox("Лимит строк", options=limit_options, index=limit_index, key="attrs_rows_limit")
        with f6:
            st.markdown("<br>", unsafe_allow_html=True)
            st.button("Сброс", key="attrs_filters_reset", on_click=_reset_attrs_filters)

    if selected_product:
        st.success(f"Выбран товар: ID {selected_product['id']} | {selected_product['article'] or '-'} | {selected_product['name'] or '-'}")
        st.caption(
            f"Поставщик: {selected_product['supplier_name'] or '-'} | "
            f"Ozon категория: {selected_product['ozon_category_path'] or '-'}"
        )
        try:
            selected_desc_id = int(selected_product["ozon_description_category_id"] or 0)
            selected_type_id = int(selected_product["ozon_type_id"] or 0)
        except Exception:
            selected_desc_id = 0
            selected_type_id = 0
        if selected_desc_id > 0 and selected_type_id > 0:
            materialized = materialize_ozon_attribute_slots_for_product(
                conn,
                product_id=int(selected_product["id"]),
                description_category_id=selected_desc_id,
                type_id=selected_type_id,
            )
            if int(materialized.get("created") or 0) > 0:
                st.info(
                    f"Для выбранного товара автоматически созданы пустые Ozon-атрибуты: "
                    f"{int(materialized.get('created') or 0)}."
                )
    else:
        st.warning("Товар не выбран. Выбери товар во вкладке `Карточка` или `Каталог`, чтобы редактировать значения атрибутов.")

    required_map: dict[str, int] = {}
    selected_scope_cached_total = 0
    selected_scope_cached_required = 0
    if category_scope != "Все":
        req_rows = conn.execute(
            """
            SELECT attribute_code, is_required
            FROM channel_attribute_requirements
            WHERE channel_code = 'ozon'
              AND category_code = ?
            """,
            (str(category_scope),),
        ).fetchall()
        required_map = {str(r["attribute_code"]): int(r["is_required"] or 0) for r in req_rows}
        parts = str(category_scope).split(":")
        if len(parts) == 3 and parts[0] == "ozon":
            try:
                selected_scope_cached_total = int(
                    conn.execute(
                        """
                        SELECT COUNT(*)
                        FROM ozon_attribute_cache
                        WHERE description_category_id = ?
                          AND type_id = ?
                        """,
                        (int(parts[1]), int(parts[2])),
                    ).fetchone()[0]
                    or 0
                )
                selected_scope_cached_required = int(
                    conn.execute(
                        """
                        SELECT COUNT(*)
                        FROM ozon_attribute_cache
                        WHERE description_category_id = ?
                          AND type_id = ?
                          AND is_required = 1
                        """,
                        (int(parts[1]), int(parts[2])),
                    ).fetchone()[0]
                    or 0
                )
            except Exception:
                selected_scope_cached_total = 0
                selected_scope_cached_required = 0

    defs = list_attribute_definitions(conn)
    defs_df = pd.DataFrame(defs) if defs else pd.DataFrame()
    if not defs_df.empty:
        if attr_source_filter == "Ozon":
            defs_df = defs_df[defs_df["code"].astype(str).str.startswith("ozon_attr_")]
        elif attr_source_filter == "Кастомные":
            defs_df = defs_df[~defs_df["code"].astype(str).str.startswith("ozon_attr_")]

        if category_scope != "Все":
            allowed_codes = set(required_map.keys())
            defs_df = defs_df[defs_df["code"].astype(str).isin(allowed_codes)]
            defs_df["is_required_for_category"] = defs_df["code"].map(lambda c: int(required_map.get(str(c), 0)))

        if only_required and category_scope != "Все":
            defs_df = defs_df[defs_df["is_required_for_category"].fillna(0).astype(int) == 1]

        if attr_search:
            q = str(attr_search).strip().lower()
            mask = (
                defs_df["code"].astype(str).str.lower().str.contains(q, na=False)
                | defs_df["name"].astype(str).str.lower().str.contains(q, na=False)
                | defs_df["description"].astype(str).str.lower().str.contains(q, na=False)
            )
            defs_df = defs_df[mask]

        data_type_ru = {"text": "Текст", "number": "Число", "boolean": "Да/Нет", "json": "JSON"}
        scope_ru = {"master": "Мастер", "channel": "Канал"}
        entity_type_ru = {"product": "Товар", "channel": "Канал", "category": "Категория"}
        defs_df = defs_df.copy()
        if "data_type" in defs_df.columns:
            defs_df["data_type"] = defs_df["data_type"].map(lambda x: data_type_ru.get(str(x), x))
        if "scope" in defs_df.columns:
            defs_df["scope"] = defs_df["scope"].map(lambda x: scope_ru.get(str(x), x))
        if "entity_type" in defs_df.columns:
            defs_df["entity_type"] = defs_df["entity_type"].map(lambda x: entity_type_ru.get(str(x), x))

    filtered_defs_count = int(len(defs_df)) if not defs_df.empty else 0
    tabs = st.tabs(["Справочник атрибутов", "Атрибуты выбранного товара", "Добавить/обновить атрибут"])

    with tabs[0]:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Атрибутов по фильтру", filtered_defs_count)
        ozon_count = int(defs_df["code"].astype(str).str.startswith("ozon_attr_").sum()) if not defs_df.empty and "code" in defs_df.columns else 0
        m2.metric("Из них Ozon", ozon_count)
        required_count = int(defs_df["is_required_for_category"].fillna(0).astype(int).sum()) if not defs_df.empty and "is_required_for_category" in defs_df.columns else 0
        m3.metric("Обязательных (scope)", required_count)
        total_requirements = conn.execute(
            "SELECT COUNT(*) FROM channel_attribute_requirements WHERE channel_code = 'ozon'"
        ).fetchone()[0]
        m4.metric("Всего Ozon requirements", int(total_requirements or 0))
        if category_scope != "Все":
            st.caption(
                f"Для выбранной Ozon-пары: в кэше Ozon {int(selected_scope_cached_total)} атрибутов, "
                f"обязательных {int(selected_scope_cached_required)}; "
                f"в PIM requirements {int(len(required_map))}."
            )
            if int(selected_scope_cached_total) > 0 and int(len(required_map)) < int(selected_scope_cached_total):
                st.warning(
                    "Для этой Ozon-категории в PIM пока меньше требований, чем в кэше Ozon. "
                    "Теперь система умеет добирать недостающие атрибуты; запусти `Подтянуть Ozon-атрибуты` для товара или выборки."
                )

        if defs_df.empty:
            st.info("По текущим фильтрам атрибуты не найдены.")
        else:
            defs_view_columns = [
                c
                for c in [
                    "name",
                    "data_type",
                    "scope",
                    "entity_type",
                    "is_required",
                    "is_multi_value",
                    "unit",
                    "description",
                    "is_required_for_category",
                    "updated_at",
                ]
                if c in defs_df.columns
            ]
            limit_n = int(rows_limit or 300)
            defs_to_show = defs_df.head(limit_n)
            st.dataframe(
                with_ru_columns(defs_to_show[defs_view_columns] if defs_view_columns else defs_to_show),
                use_container_width=True,
                hide_index=True,
            )
            if len(defs_df) > limit_n:
                st.caption(f"Показаны первые {limit_n} строк из {len(defs_df)}. Для полного списка увеличь `Лимит строк`.")

            st.download_button(
                "Скачать текущий список атрибутов (Excel)",
                data=export_current_df(with_ru_columns(defs_df[defs_view_columns] if defs_view_columns else defs_df)),
                file_name="attributes_filtered.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="attrs_export_filtered_excel",
            )

    with tabs[1]:
        if not product_id:
            st.info("Нет выбранного товара. Сначала выбери товар, затем вернись в эту вкладку.")
        else:
            st.markdown(f"### Значения атрибутов товара #{int(product_id)}")
            values = get_product_attribute_values(conn, int(product_id))
            values_df = pd.DataFrame(values) if values else pd.DataFrame()
            if not values_df.empty and category_scope != "Все":
                values_df = values_df[values_df["attribute_code"].astype(str).isin(set(required_map.keys()))]

            if not values_df.empty:
                values_df = values_df.copy()
                if "data_type" in values_df.columns:
                    values_df["data_type"] = values_df["data_type"].map(lambda x: {"text": "Текст", "number": "Число", "boolean": "Да/Нет", "json": "JSON"}.get(str(x), x))
                if "scope" in values_df.columns:
                    values_df["scope"] = values_df["scope"].map(lambda x: {"master": "Мастер", "channel": "Канал"}.get(str(x), x))
                value_columns = [
                    c
                    for c in [
                        "name",
                        "value",
                        "value_text",
                        "value_number",
                        "value_boolean",
                        "value_json",
                        "data_type",
                        "scope",
                        "unit",
                        "locale",
                        "channel_code",
                        "updated_at",
                    ]
                    if c in values_df.columns
                ]
                st.dataframe(
                    with_ru_columns(values_df[value_columns], extra_map={"value": "Значение"}),
                    use_container_width=True,
                    hide_index=True,
                )
            else:
                st.caption("По текущему фильтру значения атрибутов у выбранного товара не найдены.")

            editable_defs = list_attribute_definitions(conn)
            if category_scope != "Все":
                allowed_codes = set(required_map.keys())
                editable_defs = [d for d in editable_defs if str(d.get("code")) in allowed_codes]
            def_codes = [d["code"] for d in editable_defs] if editable_defs else []
            def_labels = {
                code: str(next((d.get("name") for d in editable_defs if str(d.get("code")) == str(code)), None) or humanize_attribute_code(code))
                for code in def_codes
            }

            st.markdown("### Быстрое редактирование значения")
            with st.form("set_product_attr"):
                attribute_code = (
                    st.selectbox("Атрибут", def_codes, format_func=lambda x: def_labels.get(x, x))
                    if def_codes
                    else st.text_input("Атрибут")
                )
                value = st.text_input("Значение")
                locale = st.text_input("Локаль (опционально)", value="")
                channel_code = st.text_input("Код канала (опционально)", value="")
                save_attr = st.form_submit_button("Сохранить значение", type="primary")

                if save_attr and attribute_code:
                    set_product_attribute_value(
                        conn=conn,
                        product_id=int(product_id),
                        attribute_code=attribute_code,
                        value=value,
                        locale=locale or None,
                        channel_code=channel_code or None,
                    )
                    save_field_source(
                        conn=conn,
                        product_id=int(product_id),
                        field_name=f"attr:{attribute_code}",
                        source_type="manual",
                        source_value_raw=value or None,
                        source_url=None,
                        confidence=1.0,
                        is_manual=True,
                    )
                    st.success("Значение атрибута сохранено.")
                    st.rerun()

    with tabs[2]:
        st.markdown("### Добавление/обновление атрибута")
        st.caption("Этот блок нужен, когда в клиентском шаблоне появилось новое поле, которого ещё нет в PIM.")
        with st.form("new_attribute_def"):
            code = st.text_input("Код атрибута", placeholder="например: rost_rebenka_ot")
            name = st.text_input("Название атрибута", placeholder="например: Рост ребёнка от")
            r1, r2 = st.columns(2)
            with r1:
                data_type = st.selectbox("Тип данных", ["text", "number", "boolean", "json"])
            with r2:
                scope = st.selectbox(
                    "Область",
                    ["master", "channel"],
                    format_func=lambda x: {"master": "Мастер", "channel": "Канал"}.get(str(x), str(x)),
                )
            unit = st.text_input("Единица измерения (опционально)")
            description = st.text_area("Описание (опционально)", height=90)
            add_def = st.form_submit_button("Добавить / обновить атрибут", type="primary")

            if add_def and code and name:
                upsert_attribute_definition(
                    conn=conn,
                    code=code.strip(),
                    name=name.strip(),
                    data_type=data_type,
                    scope=scope,
                    unit=unit or None,
                    description=description or None,
                )
                st.success("Атрибут сохранён в справочнике.")
                st.rerun()

    conn.close()


def show_template_tab():
    st.subheader("Клиентский шаблон")
    st.caption("Здесь должен быть понятный сценарий: загрузили шаблон, увидели матчинг, поняли дыры, добили данные, скачали готовый файл.")
    with st.expander("Инструкция по кнопкам раздела Клиентский шаблон", expanded=False):
        st.markdown(
            """
- `Сохранить mapping rules`: сохранить карту соответствия колонок.
- `Сохранить профиль шаблона`: запомнить тип шаблона клиента.
- `Добавить несматченные в master-атрибуты`: автоматически создать недостающие атрибуты.
- `Подтвердить значения как client_validated`: отметить значения как проверенные.
- `Скачать заполненный шаблон`: выгрузка результата в формате клиента.
- `Обогатить товар из supplier` (в Gap): быстрый переход к автодозаполнению.
            """
        )
    conn = get_db()
    product_df = load_products(conn, limit=5000)

    client_entries = list_client_channels(conn)
    client_map = {str(item.get("client_code") or "").strip(): item for item in client_entries if str(item.get("client_code") or "").strip()}
    client_codes = sorted(client_map.keys(), key=lambda x: (str(client_map.get(x, {}).get("client_name") or x).lower(), x.lower()))

    def _client_label(code: str) -> str:
        item = client_map.get(str(code or "").strip(), {})
        client_name = str(item.get("client_name") or "").strip()
        if client_name:
            return f"{client_name} ({code})"
        if item.get("is_inferred"):
            return f"{code} [из памяти]"
        return str(code)

    t1, t2 = st.columns([1, 1])
    with t1:
        selected_client_code = str(st.session_state.get("template_client_code") or "").strip()
        default_client_option = selected_client_code if selected_client_code in client_codes else None
        client_selector_options = [None] + client_codes + ["__new__"]
        client_selector = st.selectbox(
            "Клиент / канал",
            options=client_selector_options,
            index=client_selector_options.index(default_client_option) if default_client_option in client_selector_options else 0,
            format_func=lambda value: (
                "-- выбери клиента --"
                if value is None
                else "➕ Создать нового клиента"
                if value == "__new__"
                else _client_label(str(value))
            ),
            key="template_client_selector",
        )

        new_client_name = ""
        if client_selector == "__new__":
            new_client_code = st.text_input("Код нового клиента", value=str(st.session_state.get("template_new_client_code") or "").strip(), key="template_new_client_code")
            new_client_name = st.text_input("Название клиента", value=str(st.session_state.get("template_new_client_name") or "").strip(), key="template_new_client_name")
            if st.button("Сохранить клиента в базу", key="template_save_new_client"):
                normalized_new_code = str(new_client_code or "").strip()
                if not normalized_new_code:
                    st.error("Укажи код клиента.")
                else:
                    upsert_client_channel(
                        conn,
                        client_code=normalized_new_code,
                        client_name=str(new_client_name or "").strip() or None,
                    )
                    st.session_state["template_client_code"] = normalized_new_code
                    st.session_state["template_client_selector"] = normalized_new_code
                    st.success(f"Клиент `{normalized_new_code}` сохранён.")
                    st.rerun()
            channel_code = str(new_client_code or "").strip()
        else:
            channel_code = str(client_selector or "").strip()
            if channel_code:
                st.session_state["template_client_code"] = channel_code
                st.caption(f"Выбран клиент: {_client_label(channel_code)}")

    if channel_code == SPORTMASTER_CHANNEL_CODE:
        with st.expander("Импорт шаблонов Sportmaster в память", expanded=False):
            st.caption(
                f"{SPORTMASTER_CLIENT_NAME} хранит отдельный Excel-шаблон на каждую категорию. "
                "Здесь можно один раз импортировать category-specific файлы в память PIM, "
                "и потом категории, требования и профили будут открываться из базы."
            )
            sportmaster_files = st.file_uploader(
                "Выбери один или несколько шаблонов Sportmaster",
                type=["xlsx", "xls"],
                accept_multiple_files=True,
                key="sportmaster_template_batch",
            )
            if st.button("Импортировать шаблоны Sportmaster", key="sportmaster_template_import_btn", type="primary"):
                if not sportmaster_files:
                    st.error("Сначала выбери хотя бы один файл Sportmaster.")
                else:
                    imported_rows: list[dict[str, object]] = []
                    for upload in sportmaster_files:
                        result = import_sportmaster_template(
                            conn,
                            upload.getvalue(),
                            original_file_name=getattr(upload, "name", None),
                        )
                        imported_rows.append(result)
                    backup_result = backup_database_file(reason="sportmaster_templates")
                    st.success(f"Импортировано шаблонов Sportmaster: {len(imported_rows)}")
                    st.dataframe(
                        pd.DataFrame(imported_rows)[[
                            "attr_class",
                            "attr_class_id",
                            "category_code",
                            "columns_total",
                            "required_total",
                            "profile_id",
                        ]],
                        use_container_width=True,
                        hide_index=True,
                    )
                    if backup_result.get("ok"):
                        st.caption(f"Память Sportmaster зафиксирована: `{Path(str(backup_result['path'])).name}`")
                    st.rerun()

    all_profiles = list_template_profiles(conn, channel_code=None)
    profile_scope_options = ["Текущий канал", "Все каналы"]
    profile_scope_value = st.session_state.get("template_profile_scope", "Текущий канал")
    if profile_scope_value not in profile_scope_options:
        profile_scope_value = "Текущий канал"
    profile_scope = st.selectbox(
        "Показывать профили",
        options=profile_scope_options,
        index=profile_scope_options.index(profile_scope_value),
        key="template_profile_scope",
    )
    existing_profiles = (
        [p for p in all_profiles if str(p.get("channel_code") or "").strip() == str(channel_code or "").strip()]
        if profile_scope == "Текущий канал" and channel_code
        else all_profiles
    )
    category_options, category_labels = _build_ozon_template_category_options(conn, channel_code=channel_code, limit=5000)
    category_select_label = "Категория шаблона/профиля"
    if channel_code == SPORTMASTER_CHANNEL_CODE:
        category_select_label = "Категория шаблона/профиля (Sportmaster)"
    elif channel_code == "ozon":
        category_select_label = "Категория шаблона/профиля (Ozon-каталог)"
    elif channel_code == "detmir":
        category_select_label = "Категория шаблона/профиля (Detmir overlay)"
    with t2:
        category_code = st.selectbox(
            category_select_label,
            options=category_options,
            index=0,
            format_func=lambda x: category_labels.get(str(x), str(x)),
            key="template_category_select",
        )

    recent_template_files = list_uploaded_files(
        conn,
        storage_kind="client_template",
        channel_code=channel_code or None,
        category_code=category_code or None,
        limit=20,
    )
    category_profiles = (
        [p for p in existing_profiles if str(p.get("category_code") or "").strip() == str(category_code or "").strip()]
        if category_code
        else existing_profiles
    )

    p1, p2 = st.columns([1, 1])
    with p1:
        category_suffix = re.sub(r"[^a-z0-9]+", "_", str(category_code or "default").lower()).strip("_") or "default"
        profile_name = st.text_input("Имя профиля шаблона", value=f"{(channel_code or 'client')}_{category_suffix}")
    with p2:
        profile_options = [None] + [p["id"] for p in category_profiles]
        selected_profile_id = st.selectbox(
            "Загрузить сохранённый профиль",
            options=profile_options,
            format_func=lambda x: "-- нет --" if x is None else next(
                (
                    f"{p['profile_name']} | канал={p.get('channel_code') or '-'} | категория={p.get('category_code') or '-'} (#{p['id']})"
                    for p in category_profiles
                    if p["id"] == x
                ),
                str(x),
            ),
        )

    if channel_code == SPORTMASTER_CHANNEL_CODE:
        st.caption("Категория профиля берётся из сохранённого шаблона Sportmaster (`sportmaster:attr_class_id`).")
    elif channel_code == "detmir":
        st.caption("Категория профиля берётся из overlay-категории Detmir (`detmir:category_id`).")
    else:
        st.caption("Категория профиля берётся из Ozon-эталона (`ozon:description_category_id:type_id`).")
    if not client_codes and not channel_code:
        st.info("В базе пока нет клиентов. Создай первого клиента и привяжи к нему шаблон.")
    if not all_profiles:
        st.warning("В текущей БД пока нет сохранённых профилей шаблонов.")
    elif channel_code and not existing_profiles and profile_scope == "Текущий канал":
        channels = sorted(set([str(p.get("channel_code") or "") for p in all_profiles if str(p.get("channel_code") or "").strip()]))
        channels_text = ", ".join(channels[:8]) if channels else "-"
        st.info(
            f"Для канала `{channel_code or '-'}` профилей нет. "
            f"Есть профили в других каналах: {channels_text}. "
            f"Переключи `Показывать профили` на `Все каналы`."
        )
    if not channel_code:
        st.info("Сначала выбери клиента из базы или создай нового. После этого можно выбирать категорию и шаблон.")
        conn.close()
        return

    saved_template_options = [None] + [int(row["id"]) for row in recent_template_files]
    selected_saved_template_id = st.selectbox(
        "Открыть сохранённый Excel-шаблон",
        options=saved_template_options,
        format_func=lambda value: "-- нет --" if value is None else next(
            (
                f"{row.get('original_file_name') or Path(str(row.get('stored_rel_path') or '')).name} | "
                f"{row.get('category_code') or '-'} | #{row['id']}"
                for row in recent_template_files
                if int(row["id"]) == int(value)
            ),
            str(value),
        ),
        key="template_saved_file_id",
    )

    uploaded = st.file_uploader("Загрузить новый Excel-шаблон клиента", type=["xlsx", "xls"], key="client_template")
    active_template_bytes = None
    active_template_file_name = None
    saved_template_row = next((row for row in recent_template_files if int(row["id"]) == int(selected_saved_template_id)), None) if selected_saved_template_id else None
    saved_template_metadata = get_uploaded_file_metadata(saved_template_row)
    if uploaded is not None:
        active_template_bytes = uploaded.getvalue()
        active_template_file_name = getattr(uploaded, "name", None)
    elif selected_saved_template_id:
        active_template_bytes = read_uploaded_file_bytes(conn, int(selected_saved_template_id))
        if saved_template_row:
            active_template_file_name = (
                str(saved_template_row.get("original_file_name") or "").strip()
                or Path(str(saved_template_row.get("stored_rel_path") or "")).name
            )
        else:
            active_template_file_name = "client_template.xlsx"
        if active_template_bytes:
            st.success(f"Используется сохранённый шаблон: `{active_template_file_name}`.")
        else:
            st.error("Не удалось прочитать сохранённый шаблон из памяти. Выбери другой или загрузи файл заново.")

    if active_template_bytes is None:
        st.info("Выбери сохранённый шаблон или загрузи новый Excel-файл клиента.")
    if recent_template_files:
        with st.expander("Недавно загруженные шаблоны клиента", expanded=False):
            recent_template_df = pd.DataFrame(recent_template_files)
            recent_template_cols = [
                c
                for c in [
                    "created_at",
                    "original_file_name",
                    "channel_code",
                    "category_code",
                    "stored_rel_path",
                    "file_hash",
                ]
                if c in recent_template_df.columns
            ]
            st.dataframe(with_ru_columns(recent_template_df[recent_template_cols] if recent_template_cols else recent_template_df), use_container_width=True, hide_index=True)

    if active_template_bytes is not None:
        safe_uploaded_bytes = sanitize_template_xlsx_bytes(active_template_bytes)
        workbook = load_workbook(BytesIO(safe_uploaded_bytes), read_only=True, data_only=False)
        template_sheet_options = workbook.sheetnames
        workbook.close()

        default_template_sheet_name = str(saved_template_metadata.get("sheet_name") or "").strip() if saved_template_metadata else ""
        if default_template_sheet_name not in template_sheet_options:
            default_template_sheet_name = ""
        template_sheet_name = st.selectbox(
            "Лист шаблона",
            options=template_sheet_options,
            index=(
                template_sheet_options.index(default_template_sheet_name)
                if default_template_sheet_name and default_template_sheet_name in template_sheet_options
                else
                template_sheet_options.index("Товары")
                if "Товары" in template_sheet_options
                else template_sheet_options.index("Шаблон для заполнения")
                if "Шаблон для заполнения" in template_sheet_options
                else 0
            ),
            key="template_sheet_name",
        )
        suggested_data_start_row = int(saved_template_metadata.get("data_start_row") or 0) if saved_template_metadata else 0
        if suggested_data_start_row <= 0:
            suggested_data_start_row = detect_template_data_start_row(safe_uploaded_bytes, sheet_name=template_sheet_name)

        tcfg1, tcfg2 = st.columns(2)
        with tcfg1:
            template_data_start_row = st.number_input(
                "Строка начала данных",
                min_value=2,
                max_value=100000,
                value=int(suggested_data_start_row),
                step=1,
                key="template_data_start_row",
            )
        with tcfg2:
            preserve_template_workbook = st.checkbox(
                "Сохранять исходную структуру Excel",
                value=True,
                help="Оставляет исходные листы, шапку, справочники и форматирование клиентского файла.",
                key="preserve_template_workbook",
            )

        template_df = read_client_template_dataframe(safe_uploaded_bytes, sheet_name=template_sheet_name)
        template_signature = hashlib.md5(safe_uploaded_bytes).hexdigest()
        sportmaster_metadata: dict[str, object] = {}
        if channel_code == SPORTMASTER_CHANNEL_CODE:
            try:
                sportmaster_metadata = extract_sportmaster_template_metadata(safe_uploaded_bytes)
            except Exception:
                sportmaster_metadata = saved_template_metadata if isinstance(saved_template_metadata, dict) else {}
        required_by_column: dict[str, int] = {}
        mapping_by_column: dict[str, dict[str, object]] = {}
        if isinstance(sportmaster_metadata, dict) and sportmaster_metadata.get("template_kind") == "sportmaster":
            for spec in sportmaster_metadata.get("columns") or []:
                column_name = str((spec or {}).get("header") or "").strip()
                if not column_name:
                    continue
                required_by_column[column_name] = int((spec or {}).get("required") or 0)
            auto_rows = auto_match_template_columns(conn, list(template_df.columns))
            auto_rows = apply_saved_mapping_rules(
                conn,
                auto_rows,
                channel_code=channel_code,
                category_code=category_code or None,
            )
            for row in auto_rows:
                column_name = str(row.get("template_column") or "").strip()
                if not column_name:
                    continue
                mapping_by_column[column_name] = {
                    "source_type": row.get("source_type") or "attribute",
                    "source_name": row.get("source_name") or to_attribute_code(column_name),
                    "transform_rule": row.get("transform_rule"),
                    "notes": "Автодобавлено из шаблона Sportmaster",
                }
        template_upload_registry_key = f"{channel_code}|{category_code}|{template_sheet_name}|{template_signature}|upload_saved"
        if uploaded is not None and st.session_state.get(template_upload_registry_key) != True:
            if channel_code:
                upsert_client_channel(
                    conn,
                    client_code=channel_code,
                    client_name=str(client_map.get(channel_code, {}).get("client_name") or "").strip() or None,
                )
            persist_uploaded_file(
                conn=conn,
                storage_kind="client_template",
                original_file_name=active_template_file_name,
                file_bytes=safe_uploaded_bytes,
                channel_code=channel_code or None,
                category_code=category_code or None,
                metadata={
                    "sheet_name": template_sheet_name,
                    "data_start_row": int(template_data_start_row),
                    "template_signature": template_signature,
                    **(sportmaster_metadata if isinstance(sportmaster_metadata, dict) else {}),
                },
            )
            st.session_state[template_upload_registry_key] = True
        autoreg_key = f"{channel_code}|{category_code}|{template_sheet_name}|{template_signature}"
        if st.session_state.get("template_autoreg_key") != autoreg_key:
            reg = ensure_template_columns_registered(
                conn=conn,
                channel_code=channel_code,
                category_code=category_code or None,
                template_columns=list(template_df.columns),
                required_by_column=required_by_column or None,
                mapping_by_column=mapping_by_column or None,
            )
            st.session_state["template_autoreg_key"] = autoreg_key
            if (reg["attributes"] + reg["requirements"] + reg["rules"]) > 0:
                st.success(
                    f"Шаблон зарегистрирован: атрибутов {reg['attributes']}, "
                    f"требований {reg['requirements']}, правил {reg['rules']}."
                )
            else:
                st.caption("Атрибуты и требования этого шаблона уже были зарегистрированы ранее.")

        defs = list_attribute_definitions(conn)
        attr_name_map = {str(d["code"]): str(d.get("name") or humanize_attribute_code(d["code"])) for d in defs}
        source_options = [("column", c) for c in [
            "article", "internal_article", "supplier_article", "name", "barcode", "brand", "description",
            "weight", "length", "width", "height", "package_length", "package_width", "package_height",
            "gross_weight", "image_url", "ozon_category_path", "ozon_description_category_id", "ozon_type_id",
            "category", "base_category", "supplier_name", "supplier_url",
            "uom", "tnved_code", "media_gallery"
        ]] + [("attribute", d["code"]) for d in defs]
        matches = auto_match_template_columns(conn, list(template_df.columns))
        matches = apply_saved_mapping_rules(conn, matches, channel_code=channel_code, category_code=category_code or None)
        if selected_profile_id:
            profile_columns = get_template_profile_columns(conn, int(selected_profile_id))
            profile_map = {c["template_column"]: c for c in profile_columns}
            matches = [
                {
                    "template_column": m["template_column"],
                    "status": "matched" if profile_map.get(m["template_column"], {}).get("source_name") else m["status"],
                    "source_type": profile_map.get(m["template_column"], {}).get("source_type", m["source_type"]),
                    "source_name": profile_map.get(m["template_column"], {}).get("source_name", m["source_name"]),
                    "matched_by": "template_profile" if profile_map.get(m["template_column"]) else m["matched_by"],
                    "transform_rule": profile_map.get(m["template_column"], {}).get("transform_rule"),
                }
                for m in matches
            ]

        # Применяем отложенные overrides до отрисовки tmpl_* виджетов,
        # чтобы не ловить StreamlitAPIException при записи в ключ уже созданного виджета.
        pending_manual_overrides = st.session_state.pop("template_manual_overrides", None)
        if isinstance(pending_manual_overrides, dict):
            for raw_idx, payload in pending_manual_overrides.items():
                try:
                    idx = int(raw_idx)
                except Exception:
                    continue
                if idx < 0 or idx >= len(matches):
                    continue
                source_type = str(payload.get("source_type") or "attribute")
                source_name = str(payload.get("source_name") or "")
                transform_rule = str(payload.get("transform_rule") or "")
                st.session_state[f"tmpl_type_{idx}"] = source_type
                st.session_state[f"tmpl_name_{idx}"] = source_name
                if transform_rule:
                    st.session_state[f"tmpl_transform_{idx}"] = transform_rule

        match_df = pd.DataFrame(matches)
        match_df_view = match_df.copy()
        if not match_df_view.empty and "source_name" in match_df_view.columns:
            match_df_view["source_name"] = match_df_view.apply(
                lambda r: format_source_name_ui(
                    r.get("source_name"),
                    source_type=r.get("source_type"),
                    attr_name_map=attr_name_map,
                ),
                axis=1,
            )
        matched_count = int((match_df["status"] == "matched").sum()) if not match_df.empty else 0
        unmatched_count = int((match_df["status"] != "matched").sum()) if not match_df.empty else 0

        st.markdown("### Сводка по шаблону")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Колонок в шаблоне", len(template_df.columns))
        c2.metric("Сматчено", matched_count)
        c3.metric("Не сматчено", unmatched_count)
        c4.metric("Профилей найдено", len(existing_profiles))

        if unmatched_count == 0:
            st.success("По матчингу всё хорошо, можно переходить к товарам и preview.")
        else:
            st.warning("Есть несматченные колонки. Лучше сначала добить их, чтобы потом не ловить пустоты в выгрузке.")

        save_ready_rows = [
            {
                "template_column": m.get("template_column"),
                "status": m.get("status"),
                "source_type": m.get("source_type"),
                "source_name": m.get("source_name"),
                "matched_by": m.get("matched_by"),
                "transform_rule": m.get("transform_rule"),
            }
            for m in matches
            if str(m.get("template_column") or "").strip()
        ]
        save_col1, save_col2 = st.columns([1, 3])
        with save_col1:
            if st.button("Сохранить профиль шаблона (текущая схема)", key="template_save_profile_top", type="primary"):
                if channel_code:
                    upsert_client_channel(
                        conn,
                        client_code=channel_code,
                        client_name=str(client_map.get(channel_code, {}).get("client_name") or new_client_name or "").strip() or None,
                    )
                profile_id = save_template_profile(
                    conn=conn,
                    profile_name=profile_name,
                    channel_code=channel_code,
                    category_code=category_code or None,
                    file_name=active_template_file_name,
                    columns=save_ready_rows,
                )
                st.success(f"Профиль шаблона сохранён: #{profile_id}")
                backup_result = backup_database_file(reason="template_profile")
                if backup_result.get("ok"):
                    st.caption(f"Профиль и память шаблона зафиксированы: `{Path(str(backup_result['path'])).name}`")
        with save_col2:
            st.caption("Эта кнопка сохраняет тип шаблона клиента для повторного использования без повторной ручной настройки.")

        def _persist_template_mapping_rules(rows: list[dict]) -> int:
            saved = 0
            for row in rows:
                row_status = str(row.get("status") or "matched")
                if row_status != "matched":
                    continue
                target_field = str(row.get("template_column") or "").strip()
                source_type = str(row.get("source_type") or "").strip()
                source_name = str(row.get("source_name") or "").strip()
                if not target_field or not source_name or source_type not in {"attribute", "column"}:
                    continue
                upsert_channel_mapping_rule(
                    conn=conn,
                    channel_code=channel_code,
                    category_code=category_code or None,
                    target_field=target_field,
                    source_type=source_type,
                    source_name=source_name,
                    transform_rule=row.get("transform_rule"),
                    is_required=0,
                )
                saved += 1
            return saved

        def _autosave_template_mapping_rules(rows: list[dict]) -> int:
            normalized_rows = []
            for row in rows:
                if str(row.get("status") or "") != "matched":
                    continue
                source_type = str(row.get("source_type") or "").strip()
                source_name = str(row.get("source_name") or "").strip()
                template_column = str(row.get("template_column") or "").strip()
                if source_type not in {"attribute", "column"} or not source_name or not template_column:
                    continue
                normalized_rows.append(
                    {
                        "template_column": template_column,
                        "source_type": source_type,
                        "source_name": source_name,
                        "transform_rule": str(row.get("transform_rule") or ""),
                    }
                )
            signature_payload = {
                "channel_code": str(channel_code or "").strip(),
                "category_code": str(category_code or "").strip(),
                "rows": sorted(normalized_rows, key=lambda x: x["template_column"]),
            }
            signature = hashlib.md5(
                json.dumps(signature_payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
            ).hexdigest()
            signature_key = (
                f"template_mapping_signature::{str(channel_code or '').strip()}::{str(category_code or '').strip()}"
            )
            if st.session_state.get(signature_key) == signature:
                return 0
            saved = _persist_template_mapping_rules(normalized_rows)
            st.session_state[signature_key] = signature
            return saved

        current_manual_rows = []
        for idx, match in enumerate(matches):
            source_type = st.session_state.get(f"tmpl_type_{idx}", match.get("source_type") if match.get("source_type") in ["attribute", "column"] else "skip")
            source_name = st.session_state.get(f"tmpl_name_{idx}", match.get("source_name"))
            transform_rule = st.session_state.get(f"tmpl_transform_{idx}", match.get("transform_rule") or "")
            current_manual_rows.append({
                "template_column": match["template_column"],
                "status": "matched" if source_type != "skip" else "unmatched",
                "source_type": None if source_type == "skip" else source_type,
                "source_name": None if source_type == "skip" else source_name,
                "matched_by": "manual" if source_type != "skip" else None,
                "transform_rule": transform_rule or None,
            })

        product_label_map = {
            int(row["id"]): (
                f"{str(row.get('article') or row.get('supplier_article') or row.get('internal_article') or '-')} | "
                f"{_short_text(row.get('name'), 64)} | "
                f"{str(row.get('supplier_name') or '-')}"
            )
            for _, row in product_df.iterrows()
        } if not product_df.empty else {}
        template_catalog_shortlist = [
            int(x)
            for x in (st.session_state.get("template_selected_ids_from_catalog") or [])
            if int(x) in product_label_map
        ]
        existing_template_selected = [
            int(x)
            for x in (st.session_state.get("template_selected_ids") or [])
            if int(x) in product_label_map
        ]

        with st.container(border=True):
            st.markdown("### Пачка товаров для шаблона")
            st.caption("Рабочий режим: подтянули shortlist из Каталога, увидели готовность выбранной пачки и только потом выгрузили Excel.")
            ts1, ts2, ts3 = st.columns([1, 1, 3])
            with ts1:
                if st.button("Подтянуть shortlist из Каталога", key="template_pull_catalog_shortlist"):
                    st.session_state["template_selected_ids"] = template_catalog_shortlist
                    st.rerun()
            with ts2:
                if st.button("Очистить выбор шаблона", key="template_clear_selected_ids"):
                    st.session_state["template_selected_ids"] = []
                    st.rerun()
            with ts3:
                st.caption(
                    f"Shortlist из Каталога: {len(template_catalog_shortlist)} | "
                    f"Сейчас выбрано для шаблона: {len(existing_template_selected)}"
                )

            selected_ids = st.multiselect(
                "Товары для preview / readiness / экспорта",
                options=list(product_label_map.keys()),
                default=existing_template_selected,
                format_func=lambda x: product_label_map.get(int(x), f"ID {x}"),
                key="template_selected_ids",
                help="Это главный список товаров для клиентской выгрузки. Его можно подтянуть из Каталога или собрать вручную.",
            )

        filled_df = fill_template_dataframe(conn, template_df, selected_ids, current_manual_rows) if selected_ids else pd.DataFrame()
        batch_readiness = analyze_template_readiness(filled_df, current_manual_rows) if selected_ids else None
        if batch_readiness:
            summary = batch_readiness["summary"]
            rs1, rs2, rs3, rs4, rs5 = st.columns(5)
            rs1.metric("Выбрано товаров", int(len(selected_ids)))
            rs2.metric("Средняя готовность", f"{int(summary.get('avg_readiness') or 0)}%")
            rs3.metric("Готовых строк", int(summary.get("ready_rows") or 0))
            rs4.metric("Частично готовы", int(summary.get("partial_rows") or 0))
            rs5.metric("Блокеры", int(summary.get("blocked_rows") or 0))
            if int(summary.get("blocked_rows") or 0) > 0:
                st.warning("В пачке есть блокирующие строки. Лучше добить gap перед финальной выгрузкой, чтобы не получать сюрпризы в клиентском файле.")

        tab_match, tab_fill, tab_gap = st.tabs(["1. Матчинг", "2. Заполнение и preview", "3. Gap и действия"])

        with tab_match:
            st.markdown("### Колонки шаблона")
            st.dataframe(pd.DataFrame({"template_column": list(template_df.columns)}), use_container_width=True, hide_index=True)

            st.markdown("### Автоматический матчинг")
            st.dataframe(match_df_view, use_container_width=True, hide_index=True)
            st.caption("Единицы измерения конвертируются автоматически по заголовку колонки (например, см→мм, кг→г). Поле `Transform` можно вручную переопределить.")

            st.markdown("### Ручная правка матчинга")
            manual_rows = []
            for idx, match in enumerate(matches):
                c1, c2, c3, c4 = st.columns([2, 2, 2, 2])
                with c1:
                    st.text_input("Колонка", value=match["template_column"], disabled=True, key=f"tmpl_col_{idx}")
                with c2:
                    source_type = st.selectbox(
                        "Тип источника",
                        options=["attribute", "column", "skip"],
                        index=( ["attribute", "column", "skip"].index(match["source_type"]) if match["source_type"] in ["attribute", "column"] else 2 ),
                        key=f"tmpl_type_{idx}",
                    )
                with c3:
                    allowed_names = [name for stype, name in source_options if stype == source_type] if source_type != "skip" else [""]
                    current_name = match["source_name"] if match["source_name"] in allowed_names else (allowed_names[0] if allowed_names else "")
                    source_name = st.selectbox(
                        "Источник",
                        options=allowed_names,
                        index=(allowed_names.index(current_name) if current_name in allowed_names else 0),
                        key=f"tmpl_name_{idx}",
                        format_func=lambda x, stype=source_type: format_source_name_ui(
                            x,
                            source_type=stype,
                            attr_name_map=attr_name_map,
                        ),
                    ) if allowed_names else st.text_input("Источник", value="", key=f"tmpl_name_{idx}")
                with c4:
                    current_transform = match.get("transform_rule") if match.get("transform_rule") in TEMPLATE_TRANSFORM_OPTIONS else ""
                    transform_rule = st.selectbox(
                        "Transform",
                        options=TEMPLATE_TRANSFORM_OPTIONS,
                        index=TEMPLATE_TRANSFORM_OPTIONS.index(current_transform),
                        key=f"tmpl_transform_{idx}",
                    )
                manual_rows.append({
                    "template_column": match["template_column"],
                    "status": "matched" if source_type != "skip" else "unmatched",
                    "source_type": None if source_type == "skip" else source_type,
                    "source_name": None if source_type == "skip" else source_name,
                    "matched_by": "manual" if source_type != "skip" else None,
                    "transform_rule": transform_rule or None,
                })

            manual_df = pd.DataFrame(manual_rows)
            unmatched = manual_df[manual_df["status"] == "unmatched"] if not manual_df.empty else pd.DataFrame()

            s1, s2, s3 = st.columns(3)
            with s1:
                if st.button("Сохранить mapping rules", type="primary"):
                    saved = _persist_template_mapping_rules(manual_rows)
                    st.success(f"Сохранено mapping rules: {saved}")
                    if saved:
                        backup_result = backup_database_file(reason="template_mapping_rules")
                        if backup_result.get("ok"):
                            st.caption(f"Mapping rules зафиксированы: `{Path(str(backup_result['path'])).name}`")
            with s2:
                if st.button("Сохранить профиль шаблона"):
                    if channel_code:
                        upsert_client_channel(
                            conn,
                            client_code=channel_code,
                            client_name=str(client_map.get(channel_code, {}).get("client_name") or new_client_name or "").strip() or None,
                        )
                    profile_id = save_template_profile(
                        conn=conn,
                        profile_name=profile_name,
                        channel_code=channel_code,
                        category_code=category_code or None,
                        file_name=active_template_file_name,
                        columns=manual_rows,
                    )
                    st.success(f"Профиль шаблона сохранён: #{profile_id}")
                    backup_result = backup_database_file(reason="template_profile")
                    if backup_result.get("ok"):
                        st.caption(f"Профиль и память шаблона зафиксированы: `{Path(str(backup_result['path'])).name}`")
            with s3:
                if st.button("Добавить несматченные в master-атрибуты"):
                    created = 0
                    overrides: dict[str, dict[str, str]] = {}
                    for idx, row in manual_df.iterrows():
                        if row["status"] == "matched":
                            continue
                        col_name = str(row["template_column"])
                        code = to_attribute_code(col_name)
                        if not code:
                            continue
                        upsert_attribute_definition(
                            conn=conn,
                            code=code,
                            name=col_name.strip(),
                            data_type="text",
                            scope="master",
                            unit=None,
                            description=f"Автосоздано из клиентского шаблона: {col_name}",
                        )
                        overrides[str(int(idx))] = {
                            "source_type": "attribute",
                            "source_name": code,
                        }
                        created += 1
                    if overrides:
                        st.session_state["template_manual_overrides"] = overrides
                    st.success(f"Создано/обновлено master-атрибутов: {created}. Маппинг предзаполнен автоматически.")
                    st.rerun()

            if not unmatched.empty:
                st.warning(f"Не сматчено колонок: {len(unmatched)}")
                st.dataframe(unmatched[["template_column", "status"]], use_container_width=True, hide_index=True)
            else:
                st.success("Все колонки шаблона сматчены.")

        with tab_fill:
            manual_rows = []
            for idx, match in enumerate(matches):
                source_type = st.session_state.get(f"tmpl_type_{idx}", match.get("source_type") if match.get("source_type") in ["attribute", "column"] else "skip")
                source_name = st.session_state.get(f"tmpl_name_{idx}", match.get("source_name"))
                transform_rule = st.session_state.get(f"tmpl_transform_{idx}", match.get("transform_rule") or "")
                manual_rows.append({
                    "template_column": match["template_column"],
                    "status": "matched" if source_type != "skip" else "unmatched",
                    "source_type": None if source_type == "skip" else source_type,
                    "source_name": None if source_type == "skip" else source_name,
                    "matched_by": "manual" if source_type != "skip" else None,
                    "transform_rule": transform_rule or None,
                })

            if product_df.empty:
                st.info("В каталоге пока нет товаров для заполнения шаблона.")
            elif not selected_ids:
                st.info("Выбери товары в блоке `Пачка товаров для шаблона`, и я покажу preview, готовность и выгрузку.")
            else:
                st.markdown("### Предпросмотр заполнения")
                preview_limit = min(50, len(filled_df)) if not filled_df.empty else 0
                if not filled_df.empty:
                    st.dataframe(filled_df.head(preview_limit), use_container_width=True, hide_index=True)
                    if len(filled_df) > preview_limit:
                        st.caption(f"Показаны первые {preview_limit} строк из {len(filled_df)} выбранных товаров.")
                render_template_readiness(filled_df, manual_rows)

                a1, a2 = st.columns(2)
                with a1:
                    if st.button("Подтвердить значения как client_validated"):
                        autosaved = _autosave_template_mapping_rules(manual_rows)
                        result = apply_client_validated_values(conn, selected_ids, manual_rows, channel_code=channel_code or None)
                        st.success(
                            f"Применено: {result['applied']}, пропущено по приоритету: {result['skipped']}. "
                            f"Автосохранено mapping rules: {autosaved}."
                        )
                        if int(result.get("applied") or 0) > 0:
                            backup_result = backup_database_file(reason="template_client_validated")
                            if backup_result.get("ok"):
                                st.caption(f"Пачка client_validated зафиксирована: `{Path(str(backup_result['path'])).name}`")
                with a2:
                    _autosave_template_mapping_rules(manual_rows)
                    export_bytes = fill_template_workbook_bytes(
                        conn,
                        safe_uploaded_bytes,
                        selected_ids,
                        manual_rows,
                        sheet_name=template_sheet_name,
                        data_start_row=int(template_data_start_row),
                    ) if preserve_template_workbook else dataframe_to_excel_bytes(filled_df, sheet_name=template_sheet_name)
                    st.download_button(
                        "Скачать заполненный шаблон",
                        data=export_bytes,
                        file_name=f"filled_{Path(str(active_template_file_name or 'client_template.xlsx')).name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        with tab_gap:
            manual_rows = []
            for idx, match in enumerate(matches):
                source_type = st.session_state.get(f"tmpl_type_{idx}", match.get("source_type") if match.get("source_type") in ["attribute", "column"] else "skip")
                source_name = st.session_state.get(f"tmpl_name_{idx}", match.get("source_name"))
                transform_rule = st.session_state.get(f"tmpl_transform_{idx}", match.get("transform_rule") or "")
                manual_rows.append({
                    "template_column": match["template_column"],
                    "status": "matched" if source_type != "skip" else "unmatched",
                    "source_type": None if source_type == "skip" else source_type,
                    "source_name": None if source_type == "skip" else source_name,
                    "matched_by": "manual" if source_type != "skip" else None,
                    "transform_rule": transform_rule or None,
                })

            if product_df.empty:
                st.info("Сначала нужны товары в каталоге.")
            elif not selected_ids:
                st.info("Сначала собери пачку товаров в блоке `Пачка товаров для шаблона`.")
            else:
                filled_df = fill_template_dataframe(conn, template_df, selected_ids, manual_rows)
                manual_df = pd.DataFrame(manual_rows)
                gap_rows = []
                gap_actions = []
                for _, row in manual_df.iterrows():
                    if row["status"] != "matched":
                        gap_rows.append({"template_column": row["template_column"], "reason": "Нет матчинга"})
                        continue
                    if filled_df[row["template_column"]].isna().all():
                        gap_rows.append({"template_column": row["template_column"], "reason": "У выбранных товаров нет данных"})
                        for product_id in selected_ids:
                            product_row = get_product(conn, int(product_id))
                            can_supplier = bool(product_row and product_row["supplier_url"])
                            gap_actions.append({
                                "product_id": product_id,
                                "product_name": product_row["name"] if product_row else None,
                                "template_column": row["template_column"],
                                "can_supplier_enrich": can_supplier,
                            })

                if gap_rows:
                    st.markdown("### Gap-анализ")
                    st.dataframe(pd.DataFrame(gap_rows), use_container_width=True, hide_index=True)
                else:
                    st.success("Критичных gap по текущему выбору не найдено.")

                if gap_actions:
                    st.markdown("### Быстрые действия")
                    action_df = pd.DataFrame(gap_actions)
                    st.dataframe(action_df, use_container_width=True, hide_index=True)

                    action_product_id = st.selectbox(
                        "Выбери товар для быстрого действия",
                        options=sorted(set([x["product_id"] for x in gap_actions])),
                        format_func=lambda x: product_label_map.get(int(x), f"ID {x}"),
                    )
                    a1, a2 = st.columns(2)
                    with a1:
                        if st.button("Обогатить товар из supplier", key="gap_supplier_enrich"):
                            result = enrich_product_from_supplier(
                                conn,
                                int(action_product_id),
                                force=False,
                                parser_settings=load_parser_settings(conn),
                            )
                            if result["ok"]:
                                st.success(result["message"])
                                st.rerun()
                            else:
                                st.error(result["message"])
                    with a2:
                        if st.button("Открыть товар в карточке", key="gap_open_product"):
                            st.session_state["selected_product_id"] = int(action_product_id)
                            request_workspace_navigation("product")

    conn.close()


def show_ozon_tab():
    conn = get_db()
    st.subheader("Ozon")
    st.caption(
        "Ozon для нас одновременно эталон структуры и клиент-канал. "
        "Здесь синхронизируем schema-ядро, собираем publish-пачки и отправляем карточки в Ozon API."
    )
    saved_client_id, saved_api_key = load_saved_ozon_credentials(conn)
    if "ozon_client_id_input" not in st.session_state:
        st.session_state["ozon_client_id_input"] = saved_client_id or ""
    if "ozon_api_key_input" not in st.session_state:
        st.session_state["ozon_api_key_input"] = saved_api_key or ""
    with st.expander("Инструкция по разделу Ozon и расшифровка кнопок", expanded=False):
        st.markdown(
            """
**Порядок работы (рекомендуемый)**
1. `Сохранить Ozon как клиента` и `Проверить подключение`
2. `Синхронизировать дерево категорий Ozon`
3. `Запустить полную синхронизацию Ozon в фоне`
4. Проверить блок покрытия синхронизации (`Пары для проверки`, `Пары с атрибутами`, `Пропущено пар`, `%`)
5. Если есть пропуски: `Досинхронизировать пропущенные категории`
6. При необходимости импортировать в мастер: `Импортировать все атрибуты Ozon из кэша в PIM`
7. В блоке publish собрать пачку и `Отправить карточки в Ozon API`

**Кнопки верхнего блока**
- `Сохранить Ozon как клиента`: сохраняет `Client ID / API Key` в постоянную память PIM, чтобы не вставлять их заново.
- `Проверить подключение`: делает быстрый запрос к Ozon API и подтверждает, что креды рабочие.
- `Синхронизировать дерево категорий Ozon`: обновляет локальный кэш дерева категорий Ozon.
- `Запустить полную синхронизацию Ozon в фоне`: фоном проходит по категориям и подтягивает атрибуты, не блокируя UI.
- `Импортировать все атрибуты Ozon из кэша в PIM`: переносит уже загруженные атрибуты из кэша в master-слой PIM.
- `Досинхронизировать пропущенные категории`: подтягивает только те пары `cat/type`, где в кэше ещё нет атрибутов.

**Кнопки по выбранной Ozon-категории**
- `Синхронизировать атрибуты выбранной категории`: точечная синхронизация атрибутов одной пары `cat/type`.
- `Импортировать атрибуты Ozon в PIM`: перенос атрибутов выбранной пары в `attribute_definitions` и requirements.
- `Создать стартовые mapping rules для Ozon`: создаёт стартовые правила маппинга для выбранной категории.
- `Синхронизировать все справочники категории`: подтягивает dictionary-значения по всем dictionary-атрибутам категории.
- `Синхронизировать значения справочника`: подтягивает dictionary-значения только для выбранного атрибута.

**Кнопки массовой работы по товарам**
- `Загрузить список из Excel`: выбирает товары из Excel для массовых действий.
- `Проверить покрытие товара под Ozon`: отчёт по одному товару (готовность required-атрибутов).
- `Заполнить Ozon-атрибуты из мастер-карточки`: автозаполнение Ozon-атрибутов по одному товару.
- `Массовая проверка готовности по выбранным товарам`: отчёт готовности по группе товаров.
- `Массово заполнить Ozon-атрибуты для выбранных`: автозаполнение по группе товаров.
- `Сформировать dictionary gaps по выбранным (Excel)`: выгрузка проблем словарного сопоставления.
- `Отправить карточки в Ozon API`: отправка подготовленного batch в Ozon API.

**Кнопки dictionary overrides**
- `Сохранить dictionary override`: сохранить ручное правило raw -> dictionary value.
- `Импортировать overrides из Excel`: массовая загрузка overrides из файла.
- `Удалить выбранный override`: удалить сохранённый override.

**Кнопки по jobs (журнал отправок)**
- `Массово повторить jobs из Excel`: повторная отправка job_id из Excel.
- `Повторить все jobs из фильтра`: повтор всех jobs текущего фильтра статуса.
- `Повторить отправку job`: повтор одного выбранного job.
            """
        )

    c1, c2 = st.columns(2)
    with c1:
        client_id = st.text_input("Client ID Ozon", key="ozon_client_id_input")
    with c2:
        api_key = st.text_input("API Key Ozon", key="ozon_api_key_input", type="password")

    resolved_client_id, resolved_api_key = resolve_ozon_credentials(conn, client_id or None, api_key or None)
    configured = is_configured(resolved_client_id or None, resolved_api_key or None)
    cc1, cc2, cc3, cc4 = st.columns([1, 1, 1, 2])
    with cc1:
        if st.button("Сохранить Ozon как клиента", key="ozon_save_client_settings_btn"):
            if not (str(client_id or "").strip() and str(api_key or "").strip()):
                st.warning("Сначала заполни Client ID и API Key Ozon.")
            else:
                save_ozon_credentials(conn, client_id, api_key)
                st.success("Ozon-креды сохранены в постоянную память PIM.")
    with cc2:
        if st.button("Проверить подключение", key="ozon_check_connection_btn", disabled=not configured):
            check_result = check_ozon_connection(conn, client_id=client_id or None, api_key=api_key or None)
            if check_result.get("ok"):
                st.success(str(check_result.get("message") or "Подключение к Ozon API подтверждено."))
            else:
                st.error(str(check_result.get("message") or "Не удалось подключиться к Ozon API."))
    with cc3:
        if st.button("Очистить сохранённые ключи", key="ozon_clear_saved_client_settings_btn"):
            clear_saved_ozon_credentials(conn)
            st.session_state["ozon_client_id_input"] = ""
            st.session_state["ozon_api_key_input"] = ""
            st.success("Сохранённые Ozon-креды очищены из памяти PIM.")
            st.rerun()
    with cc4:
        if saved_client_id and saved_api_key:
            st.caption("Ozon сохранён в памяти PIM и может использоваться как клиент без повторного ввода ключей.")
        else:
            st.caption("Сохрани Ozon в памяти PIM, чтобы отправлять карточки и синхронизировать schema без ручного ввода ключей.")
    if configured:
        st.success("Ozon-креды доступны. Можно синхронизировать schema-ядро и отправлять карточки как в клиентский канал.")
    else:
        st.warning("Ozon-креды не заданы. Можно вставить их вручную или сохранить Ozon как клиента в памяти PIM.")

    with st.expander("Фиксация кэша Ozon (backup / restore)", expanded=False):
        st.caption(
            "Используй этот блок, чтобы не запускать полный sync заново после обновления приложения. "
            "Сначала сохрани snapshot кэша Ozon в Excel, потом при необходимости восстанови его."
        )
        snap_c1, snap_c2, snap_c3 = st.columns([1, 1, 2])
        with snap_c1:
            snapshot_include_values = st.checkbox(
                "Включать значения справочников",
                value=False,
                key="ozon_snapshot_include_values",
                help="Лист `ozon_attribute_value_cache` может быть большим. Включай только если он нужен для dictionary-сопоставления.",
            )
        with snap_c2:
            if st.button("Подготовить snapshot Ozon", key="ozon_prepare_snapshot_btn"):
                with st.spinner("Формирую snapshot кэша Ozon..."):
                    st.session_state["ozon_snapshot_bytes"] = build_ozon_cache_snapshot_excel(
                        conn,
                        include_value_cache=bool(snapshot_include_values),
                    )
                    st.session_state["ozon_snapshot_meta"] = {
                        "generated_at": _now_iso(),
                        "include_values": bool(snapshot_include_values),
                    }
                    st.session_state["ozon_snapshot_backup_result"] = backup_ozon_snapshot_bytes(
                        st.session_state["ozon_snapshot_bytes"],
                        include_value_cache=bool(snapshot_include_values),
                        source="ui_snapshot",
                    )
                st.success("Snapshot подготовлен. Ниже появилась кнопка скачивания.")
        with snap_c3:
            snapshot_bytes = st.session_state.get("ozon_snapshot_bytes")
            snapshot_meta = st.session_state.get("ozon_snapshot_meta") or {}
            if snapshot_bytes:
                meta_text = (
                    f"Snapshot готов: {snapshot_meta.get('generated_at') or '-'} | "
                    f"справочники: {'да' if snapshot_meta.get('include_values') else 'нет'}"
                )
                st.caption(meta_text)
                snapshot_backup_result = st.session_state.get("ozon_snapshot_backup_result") or {}
                if snapshot_backup_result.get("ok"):
                    st.caption(f"Копия snapshot уже сохранена в памяти сервиса: `{Path(str(snapshot_backup_result['path'])).name}`")
                st.download_button(
                    "Скачать snapshot кэша Ozon (Excel)",
                    data=snapshot_bytes,
                    file_name=f"ozon_cache_snapshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="ozon_download_snapshot_btn",
                )
            else:
                st.caption("Сначала нажми `Подготовить snapshot Ozon`.")

        restore_file = st.file_uploader(
            "Восстановить snapshot Ozon из Excel",
            type=["xlsx"],
            key="ozon_restore_snapshot_file",
        )
        if st.button("Восстановить кэш Ozon из snapshot", key="ozon_restore_snapshot_btn"):
            if restore_file is None:
                st.warning("Сначала загрузи snapshot Excel файл.")
            else:
                with st.spinner("Восстанавливаю кэш Ozon из snapshot..."):
                    restore_result = restore_ozon_cache_snapshot_excel(conn, restore_file.getvalue())
                if bool(restore_result.get("ok")):
                    msg = (
                        "Кэш Ozon восстановлен: "
                        f"категорий {int(restore_result.get('categories') or 0)}, "
                        f"атрибутов {int(restore_result.get('attributes') or 0)}"
                    )
                    if bool(restore_result.get("has_values")):
                        msg += f", значений справочников {int(restore_result.get('values') or 0)}"
                    st.success(msg)
                    backup_result = backup_database_file(reason="ozon_restore_snapshot")
                    if backup_result.get("ok"):
                        st.caption(f"Состояние Ozon-кэша зафиксировано в backup БД: `{Path(str(backup_result['path'])).name}`")
                    st.info("Далее нажми `Импортировать все атрибуты Ozon из кэша в PIM`, чтобы восстановить master-атрибуты.")
                    st.rerun()
                else:
                    st.error(str(restore_result.get("message") or "Не удалось восстановить snapshot Ozon."))

        recent_snapshots = list_ozon_snapshot_backups(limit=12)
        if recent_snapshots:
            st.markdown("#### Последние snapshot-копии Ozon в памяти сервиса")
            snapshot_options = [None] + [row["file_path"] for row in recent_snapshots]
            selected_snapshot_path = st.selectbox(
                "Открыть сохранённый snapshot сервиса",
                options=snapshot_options,
                format_func=lambda value: "-- нет --" if value is None else next(
                    (
                        f"{row['file_name']} | {row.get('created_at') or '-'} | "
                        f"values={'да' if row.get('include_value_cache') else 'нет'}"
                        for row in recent_snapshots
                        if row["file_path"] == value
                    ),
                    str(value),
                ),
                key="ozon_saved_snapshot_path",
            )
            if selected_snapshot_path:
                selected_snapshot_row = next((row for row in recent_snapshots if row["file_path"] == selected_snapshot_path), None)
                selected_snapshot_bytes = read_backup_bytes(selected_snapshot_path)
                ss1, ss2 = st.columns([1, 1])
                with ss1:
                    if selected_snapshot_bytes:
                        st.download_button(
                            "Скачать snapshot из памяти сервиса",
                            data=selected_snapshot_bytes,
                            file_name=str(selected_snapshot_row.get("file_name") if selected_snapshot_row else Path(str(selected_snapshot_path)).name),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="ozon_saved_snapshot_download_btn",
                        )
                with ss2:
                    if st.button("Восстановить выбранный snapshot из памяти сервиса", key="ozon_restore_saved_snapshot_btn"):
                        if not selected_snapshot_bytes:
                            st.error("Не удалось прочитать snapshot из памяти сервиса.")
                        else:
                            with st.spinner("Восстанавливаю snapshot Ozon из серверной памяти..."):
                                restore_result = restore_ozon_cache_snapshot_excel(conn, selected_snapshot_bytes)
                            if bool(restore_result.get("ok")):
                                st.success(
                                    "Snapshot из памяти сервиса восстановлен: "
                                    f"категорий {int(restore_result.get('categories') or 0)}, "
                                    f"атрибутов {int(restore_result.get('attributes') or 0)}."
                                )
                                backup_result = backup_database_file(reason="ozon_restore_saved_snapshot")
                                if backup_result.get("ok"):
                                    st.caption(f"Состояние БД после restore сохранено: `{Path(str(backup_result['path'])).name}`")
                                st.rerun()
                            else:
                                st.error(str(restore_result.get("message") or "Не удалось восстановить snapshot из памяти сервиса."))

    top1, top2, top3, top4 = st.columns(4)
    with top1:
        if st.button("Синхронизировать дерево категорий Ozon", type="primary", disabled=not configured, help="Обновить локальный кэш дерева категорий Ozon"):
            result = sync_category_tree(conn, client_id=resolved_client_id or None, api_key=resolved_api_key or None)
            st.success(f"Дерево категорий обновлено, записей: {result['total']}")
            backup_result = backup_database_file(reason="ozon_category_tree_sync")
            if backup_result.get("ok"):
                st.caption(f"Кэш категорий сохранён в backup БД: `{Path(str(backup_result['path'])).name}`")
            st.rerun()
    with top2:
        if st.button("Запустить полную синхронизацию Ozon в фоне", disabled=not configured, help="Фоновая загрузка атрибутов по категориям Ozon"):
            ok, message = _start_ozon_bg_sync(client_id=resolved_client_id or "", api_key=resolved_api_key or "")
            if ok:
                st.success(message)
            else:
                st.info(message)
    with top3:
        category_limit = st.number_input("Сколько категорий показать", min_value=100, max_value=10000, value=2000, step=100)
    with top4:
        if st.button("Импортировать все атрибуты Ozon из кэша в PIM", help="Перенести атрибуты из ozon_attribute_cache в master-атрибуты PIM"):
            result = import_all_cached_attributes_to_pim(conn)
            st.success(
                "Массовый импорт завершён: "
                f"пар обработано {int(result.get('pairs_processed') or 0)} из {int(result.get('pairs_total') or 0)}, "
                f"атрибутов импортировано {int(result.get('imported_total') or 0)}."
            )
            if result.get("errors"):
                st.warning(f"Ошибок при массовом импорте: {len(result['errors'])}.")
            backup_result = backup_database_file(reason="ozon_import_all_cached_attributes")
            if backup_result.get("ok"):
                st.caption(f"Master-атрибуты после импорта сохранены: `{Path(str(backup_result['path'])).name}`")
            st.rerun()

    bg_state = _get_ozon_bg_state()
    if bg_state.get("running"):
        st.info(
            "Фоновая синхронизация Ozon выполняется. "
            f"Старт: {bg_state.get('started_at') or '-'}."
        )
    elif bg_state.get("last_error"):
        st.error(f"Фоновая синхронизация Ozon завершилась с ошибкой: {bg_state.get('last_error')}")
    elif bg_state.get("result"):
        r = bg_state.get("result") or {}
        st.success(
            "Фоновая синхронизация Ozon завершена: "
            f"пар обработано {int(r.get('pairs_processed') or 0)} из {int(r.get('pairs_total') or 0)}, "
            f"атрибутов загружено {int(r.get('attributes_total') or 0)}, "
            f"импортировано в PIM {int(r.get('imported_to_pim') or 0)}."
        )
        if r.get("errors"):
            st.warning(f"Ошибок в фоновой синхронизации: {len(r['errors'])}.")
        bg_result_signature = str(bg_state.get("finished_at") or bg_state.get("started_at") or "")
        if bg_result_signature and st.session_state.get("ozon_bg_result_backup_signature") != bg_result_signature:
            backup_result = backup_database_file(reason="ozon_background_full_sync")
            if backup_result.get("ok"):
                st.caption(f"Результат фоновой синхронизации зафиксирован: `{Path(str(backup_result['path'])).name}`")
            st.session_state["ozon_bg_result_backup_signature"] = bg_result_signature
    st.caption("Полная синхронизация Ozon теперь запускается в фоне и не блокирует работу с остальными разделами.")

    stats = get_ozon_cache_stats(conn)
    s1, s2, s3, s4, s5, s6 = st.columns(6)
    s1.metric("Узлов категорий", int(stats.get("category_nodes") or 0))
    s2.metric("Уникальных пар категорий", int(stats.get("category_pairs") or 0))
    s3.metric("Атрибутов в кэше", int(stats.get("attributes_total") or 0))
    s4.metric("Обязательных", int(stats.get("attributes_required") or 0))
    s5.metric("Атрибутов в мастере", int(stats.get("attribute_defs_ozon") or 0))
    s6.metric("Требований Ozon (категорийных)", int(stats.get("ozon_requirements") or 0))
    if int(stats.get("category_pairs") or 0) == 0 and int(stats.get("attribute_pairs") or 0) > 0:
        st.warning(
            f"В кэше категорий пока 0 пар cat/type, но в кэше атрибутов уже есть {int(stats.get('attribute_pairs') or 0)} пар. "
            "Сначала нажми `Синхронизировать дерево категорий Ozon`, затем запусти полную синхронизацию."
        )

    qc1, qc2, qc3, qc4 = st.columns([1, 1, 1, 2])
    with qc1:
        coverage_only_leaf = st.checkbox(
            "Проверять только листовые категории",
            value=True,
            key="ozon_coverage_only_leaf",
        )
    with qc2:
        coverage_include_disabled = st.checkbox(
            "Включая отключённые",
            value=False,
            key="ozon_coverage_include_disabled",
        )
    with qc3:
        missing_sync_limit = st.number_input(
            "Лимит досинхронизации пропусков",
            min_value=10,
            max_value=5000,
            value=500,
            step=10,
            key="ozon_missing_sync_limit",
        )
    with qc4:
        if st.button("Досинхронизировать пропущенные категории", disabled=not configured, help="Обработать только пары cat/type без атрибутов в кэше"):
            miss_result = sync_missing_category_attributes(
                conn,
                client_id=resolved_client_id or None,
                api_key=resolved_api_key or None,
                only_leaf=bool(coverage_only_leaf),
                include_disabled=bool(coverage_include_disabled),
                limit=int(missing_sync_limit),
                import_to_pim=True,
            )
            st.success(
                "Досинхронизация завершена: "
                f"обработано пар {int(miss_result.get('pairs_processed') or 0)} из {int(miss_result.get('missing_pairs_requested') or 0)}, "
                f"атрибутов загружено {int(miss_result.get('attributes_total') or 0)}."
            )
            if miss_result.get("errors"):
                st.warning(f"Ошибок при досинхронизации: {len(miss_result['errors'])}.")
            backup_result = backup_database_file(reason="ozon_missing_categories_sync")
            if backup_result.get("ok"):
                st.caption(f"Досинхронизированный Ozon-кэш сохранён: `{Path(str(backup_result['path'])).name}`")
            st.rerun()

    coverage = get_ozon_sync_coverage(
        conn,
        only_leaf=bool(coverage_only_leaf),
        include_disabled=bool(coverage_include_disabled),
        missing_preview_limit=200,
    )
    cv1, cv2, cv3, cv4 = st.columns(4)
    cv1.metric("Пары для проверки", int(coverage.get("total_pairs") or 0))
    cv2.metric("Пары с атрибутами", int(coverage.get("pairs_with_attrs") or 0))
    cv3.metric("Пропущено пар", int(coverage.get("missing_pairs") or 0))
    cv4.metric("Покрытие синхронизации, %", float(coverage.get("coverage_percent") or 0.0))
    if int(coverage.get("missing_pairs") or 0) > 0:
        st.warning(
            f"Синхронизация Ozon покрыта не полностью: {int(coverage.get('pairs_with_attrs') or 0)} из {int(coverage.get('total_pairs') or 0)} пар. "
            "Ниже показан список первых пропусков."
        )
        missing_preview = coverage.get("missing_preview") or []
        if missing_preview:
            st.dataframe(with_ru_columns(pd.DataFrame(missing_preview)), use_container_width=True, hide_index=True)
    else:
        st.success("Покрытие синхронизации Ozon полное для выбранных условий проверки.")

    category_search = st.text_input(
        "Фильтр категорий Ozon",
        value="",
        placeholder="Например: велосипед, аксессуары, запчасти",
        key="ozon_category_search",
    )
    category_pairs = list_cached_category_pairs(conn, search=category_search or None, limit=int(category_limit))
    categories = list_cached_categories(conn, limit=min(1000, int(category_limit)))
    if not categories:
        st.warning("Кэш категорий Ozon пуст. Сначала запусти синхронизацию дерева категорий.")
    if categories:
        cat_df = pd.DataFrame(categories)
        st.markdown("### Кэш категорий Ozon")
        st.dataframe(with_ru_columns(cat_df[[c for c in ["description_category_id", "category_name", "full_path", "type_id", "type_name", "disabled", "fetched_at"] if c in cat_df.columns]]), use_container_width=True, hide_index=True)

        if category_pairs:
            pairs_df = pd.DataFrame(category_pairs)
            st.markdown("### Уникальные пары категорий Ozon (cat/type)")
            st.dataframe(
                with_ru_columns(pairs_df[[c for c in ["description_category_id", "type_id", "full_path", "type_name", "disabled", "nodes", "fetched_at"] if c in pairs_df.columns]]),
                use_container_width=True,
                hide_index=True,
            )

        valid_rows = [row for row in category_pairs if row.get("description_category_id") and row.get("type_id")]
        if valid_rows:
            category_options = [f"{row['full_path']} | cat={row['description_category_id']} | type={row['type_id']}" for row in valid_rows]
            selected_category_label = st.selectbox("Категория Ozon для загрузки атрибутов", options=category_options)
            selected_row = valid_rows[category_options.index(selected_category_label)]

            a1, a2 = st.columns(2)
            with a1:
                if st.button("Синхронизировать атрибуты выбранной категории", disabled=not configured):
                    result = sync_category_attributes(
                        conn,
                        description_category_id=int(selected_row["description_category_id"]),
                        type_id=int(selected_row["type_id"]),
                        client_id=resolved_client_id or None,
                        api_key=resolved_api_key or None,
                    )
                    st.success(f"Атрибуты обновлены: всего {result['total']}, обязательных {result['required']}")
                    backup_result = backup_database_file(reason="ozon_selected_category_sync")
                    if backup_result.get("ok"):
                        st.caption(f"Категория и её атрибуты сохранены в backup БД: `{Path(str(backup_result['path'])).name}`")
                    st.rerun()
            with a2:
                attr_limit = st.number_input("Сколько атрибутов показать", min_value=50, max_value=2000, value=300, step=50)

            attributes = list_cached_attributes(
                conn,
                description_category_id=int(selected_row["description_category_id"]),
                type_id=int(selected_row["type_id"]),
                limit=int(attr_limit),
            )
            if attributes:
                attr_df = pd.DataFrame(attributes)
                required_count = int(attr_df["is_required"].fillna(0).astype(int).sum()) if "is_required" in attr_df.columns else 0
                master_seed = ensure_ozon_master_attributes(conn)
                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Атрибутов в кэше", int(len(attr_df)))
                m2.metric("Обязательных", required_count)
                m3.metric("Справочники", int((attr_df["dictionary_id"].fillna(0).astype(float) > 0).sum()) if "dictionary_id" in attr_df.columns else 0)
                m4.metric("Базовых мастер-атрибутов", int(master_seed["total"]))

                a1, a2 = st.columns(2)
                with a1:
                    if st.button("Импортировать атрибуты Ozon в PIM"):
                        result = import_cached_attributes_to_pim(
                            conn,
                            description_category_id=int(selected_row["description_category_id"]),
                            type_id=int(selected_row["type_id"]),
                        )
                        st.success(f"В PIM импортировано {result['imported']} атрибутов, обязательных {result['required']}. category_code={result['category_code']}")
                with a2:
                    if st.button("Создать стартовые mapping rules для Ozon"):
                        result = save_suggested_mappings(
                            conn,
                            description_category_id=int(selected_row["description_category_id"]),
                            type_id=int(selected_row["type_id"]),
                        )
                        st.success(f"Сохранено стартовых mapping rules: {result['saved']}. category_code={result['category_code']}")

                mapping_df = pd.DataFrame(
                    suggest_mappings_for_cached_attributes(
                        conn,
                        description_category_id=int(selected_row["description_category_id"]),
                        type_id=int(selected_row["type_id"]),
                    )
                )
                if not mapping_df.empty:
                    defs_for_mapping = list_attribute_definitions(conn)
                    mapping_attr_name_map = {
                        str(d["code"]): str(d.get("name") or humanize_attribute_code(d["code"]))
                        for d in defs_for_mapping
                    } if defs_for_mapping else {}
                    if "source_name" in mapping_df.columns:
                        mapping_df["source_name"] = mapping_df.apply(
                            lambda r: format_source_name_ui(
                                r.get("source_name"),
                                source_type=r.get("source_type"),
                                attr_name_map=mapping_attr_name_map,
                            ),
                            axis=1,
                        )
                    mm1, mm2, mm3 = st.columns(3)
                    mm1.metric("Matched по эвристике/правилам", int((mapping_df["status"] == "matched").sum()))
                    mm2.metric("Без маппинга", int((mapping_df["status"] != "matched").sum()))
                    mm3.metric("Обязательных без маппинга", int(((mapping_df["is_required"] == 1) & (mapping_df["status"] != "matched")).sum()))

                    st.markdown("### Предлагаемые Ozon mapping rules")
                    st.dataframe(
                        with_ru_columns(mapping_df[[c for c in ["attribute_id", "name", "group_name", "is_required", "source_type", "source_name", "transform_rule", "matched_by", "status"] if c in mapping_df.columns]]),
                        use_container_width=True,
                        hide_index=True,
                    )

                st.markdown("### Атрибуты выбранной категории")
                attr_show = attr_df[[c for c in ["attribute_id", "name", "group_name", "type", "dictionary_id", "is_required", "is_collection", "max_value_count", "fetched_at"] if c in attr_df.columns]].copy()
                if "attribute_id" in attr_show.columns:
                    attr_show["attribute_code_ru"] = attr_show["attribute_id"].map(lambda x: f"Ozon атрибут ID {int(x)}" if pd.notna(x) else "")
                st.dataframe(
                    with_ru_columns(attr_show, extra_map={"attribute_code_ru": "Код атрибута (рус.)", "type": "Тип"}),
                    use_container_width=True,
                    hide_index=True,
                )

                dictionary_attrs = [row for row in attributes if int(row.get("dictionary_id") or 0) > 0]
                if dictionary_attrs:
                    st.markdown("### Справочники значений Ozon")
                    dd1, dd2, dd3 = st.columns(3)
                    dd1.metric("Атрибутов-справочников", int(len(dictionary_attrs)))
                    cached_dict_attr_count = conn.execute(
                        "SELECT COUNT(DISTINCT attribute_id) FROM ozon_attribute_value_cache WHERE description_category_id = ? AND type_id = ?",
                        (int(selected_row["description_category_id"]), int(selected_row["type_id"])),
                    ).fetchone()[0]
                    cached_dict_value_count = conn.execute(
                        "SELECT COUNT(*) FROM ozon_attribute_value_cache WHERE description_category_id = ? AND type_id = ?",
                        (int(selected_row["description_category_id"]), int(selected_row["type_id"])),
                    ).fetchone()[0]
                    dd2.metric("Справочников в кэше", int(cached_dict_attr_count or 0))
                    dd3.metric("Значений в кэше", int(cached_dict_value_count or 0))

                    if st.button("Синхронизировать все справочники категории", disabled=not configured):
                        result = sync_all_category_dictionary_values(
                            conn,
                            description_category_id=int(selected_row["description_category_id"]),
                            type_id=int(selected_row["type_id"]),
                            client_id=resolved_client_id or None,
                            api_key=resolved_api_key or None,
                        )
                        st.success(f"Синхронизировано справочников: {result['synced_attributes']}, значений: {result['synced_values']}")

                    dict_options = [f"{row['name']} | attr={row['attribute_id']} | dict={row['dictionary_id']}" for row in dictionary_attrs]
                    selected_dict_label = st.selectbox("Атрибут-справочник", options=dict_options, key="ozon_dict_attr")
                    selected_dict_row = dictionary_attrs[dict_options.index(selected_dict_label)]
                    d1, d2 = st.columns(2)
                    with d1:
                        if st.button("Синхронизировать значения справочника", disabled=not configured):
                            result = sync_attribute_dictionary_values(
                                conn,
                                description_category_id=int(selected_row["description_category_id"]),
                                type_id=int(selected_row["type_id"]),
                                attribute_id=int(selected_dict_row["attribute_id"]),
                                client_id=resolved_client_id or None,
                                api_key=resolved_api_key or None,
                            )
                            st.success(f"Значения справочника обновлены: {result['inserted']} | attr={result['attribute_id']} | dict={result['dictionary_id']}")
                    with d2:
                        dict_limit = st.number_input("Сколько значений справочника показать", min_value=50, max_value=5000, value=200, step=50, key="ozon_dict_limit")
                    dict_search = st.text_input("Фильтр по значению справочника", value="", key="ozon_dict_search")
                    dict_values = list_cached_attribute_values(
                        conn,
                        description_category_id=int(selected_row["description_category_id"]),
                        type_id=int(selected_row["type_id"]),
                        attribute_id=int(selected_dict_row["attribute_id"]),
                        search=dict_search or None,
                        limit=int(dict_limit),
                    )
                    if dict_values:
                        dict_df = pd.DataFrame(dict_values)
                        st.dataframe(with_ru_columns(dict_df[[c for c in ["value_id", "value", "info", "picture", "fetched_at"] if c in dict_df.columns]]), use_container_width=True, hide_index=True)
                    else:
                        st.caption("Значения этого справочника ещё не загружены в кэш.")

                product_rows = conn.execute(
                    "SELECT id, name, article, internal_article, supplier_article FROM products ORDER BY id DESC LIMIT 5000"
                ).fetchall()
                if product_rows:
                    product_options = [int(r["id"]) for r in product_rows]
                    selected_product_id = st.selectbox(
                        "Проверить покрытие конкретного товара под выбранную Ozon-категорию",
                        options=product_options,
                        format_func=lambda x: next((f"ID {r['id']} | {r['article'] or '-'} | {r['name'] or '-'}" for r in product_rows if int(r['id']) == int(x)), str(x)),
                        key="ozon_coverage_product_id",
                    )
                    dictionary_min_score = st.slider(
                        "Порог dictionary matching (чем выше, тем строже)",
                        min_value=0.50,
                        max_value=0.99,
                        value=0.78,
                        step=0.01,
                        key=f"ozon_dict_min_score_{selected_product_id}",
                    )
                    st.markdown("### Excel: список товаров для массовых действий")
                    excel_col1, excel_col2, excel_col3 = st.columns([1, 2, 1])
                    with excel_col1:
                        excel_lookup_field = st.selectbox(
                            "Поле поиска в Excel",
                            options=["id", "article", "internal_article", "supplier_article"],
                            index=1,
                            key=f"ozon_excel_lookup_{selected_product_id}",
                        )
                        excel_sheet_name = st.text_input(
                            "Лист Excel (опционально)",
                            value="",
                            key=f"ozon_excel_sheet_{selected_product_id}",
                            placeholder="Например: products",
                        )
                        excel_column_name = st.text_input(
                            "Колонка Excel (опционально)",
                            value="",
                            key=f"ozon_excel_column_{selected_product_id}",
                            placeholder="Например: article",
                        )
                    with excel_col2:
                        excel_file = st.file_uploader(
                            "Загрузи Excel со списком товаров",
                            type=["xlsx", "xls"],
                            key=f"ozon_excel_file_{selected_product_id}",
                        )
                    bulk_select_key = f"ozon_bulk_product_ids_{selected_product_id}"
                    if bulk_select_key not in st.session_state:
                        st.session_state[bulk_select_key] = [int(selected_product_id)]
                    catalog_shortlist_for_ozon = [
                        int(x)
                        for x in (st.session_state.get("template_selected_ids_from_catalog") or [])
                        if str(x).strip()
                    ]
                    with excel_col3:
                        st.download_button(
                            "Скачать шаблон Excel",
                            data=build_ozon_product_list_template_excel(),
                            file_name="ozon_products_list_template.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"ozon_excel_template_{selected_product_id}",
                        )
                        if st.button("Загрузить список из Excel", key=f"ozon_excel_apply_{selected_product_id}"):
                            if excel_file is None:
                                st.warning("Сначала загрузи Excel файл.")
                            else:
                                parsed = resolve_product_ids_from_excel(
                                    conn,
                                    excel_file.read(),
                                    excel_lookup_field,
                                    sheet_name=excel_sheet_name or None,
                                    column_name=excel_column_name or None,
                                )
                                st.session_state[f"ozon_excel_parse_{selected_product_id}"] = parsed
                                if parsed.get("ok"):
                                    st.session_state[bulk_select_key] = parsed.get("resolved_ids") or [int(selected_product_id)]
                                    st.success(
                                        f"Excel обработан: найдено {parsed.get('resolved_count', 0)} из {parsed.get('input_values', 0)} значений."
                                    )
                                    st.rerun()
                                else:
                                    st.error(parsed.get("message") or "Не удалось обработать Excel.")

                    parse_summary = st.session_state.get(f"ozon_excel_parse_{selected_product_id}")
                    shortlist_col1, shortlist_col2 = st.columns([1, 2])
                    with shortlist_col1:
                        if st.button("Подтянуть shortlist из Каталога", key=f"ozon_pull_catalog_shortlist_{selected_product_id}"):
                            if not catalog_shortlist_for_ozon:
                                st.warning("В каталоге пока нет shortlist для Ozon-пачки.")
                            else:
                                st.session_state[bulk_select_key] = [int(x) for x in catalog_shortlist_for_ozon]
                                st.success(f"В Ozon-пачку подтянуто товаров из Каталога: {len(catalog_shortlist_for_ozon)}.")
                                st.rerun()
                    with shortlist_col2:
                        if catalog_shortlist_for_ozon:
                            st.caption(f"Shortlist из Каталога доступен для Ozon publish: {len(catalog_shortlist_for_ozon)} товаров.")
                        else:
                            st.caption("Если уже выбрал пачку в Каталоге, здесь можно одним нажатием подтянуть её в Ozon publish-flow.")
                    if parse_summary and parse_summary.get("ok"):
                        s1, s2, s3 = st.columns(3)
                        s1.metric("Входных значений", int(parse_summary.get("input_values") or 0))
                        s2.metric("Найдено товаров", int(parse_summary.get("resolved_count") or 0))
                        s3.metric("Не найдено", int(parse_summary.get("not_found_count") or 0))
                        st.caption(
                            f"Использована колонка: {parse_summary.get('used_column')} | Поле lookup: {parse_summary.get('lookup_field')}"
                        )
                        not_found = parse_summary.get("not_found") or []
                        if not_found:
                            not_found_df = pd.DataFrame({"not_found_value": not_found})
                            st.dataframe(not_found_df, use_container_width=True, hide_index=True)
                            st.download_button(
                                "Скачать не найденные значения (Excel)",
                                data=dataframe_to_excel_bytes(not_found_df, sheet_name="not_found"),
                                file_name="ozon_excel_not_found.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_not_found_export_{selected_product_id}",
                            )

                    selected_product_ids = st.multiselect(
                        "Товары для массовых действий",
                        options=product_options,
                        format_func=lambda x: next((f"ID {r['id']} | {r['article'] or '-'} | {r['name'] or '-'}" for r in product_rows if int(r['id']) == int(x)), str(x)),
                        key=bulk_select_key,
                    )
                    required_only_mode = st.checkbox(
                        "Работать только с обязательными Ozon-атрибутами",
                        value=False,
                        key=f"ozon_required_only_{selected_product_id}",
                    )
                    offer_id_field = st.selectbox(
                        "Поле товара для Ozon offer_id",
                        options=OZON_OFFER_ID_OPTIONS,
                        index=0,
                        key=f"ozon_offer_id_field_{selected_product_id}",
                    )
                    preview_rows = build_product_ozon_payload(
                        conn,
                        product_id=int(selected_product_id),
                        description_category_id=int(selected_row["description_category_id"]),
                        type_id=int(selected_row["type_id"]),
                        required_only=required_only_mode,
                        dictionary_min_score=float(dictionary_min_score),
                    )
                    if preview_rows:
                        preview_df = pd.DataFrame(preview_rows)
                        p1, p2, p3, p4 = st.columns(4)
                        p1.metric("Готово к автозаполнению", int((preview_df["status"] == "ready").sum()))
                        p2.metric("Пусто после маппинга", int((preview_df["status"] == "empty").sum()))
                        p3.metric("Обязательных готово", int(((preview_df["status"] == "ready") & (preview_df["is_required"] == 1)).sum()))
                        p4.metric("Dictionary не сматчено", int((preview_df["status"] == "dictionary_unmatched").sum()))

                        action1, action2 = st.columns(2)
                        with action1:
                            if st.button("Проверить покрытие товара под Ozon"):
                                coverage = analyze_product_ozon_coverage(
                                    conn,
                                    product_id=int(selected_product_id),
                                    description_category_id=int(selected_row["description_category_id"]),
                                    type_id=int(selected_row["type_id"]),
                                    dictionary_min_score=float(dictionary_min_score),
                                )
                                summary = coverage["summary"]
                                cc1, cc2, cc3, cc4 = st.columns(4)
                                cc1.metric("Готовность, %", int(summary["readiness_pct"]))
                                cc2.metric("Обязательных всего", int(summary["required_total"]))
                                cc3.metric("Обязательных закрыто", int(summary["required_covered"]))
                                cc4.metric("Обязательных пусто", int(summary["required_missing"]))
                                st.caption(f"Обязательных с несопоставленным справочником: {int(summary.get('required_dictionary_unmatched') or 0)}")
                                if summary["readiness_pct"] == 100:
                                    st.success("Обязательные Ozon-атрибуты по этой категории закрыты.")
                                else:
                                    st.warning("Не все обязательные Ozon-атрибуты закрыты. Ниже видно, что именно отсутствует.")
                                coverage_df = pd.DataFrame(coverage["rows"])
                                if not coverage_df.empty:
                                    st.dataframe(coverage_df, use_container_width=True, hide_index=True)
                        with action2:
                            if st.button("Заполнить Ozon-атрибуты из мастер-карточки"):
                                result = materialize_product_ozon_attributes(
                                    conn,
                                    product_id=int(selected_product_id),
                                    description_category_id=int(selected_row["description_category_id"]),
                                    type_id=int(selected_row["type_id"]),
                                    required_only=required_only_mode,
                                    dictionary_min_score=float(dictionary_min_score),
                                )
                                st.success(
                                    f"Записано Ozon-значений: {result['applied']}, пустых пропущено: {result['skipped_empty']}, "
                                    f"dictionary без матчинга: {result.get('skipped_dictionary', 0)}. category_code={result['category_code']}"
                                )

                        b1, b2 = st.columns(2)
                        with b1:
                            if st.button("Массовая проверка готовности по выбранным товарам"):
                                if not selected_product_ids:
                                    st.warning("Выбери хотя бы один товар для массовой проверки.")
                                else:
                                    report_rows = []
                                    progress = st.progress(0)
                                    for i, pid in enumerate(selected_product_ids, start=1):
                                        coverage = analyze_product_ozon_coverage(
                                            conn,
                                            product_id=int(pid),
                                            description_category_id=int(selected_row["description_category_id"]),
                                            type_id=int(selected_row["type_id"]),
                                            dictionary_min_score=float(dictionary_min_score),
                                        )
                                        summary = coverage.get("summary", {})
                                        product_row = next((r for r in product_rows if int(r["id"]) == int(pid)), None)
                                        report_rows.append(
                                            {
                                                "product_id": int(pid),
                                                "offer_id": (product_row[offer_id_field] if product_row else None),
                                                "article": product_row["article"] if product_row else None,
                                                "name": product_row["name"] if product_row else None,
                                                "readiness_pct": int(summary.get("readiness_pct") or 0),
                                                "required_total": int(summary.get("required_total") or 0),
                                                "required_covered": int(summary.get("required_covered") or 0),
                                                "required_missing": int(summary.get("required_missing") or 0),
                                                "required_dictionary_unmatched": int(summary.get("required_dictionary_unmatched") or 0),
                                            }
                                        )
                                        progress.progress(i / len(selected_product_ids))
                                    report_df = pd.DataFrame(report_rows).sort_values(
                                        by=["readiness_pct", "required_dictionary_unmatched", "required_missing"],
                                        ascending=[False, True, True],
                                    )
                                    st.dataframe(report_df, use_container_width=True, hide_index=True)
                                    st.download_button(
                                        "Скачать отчёт готовности Ozon (Excel)",
                                        data=dataframe_to_excel_bytes(report_df, sheet_name="ozon_readiness"),
                                        file_name="ozon_readiness_report.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"ozon_readiness_export_{selected_product_id}",
                                    )
                        with b2:
                            if st.button("Массово заполнить Ozon-атрибуты для выбранных"):
                                if not selected_product_ids:
                                    st.warning("Выбери хотя бы один товар для массового заполнения.")
                                else:
                                    progress = st.progress(0)
                                    total_applied = 0
                                    total_skipped_empty = 0
                                    total_skipped_dict = 0
                                    for i, pid in enumerate(selected_product_ids, start=1):
                                        result = materialize_product_ozon_attributes(
                                            conn,
                                            product_id=int(pid),
                                            description_category_id=int(selected_row["description_category_id"]),
                                            type_id=int(selected_row["type_id"]),
                                            required_only=required_only_mode,
                                            dictionary_min_score=float(dictionary_min_score),
                                        )
                                        total_applied += int(result.get("applied") or 0)
                                        total_skipped_empty += int(result.get("skipped_empty") or 0)
                                        total_skipped_dict += int(result.get("skipped_dictionary") or 0)
                                        progress.progress(i / len(selected_product_ids))
                                    st.success(
                                        f"Массовое заполнение завершено. Записано: {total_applied}, "
                                        f"пустых пропущено: {total_skipped_empty}, dictionary без матчинга: {total_skipped_dict}."
                                    )
                        if st.button("Сформировать dictionary gaps по выбранным (Excel)", key=f"ozon_bulk_gap_export_btn_{selected_product_id}"):
                            if not selected_product_ids:
                                st.warning("Выбери хотя бы один товар.")
                            else:
                                gap_export_rows = []
                                progress = st.progress(0)
                                for i, pid in enumerate(selected_product_ids, start=1):
                                    product_row = next((r for r in product_rows if int(r["id"]) == int(pid)), None)
                                    gap_rows = preview_product_ozon_dictionary_gaps(
                                        conn=conn,
                                        product_id=int(pid),
                                        description_category_id=int(selected_row["description_category_id"]),
                                        type_id=int(selected_row["type_id"]),
                                        top_n=3,
                                        dictionary_min_score=float(dictionary_min_score),
                                    )
                                    for gap in gap_rows:
                                        gap_export_rows.append(
                                            {
                                                "product_id": int(pid),
                                                "article": product_row["article"] if product_row else None,
                                                "name": product_row["name"] if product_row else None,
                                                "attribute_id": gap.get("attribute_id"),
                                                "attribute_name": gap.get("name"),
                                                "source_name": gap.get("source_name"),
                                                "raw_value": gap.get("raw_value"),
                                                "suggestion_values": gap.get("suggestion_values"),
                                            }
                                        )
                                    progress.progress(i / len(selected_product_ids))
                                if gap_export_rows:
                                    gap_export_df = pd.DataFrame(gap_export_rows)
                                    st.dataframe(gap_export_df, use_container_width=True, hide_index=True)
                                    st.download_button(
                                        "Скачать dictionary gaps (Excel)",
                                        data=dataframe_to_excel_bytes(gap_export_df, sheet_name="dictionary_gaps"),
                                        file_name=f"ozon_dictionary_gaps_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        key=f"ozon_bulk_gap_export_{selected_product_id}",
                                    )
                                else:
                                    st.success("Dictionary gaps по выбранным товарам не найдено.")

                        payload_preview = build_product_ozon_api_attributes(
                            conn,
                            product_id=int(selected_product_id),
                            description_category_id=int(selected_row["description_category_id"]),
                            type_id=int(selected_row["type_id"]),
                            required_only=required_only_mode,
                            dictionary_min_score=float(dictionary_min_score),
                            offer_id_field=str(offer_id_field),
                        )
                        st.download_button(
                            "Скачать preview Ozon JSON",
                            data=json.dumps(payload_preview, ensure_ascii=False, indent=2).encode("utf-8"),
                            file_name=f"ozon_payload_preview_product_{int(selected_product_id)}.json",
                            mime="application/json",
                        )
                        if selected_product_ids:
                            bulk_payload = build_bulk_ozon_api_payloads(
                                conn,
                                product_ids=[int(x) for x in selected_product_ids],
                                description_category_id=int(selected_row["description_category_id"]),
                                type_id=int(selected_row["type_id"]),
                                required_only=required_only_mode,
                                dictionary_min_score=float(dictionary_min_score),
                                offer_id_field=str(offer_id_field),
                            )
                            bulk_result_df = pd.DataFrame()
                            st.download_button(
                                "Скачать bulk Ozon JSON по выбранным товарам",
                                data=json.dumps(bulk_payload, ensure_ascii=False, indent=2).encode("utf-8"),
                                file_name=f"ozon_bulk_payload_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.json",
                                mime="application/json",
                                key=f"ozon_bulk_payload_export_{selected_product_id}",
                            )
                            bulk_products = bulk_payload.get("products") or []
                            if bulk_products:
                                bulk_result_df = pd.DataFrame(
                                    [
                                        {
                                            "product_id": int(item.get("product_id") or 0),
                                            "offer_id": item.get("offer_id"),
                                            "offer_id_field": item.get("offer_id_field"),
                                            "included_attributes": int(item.get("included") or 0),
                                            "skipped_attributes": int(item.get("skipped") or 0),
                                            "description_category_id": int(item.get("description_category_id") or 0),
                                            "type_id": int(item.get("type_id") or 0),
                                        }
                                        for item in bulk_products
                                    ]
                                )
                                st.download_button(
                                    "Скачать результат bulk обработки (Excel)",
                                    data=dataframe_to_excel_bytes(bulk_result_df, sheet_name="ozon_bulk_result"),
                                    file_name=f"ozon_bulk_result_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.xlsx",
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"ozon_bulk_result_export_{selected_product_id}",
                                )
                            update_request = build_ozon_attributes_update_request(bulk_payload)
                            st.download_button(
                                "Скачать request JSON для /v1/product/attributes/update",
                                data=json.dumps(update_request, ensure_ascii=False, indent=2).encode("utf-8"),
                                file_name=f"ozon_attributes_update_request_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.json",
                                mime="application/json",
                                key=f"ozon_update_request_export_{selected_product_id}",
                            )
                            update_items = update_request.get("items") or []
                            update_items_df = pd.DataFrame(
                                [
                                    {
                                        "offer_id": item.get("offer_id"),
                                        "description_category_id": item.get("description_category_id"),
                                        "type_id": item.get("type_id"),
                                        "attributes_count": len(item.get("attributes") or []),
                                    }
                                    for item in update_items
                                ]
                            )
                            update_summary_df = pd.DataFrame(
                                [
                                    {
                                        "products_total": int((bulk_payload.get("summary") or {}).get("products_total") or 0),
                                        "attributes_included": int((bulk_payload.get("summary") or {}).get("attributes_included") or 0),
                                        "attributes_skipped": int((bulk_payload.get("summary") or {}).get("attributes_skipped") or 0),
                                        "missing_offer_id": int((bulk_payload.get("summary") or {}).get("missing_offer_id") or 0),
                                        "request_items": int((update_request.get("summary") or {}).get("items_total") or 0),
                                        "request_skipped_missing_offer": int((update_request.get("summary") or {}).get("skipped_missing_offer") or 0),
                                        "request_skipped_empty_attrs": int((update_request.get("summary") or {}).get("skipped_empty_attrs") or 0),
                                    }
                                ]
                            )
                            st.download_button(
                                "Скачать Ozon bulk пакет (Excel)",
                                data=dataframes_to_excel_bytes(
                                    {
                                        "bulk_result": bulk_result_df,
                                        "update_items": update_items_df,
                                        "update_summary": update_summary_df,
                                    }
                                ),
                                file_name=f"ozon_bulk_package_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_bulk_package_export_{selected_product_id}",
                            )
                            if st.button(
                                "Отправить карточки в Ozon API",
                                disabled=(not configured),
                                key=f"ozon_send_update_{selected_product_id}",
                            ):
                                send_result = submit_ozon_attributes_update(
                                    conn=conn,
                                    product_ids=[int(x) for x in selected_product_ids],
                                    description_category_id=int(selected_row["description_category_id"]),
                                    type_id=int(selected_row["type_id"]),
                                    required_only=required_only_mode,
                                    dictionary_min_score=float(dictionary_min_score),
                                    offer_id_field=str(offer_id_field),
                                    client_id=resolved_client_id or None,
                                    api_key=resolved_api_key or None,
                                )
                                if send_result.get("ok"):
                                    response = send_result.get("response") or {}
                                    result_part = response.get("result") if isinstance(response, dict) else None
                                    task_id = result_part.get("task_id") if isinstance(result_part, dict) else None
                                    st.success(
                                        f"Batch отправлен в Ozon. items={send_result.get('request', {}).get('summary', {}).get('items_total', 0)}"
                                        + (f", task_id={task_id}" if task_id else "")
                                    )
                                else:
                                    st.error(send_result.get("message") or "Не удалось отправить batch в Ozon")
                            st.caption(
                                f"Bulk summary: products={bulk_payload.get('summary', {}).get('products_total', 0)}, "
                                f"included={bulk_payload.get('summary', {}).get('attributes_included', 0)}, "
                                f"skipped={bulk_payload.get('summary', {}).get('attributes_skipped', 0)}, "
                                f"missing_offer_id={bulk_payload.get('summary', {}).get('missing_offer_id', 0)} | "
                                f"request_items={update_request.get('summary', {}).get('items_total', 0)}"
                            )
                            st.caption(
                                "Текущий publish-flow Ozon использует `/v1/product/attributes/update`: "
                                "отправляет category/type и собранные атрибуты по пачке товаров."
                            )

                        st.markdown("### Preview полуавтозаполнения Ozon")
                        st.dataframe(
                            preview_df[
                                [
                                    c
                                    for c in [
                                        "attribute_id",
                                        "dictionary_id",
                                        "name",
                                        "is_required",
                                        "source_type",
                                        "source_name",
                                        "transform_rule",
                                        "status",
                                        "value",
                                        "dictionary_value_id",
                                        "dictionary_match_score",
                                        "dictionary_match_by",
                                    ]
                                    if c in preview_df.columns
                                ]
                            ],
                            use_container_width=True,
                            hide_index=True,
                        )

                        dict_unmatched_count = int((preview_df["status"] == "dictionary_unmatched").sum())
                        if dict_unmatched_count > 0:
                            st.markdown("### Подсказки по dictionary mismatch")
                            st.caption("Для несопоставленных значений система предлагает ближайшие варианты из кэша справочника Ozon.")
                            top_n = st.number_input(
                                "Сколько вариантов показывать на один атрибут",
                                min_value=1,
                                max_value=10,
                                value=3,
                                step=1,
                                key=f"ozon_dict_gap_topn_{selected_product_id}",
                            )
                            gap_rows = preview_product_ozon_dictionary_gaps(
                                conn,
                                product_id=int(selected_product_id),
                                description_category_id=int(selected_row["description_category_id"]),
                                type_id=int(selected_row["type_id"]),
                                top_n=int(top_n),
                                dictionary_min_score=float(dictionary_min_score),
                            )
                            if gap_rows:
                                gap_df = pd.DataFrame(gap_rows)
                                st.dataframe(
                                    gap_df[
                                        [
                                            c
                                            for c in [
                                                "attribute_id",
                                                "name",
                                                "source_name",
                                                "raw_value",
                                                "suggestion_values",
                                            ]
                                            if c in gap_df.columns
                                        ]
                                    ],
                                    use_container_width=True,
                                    hide_index=True,
                                )

                                gap_options = list(range(len(gap_rows)))
                                selected_gap_idx = st.selectbox(
                                    "Выбери проблемное значение для dictionary override",
                                    options=gap_options,
                                    format_func=lambda idx: (
                                        f"attr={gap_rows[idx].get('attribute_id')} | "
                                        f"{gap_rows[idx].get('name')} | raw={gap_rows[idx].get('raw_value')}"
                                    ),
                                    key=f"ozon_override_gap_idx_{selected_product_id}",
                                )
                                selected_gap = gap_rows[int(selected_gap_idx)]
                                selected_gap_suggestions = selected_gap.get("suggestions") or []
                                if selected_gap_suggestions:
                                    suggestion_options = list(range(len(selected_gap_suggestions)))
                                    selected_suggestion_idx = st.selectbox(
                                        "Подходящее значение из справочника",
                                        options=suggestion_options,
                                        format_func=lambda idx: (
                                            f"{selected_gap_suggestions[idx].get('value')} "
                                            f"(id={selected_gap_suggestions[idx].get('value_id')}, s={selected_gap_suggestions[idx].get('score')})"
                                        ),
                                        key=f"ozon_override_suggestion_idx_{selected_product_id}",
                                    )
                                    override_comment = st.text_input(
                                        "Комментарий к override (необязательно)",
                                        value="Сохранено из блока dictionary mismatch",
                                        key=f"ozon_override_comment_{selected_product_id}",
                                    )
                                    if st.button("Сохранить dictionary override", key=f"ozon_save_override_{selected_product_id}"):
                                        picked = selected_gap_suggestions[int(selected_suggestion_idx)]
                                        save_dictionary_override(
                                            conn=conn,
                                            description_category_id=int(selected_row["description_category_id"]),
                                            type_id=int(selected_row["type_id"]),
                                            attribute_id=int(selected_gap.get("attribute_id")),
                                            raw_value=selected_gap.get("raw_value"),
                                            value_id=int(picked.get("value_id")),
                                            value=picked.get("value"),
                                            comment=override_comment or None,
                                        )
                                        st.success(
                                            f"Override сохранён: raw='{selected_gap.get('raw_value')}' -> "
                                            f"id={picked.get('value_id')} ({picked.get('value')})"
                                        )
                                        st.rerun()
                                else:
                                    st.info("Для выбранного raw-значения пока нет кандидатов из справочника.")
                            else:
                                st.info("Несматченные dictionary-значения не найдены.")

                        overrides = list_dictionary_overrides(
                            conn=conn,
                            description_category_id=int(selected_row["description_category_id"]),
                            type_id=int(selected_row["type_id"]),
                            limit=200,
                        )
                        st.markdown("### Excel: массовый импорт dictionary overrides")
                        ov1, ov2 = st.columns([2, 1])
                        with ov1:
                            overrides_excel = st.file_uploader(
                                "Загрузи Excel с overrides",
                                type=["xlsx", "xls"],
                                key=f"ozon_overrides_excel_{selected_product_id}",
                            )
                        with ov2:
                            st.download_button(
                                "Скачать шаблон overrides",
                                data=build_ozon_dictionary_overrides_template_excel(),
                                file_name="ozon_dictionary_overrides_template.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_overrides_template_{selected_product_id}",
                            )
                            if st.button("Импортировать overrides из Excel", key=f"ozon_import_overrides_{selected_product_id}"):
                                if overrides_excel is None:
                                    st.warning("Сначала загрузи Excel файл с overrides.")
                                else:
                                    import_result = import_dictionary_overrides_from_excel(
                                        conn=conn,
                                        file_bytes=overrides_excel.read(),
                                        description_category_id=int(selected_row["description_category_id"]),
                                        type_id=int(selected_row["type_id"]),
                                    )
                                    if import_result.get("ok"):
                                        st.success(
                                            f"Импорт завершён: применено {import_result.get('applied', 0)}, "
                                            f"пропущено {import_result.get('skipped', 0)}."
                                        )
                                        errors = import_result.get("errors") or []
                                        if errors:
                                            st.dataframe(pd.DataFrame(errors), use_container_width=True, hide_index=True)
                                        st.rerun()
                                    else:
                                        st.error(import_result.get("message") or "Не удалось импортировать overrides.")

                        if overrides:
                            st.markdown("### Сохранённые dictionary overrides")
                            overrides_df = pd.DataFrame(overrides)
                            st.download_button(
                                "Скачать overrides (Excel)",
                                data=dataframe_to_excel_bytes(overrides_df, sheet_name="overrides"),
                                file_name=f"ozon_dictionary_overrides_{int(selected_row['description_category_id'])}_{int(selected_row['type_id'])}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_overrides_export_{selected_product_id}",
                            )
                            st.dataframe(
                                overrides_df[
                                    [
                                        c
                                        for c in [
                                            "attribute_id",
                                            "raw_value",
                                            "value_id",
                                            "value",
                                            "comment",
                                            "updated_at",
                                        ]
                                        if c in overrides_df.columns
                                    ]
                                ],
                                use_container_width=True,
                                hide_index=True,
                            )
                            selected_override_idx = st.selectbox(
                                "Выбери override для удаления",
                                options=list(range(len(overrides))),
                                format_func=lambda idx: (
                                    f"attr={overrides[idx].get('attribute_id')} | "
                                    f"raw={overrides[idx].get('raw_value')} -> "
                                    f"id={overrides[idx].get('value_id')} ({overrides[idx].get('value')})"
                                ),
                                key=f"ozon_override_delete_idx_{selected_product_id}",
                            )
                            if st.button("Удалить выбранный override", key=f"ozon_delete_override_{selected_product_id}"):
                                item = overrides[int(selected_override_idx)]
                                result = delete_dictionary_override(
                                    conn=conn,
                                    description_category_id=int(selected_row["description_category_id"]),
                                    type_id=int(selected_row["type_id"]),
                                    attribute_id=int(item.get("attribute_id")),
                                    raw_value=item.get("raw_value"),
                                )
                                st.success(f"Удалено overrides: {int(result.get('deleted') or 0)}")
                                st.rerun()

                        st.markdown("### Журнал отправок в Ozon")
                        retry_col1, retry_col2 = st.columns([2, 1])
                        with retry_col1:
                            retry_excel_file = st.file_uploader(
                                "Excel со списком job_id для повторной отправки",
                                type=["xlsx", "xls"],
                                key=f"ozon_retry_jobs_file_{selected_product_id}",
                            )
                            retry_excel_column = st.text_input(
                                "Колонка job_id (опционально)",
                                value="",
                                key=f"ozon_retry_jobs_column_{selected_product_id}",
                                placeholder="job_id",
                            )
                        with retry_col2:
                            st.download_button(
                                "Скачать шаблон job_id",
                                data=build_ozon_retry_jobs_template_excel(),
                                file_name="ozon_retry_jobs_template.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_retry_jobs_template_{selected_product_id}",
                            )
                            if st.button(
                                "Массово повторить jobs из Excel",
                                disabled=(not configured),
                                key=f"ozon_retry_jobs_apply_{selected_product_id}",
                            ):
                                if retry_excel_file is None:
                                    st.warning("Сначала загрузи Excel со списком job_id.")
                                else:
                                    parsed_jobs = resolve_job_ids_from_excel(
                                        retry_excel_file.read(),
                                        column_name=retry_excel_column or None,
                                    )
                                    if not parsed_jobs.get("ok"):
                                        st.error(parsed_jobs.get("message") or "Не удалось прочитать job_id из Excel.")
                                    else:
                                        job_ids = parsed_jobs.get("job_ids") or []
                                        if not job_ids:
                                            st.warning("В Excel не найдено корректных job_id.")
                                        else:
                                            progress = st.progress(0)
                                            ok_count = 0
                                            err_count = 0
                                            errors = []
                                            retry_rows = []
                                            for i, job_id in enumerate(job_ids, start=1):
                                                result = retry_ozon_update_job(
                                                    conn=conn,
                                                    job_id=int(job_id),
                                                    client_id=resolved_client_id or None,
                                                    api_key=resolved_api_key or None,
                                                )
                                                if result.get("ok"):
                                                    ok_count += 1
                                                    retry_rows.append(
                                                        {
                                                            "job_id": int(job_id),
                                                            "status": "success",
                                                            "task_id": result.get("task_id"),
                                                            "error": None,
                                                        }
                                                    )
                                                else:
                                                    err_count += 1
                                                    err_msg = result.get("message") or "Ошибка"
                                                    errors.append({"job_id": job_id, "error": err_msg})
                                                    retry_rows.append(
                                                        {
                                                            "job_id": int(job_id),
                                                            "status": "error",
                                                            "task_id": None,
                                                            "error": err_msg,
                                                        }
                                                    )
                                                progress.progress(i / len(job_ids))
                                            st.success(f"Массовый retry завершён. Успешно: {ok_count}, с ошибкой: {err_count}.")
                                            if retry_rows:
                                                retry_df = pd.DataFrame(retry_rows)
                                                st.dataframe(retry_df, use_container_width=True, hide_index=True)
                                                st.download_button(
                                                    "Скачать результат retry (Excel)",
                                                    data=dataframe_to_excel_bytes(retry_df, sheet_name="retry_result"),
                                                    file_name="ozon_retry_result.xlsx",
                                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                    key=f"ozon_retry_result_export_{selected_product_id}",
                                                )
                                            if parsed_jobs.get("errors"):
                                                st.dataframe(pd.DataFrame(parsed_jobs.get("errors")), use_container_width=True, hide_index=True)
                                            if errors:
                                                err_df = pd.DataFrame(errors)
                                                st.dataframe(err_df, use_container_width=True, hide_index=True)
                                                st.download_button(
                                                    "Скачать ошибки retry (Excel)",
                                                    data=dataframe_to_excel_bytes(err_df, sheet_name="retry_errors"),
                                                    file_name="ozon_retry_errors.xlsx",
                                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                    key=f"ozon_retry_errors_export_{selected_product_id}",
                                                )
                                            st.rerun()
                        jobs_limit = st.number_input(
                            "Сколько последних отправок показывать",
                            min_value=10,
                            max_value=500,
                            value=50,
                            step=10,
                            key=f"ozon_jobs_limit_{selected_product_id}",
                        )
                        jobs = list_ozon_update_jobs(conn, limit=int(jobs_limit))
                        if jobs:
                            jobs_df = pd.DataFrame(jobs)
                            st.download_button(
                                "Скачать журнал jobs (Excel)",
                                data=dataframe_to_excel_bytes(jobs_df, sheet_name="ozon_jobs"),
                                file_name="ozon_update_jobs.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"ozon_jobs_export_{selected_product_id}",
                            )
                            jm1, jm2, jm3, jm4 = st.columns(4)
                            jm1.metric("Всего jobs", int(len(jobs_df)))
                            jm2.metric("Успех", int((jobs_df["status"] == "success").sum()) if "status" in jobs_df.columns else 0)
                            jm3.metric("Ошибка", int((jobs_df["status"] == "error").sum()) if "status" in jobs_df.columns else 0)
                            jm4.metric("Skipped", int((jobs_df["status"] == "skipped").sum()) if "status" in jobs_df.columns else 0)

                            status_filter = st.selectbox(
                                "Фильтр jobs по статусу",
                                options=["Все", "success", "error", "skipped"],
                                index=0,
                                key=f"ozon_jobs_status_filter_{selected_product_id}",
                            )
                            if status_filter != "Все":
                                jobs_df = jobs_df[jobs_df["status"] == status_filter]

                            st.dataframe(
                                jobs_df[
                                    [
                                        c
                                        for c in [
                                            "id",
                                            "status",
                                            "items_count",
                                            "description_category_id",
                                            "type_id",
                                            "offer_id_field",
                                            "task_id",
                                            "retry_of_job_id",
                                            "error_message",
                                            "created_at",
                                        ]
                                        if c in jobs_df.columns
                                    ]
                                ],
                                use_container_width=True,
                                hide_index=True,
                            )
                            if jobs_df.empty:
                                st.info("По текущему фильтру jobs не найдено.")
                            else:
                                retry_all_col1, retry_all_col2 = st.columns([1, 2])
                                with retry_all_col1:
                                    if st.button(
                                        "Повторить все jobs из фильтра",
                                        disabled=(not configured),
                                        key=f"ozon_retry_filtered_jobs_{selected_product_id}",
                                    ):
                                        filtered_ids = [int(jid) for jid in jobs_df["id"].tolist()]
                                        progress = st.progress(0)
                                        ok_count = 0
                                        err_rows = []
                                        for i, jid in enumerate(filtered_ids, start=1):
                                            res = retry_ozon_update_job(
                                                conn=conn,
                                                job_id=int(jid),
                                                client_id=resolved_client_id or None,
                                                api_key=resolved_api_key or None,
                                            )
                                            if res.get("ok"):
                                                ok_count += 1
                                            else:
                                                err_rows.append({"job_id": int(jid), "error": res.get("message") or "Ошибка"})
                                            progress.progress(i / len(filtered_ids))
                                        st.success(
                                            f"Retry по фильтру завершён. Успешно: {ok_count}, ошибок: {len(err_rows)}."
                                        )
                                        if err_rows:
                                            err_df = pd.DataFrame(err_rows)
                                            st.dataframe(err_df, use_container_width=True, hide_index=True)
                                            st.download_button(
                                                "Скачать ошибки retry по фильтру (Excel)",
                                                data=dataframe_to_excel_bytes(err_df, sheet_name="retry_errors"),
                                                file_name="ozon_retry_filtered_errors.xlsx",
                                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                                key=f"ozon_retry_filtered_errors_{selected_product_id}",
                                            )
                                        st.rerun()
                                with retry_all_col2:
                                    st.caption("Кнопка повторяет все jobs, которые видны после текущего фильтра статуса.")

                                selected_job_id = st.selectbox(
                                    "Job для действий",
                                    options=[int(jid) for jid in jobs_df["id"].tolist()],
                                    format_func=lambda jid: next(
                                        (
                                            f"#{j['id']} | {j.get('status')} | items={j.get('items_count')} | created={j.get('created_at')}"
                                            for j in jobs
                                            if int(j["id"]) == int(jid)
                                        ),
                                        str(jid),
                                    ),
                                    key=f"ozon_job_action_{selected_product_id}",
                                )
                                job_item = get_ozon_update_job(conn, int(selected_job_id))
                                if job_item:
                                    a1, a2, a3 = st.columns(3)
                                    with a1:
                                        request_bytes = (job_item.get("request_json") or "{}").encode("utf-8")
                                        st.download_button(
                                            "Скачать request job",
                                            data=request_bytes,
                                            file_name=f"ozon_job_{int(selected_job_id)}_request.json",
                                            mime="application/json",
                                            key=f"ozon_job_req_dl_{selected_product_id}",
                                        )
                                    with a2:
                                        response_bytes = (job_item.get("response_json") or "{}").encode("utf-8")
                                        st.download_button(
                                            "Скачать response job",
                                            data=response_bytes,
                                            file_name=f"ozon_job_{int(selected_job_id)}_response.json",
                                            mime="application/json",
                                            key=f"ozon_job_resp_dl_{selected_product_id}",
                                        )
                                    with a3:
                                        if st.button(
                                            "Повторить отправку job",
                                            disabled=(not configured),
                                            key=f"ozon_job_retry_{selected_product_id}",
                                        ):
                                            retry_result = retry_ozon_update_job(
                                                conn=conn,
                                                job_id=int(selected_job_id),
                                                client_id=resolved_client_id or None,
                                                api_key=resolved_api_key or None,
                                            )
                                            if retry_result.get("ok"):
                                                st.success(
                                                    "Повторная отправка выполнена"
                                                    + (f", task_id={retry_result.get('task_id')}" if retry_result.get("task_id") else "")
                                                )
                                            else:
                                                st.error(retry_result.get("message") or "Не удалось повторить отправку job")
                                            st.rerun()
                                    job_items = list_ozon_update_job_items(conn, int(selected_job_id), limit=10000)
                                    try:
                                        job_response = json.loads(job_item.get("response_json") or "{}")
                                    except Exception:
                                        job_response = {}
                                    if job_items:
                                        job_items_df = pd.DataFrame(job_items)
                                        st.download_button(
                                            "Скачать selected job items (Excel)",
                                            data=dataframe_to_excel_bytes(job_items_df, sheet_name="job_items"),
                                            file_name=f"ozon_job_{int(selected_job_id)}_items.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key=f"ozon_job_items_excel_{selected_product_id}",
                                        )
                                        st.dataframe(job_items_df, use_container_width=True, hide_index=True)
                                    if job_response:
                                        response_df = pd.DataFrame([job_response])
                                        st.download_button(
                                            "Скачать selected job response (Excel)",
                                            data=dataframe_to_excel_bytes(response_df, sheet_name="job_response"),
                                            file_name=f"ozon_job_{int(selected_job_id)}_response.xlsx",
                                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                            key=f"ozon_job_response_excel_{selected_product_id}",
                                        )
                        else:
                            st.info("Отправок в Ozon пока не было.")
            else:
                st.info("По этой категории атрибуты ещё не загружались.")
    else:
        st.info("Кэш категорий пока пуст. Сначала синхронизируй дерево Ozon.")

    conn.close()


def show_settings_tab():
    conn = get_db()
    summary = build_workspace_summary(conn)
    st.subheader("Настройки PIM")
    st.caption("Все общие настройки сервиса собраны здесь: AI, парсинг, фото и рабочая конфигурация. Каналы больше не должны играть роль системного шкафа со всем подряд.")

    top1, top2, top3, top4 = st.columns(4)
    top1.metric("Товаров в памяти", int(summary.get("products_total") or 0))
    top2.metric("Клиентов", int(summary.get("clients_total") or 0))
    top3.metric("Detmir категорий", int(summary.get("detmir_categories_total") or 0))
    top4.metric("Ozon cat/type", int(summary.get("ozon_pairs_total") or 0))

    with st.expander("Быстрый маршрут по настройкам", expanded=False):
        st.markdown(
            """
1. `AI`: выбрать провайдера, активировать профиль и проверить доступ к модели.
2. `Парсинг`: задать стратегию enrichment и защиту от ложных совпадений.
3. `Фото`: указать публичный base URL, если в прайсах приходят локальные пути.
4. `Система`: проверить активную БД и понять, где сейчас живёт память PIM.
            """
        )

    tab_overview, tab_ai, tab_parser, tab_media = st.tabs(["Обзор", "AI", "Парсинг", "Фото"])
    with tab_overview:
        st.markdown("### Состояние памяти и активной конфигурации")
        ai_settings = load_ai_settings(conn)
        parser_settings = load_parser_settings(conn)
        media_settings = load_media_settings(conn)
        ov1, ov2, ov3 = st.columns(3)
        with ov1:
            st.markdown(
                f"""
<div class="pim-soft-card">
  <strong>AI</strong><br/>
  Провайдер: <span class="pim-mono">{str(ai_settings.get('provider') or '-')}</span><br/>
  Модель: <span class="pim-mono">{str(ai_settings.get('chat_model') or '-')}</span><br/>
  Активный ключ: <strong>{'есть' if str(ai_settings.get('api_key') or '').strip() else ('env' if bool(ai_settings.get('use_env_api_key', True)) else 'нет')}</strong>
</div>
                """,
                unsafe_allow_html=True,
            )
        with ov2:
            st.markdown(
                f"""
<div class="pim-soft-card">
  <strong>Парсинг</strong><br/>
  Стратегия: <span class="pim-mono">{str(parser_settings.get('source_strategy') or '-')}</span><br/>
  Таймаут: <strong>{float(parser_settings.get('timeout_seconds') or 0):.0f} сек</strong><br/>
  Домены fallback: <span class="pim-mono">{str(parser_settings.get('extra_fallback_domains') or '-')}</span>
</div>
                """,
                unsafe_allow_html=True,
            )
        with ov3:
            st.markdown(
                f"""
<div class="pim-soft-card">
  <strong>Фото</strong><br/>
  Public base URL:<br/>
  <span class="pim-mono">{str(media_settings.get('public_base_url') or '-')}</span><br/><br/>
  База данных:<br/>
  <span class="pim-mono">{str(summary.get('active_db_path') or '-')}</span>
</div>
                """,
                unsafe_allow_html=True,
            )
    with tab_ai:
        render_ai_settings_panel(conn)
    with tab_parser:
        render_parser_settings_panel(conn)
    with tab_media:
        render_media_settings_panel(conn)

    conn.close()


def show_channels_tab():
    conn = get_db()
    st.subheader("Каналы")
    st.caption("Здесь живут channel requirements, mapping rules и клиентские overlay-интеграции. Общие настройки PIM теперь вынесены в отдельный раздел `Настройки`.")

    with st.expander("Детский Мир API: схема клиента и карточки", expanded=False):
        st.caption(
            "Детский Мир для PIM остаётся клиентским overlay поверх Ozon-ядра. "
            "Здесь мы read-only подтягиваем схему клиента: категории, атрибуты, значения справочников и текущие карточки товаров."
        )
        detmir_settings = load_detmir_settings(conn)
        detmir_stats = get_detmir_cache_stats(conn)
        det1, det2, det3, det4 = st.columns(4)
        with det1:
            detmir_client_id = st.text_input(
                "Detmir Client ID",
                value=str(detmir_settings.get("client_id") or ""),
                key="detmir_client_id_input",
            )
        with det2:
            detmir_api_key = st.text_input(
                "Detmir API key",
                value=str(detmir_settings.get("api_key") or ""),
                type="password",
                key="detmir_api_key_input",
            )
        with det3:
            detmir_use_env = st.checkbox(
                "Брать key из env",
                value=bool(detmir_settings.get("use_env_api_key", True)),
                key="detmir_use_env_api_key",
            )
        with det4:
            if detmir_is_configured(
                conn=conn,
                client_id=str(detmir_client_id or "").strip() or None,
                api_key=str(detmir_api_key or "").strip() or None,
                use_env_api_key=bool(detmir_use_env),
            ):
                st.success("Detmir API готов")
            else:
                st.warning("Нужны client_id и api_key")

        dact1, dact2, dact3 = st.columns([1, 1, 2])
        with dact1:
            if st.button("Сохранить Detmir-настройки", key="detmir_save_settings_btn"):
                save_detmir_settings(
                    conn,
                    {
                        "client_id": str(detmir_client_id or "").strip(),
                        "api_key": str(detmir_api_key or "").strip(),
                        "use_env_api_key": bool(detmir_use_env),
                    },
                )
                st.success("Detmir-настройки сохранены.")
                st.rerun()
        with dact2:
            if st.button("Проверить Detmir API", key="detmir_check_btn"):
                result = check_detmir_connection(
                    conn=conn,
                    client_id=str(detmir_client_id or "").strip() or None,
                    api_key=str(detmir_api_key or "").strip() or None,
                    use_env_api_key=bool(detmir_use_env),
                )
                if result.get("ok"):
                    st.success(f"Detmir API отвечает. Корневых категорий: {int(result.get('categories_root') or 0)}.")
                else:
                    st.error(result.get("error") or "Не удалось подключиться к Detmir API.")
        with dact3:
            st.caption(
                f"В памяти: категорий {int(detmir_stats.get('detmir_category_cache') or 0)}, "
                f"атрибутов {int(detmir_stats.get('detmir_attribute_cache') or 0)}, "
                f"значений {int(detmir_stats.get('detmir_attribute_value_cache') or 0)}, "
                f"карточек {int(detmir_stats.get('detmir_product_cache') or 0)}."
            )
            if detmir_stats.get("last_schema_sync_at") or detmir_stats.get("last_product_sync_at"):
                st.caption(
                    f"Последний schema sync: {detmir_stats.get('last_schema_sync_at') or '-'} | "
                    f"Последний product sync: {detmir_stats.get('last_product_sync_at') or '-'}"
                )

        sync_col1, sync_col2, sync_col3, sync_col4 = st.columns([1.2, 1.4, 1.5, 1.6])
        with sync_col1:
            if st.button("Синхронизировать дерево категорий", key="detmir_sync_tree_btn"):
                with st.spinner("Синхронизирую дерево категорий Детского Мира..."):
                    result = sync_detmir_category_tree(
                        conn,
                        client_id=str(detmir_client_id or "").strip() or None,
                        api_key=str(detmir_api_key or "").strip() or None,
                        use_env_api_key=bool(detmir_use_env),
                    )
                st.success(f"Дерево категорий обновлено: {int(result.get('categories') or 0)} узлов.")
                backup_result = backup_database_file(reason="detmir_category_tree_sync")
                if backup_result.get("ok"):
                    st.caption(f"Память Detmir зафиксирована: `{Path(str(backup_result['path'])).name}`")
                st.rerun()
        with sync_col2:
            detmir_schema_pages = st.number_input(
                "Max pages schema",
                min_value=1,
                max_value=500,
                value=10,
                step=1,
                key="detmir_schema_pages_limit",
                help="Ограничь число страниц, если нужен быстрый тестовый sync.",
            )
            if st.button("Синхронизировать категории и атрибуты", key="detmir_sync_schema_btn"):
                with st.spinner("Синхронизирую категории и атрибуты Детского Мира..."):
                    result = sync_categories_with_attributes(
                        conn,
                        client_id=str(detmir_client_id or "").strip() or None,
                        api_key=str(detmir_api_key or "").strip() or None,
                        use_env_api_key=bool(detmir_use_env),
                        max_pages=int(detmir_schema_pages),
                    )
                st.success(
                    f"Schema sync завершён: категорий {int(result.get('categories') or 0)}, "
                    f"атрибутов {int(result.get('attributes') or 0)}, variant {int(result.get('variant_attributes') or 0)}."
                )
                backup_result = backup_database_file(reason="detmir_schema_sync")
                if backup_result.get("ok"):
                    st.caption(f"Память Detmir зафиксирована: `{Path(str(backup_result['path'])).name}`")
                st.rerun()
        with sync_col3:
            detmir_value_attrs = st.number_input(
                "Max attrs values",
                min_value=1,
                max_value=5000,
                value=40,
                step=1,
                key="detmir_values_attr_limit",
                help="Сколько attribute keys брать в пакетный sync значений справочников.",
            )
            if st.button("Синхронизировать значения атрибутов", key="detmir_sync_values_btn"):
                with st.spinner("Синхронизирую значения справочников Детского Мира..."):
                    result = sync_all_detmir_attribute_values(
                        conn,
                        client_id=str(detmir_client_id or "").strip() or None,
                        api_key=str(detmir_api_key or "").strip() or None,
                        use_env_api_key=bool(detmir_use_env),
                        max_attributes=int(detmir_value_attrs),
                        dictionary_only=True,
                    )
                st.success(
                    f"Values sync завершён: атрибутов {int(result.get('attributes') or 0)}, "
                    f"значений {int(result.get('values') or 0)}."
                )
                if result.get("errors"):
                    st.dataframe(pd.DataFrame([{"error": item} for item in result.get("errors") or []]), use_container_width=True, hide_index=True)
                backup_result = backup_database_file(reason="detmir_attribute_values_sync")
                if backup_result.get("ok"):
                    st.caption(f"Память Detmir зафиксирована: `{Path(str(backup_result['path'])).name}`")
                st.rerun()
        with sync_col4:
            detmir_product_pages = st.number_input(
                "Max pages products",
                min_value=1,
                max_value=500,
                value=5,
                step=1,
                key="detmir_products_pages_limit",
                help="Сколько страниц карточек товаров читать за один пакетный sync.",
            )
            if st.button("Синхронизировать карточки товаров", key="detmir_sync_products_btn"):
                with st.spinner("Синхронизирую карточки товаров Детского Мира..."):
                    result = sync_detmir_products(
                        conn,
                        client_id=str(detmir_client_id or "").strip() or None,
                        api_key=str(detmir_api_key or "").strip() or None,
                        use_env_api_key=bool(detmir_use_env),
                        max_pages=int(detmir_product_pages),
                    )
                st.success(
                    f"Синк карточек завершён: товаров {int(result.get('products') or 0)}, "
                    f"страниц {int(result.get('pages') or 0)}."
                )
                backup_result = backup_database_file(reason="detmir_products_sync")
                if backup_result.get("ok"):
                    st.caption(f"Память Detmir зафиксирована: `{Path(str(backup_result['path'])).name}`")
                st.rerun()

        st.markdown("### Категории Детского Мира")
        cat_f1, cat_f2 = st.columns([2, 1])
        with cat_f1:
            detmir_category_search = st.text_input(
                "Поиск категории Detmir",
                value="",
                placeholder="Например: велосипед, самокат, беговел",
                key="detmir_category_search",
            )
        with cat_f2:
            detmir_only_leaf = st.checkbox("Только листовые", value=True, key="detmir_only_leaf")
        detmir_categories = list_detmir_cached_categories(
            conn,
            search=detmir_category_search or None,
            only_leaf=bool(detmir_only_leaf),
            limit=500,
        )
        detmir_categories_with_schema = [
            row
            for row in detmir_categories
            if int(row.get("attributes_count") or 0) > 0 or int(row.get("variant_attributes_count") or 0) > 0
        ]
        if detmir_categories_with_schema:
            detmir_categories = detmir_categories_with_schema
        if detmir_categories:
            detmir_category_options = [int(row["category_id"]) for row in detmir_categories]
            detmir_category_map = {int(row["category_id"]): row for row in detmir_categories}
            selected_detmir_category_id = st.selectbox(
                "Категория Detmir",
                options=detmir_category_options,
                format_func=lambda cid: (
                    f"{detmir_category_map[int(cid)].get('full_path') or detmir_category_map[int(cid)].get('name')} "
                    f"| cat={cid} | attrs={int(detmir_category_map[int(cid)].get('attributes_count') or 0)}"
                ),
                key="detmir_selected_category_id",
            )
            selected_detmir_category = detmir_category_map[int(selected_detmir_category_id)]
            cat_info_df = pd.DataFrame(
                [
                    {
                        "category_id": int(selected_detmir_category.get("category_id") or 0),
                        "path": selected_detmir_category.get("full_path"),
                        "product_type": selected_detmir_category.get("product_type_name"),
                        "dimension_type": selected_detmir_category.get("dimension_type"),
                        "published": bool(selected_detmir_category.get("published")),
                        "attrs": int(selected_detmir_category.get("attributes_count") or 0),
                        "variant_attrs": int(selected_detmir_category.get("variant_attributes_count") or 0),
                        "updated_remote_at": selected_detmir_category.get("updated_remote_at"),
                    }
                ]
            )
            st.dataframe(cat_info_df, use_container_width=True, hide_index=True)

            detmir_attrs = list_detmir_cached_attributes(
                conn,
                category_id=int(selected_detmir_category_id),
                include_variant=True,
                limit=10000,
            )
            act1, act2 = st.columns([1, 1])
            with act1:
                if st.button("Импортировать требования категории Detmir в PIM", key="detmir_import_requirements_btn"):
                    result = import_detmir_category_requirements_to_pim(
                        conn,
                        category_id=int(selected_detmir_category_id),
                        create_mapping_rules=True,
                    )
                    st.success(
                        f"В PIM импортировано {int(result.get('imported') or 0)} требований, "
                        f"обязательных {int(result.get('required') or 0)}, "
                        f"mapping rules {int(result.get('mapping_saved') or 0)}."
                    )
                    backup_result = backup_database_file(reason="detmir_import_requirements")
                    if backup_result.get("ok"):
                        st.caption(f"Overlay Detmir зафиксирован: `{Path(str(backup_result['path'])).name}`")
                    st.rerun()
            with act2:
                if st.button("Синхронизировать значения только для этой категории", key="detmir_sync_selected_values_btn"):
                    attr_keys = sorted(
                        {
                            str(row.get("attribute_key") or "").strip()
                            for row in detmir_attrs
                            if str(row.get("data_type") or "").strip().upper() in {"SELECT", "SELECT_MULTIPLE", "EXTENDED_DICTIONARY"}
                            and str(row.get("attribute_key") or "").strip()
                        }
                    )
                    synced_attr_total = 0
                    synced_values_total = 0
                    sync_errors: list[str] = []
                    for attr_key in attr_keys:
                        try:
                            one_result = sync_detmir_attribute_values(
                                conn,
                                attribute_key=str(attr_key),
                                client_id=str(detmir_client_id or "").strip() or None,
                                api_key=str(detmir_api_key or "").strip() or None,
                                use_env_api_key=bool(detmir_use_env),
                            )
                            synced_attr_total += 1
                            synced_values_total += int(one_result.get("values") or 0)
                        except Exception as e:
                            sync_errors.append(f"{attr_key}: {e}")
                    st.success(
                        f"По категории синхронизированы значения: атрибутов {synced_attr_total}, "
                        f"значений {synced_values_total}."
                    )
                    if sync_errors:
                        st.dataframe(pd.DataFrame([{"error": item} for item in sync_errors]), use_container_width=True, hide_index=True)
                    st.rerun()

            if detmir_attrs:
                attr_df = pd.DataFrame(detmir_attrs)
                show_cols = [
                    c
                    for c in [
                        "attribute_key",
                        "attribute_name",
                        "data_type",
                        "is_required",
                        "is_variant_attribute",
                        "restriction_type",
                        "feature_type",
                    ]
                    if c in attr_df.columns
                ]
                st.dataframe(attr_df[show_cols], use_container_width=True, hide_index=True)
                dict_attr_options = [
                    str(row.get("attribute_key"))
                    for row in detmir_attrs
                    if str(row.get("data_type") or "").strip().upper() in {"SELECT", "SELECT_MULTIPLE", "EXTENDED_DICTIONARY"}
                ]
                if dict_attr_options:
                    selected_detmir_attr_key = st.selectbox(
                        "Справочный атрибут категории",
                        options=dict_attr_options,
                        key="detmir_selected_attr_key",
                    )
                    val_a1, val_a2 = st.columns([1, 2])
                    with val_a1:
                        if st.button("Синхронизировать значения выбранного атрибута", key="detmir_sync_one_attr_values_btn"):
                            result = sync_detmir_attribute_values(
                                conn,
                                attribute_key=str(selected_detmir_attr_key),
                                client_id=str(detmir_client_id or "").strip() or None,
                                api_key=str(detmir_api_key or "").strip() or None,
                                use_env_api_key=bool(detmir_use_env),
                            )
                            st.success(
                                f"Значения атрибута `{selected_detmir_attr_key}` синхронизированы: {int(result.get('values') or 0)}."
                            )
                            st.rerun()
                    with val_a2:
                        detmir_attr_values = list_detmir_cached_attribute_values(
                            conn,
                            attribute_key=str(selected_detmir_attr_key),
                            limit=400,
                        )
                        if detmir_attr_values:
                            st.dataframe(pd.DataFrame(detmir_attr_values), use_container_width=True, hide_index=True)
                        else:
                            st.info("Для выбранного атрибута значения пока не синхронизировались.")
            else:
                st.info("По выбранной категории атрибуты пока не синхронизированы.")
        else:
            st.info("Кэш категорий Detmir пока пуст. Сначала выполни sync дерева или schema sync.")

        st.markdown("### Карточки товаров Детского Мира")
        p1, p2, p3 = st.columns([2, 1, 1])
        with p1:
            detmir_product_search = st.text_input(
                "Поиск по карточкам Detmir",
                value="",
                placeholder="title, siteName, productCode, mastercardId",
                key="detmir_product_search",
            )
        with p2:
            detmir_status_filter = st.text_input("Статус", value="", placeholder="ACCEPTED / DRAFT / ...", key="detmir_status_filter")
        with p3:
            detmir_products_limit = st.number_input("Лимит карточек", min_value=10, max_value=500, value=50, step=10, key="detmir_products_limit")
        detmir_products = list_detmir_cached_products(
            conn,
            search=detmir_product_search or None,
            status=(detmir_status_filter or None),
            limit=int(detmir_products_limit),
        )
        if detmir_products:
            detmir_products_df = pd.DataFrame(detmir_products)
            visible_cols = [
                c
                for c in [
                    "product_id",
                    "product_code",
                    "mastercard_id",
                    "category_id",
                    "title",
                    "site_name",
                    "status",
                    "fbo_stock_level",
                    "fbs_stock_level",
                    "updated_remote_at",
                ]
                if c in detmir_products_df.columns
            ]
            st.dataframe(detmir_products_df[visible_cols], use_container_width=True, hide_index=True)
            detmir_product_options = [int(row["product_id"]) for row in detmir_products]
            detmir_product_map = {int(row["product_id"]): row for row in detmir_products}
            selected_detmir_product_id = st.selectbox(
                "Карточка товара Detmir",
                options=detmir_product_options,
                format_func=lambda pid: (
                    f"{detmir_product_map[int(pid)].get('product_code') or '-'} | "
                    f"{detmir_product_map[int(pid)].get('title') or '-'}"
                ),
                key="detmir_selected_product_id",
            )
            selected_detmir_product = detmir_product_map[int(selected_detmir_product_id)]
            detail_df = pd.DataFrame(
                [
                    {
                        "product_id": int(selected_detmir_product.get("product_id") or 0),
                        "product_code": selected_detmir_product.get("product_code"),
                        "mastercard_id": selected_detmir_product.get("mastercard_id"),
                        "category_id": selected_detmir_product.get("category_id"),
                        "title": selected_detmir_product.get("title"),
                        "site_name": selected_detmir_product.get("site_name"),
                        "status": selected_detmir_product.get("status"),
                        "photos": len(json.loads(selected_detmir_product.get("photos_json") or "[]")),
                        "attributes": len(json.loads(selected_detmir_product.get("attributes_json") or "[]")),
                        "updated_remote_at": selected_detmir_product.get("updated_remote_at"),
                    }
                ]
            )
            st.dataframe(detail_df, use_container_width=True, hide_index=True)
            product_raw_json = str(selected_detmir_product.get("raw_json") or "{}")
            st.download_button(
                "Скачать raw JSON выбранной карточки Detmir",
                data=product_raw_json.encode("utf-8"),
                file_name=f"detmir_product_{int(selected_detmir_product_id)}.json",
                mime="application/json",
                key="detmir_product_raw_download_btn",
            )
        else:
            st.info("Карточки Detmir пока не синхронизированы или не подходят под текущий фильтр.")

    with st.expander("Wildberries API: клиент и карточки", expanded=False):
        st.caption(
            "Wildberries для PIM — ещё один клиентский канал. "
            "Здесь сохраняем токен, читаем schema WB и готовим draft-карточки для отправки в Content API."
        )
        wb_settings = load_wb_settings(conn)
        if "wb_api_token_input" not in st.session_state:
            st.session_state["wb_api_token_input"] = str(wb_settings.get("api_token") or "")
        wb_token = st.text_input(
            "WB API token",
            key="wb_api_token_input",
            type="password",
        )
        wb_ready = wb_is_configured(conn=conn, api_token=str(wb_token or "").strip() or None)
        wb1, wb2, wb3, wb4 = st.columns([1, 1, 1, 2])
        with wb1:
            if st.button("Сохранить Wildberries", key="wb_save_settings_btn"):
                if not str(wb_token or "").strip():
                    st.warning("Сначала вставь WB API token.")
                else:
                    save_wb_settings(conn, {"api_token": str(wb_token or "").strip()})
                    st.success("WB token сохранён в памяти PIM.")
        with wb2:
            if st.button("Проверить WB API", key="wb_check_btn", disabled=not wb_ready):
                wb_check = check_wb_connection(conn=conn, api_token=str(wb_token or "").strip() or None)
                if wb_check.get("ok"):
                    st.success(
                        f"WB API отвечает. Продавец: {wb_check.get('seller_name') or '-'}"
                        + (f" | ТМ: {wb_check.get('trade_mark')}" if wb_check.get("trade_mark") else "")
                    )
                else:
                    st.error(str(wb_check.get("message") or "Не удалось подключиться к WB API."))
        with wb3:
            if st.button("Очистить WB token", key="wb_clear_settings_btn"):
                clear_wb_settings(conn)
                st.session_state["wb_api_token_input"] = ""
                st.success("WB token очищен из памяти PIM.")
                st.rerun()
        with wb4:
            if wb_ready:
                st.caption("WB-клиент подключён. Можно читать schema и отправлять draft-карточки в Content API.")
            else:
                st.caption("Сохрани токен WB, чтобы использовать этот канал как клиента внутри PIM.")

        schema_c1, schema_c2 = st.columns([1, 2])
        with schema_c1:
            if st.button("Загрузить parent-категории WB", key="wb_load_parents_btn", disabled=not wb_ready):
                try:
                    st.session_state["wb_parent_categories"] = list_wb_parent_categories(
                        conn=conn,
                        api_token=str(wb_token or "").strip() or None,
                    )
                    st.success(f"Parent-категории WB загружены: {len(st.session_state['wb_parent_categories'])}.")
                except Exception as e:
                    st.error(str(e))
        with schema_c2:
            parent_categories = st.session_state.get("wb_parent_categories") or []
            if parent_categories:
                st.caption(f"В памяти сессии WB parent-категорий: {len(parent_categories)}.")
            else:
                st.caption("Сначала подтяни parent-категории WB, затем выбери предмет и посмотри характеристики.")

        parent_categories = st.session_state.get("wb_parent_categories") or []
        selected_parent_id = None
        if parent_categories:
            parent_options = [int(x.get("parentID") or x.get("id") or 0) for x in parent_categories if int(x.get("parentID") or x.get("id") or 0) > 0]
            parent_map = {
                int(x.get("parentID") or x.get("id") or 0): x
                for x in parent_categories
                if int(x.get("parentID") or x.get("id") or 0) > 0
            }
            if parent_options:
                selected_parent_id = st.selectbox(
                    "Parent-категория WB",
                    options=parent_options,
                    format_func=lambda pid: str(parent_map[int(pid)].get("name") or parent_map[int(pid)].get("parentName") or f"parentID={pid}"),
                    key="wb_selected_parent_id",
                )

        sb1, sb2, sb3 = st.columns([2, 1, 1])
        with sb1:
            wb_subject_search = st.text_input(
                "Поиск предмета WB",
                value="",
                placeholder="Например: велосипед, шлем, велосумка",
                key="wb_subject_search",
            )
        with sb2:
            wb_subject_limit = st.number_input("Лимит предметов", min_value=10, max_value=500, value=50, step=10, key="wb_subject_limit")
        with sb3:
            if st.button("Найти предметы WB", key="wb_find_subjects_btn", disabled=not wb_ready):
                try:
                    st.session_state["wb_subjects"] = search_wb_subjects(
                        conn=conn,
                        api_token=str(wb_token or "").strip() or None,
                        name=wb_subject_search or None,
                        parent_id=int(selected_parent_id) if selected_parent_id else None,
                        limit=int(wb_subject_limit),
                    )
                    st.success(f"Предметов WB найдено: {len(st.session_state['wb_subjects'])}.")
                except Exception as e:
                    st.error(str(e))

        wb_subjects = st.session_state.get("wb_subjects") or []
        selected_subject_id = None
        if wb_subjects:
            subject_options = [int(x.get("subjectID") or 0) for x in wb_subjects if int(x.get("subjectID") or 0) > 0]
            subject_map = {int(x.get("subjectID") or 0): x for x in wb_subjects if int(x.get("subjectID") or 0) > 0}
            if subject_options:
                selected_subject_id = st.selectbox(
                    "Предмет WB",
                    options=subject_options,
                    format_func=lambda sid: (
                        f"{subject_map[int(sid)].get('subjectName') or '-'} | subjectID={sid}"
                        + (
                            f" | parent={subject_map[int(sid)].get('parentName')}"
                            if subject_map[int(sid)].get("parentName")
                            else ""
                        )
                    ),
                    key="wb_selected_subject_id",
                )
                if st.button("Показать характеристики предмета WB", key="wb_load_subject_charcs_btn", disabled=not wb_ready):
                    try:
                        st.session_state["wb_subject_characteristics"] = get_wb_subject_characteristics(
                            conn=conn,
                            api_token=str(wb_token or "").strip() or None,
                            subject_id=int(selected_subject_id),
                        )
                        st.success(
                            f"Характеристик предмета WB загружено: {len(st.session_state['wb_subject_characteristics'])}."
                        )
                    except Exception as e:
                        st.error(str(e))

        wb_charcs = st.session_state.get("wb_subject_characteristics") or []
        if wb_charcs:
            wb_char_df = pd.DataFrame(wb_charcs)
            show_cols = [c for c in ["charcID", "name", "required", "unitName", "maxCount", "charcType"] if c in wb_char_df.columns]
            st.dataframe(wb_char_df[show_cols], use_container_width=True, hide_index=True)

        st.markdown("### Draft-пачка Wildberries")
        wb_catalog_shortlist = [
            int(x)
            for x in (st.session_state.get("template_selected_ids_from_catalog") or [])
            if str(x).strip()
        ]
        wb_products = conn.execute(
            "SELECT id, article, name FROM products ORDER BY id DESC LIMIT 1000"
        ).fetchall()
        wb_product_options = [int(r["id"]) for r in wb_products]
        wb_product_label_map = {
            int(r["id"]): f"{r['article'] or '-'} | {r['name'] or '-'}"
            for r in wb_products
        }
        wb_select_key = "wb_selected_product_ids"
        if wb_select_key not in st.session_state:
            st.session_state[wb_select_key] = []
        wbp1, wbp2, wbp3 = st.columns([1, 1, 2])
        with wbp1:
            if st.button("Подтянуть shortlist из Каталога", key="wb_pull_catalog_shortlist_btn"):
                st.session_state[wb_select_key] = [int(x) for x in wb_catalog_shortlist]
                st.success(f"В WB-пачку подтянуто товаров: {len(wb_catalog_shortlist)}.")
                st.rerun()
        with wbp2:
            if st.button("Очистить WB-пачку", key="wb_clear_selected_ids_btn"):
                st.session_state[wb_select_key] = []
                st.rerun()
        with wbp3:
            st.caption(
                f"Shortlist из Каталога: {len(wb_catalog_shortlist)} | "
                f"Сейчас выбрано для WB: {len(st.session_state.get(wb_select_key) or [])}"
            )

        selected_wb_product_ids = st.multiselect(
            "Товары для WB draft / publish",
            options=wb_product_options,
            default=st.session_state.get(wb_select_key) or [],
            format_func=lambda x: wb_product_label_map.get(int(x), f"ID {x}"),
            key=wb_select_key,
        )
        draft_characteristics_json = st.text_area(
            "Доп. характеристики WB JSON",
            value="[]",
            height=120,
            key="wb_draft_characteristics_json",
            help='Опционально: [{"id": 123, "value": ["значение"]}]',
        )
        wb_draft_cards: list[dict[str, Any]] = []
        if selected_subject_id and selected_wb_product_ids:
            try:
                extra_characteristics = json.loads(draft_characteristics_json or "[]")
                if not isinstance(extra_characteristics, list):
                    extra_characteristics = []
            except Exception:
                extra_characteristics = []
            for pid in selected_wb_product_ids:
                try:
                    wb_draft_cards.append(
                        build_wb_card_draft(
                            conn,
                            product_id=int(pid),
                            subject_id=int(selected_subject_id),
                            extra_characteristics=extra_characteristics,
                        )
                    )
                except Exception:
                    continue
        if wb_draft_cards:
            st.caption(
                "Это первый draft-поток WB: title/description/brand/dimensions/sizes берутся из master-карточки. "
                "Для category-specific атрибутов можно временно подмешать JSON выше."
            )
            st.download_button(
                "Скачать draft JSON для Wildberries",
                data=json.dumps(wb_draft_cards, ensure_ascii=False, indent=2).encode("utf-8"),
                file_name="wildberries_cards_draft.json",
                mime="application/json",
                key="wb_download_draft_json_btn",
            )
            st.code(json.dumps(wb_draft_cards[:2], ensure_ascii=False, indent=2), language="json")
            if st.button("Отправить draft-карточки в Wildberries", key="wb_upload_cards_btn", disabled=not wb_ready):
                upload_result = upload_wb_product_cards(
                    conn=conn,
                    api_token=str(wb_token or "").strip() or None,
                    cards=wb_draft_cards,
                )
                if upload_result.get("ok"):
                    st.success("Draft-карточки отправлены в Wildberries Content API.")
                    st.json(upload_result.get("response") or {})
                else:
                    st.error(str(upload_result.get("message") or "Не удалось отправить карточки в Wildberries."))

        ferr1, ferr2 = st.columns([1, 3])
        with ferr1:
            if st.button("Проверить failed drafts WB", key="wb_failed_cards_btn", disabled=not wb_ready):
                try:
                    st.session_state["wb_failed_cards_result"] = list_wb_failed_cards(
                        conn=conn,
                        api_token=str(wb_token or "").strip() or None,
                        limit=100,
                    )
                    st.success("Список failed product cards WB обновлён.")
                except Exception as e:
                    st.error(str(e))
        with ferr2:
            st.caption(
                "WB создаёт и обновляет карточки асинхронно. Если ответ `200`, но карточка не создалась, "
                "нужно смотреть `List of Failed Product Cards with Errors`."
            )
        wb_failed_cards_result = st.session_state.get("wb_failed_cards_result") or {}
        wb_failed_cards_data = wb_failed_cards_result.get("data") or []
        if wb_failed_cards_data:
            st.dataframe(pd.DataFrame(wb_failed_cards_data), use_container_width=True, hide_index=True)

    channels = conn.execute(
        "SELECT channel_code, channel_name, is_active FROM channel_profiles ORDER BY channel_name"
    ).fetchall()
    channel_df = pd.DataFrame([dict(r) for r in channels]) if channels else pd.DataFrame()
    st.markdown("### Реестр каналов и правила")
    if not channel_df.empty:
        st.dataframe(channel_df, use_container_width=True, hide_index=True)

    channel_code = st.text_input("Channel code", value="detmir")
    category_code = st.text_input("Category code", value="bicycle")

    col1, col2 = st.columns(2)

    with col1:
        st.markdown("### Требования канала")
        reqs = list_channel_requirements(conn, channel_code=channel_code, category_code=category_code or None)
        defs = list_attribute_definitions(conn)
        attr_name_map = {str(d["code"]): str(d.get("name") or humanize_attribute_code(d["code"])) for d in defs} if defs else {}
        if reqs:
            req_df = pd.DataFrame(reqs)
            if "attribute_code" in req_df.columns:
                req_df["attribute_name"] = req_df["attribute_code"].map(lambda x: format_source_name_ui(x, "attribute", attr_name_map))
                req_df = req_df.drop(columns=["attribute_code"], errors="ignore")
            st.dataframe(with_ru_columns(req_df, extra_map={"attribute_name": "Атрибут"}), use_container_width=True, hide_index=True)

        def_codes = [d["code"] for d in defs] if defs else []

        with st.form("channel_req_form"):
            attribute_code = st.selectbox(
                "Обязательный атрибут",
                def_codes,
                format_func=lambda x: format_source_name_ui(x, "attribute", attr_name_map),
            ) if def_codes else st.text_input("Атрибут")
            is_required = st.checkbox("Обязательный", value=True)
            sort_order = st.number_input("Порядок", min_value=1, value=100, step=1)
            notes = st.text_input("Комментарий")
            save_req = st.form_submit_button("Сохранить требование")

            if save_req and attribute_code:
                upsert_channel_attribute_requirement(
                    conn=conn,
                    channel_code=channel_code,
                    category_code=category_code or None,
                    attribute_code=attribute_code,
                    is_required=1 if is_required else 0,
                    sort_order=int(sort_order),
                    notes=notes or None,
                )
                st.success("Требование сохранено")
                st.rerun()

    with col2:
        st.markdown("### Mapping rules")
        rules = list_channel_mapping_rules(conn, channel_code=channel_code, category_code=category_code or None)
        if rules:
            rules_df = pd.DataFrame(rules)
            if not rules_df.empty and "source_name" in rules_df.columns:
                rules_df["source_name"] = rules_df.apply(
                    lambda r: format_source_name_ui(
                        r.get("source_name"),
                        source_type=r.get("source_type"),
                        attr_name_map=attr_name_map,
                    ),
                    axis=1,
                )
            st.dataframe(with_ru_columns(rules_df), use_container_width=True, hide_index=True)

        def_codes = [d["code"] for d in defs] if defs else []

        with st.form("channel_rule_form"):
            target_field = st.text_input("Поле канала")
            source_type = st.selectbox("Источник", ["attribute", "column", "constant"])
            source_name = st.selectbox(
                "Source name",
                def_codes,
                format_func=lambda x: format_source_name_ui(x, "attribute", attr_name_map),
            ) if source_type == "attribute" and def_codes else st.text_input("Source name")
            transform_rule = st.text_input("Transform rule")
            is_required = st.checkbox("Обязательное поле", value=False)
            save_rule = st.form_submit_button("Сохранить mapping")

            if save_rule and target_field and source_name:
                upsert_channel_mapping_rule(
                    conn=conn,
                    channel_code=channel_code,
                    category_code=category_code or None,
                    target_field=target_field.strip(),
                    source_type=source_type,
                    source_name=str(source_name).strip(),
                    transform_rule=transform_rule or None,
                    is_required=1 if is_required else 0,
                )
                st.success("Mapping сохранён")
                st.rerun()

    conn.close()


def main():
    apply_app_theme()
    pending_nav_target = st.session_state.pop("workspace_nav_target", None)
    if pending_nav_target:
        resolved_pending_section = _resolve_workspace_nav_key(pending_nav_target, "catalog")
        st.session_state["workspace_nav_section"] = str(resolved_pending_section)
        st.session_state["workspace_nav_header_label"] = WORKSPACE_NAV_LABEL_BY_KEY.get(str(resolved_pending_section), "▦ Каталог")
    shell_conn = get_db()
    summary = build_workspace_summary(shell_conn)
    shell_conn.close()
    selected_section = render_workspace_top_navigation()
    render_sidebar_navigation(summary, selected_section)
    active_db = summary.get("active_db_path") or str(Path("data/catalog.db"))
    low_db = str(active_db).lower()
    if "\\temp\\pim\\catalog.db" in low_db or "/tmp/pim/catalog.db" in low_db:
        st.warning("Сейчас используется временная БД. Чтобы каталог не пропадал, задай постоянный путь через переменную окружения `PIM_DB_PATH`.")
    if "/mount/src/" in low_db and low_db.endswith("/data/catalog.db"):
        st.warning("Используется БД внутри папки приложения. На Streamlit Cloud она может сбрасываться при redeploy. Рекомендуется `PIM_DB_PATH=/home/adminuser/.pim/catalog.db`.")

    with st.expander("Маршрут менеджера", expanded=False):
        st.markdown(
            """
1. **Настройки**: один раз подключи AI, парсинг и фото-базу.
2. **Импорт**: загрузи каталог поставщика.
3. **Ozon**: держи эталон категорий и атрибутов в памяти.
4. **Каталог**: запусти массовое наполнение по выборке.
5. **Карточка**: доведи спорные товары, Detmir и фото.
6. **Клиентский шаблон**: подтяни shortlist и выгрузи готовую пачку.
            """
        )
    if selected_section == "import":
        show_import_tab()
    elif selected_section == "catalog":
        show_catalog_tab()
    elif selected_section == "product":
        show_product_tab(summary)
    elif selected_section == "attributes":
        show_attributes_tab()
    elif selected_section == "template":
        show_template_tab()
    elif selected_section == "ozon":
        show_ozon_tab()
    elif selected_section == "settings":
        show_settings_tab()
    else:
        show_channels_tab()


if __name__ == "__main__":
    main()
